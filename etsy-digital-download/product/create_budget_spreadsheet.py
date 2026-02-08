"""
Create a Complete Personal Finance Budget Spreadsheet Bundle
for Etsy Digital Download
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from copy import copy

# ── Color Palette (Modern, clean, gender-neutral) ──
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "52B788"
PALE_GREEN = "D8F3DC"
ACCENT_MINT = "B7E4C7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "E0E0E0"
DARK_TEXT = "212121"
RED_ACCENT = "E63946"
GOLD_ACCENT = "F4A261"

# ── Reusable style helpers ──
def header_font(size=12, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=11, bold=False, color=DARK_TEXT):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    side = Side(style="thin", color=MED_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def currency_format():
    return '"$"#,##0.00'

def pct_format():
    return '0.0%'

def style_header_row(ws, row, cols, bg_color=DARK_GREEN):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = fill(bg_color)
        cell.alignment = center_align()
        cell.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    bg = LIGHT_GRAY if alt else WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.fill = fill(bg)
        cell.alignment = center_align()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title_block(ws, title, subtitle, merge_end_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=DARK_GREEN)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = fill(PALE_GREEN)
    ws.row_dimensions[1].height = 45

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end_col)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color=MED_GREEN)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.fill = fill(PALE_GREEN)
    ws.row_dimensions[2].height = 25


# ═══════════════════════════════════════════════════
#  CREATE WORKBOOK
# ═══════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────
#  SHEET 1: MONTHLY BUDGET DASHBOARD
# ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Monthly Budget"
ws1.sheet_properties.tabColor = DARK_GREEN

add_title_block(ws1, "MONTHLY BUDGET PLANNER", "Track your income & expenses each month", 8)

# Income section
row = 4
ws1.cell(row=row, column=1, value="INCOME").font = header_font(14, True, DARK_GREEN)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws1.cell(row=row, column=1).fill = fill(ACCENT_MINT)
ws1.cell(row=row, column=1).alignment = center_align()

row = 5
income_headers = ["Source", "Expected", "Actual", "Difference"]
for i, h in enumerate(income_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 4, MED_GREEN)

income_sources = ["Primary Salary", "Side Income", "Freelance/Gig Work", "Investment Income", "Other Income"]
for idx, src in enumerate(income_sources):
    r = 6 + idx
    ws1.cell(row=r, column=1, value=src)
    ws1.cell(row=r, column=2).number_format = currency_format()
    ws1.cell(row=r, column=3).number_format = currency_format()
    ws1.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = currency_format()
    style_data_row(ws1, r, 4, idx % 2 == 1)

total_income_row = 6 + len(income_sources)
ws1.cell(row=total_income_row, column=1, value="TOTAL INCOME")
ws1.cell(row=total_income_row, column=2, value=f"=SUM(B6:B{total_income_row-1})").number_format = currency_format()
ws1.cell(row=total_income_row, column=3, value=f"=SUM(C6:C{total_income_row-1})").number_format = currency_format()
ws1.cell(row=total_income_row, column=4, value=f"=C{total_income_row}-B{total_income_row}").number_format = currency_format()
style_header_row(ws1, total_income_row, 4, DARK_GREEN)

# Expense section
exp_start = total_income_row + 2
ws1.cell(row=exp_start, column=1, value="EXPENSES").font = header_font(14, True, DARK_GREEN)
ws1.merge_cells(start_row=exp_start, start_column=1, end_row=exp_start, end_column=4)
ws1.cell(row=exp_start, column=1).fill = fill(ACCENT_MINT)
ws1.cell(row=exp_start, column=1).alignment = center_align()

exp_header_row = exp_start + 1
exp_headers = ["Category", "Budgeted", "Actual", "Difference"]
for i, h in enumerate(exp_headers, 1):
    ws1.cell(row=exp_header_row, column=i, value=h)
style_header_row(ws1, exp_header_row, 4, MED_GREEN)

expense_categories = [
    "Housing (Rent/Mortgage)", "Utilities", "Groceries", "Transportation",
    "Insurance", "Healthcare", "Phone/Internet", "Subscriptions",
    "Dining Out", "Entertainment", "Clothing", "Personal Care",
    "Education", "Childcare", "Pet Expenses", "Gifts/Donations",
    "Savings/Investments", "Debt Payments", "Miscellaneous"
]

for idx, cat in enumerate(expense_categories):
    r = exp_header_row + 1 + idx
    ws1.cell(row=r, column=1, value=cat)
    ws1.cell(row=r, column=2).number_format = currency_format()
    ws1.cell(row=r, column=3).number_format = currency_format()
    ws1.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = currency_format()
    style_data_row(ws1, r, 4, idx % 2 == 1)

exp_total_row = exp_header_row + 1 + len(expense_categories)
ws1.cell(row=exp_total_row, column=1, value="TOTAL EXPENSES")
first_exp = exp_header_row + 1
ws1.cell(row=exp_total_row, column=2, value=f"=SUM(B{first_exp}:B{exp_total_row-1})").number_format = currency_format()
ws1.cell(row=exp_total_row, column=3, value=f"=SUM(C{first_exp}:C{exp_total_row-1})").number_format = currency_format()
ws1.cell(row=exp_total_row, column=4, value=f"=B{exp_total_row}-C{exp_total_row}").number_format = currency_format()
style_header_row(ws1, exp_total_row, 4, DARK_GREEN)

# Summary block (right side)
summary_col = 6
ws1.cell(row=4, column=summary_col, value="MONTHLY SUMMARY").font = header_font(14, True, DARK_GREEN)
ws1.merge_cells(start_row=4, start_column=summary_col, end_row=4, end_column=summary_col+2)
ws1.cell(row=4, column=summary_col).fill = fill(ACCENT_MINT)
ws1.cell(row=4, column=summary_col).alignment = center_align()

summary_items = [
    ("Total Income", f"=C{total_income_row}"),
    ("Total Expenses", f"=C{exp_total_row}"),
    ("NET BALANCE", f"=C{total_income_row}-C{exp_total_row}"),
    ("Savings Rate", f"=IF(C{total_income_row}=0,0,(C{total_income_row}-C{exp_total_row})/C{total_income_row})"),
]

for idx, (label, formula) in enumerate(summary_items):
    r = 5 + idx
    ws1.cell(row=r, column=summary_col, value=label).font = body_font(11, True)
    ws1.cell(row=r, column=summary_col).fill = fill(PALE_GREEN)
    ws1.cell(row=r, column=summary_col).border = thin_border()
    val_cell = ws1.cell(row=r, column=summary_col+1, value=formula)
    if idx == 3:
        val_cell.number_format = pct_format()
    else:
        val_cell.number_format = currency_format()
    val_cell.font = body_font(12, True, DARK_GREEN if idx < 3 else MED_GREEN)
    val_cell.fill = fill(WHITE)
    val_cell.border = thin_border()
    val_cell.alignment = center_align()

# Add pie chart for expenses
pie = PieChart()
pie.title = "Expense Breakdown"
pie.style = 10
cats = Reference(ws1, min_col=1, min_row=first_exp, max_row=exp_total_row-1)
data = Reference(ws1, min_col=3, min_row=first_exp, max_row=exp_total_row-1)
pie.add_data(data)
pie.set_categories(cats)
pie.width = 18
pie.height = 14
ws1.add_chart(pie, f"{get_column_letter(summary_col)}10")

set_col_widths(ws1, [25, 14, 14, 14, 3, 20, 16, 16])

# ──────────────────────────────────────────────────
#  SHEET 2: EXPENSE TRACKER (Daily Log)
# ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Expense Tracker")
ws2.sheet_properties.tabColor = MED_GREEN

add_title_block(ws2, "DAILY EXPENSE TRACKER", "Log every purchase to stay on top of your spending", 7)

row = 4
headers2 = ["Date", "Description", "Category", "Payment Method", "Amount", "Notes", "Running Total"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, 7)

# Pre-format 100 rows for data entry
for r in range(5, 105):
    ws2.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    ws2.cell(row=r, column=5).number_format = currency_format()
    ws2.cell(row=r, column=7, value=f"=IF(E{r}=\"\",\"\",SUM(E$5:E{r}))").number_format = currency_format()
    style_data_row(ws2, r, 7, (r - 5) % 2 == 1)

# Add data validation for Category column
from openpyxl.worksheet.datavalidation import DataValidation
cat_list = ",".join(expense_categories)
dv = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
dv.error = "Please select a valid category"
dv.errorTitle = "Invalid Category"
ws2.add_data_validation(dv)
dv.add(f"C5:C104")

# Payment method validation
pm_dv = DataValidation(type="list", formula1='"Cash,Debit Card,Credit Card,Bank Transfer,Venmo,PayPal,Zelle,Other"', allow_blank=True)
ws2.add_data_validation(pm_dv)
pm_dv.add(f"D5:D104")

set_col_widths(ws2, [14, 30, 25, 18, 14, 25, 16])

# ──────────────────────────────────────────────────
#  SHEET 3: SAVINGS GOALS TRACKER
# ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Savings Goals")
ws3.sheet_properties.tabColor = LIGHT_GREEN

add_title_block(ws3, "SAVINGS GOALS TRACKER", "Set goals, track progress, and celebrate milestones", 8)

row = 4
headers3 = ["Goal Name", "Target Amount", "Current Saved", "Remaining", "% Complete", "Target Date", "Monthly Needed", "Status"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 8)

sample_goals = [
    "Emergency Fund (3-6 months)",
    "Vacation Fund",
    "New Car Down Payment",
    "Home Down Payment",
    "Holiday Gift Fund",
    "Education/Course Fund",
    "Tech Upgrade Fund",
    "Wedding Fund",
    "Retirement Boost",
    "Custom Goal"
]

for idx, goal in enumerate(sample_goals):
    r = 5 + idx
    ws3.cell(row=r, column=1, value=goal)
    ws3.cell(row=r, column=2).number_format = currency_format()
    ws3.cell(row=r, column=3).number_format = currency_format()
    ws3.cell(row=r, column=4, value=f"=IF(B{r}=\"\",\"\",B{r}-C{r})").number_format = currency_format()
    ws3.cell(row=r, column=5, value=f"=IF(B{r}=0,0,C{r}/B{r})").number_format = pct_format()
    ws3.cell(row=r, column=6).number_format = "MM/DD/YYYY"
    ws3.cell(row=r, column=7, value=f'=IF(OR(B{r}="",F{r}=""),"",IF((F{r}-TODAY())/30<=0,"Past Due",(B{r}-C{r})/MAX(1,(F{r}-TODAY())/30)))').number_format = currency_format()
    ws3.cell(row=r, column=8, value=f'=IF(B{r}="","",IF(C{r}>=B{r},"COMPLETE!",IF(E{r}>=0.75,"Almost There!",IF(E{r}>=0.5,"Halfway!",IF(E{r}>=0.25,"Getting There","Just Started")))))')
    style_data_row(ws3, r, 8, idx % 2 == 1)

# Progress bar chart
bar = BarChart()
bar.type = "bar"
bar.title = "Savings Goal Progress"
bar.style = 10
bar.y_axis.title = ""
bar.x_axis.title = "% Complete"
cats = Reference(ws3, min_col=1, min_row=5, max_row=14)
data = Reference(ws3, min_col=5, min_row=4, max_row=14)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.shape = 4
bar.width = 22
bar.height = 14
ws3.add_chart(bar, "A17")

set_col_widths(ws3, [28, 16, 16, 16, 14, 14, 16, 16])

# ──────────────────────────────────────────────────
#  SHEET 4: DEBT PAYOFF PLANNER
# ──────────────────────────────────────────────────
ws4 = wb.create_sheet("Debt Payoff")
ws4.sheet_properties.tabColor = GOLD_ACCENT

add_title_block(ws4, "DEBT PAYOFF PLANNER", "Crush your debt with a clear payoff strategy", 9)

row = 4
headers4 = ["Debt Name", "Total Owed", "Interest Rate", "Min Payment", "Extra Payment", "Total Payment", "Months to Payoff", "Total Interest", "Payoff Date"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 9)

debt_examples = [
    "Credit Card 1",
    "Credit Card 2",
    "Student Loan",
    "Car Loan",
    "Personal Loan",
    "Medical Debt",
    "Other Debt 1",
    "Other Debt 2",
]

for idx, debt in enumerate(debt_examples):
    r = 5 + idx
    ws4.cell(row=r, column=1, value=debt)
    ws4.cell(row=r, column=2).number_format = currency_format()
    ws4.cell(row=r, column=3).number_format = '0.00%'
    ws4.cell(row=r, column=4).number_format = currency_format()
    ws4.cell(row=r, column=5).number_format = currency_format()
    ws4.cell(row=r, column=6, value=f"=IF(D{r}=\"\",\"\",D{r}+IF(E{r}=\"\",0,E{r}))").number_format = currency_format()
    ws4.cell(row=r, column=7, value=f'=IF(OR(B{r}="",F{r}="",F{r}=0),"",ROUNDUP(B{r}/F{r},0))')
    ws4.cell(row=r, column=8, value=f'=IF(OR(B{r}="",G{r}=""),"",G{r}*F{r}-B{r})').number_format = currency_format()
    ws4.cell(row=r, column=9, value=f'=IF(G{r}="","",DATE(YEAR(TODAY()),MONTH(TODAY())+G{r},DAY(TODAY())))').number_format = "MMM YYYY"
    style_data_row(ws4, r, 9, idx % 2 == 1)

debt_total_row = 5 + len(debt_examples)
ws4.cell(row=debt_total_row, column=1, value="TOTALS")
ws4.cell(row=debt_total_row, column=2, value=f"=SUM(B5:B{debt_total_row-1})").number_format = currency_format()
ws4.cell(row=debt_total_row, column=4, value=f"=SUM(D5:D{debt_total_row-1})").number_format = currency_format()
ws4.cell(row=debt_total_row, column=5, value=f"=SUM(E5:E{debt_total_row-1})").number_format = currency_format()
ws4.cell(row=debt_total_row, column=6, value=f"=SUM(F5:F{debt_total_row-1})").number_format = currency_format()
ws4.cell(row=debt_total_row, column=8, value=f"=SUM(H5:H{debt_total_row-1})").number_format = currency_format()
style_header_row(ws4, debt_total_row, 9, DARK_GREEN)

# Debt strategy tips
tips_row = debt_total_row + 2
ws4.merge_cells(start_row=tips_row, start_column=1, end_row=tips_row, end_column=9)
ws4.cell(row=tips_row, column=1, value="PAYOFF STRATEGIES").font = header_font(14, True, DARK_GREEN)
ws4.cell(row=tips_row, column=1).fill = fill(ACCENT_MINT)
ws4.cell(row=tips_row, column=1).alignment = center_align()

strategies = [
    "AVALANCHE METHOD: Pay minimums on all debts, put extra $ toward the highest interest rate debt first. Saves the most money.",
    "SNOWBALL METHOD: Pay minimums on all debts, put extra $ toward the smallest balance first. Builds momentum with quick wins.",
    "TIP: Even $25/month extra toward debt can save you hundreds in interest and months of payments!",
]
for idx, tip in enumerate(strategies):
    r = tips_row + 1 + idx
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws4.cell(row=r, column=1, value=f"  {tip}")
    c.font = body_font(10, False, MED_GREEN)
    c.fill = fill(PALE_GREEN)
    c.alignment = left_align()

set_col_widths(ws4, [20, 16, 14, 14, 14, 14, 16, 16, 14])

# ──────────────────────────────────────────────────
#  SHEET 5: NET WORTH TRACKER
# ──────────────────────────────────────────────────
ws5 = wb.create_sheet("Net Worth")
ws5.sheet_properties.tabColor = "40916C"

add_title_block(ws5, "NET WORTH TRACKER", "See your full financial picture at a glance", 7)

# Assets
row = 4
ws5.cell(row=row, column=1, value="ASSETS (What You Own)").font = header_font(14, True, DARK_GREEN)
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws5.cell(row=row, column=1).fill = fill(ACCENT_MINT)
ws5.cell(row=row, column=1).alignment = center_align()

row = 5
asset_headers = ["Asset", "Current Value", "Notes"]
for i, h in enumerate(asset_headers, 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, 3, MED_GREEN)

assets = [
    "Checking Account(s)", "Savings Account(s)", "Cash on Hand",
    "401(k)/Retirement", "IRA/Roth IRA", "Brokerage/Investments",
    "HSA/FSA", "Real Estate (Market Value)", "Vehicle(s)",
    "Other Valuable Assets", "Crypto/Alternative Investments"
]

for idx, asset in enumerate(assets):
    r = 6 + idx
    ws5.cell(row=r, column=1, value=asset)
    ws5.cell(row=r, column=2).number_format = currency_format()
    style_data_row(ws5, r, 3, idx % 2 == 1)

asset_total_row = 6 + len(assets)
ws5.cell(row=asset_total_row, column=1, value="TOTAL ASSETS")
ws5.cell(row=asset_total_row, column=2, value=f"=SUM(B6:B{asset_total_row-1})").number_format = currency_format()
style_header_row(ws5, asset_total_row, 3, DARK_GREEN)

# Liabilities
liab_start = asset_total_row + 2
ws5.cell(row=liab_start, column=1, value="LIABILITIES (What You Owe)").font = header_font(14, True, RED_ACCENT)
ws5.merge_cells(start_row=liab_start, start_column=1, end_row=liab_start, end_column=3)
ws5.cell(row=liab_start, column=1).fill = fill("FFE5E5")
ws5.cell(row=liab_start, column=1).alignment = center_align()

liab_header_row = liab_start + 1
liab_headers = ["Liability", "Amount Owed", "Notes"]
for i, h in enumerate(liab_headers, 1):
    ws5.cell(row=liab_header_row, column=i, value=h)
style_header_row(ws5, liab_header_row, 3, RED_ACCENT)

liabilities = [
    "Mortgage", "Car Loan(s)", "Student Loans", "Credit Card Debt",
    "Personal Loans", "Medical Debt", "Other Debt"
]

for idx, liab in enumerate(liabilities):
    r = liab_header_row + 1 + idx
    ws5.cell(row=r, column=1, value=liab)
    ws5.cell(row=r, column=2).number_format = currency_format()
    style_data_row(ws5, r, 3, idx % 2 == 1)

liab_total_row = liab_header_row + 1 + len(liabilities)
ws5.cell(row=liab_total_row, column=1, value="TOTAL LIABILITIES")
liab_first = liab_header_row + 1
ws5.cell(row=liab_total_row, column=2, value=f"=SUM(B{liab_first}:B{liab_total_row-1})").number_format = currency_format()
style_header_row(ws5, liab_total_row, 3, RED_ACCENT)

# Net Worth Summary (right side)
nw_col = 5
ws5.cell(row=4, column=nw_col, value="NET WORTH SUMMARY").font = header_font(14, True, DARK_GREEN)
ws5.merge_cells(start_row=4, start_column=nw_col, end_row=4, end_column=nw_col+2)
ws5.cell(row=4, column=nw_col).fill = fill(ACCENT_MINT)
ws5.cell(row=4, column=nw_col).alignment = center_align()

nw_items = [
    ("Total Assets", f"=B{asset_total_row}"),
    ("Total Liabilities", f"=B{liab_total_row}"),
    ("NET WORTH", f"=B{asset_total_row}-B{liab_total_row}"),
]

for idx, (label, formula) in enumerate(nw_items):
    r = 5 + idx
    ws5.cell(row=r, column=nw_col, value=label).font = body_font(12, True)
    ws5.cell(row=r, column=nw_col).fill = fill(PALE_GREEN)
    ws5.cell(row=r, column=nw_col).border = thin_border()
    val_cell = ws5.cell(row=r, column=nw_col+1, value=formula)
    val_cell.number_format = currency_format()
    val_cell.font = Font(name="Calibri", size=14, bold=True, color=DARK_GREEN)
    val_cell.fill = fill(WHITE)
    val_cell.border = thin_border()
    val_cell.alignment = center_align()

# Monthly net worth tracking
track_row = 10
ws5.cell(row=track_row, column=nw_col, value="MONTHLY NET WORTH HISTORY").font = header_font(12, True, DARK_GREEN)
ws5.merge_cells(start_row=track_row, start_column=nw_col, end_row=track_row, end_column=nw_col+2)
ws5.cell(row=track_row, column=nw_col).fill = fill(ACCENT_MINT)
ws5.cell(row=track_row, column=nw_col).alignment = center_align()

months = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

ws5.cell(row=11, column=nw_col, value="Month").font = header_font(11)
ws5.cell(row=11, column=nw_col).fill = fill(MED_GREEN)
ws5.cell(row=11, column=nw_col).border = thin_border()
ws5.cell(row=11, column=nw_col+1, value="Net Worth").font = header_font(11)
ws5.cell(row=11, column=nw_col).alignment = center_align()
ws5.cell(row=11, column=nw_col+1).fill = fill(MED_GREEN)
ws5.cell(row=11, column=nw_col+1).border = thin_border()
ws5.cell(row=11, column=nw_col+1).alignment = center_align()
ws5.cell(row=11, column=nw_col+2, value="Change").font = header_font(11)
ws5.cell(row=11, column=nw_col+2).fill = fill(MED_GREEN)
ws5.cell(row=11, column=nw_col+2).border = thin_border()
ws5.cell(row=11, column=nw_col+2).alignment = center_align()

for idx, month in enumerate(months):
    r = 12 + idx
    ws5.cell(row=r, column=nw_col, value=month).font = body_font()
    ws5.cell(row=r, column=nw_col).border = thin_border()
    ws5.cell(row=r, column=nw_col).alignment = center_align()
    ws5.cell(row=r, column=nw_col+1).number_format = currency_format()
    ws5.cell(row=r, column=nw_col+1).border = thin_border()
    ws5.cell(row=r, column=nw_col+1).alignment = center_align()
    if idx == 0:
        ws5.cell(row=r, column=nw_col+2, value="—")
    else:
        ws5.cell(row=r, column=nw_col+2, value=f"=IF(OR(F{r}=\"\",F{r-1}=\"\"),\"\",F{r}-F{r-1})")
    ws5.cell(row=r, column=nw_col+2).number_format = currency_format()
    ws5.cell(row=r, column=nw_col+2).border = thin_border()
    ws5.cell(row=r, column=nw_col+2).alignment = center_align()
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    for c in range(nw_col, nw_col + 3):
        ws5.cell(row=r, column=c).fill = fill(bg)

# Line chart for net worth over time
line = LineChart()
line.title = "Net Worth Over Time"
line.style = 10
line.y_axis.title = "Net Worth ($)"
line.x_axis.title = "Month"
cats = Reference(ws5, min_col=nw_col, min_row=12, max_row=23)
data = Reference(ws5, min_col=nw_col+1, min_row=11, max_row=23)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.width = 22
line.height = 12
ws5.add_chart(line, f"{get_column_letter(nw_col)}25")

set_col_widths(ws5, [26, 18, 25, 3, 16, 18, 16])

# ──────────────────────────────────────────────────
#  SHEET 6: ANNUAL OVERVIEW (12-month summary)
# ──────────────────────────────────────────────────
ws6 = wb.create_sheet("Annual Overview")
ws6.sheet_properties.tabColor = "74C69D"

add_title_block(ws6, "ANNUAL FINANCIAL OVERVIEW", "Track your income, expenses, and savings month by month across the entire year", 14)

row = 4
headers6 = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "TOTAL"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 14)

metrics = ["Total Income", "Total Expenses", "Net Savings", "Savings Rate (%)"]
for idx, metric in enumerate(metrics):
    r = 5 + idx
    ws6.cell(row=r, column=1, value=metric).font = body_font(11, True, DARK_GREEN)
    ws6.cell(row=r, column=1).fill = fill(PALE_GREEN)
    ws1.cell(row=r, column=1).border = thin_border()
    for col in range(2, 14):
        c = ws6.cell(row=r, column=col)
        if idx < 3:
            c.number_format = currency_format()
        else:
            c.number_format = pct_format()
        c.border = thin_border()
        c.alignment = center_align()
        c.fill = fill(LIGHT_GRAY if (col - 2) % 2 == 1 else WHITE)

    # Total column
    total_cell = ws6.cell(row=r, column=14)
    if idx < 3:
        total_cell.value = f"=SUM(B{r}:M{r})"
        total_cell.number_format = currency_format()
    else:
        total_cell.value = f"=IF(B5=0,0,B7/B5)"
        total_cell.number_format = pct_format()
    total_cell.font = body_font(11, True, DARK_GREEN)
    total_cell.fill = fill(ACCENT_MINT)
    total_cell.border = thin_border()
    total_cell.alignment = center_align()

# Net savings formula
for col in range(2, 14):
    col_letter = get_column_letter(col)
    ws6.cell(row=7, column=col, value=f"={col_letter}5-{col_letter}6")
    ws6.cell(row=8, column=col, value=f"=IF({col_letter}5=0,0,{col_letter}7/{col_letter}5)")

# Annual bar chart
bar2 = BarChart()
bar2.type = "col"
bar2.title = "Monthly Income vs Expenses"
bar2.style = 10
bar2.y_axis.title = "Amount ($)"
cats = Reference(ws6, min_col=2, min_row=4, max_col=13)
income_data = Reference(ws6, min_col=2, min_row=5, max_col=13)
expense_data = Reference(ws6, min_col=2, min_row=6, max_col=13)
# Just use both rows with from_rows
data = Reference(ws6, min_col=1, min_row=5, max_col=13, max_row=6)
bar2.add_data(data, from_rows=True, titles_from_data=True)
bar2.set_categories(cats)
bar2.width = 28
bar2.height = 14
ws6.add_chart(bar2, "A11")

set_col_widths(ws6, [18] + [11]*12 + [14])

# ──────────────────────────────────────────────────
#  SHEET 7: INSTRUCTIONS
# ──────────────────────────────────────────────────
ws7 = wb.create_sheet("How to Use")
ws7.sheet_properties.tabColor = PALE_GREEN

add_title_block(ws7, "HOW TO USE YOUR BUDGET BUNDLE", "Quick-start guide to get the most out of every tab", 6)

instructions = [
    ("GETTING STARTED", [
        "1. Save a copy of this file to your computer or Google Drive before editing.",
        "2. Start with the Monthly Budget tab to set up your income and expense categories.",
        "3. Customize the categories to match YOUR actual spending (rename any row).",
        "4. Fill in your expected/budgeted amounts first, then update actuals throughout the month.",
    ]),
    ("MONTHLY BUDGET TAB", [
        "- Enter your income sources and expected amounts in the 'Expected' column.",
        "- As you receive income, enter actual amounts in the 'Actual' column.",
        "- Enter budgeted amounts for each expense category.",
        "- The 'Difference' column automatically shows if you're over or under budget.",
        "- The pie chart updates automatically to show your expense breakdown.",
    ]),
    ("EXPENSE TRACKER TAB", [
        "- Log every purchase as it happens (or at the end of each day).",
        "- Use the dropdown menus for Category and Payment Method.",
        "- The Running Total column tracks your cumulative spending.",
        "- TIP: Review this weekly to catch spending patterns early!",
    ]),
    ("SAVINGS GOALS TAB", [
        "- Enter each savings goal with a target amount and date.",
        "- Update your 'Current Saved' as you make deposits.",
        "- The 'Monthly Needed' column tells you exactly how much to save per month.",
        "- Watch your status change from 'Just Started' to 'COMPLETE!'",
    ]),
    ("DEBT PAYOFF TAB", [
        "- List all your debts with balances, interest rates, and minimum payments.",
        "- Add any extra payments you can afford in the 'Extra Payment' column.",
        "- Choose your strategy: Avalanche (highest interest first) or Snowball (smallest balance first).",
        "- The payoff date automatically calculates when you'll be debt-free!",
    ]),
    ("NET WORTH TAB", [
        "- List all your assets (what you own) and their current values.",
        "- List all your liabilities (what you owe).",
        "- Your net worth is calculated automatically: Assets - Liabilities.",
        "- Update the Monthly Net Worth History to track your progress over the year.",
    ]),
    ("ANNUAL OVERVIEW TAB", [
        "- Enter monthly totals for income and expenses.",
        "- Net savings and savings rate calculate automatically.",
        "- Use this for a bird's-eye view of your financial year.",
        "- Great for tax planning and year-over-year comparisons!",
    ]),
    ("PRO TIPS", [
        "- Update your budget WEEKLY for best results (takes 5-10 minutes).",
        "- Set a calendar reminder to review your finances every Sunday evening.",
        "- Make a copy each month to keep a history of your progress.",
        "- Share with a partner or accountability buddy for better results.",
        "- Remember: A budget isn't about restriction - it's about intention!",
    ]),
]

current_row = 4
for section_title, items in instructions:
    ws7.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws7.cell(row=current_row, column=1, value=section_title).font = header_font(13, True, DARK_GREEN)
    ws7.cell(row=current_row, column=1).fill = fill(ACCENT_MINT)
    ws7.cell(row=current_row, column=1).alignment = left_align()
    ws7.row_dimensions[current_row].height = 28
    current_row += 1
    for item in items:
        ws7.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws7.cell(row=current_row, column=1, value=f"  {item}").font = body_font(10)
        ws7.cell(row=current_row, column=1).alignment = left_align()
        current_row += 1
    current_row += 1  # Blank row between sections

set_col_widths(ws7, [80, 15, 15, 15, 15, 15])

# ── Move Instructions sheet to the front ──
wb.move_sheet("How to Use", offset=-6)

# ── Save ──
output_path = "/home/user/claude-code/etsy-digital-download/product/Complete_Budget_Finance_Bundle_2026.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
