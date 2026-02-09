#!/usr/bin/env python3
"""Create Social Media Marketing Planner 2026 Excel spreadsheet with 8 professional tabs."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime, timedelta
import random

# --- COLOR PALETTE ---
HOT_PINK = "D53F8C"
PURPLE = "6B46C1"
ELECTRIC_BLUE = "3182CE"
CORAL = "FC8181"
MINT = "68D391"
PALE_PINK = "FED7E2"
DARK_TEXT = "1A202C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F7FAFC"
MED_GRAY = "E2E8F0"
VERY_LIGHT_PURPLE = "FAF5FF"
VERY_LIGHT_PINK = "FFF5F7"
VERY_LIGHT_BLUE = "EBF8FF"

OUTPUT = "/home/user/claude-code/etsy-digital-download-6/product/Social_Media_Marketing_Planner_2026.xlsx"

wb = openpyxl.Workbook()

# --- SHARED STYLES ---
header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
title_font = Font(name="Calibri", bold=True, size=16, color=DARK_TEXT)
subtitle_font = Font(name="Calibri", bold=True, size=11, color=PURPLE)
body_font = Font(name="Calibri", size=11, color=DARK_TEXT)
link_font = Font(name="Calibri", size=11, color=ELECTRIC_BLUE, underline="single")
number_font = Font(name="Calibri", size=11, color=DARK_TEXT)
currency_font = Font(name="Calibri", bold=True, size=11, color="2D7D46")
total_font = Font(name="Calibri", bold=True, size=12, color=WHITE)

header_fill_pink = PatternFill(start_color=HOT_PINK, end_color=HOT_PINK, fill_type="solid")
header_fill_purple = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
header_fill_blue = PatternFill(start_color=ELECTRIC_BLUE, end_color=ELECTRIC_BLUE, fill_type="solid")
header_fill_coral = PatternFill(start_color="E53E3E", end_color="E53E3E", fill_type="solid")
header_fill_mint = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid")
header_fill_dark = PatternFill(start_color=DARK_TEXT, end_color=DARK_TEXT, fill_type="solid")

alt_row_pink = PatternFill(start_color=VERY_LIGHT_PINK, end_color=VERY_LIGHT_PINK, fill_type="solid")
alt_row_purple = PatternFill(start_color=VERY_LIGHT_PURPLE, end_color=VERY_LIGHT_PURPLE, fill_type="solid")
alt_row_blue = PatternFill(start_color=VERY_LIGHT_BLUE, end_color=VERY_LIGHT_BLUE, fill_type="solid")
pale_pink_fill = PatternFill(start_color=PALE_PINK, end_color=PALE_PINK, fill_type="solid")
total_fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
mint_fill = PatternFill(start_color=MINT, end_color=MINT, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
header_border = Border(
    left=Side(style="thin", color=WHITE),
    right=Side(style="thin", color=WHITE),
    top=Side(style="thin", color=WHITE),
    bottom=Side(style="medium", color=WHITE),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")


def style_header_row(ws, row, num_cols, fill, freeze_col=None):
    """Style a header row with given fill."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = header_border
    ws.row_dimensions[row].height = 36
    if freeze_col:
        ws.freeze_panes = f"A{row + 1}"
    else:
        ws.freeze_panes = f"A{row + 1}"


def style_data_rows(ws, start_row, end_row, num_cols, alt_fill, wrap_cols=None):
    """Style data rows with alternating colors."""
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = thin_border
            if wrap_cols and c in wrap_cols:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill
        ws.row_dimensions[r].height = 28


def add_title_block(ws, title, subtitle, num_cols):
    """Add a branded title block at the top of a sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"✦ {title}"
    title_cell.font = Font(name="Calibri", bold=True, size=18, color=HOT_PINK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color=VERY_LIGHT_PINK, end_color=VERY_LIGHT_PINK, fill_type="solid")
    ws.row_dimensions[1].height = 48

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = subtitle
    sub_cell.font = Font(name="Calibri", size=11, color=PURPLE, italic=True)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.fill = PatternFill(start_color=VERY_LIGHT_PINK, end_color=VERY_LIGHT_PINK, fill_type="solid")
    ws.row_dimensions[2].height = 28


# ============================================================
# TAB 1: CONTENT CALENDAR
# ============================================================
ws1 = wb.active
ws1.title = "Content Calendar"
ws1.sheet_properties.tabColor = HOT_PINK

headers1 = ["Date", "Day", "Platform", "Content Type", "Topic / Caption",
            "Hashtags", "Status", "Engagement", "Link", "Notes"]
widths1 = [14, 12, 16, 16, 40, 30, 14, 14, 25, 30]
num_cols1 = len(headers1)

add_title_block(ws1, "SOCIAL MEDIA CONTENT CALENDAR 2026", "Plan, schedule, and track every post across all platforms", num_cols1)

# Header row at row 3
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, num_cols1, header_fill_pink)
ws1.freeze_panes = "A4"

# Column widths
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Dropdowns
platform_dv = DataValidation(type="list", formula1='"Instagram,TikTok,YouTube,Pinterest,Facebook,Twitter,LinkedIn,Blog,Email"', allow_blank=True)
platform_dv.error = "Please select a valid platform"
platform_dv.errorTitle = "Invalid Platform"
ws1.add_data_validation(platform_dv)

content_type_dv = DataValidation(type="list", formula1='"Photo,Video,Carousel,Reel,Story,Blog,Newsletter"', allow_blank=True)
ws1.add_data_validation(content_type_dv)

status_dv = DataValidation(type="list", formula1='"Idea,Draft,Scheduled,Published,Boosted"', allow_blank=True)
ws1.add_data_validation(status_dv)

# Sample data for 100 rows
platforms = ["Instagram", "TikTok", "YouTube", "Pinterest", "Facebook", "Twitter", "LinkedIn", "Blog", "Email"]
content_types = ["Photo", "Video", "Carousel", "Reel", "Story", "Blog", "Newsletter"]
statuses = ["Idea", "Draft", "Scheduled", "Published", "Boosted"]
sample_topics = [
    "Monday Motivation - Share your 2026 goals",
    "Behind the scenes: Morning routine",
    "Product launch teaser video",
    "5 Tips for growing your audience",
    "Client testimonial spotlight",
    "New blog post: SEO strategy guide",
    "Weekly newsletter: Industry roundup",
    "Trending audio reel - Brand twist",
    "Infographic: Social media stats 2026",
    "Q&A session: Ask me anything",
    "Tutorial: How to create Reels",
    "User-generated content reshare",
    "Pinterest pin: Free checklist",
    "Brand partnership announcement",
    "Story poll: Which design do you prefer?",
    "Throwback Thursday post",
    "Carousel: Step-by-step branding guide",
    "YouTube shorts: Quick tip",
    "Engagement post: This or That?",
    "Live session: Content planning workshop",
    "Holiday content: Valentine's Day promo",
    "Spring collection preview",
    "Email blast: Exclusive offer",
    "LinkedIn article: Industry trends",
    "Community spotlight post",
]
sample_hashtags = [
    "#socialmedia #marketing #2026",
    "#contentcreator #digitalmarketing",
    "#brandstrategy #growyourbrand",
    "#reelsinstagram #viralcontent",
    "#tiktokmarketing #trending",
    "#pinterestmarketing #pinit",
    "#smallbusiness #entrepreneur",
    "#contentcalendar #planning",
    "#influencer #collab #sponsored",
    "#emailmarketing #newsletter",
]

start_date = datetime(2026, 1, 5)
for row_idx in range(4, 104):
    r = row_idx - 4
    date_val = start_date + timedelta(days=r)
    ws1.cell(row=row_idx, column=1, value=date_val).number_format = "MM/DD/YYYY"

    # Day formula
    ws1.cell(row=row_idx, column=2, value=f'=TEXT(A{row_idx},"dddd")')

    plat = platforms[r % len(platforms)]
    ws1.cell(row=row_idx, column=3, value=plat)
    platform_dv.add(ws1.cell(row=row_idx, column=3))

    ct = content_types[r % len(content_types)]
    ws1.cell(row=row_idx, column=4, value=ct)
    content_type_dv.add(ws1.cell(row=row_idx, column=4))

    ws1.cell(row=row_idx, column=5, value=sample_topics[r % len(sample_topics)])
    ws1.cell(row=row_idx, column=6, value=sample_hashtags[r % len(sample_hashtags)])

    st = statuses[min(r // 20, 4)]
    ws1.cell(row=row_idx, column=7, value=st)
    status_dv.add(ws1.cell(row=row_idx, column=7))

    if st in ("Published", "Boosted"):
        ws1.cell(row=row_idx, column=8, value=random.randint(50, 5000))
    ws1.cell(row=row_idx, column=9, value="")
    ws1.cell(row=row_idx, column=10, value="")

style_data_rows(ws1, 4, 103, num_cols1, alt_row_pink, wrap_cols={5, 6, 10})

# Conditional formatting for status
green_fill = PatternFill(start_color=MINT, end_color=MINT, fill_type="solid")
coral_fill_cf = PatternFill(start_color=CORAL, end_color=CORAL, fill_type="solid")
blue_fill_cf = PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")

ws1.conditional_formatting.add("G4:G103",
    CellIsRule(operator="equal", formula=['"Published"'], fill=green_fill))
ws1.conditional_formatting.add("G4:G103",
    CellIsRule(operator="equal", formula=['"Boosted"'], fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")))
ws1.conditional_formatting.add("G4:G103",
    CellIsRule(operator="equal", formula=['"Scheduled"'], fill=blue_fill_cf))


# ============================================================
# TAB 2: ANALYTICS DASHBOARD
# ============================================================
ws2 = wb.create_sheet("Analytics Dashboard")
ws2.sheet_properties.tabColor = PURPLE

headers2 = ["Month", "Followers Gained", "Total Followers", "Engagement Rate %",
            "Reach", "Impressions", "Website Clicks", "Top Post", "Revenue from Social"]
widths2 = [16, 18, 18, 18, 16, 16, 16, 35, 20]
num_cols2 = len(headers2)

add_title_block(ws2, "ANALYTICS DASHBOARD 2026", "Track your growth metrics month by month across all platforms", num_cols2)

for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, num_cols2, header_fill_purple)
ws2.freeze_panes = "A4"

for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

months = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
base_followers = 5200
top_posts = [
    "Reel: Morning Routine", "Carousel: Branding Tips", "TikTok: Viral Trend",
    "Blog Post: SEO Guide", "IG Story: Q&A", "YouTube: Tutorial",
    "Pin: Free Checklist", "Reel: Product Launch", "Carousel: Case Study",
    "TikTok: Behind Scenes", "IG Post: Testimonial", "Year in Review Reel"
]

for i, month in enumerate(months):
    row = 4 + i
    ws2.cell(row=row, column=1, value=month)
    gained = random.randint(150, 800)
    base_followers += gained
    ws2.cell(row=row, column=2, value=gained)
    ws2.cell(row=row, column=3, value=base_followers)
    ws2.cell(row=row, column=4, value=round(random.uniform(2.5, 8.5), 1))
    ws2.cell(row=row, column=5, value=random.randint(15000, 95000))
    ws2.cell(row=row, column=6, value=random.randint(25000, 150000))
    ws2.cell(row=row, column=7, value=random.randint(200, 2500))
    ws2.cell(row=row, column=8, value=top_posts[i])
    rev = round(random.uniform(150, 3500), 2)
    ws2.cell(row=row, column=9, value=rev)
    ws2.cell(row=row, column=9).number_format = '$#,##0.00'

style_data_rows(ws2, 4, 15, num_cols2, alt_row_purple, wrap_cols={8})

# Totals row
total_row = 16
ws2.cell(row=total_row, column=1, value="TOTALS / AVG")
ws2.cell(row=total_row, column=2, value=f"=SUM(B4:B15)")
ws2.cell(row=total_row, column=3, value=f"=B16+C3")  # last total followers
ws2.cell(row=total_row, column=3).value = f"=MAX(C4:C15)"
ws2.cell(row=total_row, column=4, value=f"=AVERAGE(D4:D15)")
ws2.cell(row=total_row, column=4).number_format = '0.0'
ws2.cell(row=total_row, column=5, value=f"=SUM(E4:E15)")
ws2.cell(row=total_row, column=6, value=f"=SUM(F4:F15)")
ws2.cell(row=total_row, column=7, value=f"=SUM(G4:G15)")
ws2.cell(row=total_row, column=8, value="—")
ws2.cell(row=total_row, column=9, value=f"=SUM(I4:I15)")
ws2.cell(row=total_row, column=9).number_format = '$#,##0.00'

for c in range(1, num_cols2 + 1):
    cell = ws2.cell(row=total_row, column=c)
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = center_align
    cell.border = header_border
ws2.row_dimensions[total_row].height = 36


# ============================================================
# TAB 3: HASHTAG RESEARCH
# ============================================================
ws3 = wb.create_sheet("Hashtag Research")
ws3.sheet_properties.tabColor = ELECTRIC_BLUE

headers3 = ["Hashtag", "Category", "Post Count / Popularity", "Relevance (1-10)",
            "Last Used", "Performance Rating", "Notes"]
widths3 = [25, 18, 22, 16, 14, 18, 30]
num_cols3 = len(headers3)

add_title_block(ws3, "HASHTAG RESEARCH TRACKER", "Organize, test, and optimize your hashtag strategy for maximum reach", num_cols3)

for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, num_cols3, header_fill_blue)
ws3.freeze_panes = "A4"

for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

perf_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws3.add_data_validation(perf_dv)

hashtag_data = [
    ("#socialmediamarketing", "General", "45.2M", 10), ("#contentcreator", "General", "38.7M", 9),
    ("#digitalmarketing", "General", "52.1M", 9), ("#marketingtips", "Education", "12.4M", 8),
    ("#instagramgrowth", "Instagram", "8.9M", 9), ("#tiktokmarketing", "TikTok", "6.3M", 8),
    ("#socialmediatips", "Education", "15.1M", 10), ("#brandstrategy", "Branding", "3.2M", 8),
    ("#contentmarketing", "General", "22.8M", 9), ("#influencermarketing", "Collab", "9.5M", 7),
    ("#growyourbrand", "Branding", "4.1M", 8), ("#reelsinstagram", "Instagram", "18.9M", 9),
    ("#viralcontent", "General", "7.6M", 7), ("#marketingstrategy", "Strategy", "11.2M", 9),
    ("#smallbusinessmarketing", "Niche", "5.4M", 8), ("#pinterestmarketing", "Pinterest", "2.1M", 7),
    ("#emailmarketing", "Email", "6.8M", 7), ("#seomarketing", "SEO", "4.5M", 6),
    ("#youtubemarketing", "YouTube", "3.9M", 7), ("#linkedinmarketing", "LinkedIn", "2.8M", 6),
    ("#socialmediamanager", "Career", "7.2M", 8), ("#contentcalendar", "Planning", "1.5M", 9),
    ("#engagementtips", "Education", "2.3M", 8), ("#aestheticfeed", "Instagram", "5.7M", 6),
    ("#canvatemplates", "Tools", "3.1M", 5), ("#2026goals", "Seasonal", "1.2M", 7),
    ("#morningroutine", "Lifestyle", "9.8M", 5), ("#behindthescenes", "BTS", "6.4M", 7),
    ("#creativebusiness", "Niche", "3.8M", 7), ("#onlinebusiness", "Business", "11.5M", 7),
    ("#passiveincome", "Business", "8.3M", 5), ("#freelancer", "Career", "7.9M", 6),
    ("#womeninbusiness", "Niche", "6.1M", 6), ("#bossbabe", "Niche", "12.3M", 5),
    ("#entrepreneurlife", "Business", "9.7M", 6), ("#marketingagency", "Business", "2.4M", 5),
    ("#socialmedia2026", "Seasonal", "0.8M", 9), ("#contentideas", "Planning", "3.5M", 8),
    ("#hashtagstrategy", "Strategy", "1.1M", 10), ("#analyticstips", "Data", "0.9M", 7),
    ("#ugccreator", "UGC", "2.7M", 7), ("#brandcollab", "Collab", "1.8M", 8),
    ("#affiliatemarketing", "Revenue", "5.2M", 6), ("#podcastmarketing", "Audio", "1.3M", 5),
    ("#videomarketing", "Video", "4.6M", 8), ("#storytelling", "Content", "7.1M", 7),
    ("#marketingtools", "Tools", "2.0M", 6), ("#growoninstagram", "Instagram", "4.3M", 8),
    ("#fyp", "TikTok", "200M+", 6), ("#trending", "General", "150M+", 5),
]

perfs = ["High", "Medium", "Low"]
for idx, (tag, cat, count, rel) in enumerate(hashtag_data):
    row = 4 + idx
    ws3.cell(row=row, column=1, value=tag)
    ws3.cell(row=row, column=2, value=cat)
    ws3.cell(row=row, column=3, value=count)
    ws3.cell(row=row, column=4, value=rel)
    ws3.cell(row=row, column=5, value=(start_date + timedelta(days=random.randint(0, 60))).strftime("%m/%d/%Y"))
    perf = perfs[0] if rel >= 9 else (perfs[1] if rel >= 7 else perfs[2])
    ws3.cell(row=row, column=6, value=perf)
    perf_dv.add(ws3.cell(row=row, column=6))
    ws3.cell(row=row, column=7, value="")

style_data_rows(ws3, 4, 53, num_cols3, alt_row_blue, wrap_cols={1, 7})

# Conditional formatting for performance
ws3.conditional_formatting.add("F4:F53",
    CellIsRule(operator="equal", formula=['"High"'], fill=green_fill))
ws3.conditional_formatting.add("F4:F53",
    CellIsRule(operator="equal", formula=['"Low"'], fill=coral_fill_cf))


# ============================================================
# TAB 4: BRAND COLLABORATIONS
# ============================================================
ws4 = wb.create_sheet("Brand Collaborations")
ws4.sheet_properties.tabColor = CORAL

headers4 = ["Brand Name", "Contact", "Platform", "Deliverables", "Deadline",
            "Payment", "Status", "Contract Link", "Notes"]
widths4 = [22, 25, 16, 35, 14, 16, 16, 25, 30]
num_cols4 = len(headers4)

add_title_block(ws4, "BRAND COLLABORATIONS TRACKER", "Manage partnerships from pitch to payment — never miss a deliverable", num_cols4)

for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, num_cols4, header_fill_coral)
ws4.freeze_panes = "A4"

for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

collab_status_dv = DataValidation(type="list",
    formula1='"Pitched,Negotiating,Confirmed,In Progress,Completed,Paid"', allow_blank=True)
ws4.add_data_validation(collab_status_dv)

collab_data = [
    ("GlowUp Skincare", "sarah@glowup.com", "Instagram", "3 Reels + 5 Stories + 1 Post", "02/15/2026", 2500, "Confirmed"),
    ("TechFlow App", "partnerships@techflow.io", "TikTok", "2 Videos + Blog Review", "03/01/2026", 1800, "In Progress"),
    ("EcoStyle Fashion", "collab@ecostyle.com", "YouTube", "Dedicated Video + 2 Shorts", "03/20/2026", 3500, "Negotiating"),
    ("FitLife Supplements", "brand@fitlife.com", "Instagram", "Carousel + 3 Stories + Reel", "01/30/2026", 1200, "Completed"),
    ("BookNest Library", "hello@booknest.co", "Pinterest", "10 Pins + Blog Post", "04/10/2026", 800, "Pitched"),
    ("CloudSync Pro", "marketing@cloudsync.com", "LinkedIn", "3 Posts + Article", "02/28/2026", 2000, "Confirmed"),
    ("ArtisanBrew Co.", "collab@artisanbrew.com", "TikTok", "5 TikToks + 1 Live", "03/15/2026", 2800, "Negotiating"),
    ("WanderLust Travel", "partners@wanderlust.travel", "YouTube", "Travel Vlog + 3 Shorts", "05/01/2026", 4000, "Pitched"),
    ("MindfulMe App", "brand@mindfulme.app", "Instagram", "2 Reels + UGC Content Pack", "02/20/2026", 1500, "In Progress"),
    ("PlannerPerfect", "hello@plannerperfect.com", "Blog", "Sponsored Blog + Social Bundle", "03/30/2026", 1000, "Paid"),
]

for idx, (brand, contact, plat, deliv, deadline, pay, status) in enumerate(collab_data):
    row = 4 + idx
    ws4.cell(row=row, column=1, value=brand)
    ws4.cell(row=row, column=2, value=contact)
    ws4.cell(row=row, column=3, value=plat)
    ws4.cell(row=row, column=4, value=deliv)
    ws4.cell(row=row, column=5, value=deadline)
    ws4.cell(row=row, column=6, value=pay)
    ws4.cell(row=row, column=6).number_format = '$#,##0'
    ws4.cell(row=row, column=7, value=status)
    collab_status_dv.add(ws4.cell(row=row, column=7))
    ws4.cell(row=row, column=8, value="")
    ws4.cell(row=row, column=9, value="")

# Add empty rows for user
for row in range(14, 34):
    for c in range(1, num_cols4 + 1):
        ws4.cell(row=row, column=c, value="")
    collab_status_dv.add(ws4.cell(row=row, column=7))

style_data_rows(ws4, 4, 33, num_cols4, alt_row_pink, wrap_cols={4, 9})

# Total payment row
tr = 34
ws4.cell(row=tr, column=1, value="TOTAL EARNINGS")
ws4.cell(row=tr, column=6, value="=SUM(F4:F33)")
ws4.cell(row=tr, column=6).number_format = '$#,##0'
for c in range(1, num_cols4 + 1):
    cell = ws4.cell(row=tr, column=c)
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = center_align
    cell.border = header_border
ws4.row_dimensions[tr].height = 36


# ============================================================
# TAB 5: CONTENT IDEAS BANK
# ============================================================
ws5 = wb.create_sheet("Content Ideas Bank")
ws5.sheet_properties.tabColor = MINT

headers5 = ["Idea", "Platform", "Content Type", "Season / Trending",
            "Priority", "Used?", "Date Used", "Performance"]
widths5 = [40, 16, 16, 20, 14, 10, 14, 16]
num_cols5 = len(headers5)

add_title_block(ws5, "CONTENT IDEAS BANK", "Never run out of content — capture every spark of inspiration here", num_cols5)

for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, num_cols5, header_fill_mint)
ws5.freeze_panes = "A4"

for i, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

priority_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws5.add_data_validation(priority_dv)
used_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws5.add_data_validation(used_dv)

content_ideas = [
    ("Day in my life as a social media manager", "TikTok", "Video", "Evergreen", "High"),
    ("10 Instagram Reels ideas for small businesses", "Instagram", "Carousel", "Evergreen", "High"),
    ("How I plan a month of content in 2 hours", "YouTube", "Video", "Evergreen", "High"),
    ("Pinterest SEO secrets most creators miss", "Pinterest", "Video", "Trending", "Medium"),
    ("My favorite content creation tools for 2026", "Blog", "Blog", "Q1 2026", "High"),
    ("Content repurposing: 1 video → 10 pieces", "Instagram", "Reel", "Evergreen", "High"),
    ("Client onboarding process walkthrough", "LinkedIn", "Carousel", "Evergreen", "Medium"),
    ("Seasonal content calendar template (free)", "Email", "Newsletter", "Q1 2026", "High"),
    ("How to pitch brands as a micro-influencer", "TikTok", "Video", "Trending", "High"),
    ("My morning routine as a content creator", "Instagram", "Reel", "Evergreen", "Medium"),
    ("5 email subject lines that get 40%+ open rate", "Blog", "Blog", "Evergreen", "Medium"),
    ("Behind the scenes of a brand photoshoot", "Instagram", "Story", "Evergreen", "Low"),
    ("Valentine's Day marketing ideas for brands", "Instagram", "Carousel", "February", "High"),
    ("How to use ChatGPT for content planning", "YouTube", "Video", "Trending", "High"),
    ("Spring collection flat lay styling tips", "Pinterest", "Photo", "Spring", "Medium"),
    ("LinkedIn engagement strategy that works", "LinkedIn", "Carousel", "Evergreen", "Medium"),
    ("Weekly analytics review template walkthrough", "YouTube", "Video", "Evergreen", "Medium"),
    ("UGC portfolio: What brands want to see", "TikTok", "Video", "Trending", "High"),
    ("How I grew to 10K followers in 3 months", "Instagram", "Reel", "Evergreen", "High"),
    ("Email welcome sequence breakdown", "Blog", "Blog", "Evergreen", "Medium"),
    ("Batch content creation workshop (LIVE)", "Instagram", "Video", "Monthly", "High"),
    ("Travel content creation essentials", "TikTok", "Video", "Summer", "Low"),
    ("How to create scroll-stopping thumbnails", "YouTube", "Video", "Evergreen", "Medium"),
    ("Monthly social media audit checklist", "Blog", "Blog", "Monthly", "High"),
    ("Year-end social media wrapped / review", "Instagram", "Carousel", "December", "Medium"),
]

for idx, (idea, plat, ct, season, prio) in enumerate(content_ideas):
    row = 4 + idx
    ws5.cell(row=row, column=1, value=idea)
    ws5.cell(row=row, column=2, value=plat)
    ws5.cell(row=row, column=3, value=ct)
    ws5.cell(row=row, column=4, value=season)
    ws5.cell(row=row, column=5, value=prio)
    priority_dv.add(ws5.cell(row=row, column=5))
    ws5.cell(row=row, column=6, value="No")
    used_dv.add(ws5.cell(row=row, column=6))
    ws5.cell(row=row, column=7, value="")
    ws5.cell(row=row, column=8, value="")

# Add empty rows for user to fill
for row in range(29, 54):
    for c in range(1, num_cols5 + 1):
        ws5.cell(row=row, column=c, value="")
    priority_dv.add(ws5.cell(row=row, column=5))
    used_dv.add(ws5.cell(row=row, column=6))

style_data_rows(ws5, 4, 53, num_cols5, PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid"), wrap_cols={1, 4})


# ============================================================
# TAB 6: AUDIENCE INSIGHTS
# ============================================================
ws6 = wb.create_sheet("Audience Insights")
ws6.sheet_properties.tabColor = "805AD5"

headers6 = ["Platform", "Total Followers", "Growth Rate %", "Top Demographics",
            "Best Post Time", "Best Day", "Top Content Type", "Avg Engagement %", "Notes"]
widths6 = [18, 18, 16, 28, 16, 14, 20, 18, 30]
num_cols6 = len(headers6)

add_title_block(ws6, "AUDIENCE INSIGHTS BY PLATFORM", "Know your audience — optimize posting times, content types, and growth strategies", num_cols6)

for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header_row(ws6, 3, num_cols6, PatternFill(start_color="805AD5", end_color="805AD5", fill_type="solid"))
ws6.freeze_panes = "A4"

for i, w in enumerate(widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

audience_data = [
    ("Instagram", 12500, 4.2, "Women 25-34 (42%), US/UK", "7:00 PM", "Tuesday", "Reels", 5.8),
    ("TikTok", 8300, 8.7, "Women 18-24 (38%), US", "12:00 PM", "Thursday", "Short Video", 7.2),
    ("YouTube", 3200, 3.1, "Men 25-44 (35%), Global", "2:00 PM", "Saturday", "Tutorial", 4.5),
    ("Pinterest", 6700, 5.5, "Women 25-44 (55%), US", "8:00 PM", "Sunday", "Infographic", 3.2),
    ("Facebook", 4100, 1.8, "Mixed 35-54 (40%), US", "10:00 AM", "Wednesday", "Video", 2.1),
    ("Twitter/X", 2900, 2.4, "Men 25-34 (33%), Global", "9:00 AM", "Monday", "Thread", 3.8),
    ("LinkedIn", 5600, 6.3, "Mixed 30-45 (45%), US/EU", "8:00 AM", "Tuesday", "Carousel", 6.1),
    ("Blog", 1800, 3.9, "Mixed 25-44, Organic Search", "N/A", "N/A", "Long-form", 2.8),
    ("Email List", 4500, 4.8, "Subscribers, 42% open rate", "6:00 AM", "Tuesday", "Newsletter", 3.5),
]

for idx, (plat, foll, growth, demo, time, day, ct, eng) in enumerate(audience_data):
    row = 4 + idx
    ws6.cell(row=row, column=1, value=plat)
    ws6.cell(row=row, column=2, value=foll)
    ws6.cell(row=row, column=2).number_format = '#,##0'
    ws6.cell(row=row, column=3, value=growth)
    ws6.cell(row=row, column=3).number_format = '0.0'
    ws6.cell(row=row, column=4, value=demo)
    ws6.cell(row=row, column=5, value=time)
    ws6.cell(row=row, column=6, value=day)
    ws6.cell(row=row, column=7, value=ct)
    ws6.cell(row=row, column=8, value=eng)
    ws6.cell(row=row, column=8).number_format = '0.0'
    ws6.cell(row=row, column=9, value="")

style_data_rows(ws6, 4, 12, num_cols6, alt_row_purple, wrap_cols={4, 9})


# ============================================================
# TAB 7: REVENUE TRACKER
# ============================================================
ws7 = wb.create_sheet("Revenue Tracker")
ws7.sheet_properties.tabColor = "38A169"

headers7 = ["Date", "Source", "Brand / Platform", "Amount", "Payment Status", "Notes"]
widths7 = [14, 20, 25, 16, 18, 35]
num_cols7 = len(headers7)

add_title_block(ws7, "REVENUE TRACKER 2026", "Track every dollar earned from your social media empire", num_cols7)

for i, h in enumerate(headers7, 1):
    ws7.cell(row=3, column=i, value=h)
style_header_row(ws7, 3, num_cols7, PatternFill(start_color="38A169", end_color="38A169", fill_type="solid"))
ws7.freeze_panes = "A4"

for i, w in enumerate(widths7, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

source_dv = DataValidation(type="list",
    formula1='"Sponsorship,Affiliate,Product,Ad Revenue,Coaching,Other"', allow_blank=True)
ws7.add_data_validation(source_dv)

pay_status_dv = DataValidation(type="list",
    formula1='"Pending,Invoiced,Paid,Overdue"', allow_blank=True)
ws7.add_data_validation(pay_status_dv)

revenue_data = [
    ("01/05/2026", "Sponsorship", "GlowUp Skincare", 2500, "Paid"),
    ("01/12/2026", "Affiliate", "Amazon Associates", 342.50, "Paid"),
    ("01/20/2026", "Product", "Content Calendar Template", 189.00, "Paid"),
    ("01/28/2026", "Ad Revenue", "YouTube AdSense", 567.80, "Paid"),
    ("02/03/2026", "Sponsorship", "TechFlow App", 1800, "Invoiced"),
    ("02/10/2026", "Coaching", "1:1 Strategy Session", 350, "Paid"),
    ("02/15/2026", "Affiliate", "Canva Pro Referral", 125.00, "Pending"),
    ("02/22/2026", "Product", "Hashtag Guide eBook", 95.00, "Paid"),
    ("03/01/2026", "Sponsorship", "ArtisanBrew Co.", 2800, "Pending"),
    ("03/08/2026", "Ad Revenue", "TikTok Creator Fund", 234.50, "Paid"),
    ("03/15/2026", "Coaching", "Group Workshop (10 ppl)", 1500, "Invoiced"),
    ("03/20/2026", "Affiliate", "ConvertKit Referral", 210.00, "Pending"),
    ("03/28/2026", "Product", "Social Media Planner", 445.50, "Paid"),
    ("04/05/2026", "Sponsorship", "WanderLust Travel", 4000, "Pending"),
    ("04/12/2026", "Other", "Speaking Fee - Conference", 2000, "Invoiced"),
]

for idx, (date, source, brand, amount, status) in enumerate(revenue_data):
    row = 4 + idx
    ws7.cell(row=row, column=1, value=date)
    ws7.cell(row=row, column=2, value=source)
    source_dv.add(ws7.cell(row=row, column=2))
    ws7.cell(row=row, column=3, value=brand)
    ws7.cell(row=row, column=4, value=amount)
    ws7.cell(row=row, column=4).number_format = '$#,##0.00'
    ws7.cell(row=row, column=5, value=status)
    pay_status_dv.add(ws7.cell(row=row, column=5))
    ws7.cell(row=row, column=6, value="")

# Empty rows for user
for row in range(19, 54):
    for c in range(1, num_cols7 + 1):
        ws7.cell(row=row, column=c, value="")
    source_dv.add(ws7.cell(row=row, column=2))
    pay_status_dv.add(ws7.cell(row=row, column=5))

style_data_rows(ws7, 4, 53, num_cols7, PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid"), wrap_cols={6})

# Totals
tr = 54
ws7.cell(row=tr, column=1, value="TOTAL REVENUE")
ws7.cell(row=tr, column=4, value="=SUM(D4:D53)")
ws7.cell(row=tr, column=4).number_format = '$#,##0.00'
for c in range(1, num_cols7 + 1):
    cell = ws7.cell(row=tr, column=c)
    cell.font = total_font
    cell.fill = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid")
    cell.alignment = center_align
    cell.border = header_border
ws7.row_dimensions[tr].height = 36


# ============================================================
# TAB 8: HOW TO USE
# ============================================================
ws8 = wb.create_sheet("How to Use")
ws8.sheet_properties.tabColor = "E53E3E"

ws8.column_dimensions["A"].width = 6
ws8.column_dimensions["B"].width = 80
ws8.column_dimensions["C"].width = 6

# Title
ws8.merge_cells("A1:C1")
c1 = ws8.cell(row=1, column=1)
c1.value = "✦ HOW TO USE YOUR SOCIAL MEDIA MARKETING PLANNER"
c1.font = Font(name="Calibri", bold=True, size=20, color=HOT_PINK)
c1.alignment = Alignment(horizontal="center", vertical="center")
c1.fill = PatternFill(start_color=VERY_LIGHT_PINK, end_color=VERY_LIGHT_PINK, fill_type="solid")
ws8.row_dimensions[1].height = 56

ws8.merge_cells("A2:C2")
c2 = ws8.cell(row=2, column=1)
c2.value = "Welcome! This planner is designed to streamline your entire social media workflow."
c2.font = Font(name="Calibri", size=12, color=PURPLE, italic=True)
c2.alignment = Alignment(horizontal="center", vertical="center")
c2.fill = PatternFill(start_color=VERY_LIGHT_PINK, end_color=VERY_LIGHT_PINK, fill_type="solid")
ws8.row_dimensions[2].height = 32

instructions = [
    ("", ""),
    ("📋 TAB-BY-TAB GUIDE", ""),
    ("", ""),
    ("1. Content Calendar", "Your command center! Enter dates, select platforms from dropdowns, type captions, and track status. The Day column auto-fills from the date. Use the Status dropdown to move posts through your workflow: Idea → Draft → Scheduled → Published → Boosted."),
    ("2. Analytics Dashboard", "Enter monthly metrics to spot trends. Total/average rows calculate automatically. Track followers gained, engagement rate, reach, and revenue month-over-month."),
    ("3. Hashtag Research", "Organize hashtags by category and track which ones perform best. Rate them High/Medium/Low and note when you last used each set. Rotate hashtags regularly for best results."),
    ("4. Brand Collaborations", "Track every brand deal from first pitch to final payment. Use the Status dropdown to manage your pipeline. The total row auto-calculates your earnings."),
    ("5. Content Ideas Bank", "Never run out of ideas! 25 pre-loaded ideas included. Add your own, mark priority level, and check them off when used. Great for batch planning sessions."),
    ("6. Audience Insights", "One row per platform — track followers, growth rate, demographics, and best posting times. Update monthly for accurate data."),
    ("7. Revenue Tracker", "Log every income source: sponsorships, affiliates, product sales, ad revenue, coaching. Auto-totals at the bottom. Track payment status so nothing falls through."),
    ("", ""),
    ("💡 PRO TIPS", ""),
    ("", ""),
    ("Tip 1:", "Start each month by filling in Content Calendar dates and themes. Batch-plan at least 2 weeks ahead."),
    ("Tip 2:", "Update Analytics Dashboard on the 1st of each month — consistency is key for spotting growth patterns."),
    ("Tip 3:", "Keep 3 sets of hashtags (branded, niche, trending) and rotate them to avoid shadowbans."),
    ("Tip 4:", "Screenshot your Audience Insights from each platform monthly and log the numbers here."),
    ("Tip 5:", "Use the Revenue Tracker religiously — knowing your income per platform helps you prioritize."),
    ("Tip 6:", "Color-code your Content Calendar rows by platform for quick visual scanning."),
    ("Tip 7:", "Back up this file regularly! Save a copy to Google Drive or Dropbox."),
    ("", ""),
    ("🎯 BONUS: 20 CONTENT IDEAS TO GET STARTED", ""),
    ("", ""),
    ("1.", "Share your \"why\" story — why you started your brand/business"),
    ("2.", "Create a \"tools I use daily\" carousel post"),
    ("3.", "Film a \"get ready with me\" while discussing business tips"),
    ("4.", "Post a client transformation or before/after"),
    ("5.", "Make a trending audio Reel with your own twist"),
    ("6.", "Share 3 mistakes you made early on (and what you learned)"),
    ("7.", "Create a step-by-step tutorial in your niche"),
    ("8.", "Do a \"this or that\" Story poll for engagement"),
    ("9.", "Share your workspace/desk setup tour"),
    ("10.", "Post your monthly goals and invite followers to share theirs"),
    ("11.", "Create a \"myth vs. reality\" post about your industry"),
    ("12.", "Film a day-in-the-life TikTok/Reel"),
    ("13.", "Share a free resource or checklist (great for email signups)"),
    ("14.", "Do a live Q&A session — repurpose clips later"),
    ("15.", "Post a testimonial or review from a happy client"),
    ("16.", "Create an infographic with surprising industry stats"),
    ("17.", "Share your content creation process behind the scenes"),
    ("18.", "Write a LinkedIn article about an industry trend"),
    ("19.", "Create a Pinterest pin linking to your latest blog post"),
    ("20.", "Host a giveaway or challenge to boost engagement"),
]

for idx, (label, text) in enumerate(instructions):
    row = 3 + idx
    cell_a = ws8.cell(row=row, column=1)
    cell_b = ws8.cell(row=row, column=2)

    if "TAB-BY-TAB" in label or "PRO TIPS" in label or "BONUS" in label:
        ws8.merge_cells(f"A{row}:C{row}")
        cell_a.value = label
        cell_a.font = Font(name="Calibri", bold=True, size=14, color=PURPLE)
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_a.fill = PatternFill(start_color=PALE_PINK, end_color=PALE_PINK, fill_type="solid")
        ws8.row_dimensions[row].height = 36
    elif label.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.")):
        if "Content Calendar" in label or "Analytics" in label or "Hashtag" in label or \
           "Brand" in label or "Content Ideas" in label or "Audience" in label or "Revenue" in label:
            cell_a.value = ""
            cell_b.value = f"{label}: {text}"
            cell_b.font = Font(name="Calibri", size=11, color=DARK_TEXT)
            cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws8.row_dimensions[row].height = 56
            # Bold the tab name
            cell_b.font = Font(name="Calibri", size=11, color=DARK_TEXT)
        else:
            cell_a.value = label
            cell_a.font = Font(name="Calibri", bold=True, size=11, color=HOT_PINK)
            cell_a.alignment = Alignment(horizontal="right", vertical="top")
            cell_b.value = text
            cell_b.font = body_font
            cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws8.row_dimensions[row].height = 28
    elif label.startswith("Tip"):
        cell_a.value = label
        cell_a.font = Font(name="Calibri", bold=True, size=11, color=HOT_PINK)
        cell_a.alignment = Alignment(horizontal="right", vertical="top")
        cell_b.value = text
        cell_b.font = body_font
        cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws8.row_dimensions[row].height = 28
    elif label == "" and text == "":
        ws8.row_dimensions[row].height = 12
    else:
        cell_a.value = label
        cell_b.value = text


# Save
wb.save(OUTPUT)
print(f"Spreadsheet saved to {OUTPUT}")
print(f"Tabs: {wb.sheetnames}")
