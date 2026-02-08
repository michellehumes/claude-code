#!/usr/bin/env python3
"""
Wedding Planner & Budget Spreadsheet Bundle for Etsy
Generates a comprehensive, professionally-styled Excel workbook.
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Color Palette ──────────────────────────────────────────────────────────
BLUSH       = "E8C4C8"
ROSE_GOLD   = "B76E79"
CHAMPAGNE   = "F7E7CE"
IVORY       = "FFFFF0"
SAGE_GREEN  = "9CAF88"
DUSTY_MAUVE = "C9A9A6"
DEEP_PLUM   = "4A2040"
GOLD        = "C5A55A"
WHITE       = "FFFFFF"

# ── Reusable style objects ─────────────────────────────────────────────────
FILL_DEEP_PLUM  = PatternFill("solid", fgColor=DEEP_PLUM)
FILL_ROSE_GOLD  = PatternFill("solid", fgColor=ROSE_GOLD)
FILL_CHAMPAGNE  = PatternFill("solid", fgColor=CHAMPAGNE)
FILL_IVORY      = PatternFill("solid", fgColor=IVORY)
FILL_WHITE      = PatternFill("solid", fgColor=WHITE)
FILL_BLUSH      = PatternFill("solid", fgColor=BLUSH)
FILL_SAGE       = PatternFill("solid", fgColor=SAGE_GREEN)
FILL_GOLD       = PatternFill("solid", fgColor=GOLD)
FILL_MAUVE      = PatternFill("solid", fgColor=DUSTY_MAUVE)

FONT_TITLE    = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=13, bold=True, color=DEEP_PLUM)
FONT_HEADER   = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_HEADER_DARK = Font(name="Calibri", size=11, bold=True, color=DEEP_PLUM)
FONT_BODY     = Font(name="Calibri", size=11, color=DEEP_PLUM)
FONT_BODY_SM  = Font(name="Calibri", size=10, color=DEEP_PLUM)
FONT_LABEL    = Font(name="Calibri", size=11, bold=True, color=ROSE_GOLD)
FONT_GOLD_HDR = Font(name="Calibri", size=11, bold=True, color=GOLD)
FONT_WHITE_SM = Font(name="Calibri", size=10, color=WHITE)

THIN_BORDER = Border(
    left=Side(style="thin", color=DUSTY_MAUVE),
    right=Side(style="thin", color=DUSTY_MAUVE),
    top=Side(style="thin", color=DUSTY_MAUVE),
    bottom=Side(style="thin", color=DUSTY_MAUVE),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)

MONEY_FMT = '"$"#,##0.00'
DATE_FMT  = "MM/DD/YYYY"
PCT_FMT   = "0.0%"

# ── Helper functions ───────────────────────────────────────────────────────

def set_col_widths(ws, widths: dict):
    """Set column widths from {col_letter: width} dict."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_title_block(ws, title, subtitle, merge_end_col="N", row_start=1):
    """Write a two-row title block across the top of the sheet."""
    end = merge_end_col
    r1 = row_start
    r2 = row_start + 1

    ws.merge_cells(f"A{r1}:{end}{r1}")
    cell = ws[f"A{r1}"]
    cell.value = title
    cell.font = FONT_TITLE
    cell.fill = FILL_DEEP_PLUM
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r1].height = 42

    ws.merge_cells(f"A{r2}:{end}{r2}")
    cell2 = ws[f"A{r2}"]
    cell2.value = subtitle
    cell2.font = FONT_SUBTITLE
    cell2.fill = FILL_CHAMPAGNE
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r2].height = 26
    return r2 + 1  # next available row


def write_header_row(ws, row, headers, fill=None, font=None):
    """Write a styled header row."""
    if fill is None:
        fill = FILL_ROSE_GOLD
    if font is None:
        font = FONT_HEADER
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = font
        c.fill = fill
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def style_data_row(ws, row, num_cols, is_even=False):
    """Apply alternating row fill and borders."""
    fill = FILL_IVORY if is_even else FILL_WHITE
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = fill
        c.border = THIN_BORDER
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER


def add_dropdown(ws, validation_list, cell_range):
    """Add a dropdown data validation."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(validation_list)}"',
        allow_blank=True,
    )
    dv.error = "Please select from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select an option"
    dv.promptTitle = "Choose"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def format_money_col(ws, col, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).number_format = MONEY_FMT


def format_date_col(ws, col, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).number_format = DATE_FMT


def apply_conditional_status(ws, col_letter, start_row, end_row):
    """Color cells by status keywords."""
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Done"'],
                   fill=PatternFill("solid", fgColor=SAGE_GREEN),
                   font=Font(color=WHITE, bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"In Progress"'],
                   fill=PatternFill("solid", fgColor=GOLD),
                   font=Font(color=DEEP_PLUM, bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Not Started"'],
                   fill=PatternFill("solid", fgColor=DUSTY_MAUVE),
                   font=Font(color=WHITE, bold=True)),
    )


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def build_how_to_use(wb):
    """Tab 9 → moved to first position: comprehensive guide."""
    ws = wb.active
    ws.title = "How to Use"
    ws.sheet_properties.tabColor = DEEP_PLUM

    next_row = write_title_block(
        ws,
        "Wedding Planner Bundle — How to Use",
        "Your Complete Guide to Every Tab  |  Tips & Best Practices",
        merge_end_col="G",
    )

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 80
    ws.column_dimensions["G"].width = 4

    sections = [
        ("Wedding Dashboard", [
            "This is your command center — start here every time you open the file.",
            "Enter your wedding date in cell C6; the countdown formula updates automatically.",
            "The budget summary pulls totals from the Budget Tracker tab via formulas.",
            "The RSVP summary also links to the Guest List tab for real-time counts.",
            "Use the Key Contacts section to keep your most important numbers handy.",
            "Review the countdown checklist to stay on track with planning milestones.",
        ]),
        ("Budget Tracker", [
            "All 23 common wedding categories are pre-loaded — add more rows as needed.",
            "Enter Estimated Cost first, then update Actual Cost as you book vendors.",
            "The Deposit Paid and Balance Due columns help you track payment schedules.",
            "Mark 'Yes' in the Paid In Full column when a vendor is fully settled.",
            "The pie chart auto-updates to show your budget breakdown visually.",
            "Totals at the bottom auto-calculate — never manually change the TOTAL row.",
        ]),
        ("Guest List & RSVPs", [
            "Pre-formatted for 150 guests — insert rows above the totals to add more.",
            "Use the Side dropdown (Bride/Groom/Both) to balance your guest list.",
            "Track Plus Ones separately so catering counts are accurate.",
            "Update RSVP Status as responses arrive (Invited → Attending/Declined/Pending).",
            "The auto-count totals at the bottom give you instant RSVP summaries.",
            "Mark Thank You Sent to keep track of your gratitude notes.",
        ]),
        ("Vendor Tracker", [
            "20 vendor categories are pre-filled — customize names and details.",
            "Store ALL vendor contact info here for easy reference on the wedding day.",
            "Track contracts, deposits, and balances to avoid missed payments.",
            "Balance Due auto-calculates as (Total Cost − Deposit Amount).",
            "Use the Notes column for special arrangements or package details.",
        ]),
        ("Timeline / Checklist", [
            "Organized from 12 months out to Wedding Day in logical timeframes.",
            "Each task has Priority (High/Medium/Low) and Status tracking.",
            "Assign tasks to specific people (Bride, Groom, MOH, Best Man, etc.).",
            "Status conditional formatting: Green=Done, Gold=In Progress, Mauve=Not Started.",
            "Add your own tasks as needed — keep the Timeframe column consistent.",
        ]),
        ("Seating Chart", [
            "Pre-set for 20 tables of 8 guests each (160 seats).",
            "Use Table Name/Theme for creative table identifiers.",
            "Enter guest names from your Guest List to cross-reference table assignments.",
            "Note dietary restrictions and special needs for catering and venue staff.",
            "Update the Table # column on the Guest List tab to match.",
        ]),
        ("Day-Of Timeline", [
            "A minute-by-minute schedule from 8 AM to 11 PM.",
            "Assign a Person Responsible for each event so everyone knows their role.",
            "Print copies for your coordinator, DJ, photographer, and bridal party.",
            "Adjust times to match your ceremony and reception schedule.",
            "Use Status column day-of to mark events as they happen.",
        ]),
        ("Gift & Thank You Tracker", [
            "Log each gift as it arrives — don't rely on memory!",
            "Estimated Value helps with insurance and registry tracking.",
            "Mark Thank You Written and Date Sent to ensure no one is overlooked.",
            "Etiquette tip: send thank-you notes within 2-3 months of receiving a gift.",
            "Use Notes for personal details to mention in your thank-you card.",
        ]),
    ]

    r = next_row + 1

    # General tips header
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(row=r, column=2, value="General Tips")
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = FILL_SAGE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[r].height = 30
    r += 1

    general_tips = [
        "Save a backup copy before making major changes.",
        "Use Ctrl+S frequently — spreadsheets do not auto-save.",
        "Do NOT delete formula cells (they are in the TOTAL/summary rows).",
        "Dropdown menus appear when you click a cell — look for the small arrow.",
        "Colored cells indicate formulas or conditional formatting — avoid overwriting.",
        "Print individual tabs by right-clicking the tab → Print.",
        "This workbook is fully customizable — change colors, add rows, adjust as needed!",
    ]

    for tip in general_tips:
        ws.merge_cells(f"B{r}:F{r}")
        tc = ws.cell(row=r, column=2, value=f"  •  {tip}")
        tc.font = FONT_BODY
        tc.fill = FILL_IVORY if (r % 2 == 0) else FILL_WHITE
        tc.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1  # spacer

    # Tab-by-tab sections in two-column layout
    col_pairs = [(2, 3), (5, 6)]  # (label_col, desc_col)
    pair_idx = 0
    row_anchor = r

    for si, (tab_name, bullets) in enumerate(sections):
        lc, dc = col_pairs[pair_idx]
        cr = row_anchor

        # Section header
        ws.cell(row=cr, column=lc, value=tab_name).font = Font(
            name="Calibri", size=13, bold=True, color=WHITE
        )
        ws.cell(row=cr, column=lc).fill = FILL_ROSE_GOLD
        ws.cell(row=cr, column=lc).alignment = ALIGN_CENTER
        ws.cell(row=cr, column=dc).fill = FILL_ROSE_GOLD
        ws.row_dimensions[cr].height = 28
        cr += 1

        for bi, bullet in enumerate(bullets):
            ws.merge_cells(
                start_row=cr, start_column=lc, end_row=cr, end_column=dc
            )
            bc = ws.cell(row=cr, column=lc, value=f"  •  {bullet}")
            bc.font = FONT_BODY_SM
            bc.fill = FILL_IVORY if bi % 2 == 0 else FILL_WHITE
            bc.alignment = ALIGN_LEFT
            ws.row_dimensions[cr].height = 22
            cr += 1

        cr += 1  # spacer between sections

        if pair_idx == 0:
            pair_idx = 1
        else:
            pair_idx = 0
            row_anchor = max(row_anchor, cr)

        if pair_idx == 0 and si < len(sections) - 1:
            row_anchor = cr

    # Footer
    fr = max(row_anchor, cr) + 2
    ws.merge_cells(f"B{fr}:F{fr}")
    fc = ws.cell(row=fr, column=2,
                 value="Congratulations on your engagement! Wishing you a beautiful wedding day.")
    fc.font = Font(name="Calibri", size=12, italic=True, color=ROSE_GOLD)
    fc.alignment = Alignment(horizontal="center")

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_dashboard(wb):
    ws = wb.create_sheet("Wedding Dashboard")
    ws.sheet_properties.tabColor = ROSE_GOLD

    next_row = write_title_block(
        ws,
        "Wedding Dashboard",
        "Your Wedding At A Glance  |  Key Metrics & Countdown",
        merge_end_col="J",
    )

    set_col_widths(ws, {
        "A": 4, "B": 24, "C": 22, "D": 4,
        "E": 24, "F": 22, "G": 4,
        "H": 24, "I": 22, "J": 4,
    })

    r = next_row + 1

    # ── Section 1: Wedding Date & Countdown ──
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(row=r, column=2, value="Wedding Date & Countdown").font = Font(
        name="Calibri", size=14, bold=True, color=WHITE
    )
    ws.cell(row=r, column=2).fill = FILL_ROSE_GOLD
    ws.cell(row=r, column=2).alignment = ALIGN_CENTER
    ws.cell(row=r, column=3).fill = FILL_ROSE_GOLD
    ws.row_dimensions[r].height = 30
    r += 1

    labels_vals = [
        ("Wedding Date:", "06/20/2026"),
        ("Days Until Wedding:", '=IF(C{0}="","",C{0}-TODAY())'.format(r - 1)),
        ("Weeks Until Wedding:", '=IF(C{0}="","",INT((C{0}-TODAY())/7))'.format(r - 1)),
    ]

    for li, (lbl, val) in enumerate(labels_vals):
        ws.cell(row=r, column=2, value=lbl).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_IVORY if li % 2 == 0 else FILL_WHITE
        ws.cell(row=r, column=2).alignment = ALIGN_RIGHT
        ws.cell(row=r, column=2).border = THIN_BORDER
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = Font(name="Calibri", size=14, bold=True, color=DEEP_PLUM)
        vc.fill = FILL_IVORY if li % 2 == 0 else FILL_WHITE
        vc.alignment = ALIGN_CENTER
        vc.border = THIN_BORDER
        if li == 0:
            vc.number_format = DATE_FMT
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1  # spacer

    # ── Section 2: Overall Budget Summary ──
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(row=r, column=2, value="Budget Summary").font = Font(
        name="Calibri", size=14, bold=True, color=WHITE
    )
    ws.cell(row=r, column=2).fill = FILL_DEEP_PLUM
    ws.cell(row=r, column=2).alignment = ALIGN_CENTER
    ws.cell(row=r, column=3).fill = FILL_DEEP_PLUM
    ws.row_dimensions[r].height = 30
    r += 1

    budget_items = [
        ("Total Budget (Estimated):", "='Budget Tracker'!C27"),
        ("Total Spent (Actual):",     "='Budget Tracker'!D27"),
        ("Remaining:",                "='Budget Tracker'!C27-'Budget Tracker'!D27"),
        ("% Used:",                   "=IF('Budget Tracker'!C27=0,0,'Budget Tracker'!D27/'Budget Tracker'!C27)"),
    ]
    for bi, (lbl, val) in enumerate(budget_items):
        ws.cell(row=r, column=2, value=lbl).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_IVORY if bi % 2 == 0 else FILL_WHITE
        ws.cell(row=r, column=2).alignment = ALIGN_RIGHT
        ws.cell(row=r, column=2).border = THIN_BORDER
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = Font(name="Calibri", size=13, bold=True, color=DEEP_PLUM)
        vc.fill = FILL_IVORY if bi % 2 == 0 else FILL_WHITE
        vc.alignment = ALIGN_CENTER
        vc.border = THIN_BORDER
        if bi < 3:
            vc.number_format = MONEY_FMT
        else:
            vc.number_format = PCT_FMT
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1

    # ── Section 3: RSVP Summary ──
    ws.merge_cells(f"E{next_row+1}:F{next_row+1}")
    rsvp_hdr_row = next_row + 1
    ws.cell(row=rsvp_hdr_row, column=5, value="RSVP Summary").font = Font(
        name="Calibri", size=14, bold=True, color=WHITE
    )
    ws.cell(row=rsvp_hdr_row, column=5).fill = FILL_SAGE
    ws.cell(row=rsvp_hdr_row, column=5).alignment = ALIGN_CENTER
    ws.cell(row=rsvp_hdr_row, column=6).fill = FILL_SAGE

    rsvp_items = [
        ("Total Invited:",  "=COUNTA('Guest List & RSVPs'!B5:B154)"),
        ("Attending:",      '=COUNTIF(\'Guest List & RSVPs\'!K5:K154,"Attending")'),
        ("Declined:",       '=COUNTIF(\'Guest List & RSVPs\'!K5:K154,"Declined")'),
        ("Pending:",        '=COUNTIF(\'Guest List & RSVPs\'!K5:K154,"Pending")'),
    ]
    rr = rsvp_hdr_row + 1
    for ri, (lbl, val) in enumerate(rsvp_items):
        ws.cell(row=rr, column=5, value=lbl).font = FONT_LABEL
        ws.cell(row=rr, column=5).fill = FILL_IVORY if ri % 2 == 0 else FILL_WHITE
        ws.cell(row=rr, column=5).alignment = ALIGN_RIGHT
        ws.cell(row=rr, column=5).border = THIN_BORDER
        vc = ws.cell(row=rr, column=6, value=val)
        vc.font = Font(name="Calibri", size=13, bold=True, color=DEEP_PLUM)
        vc.fill = FILL_IVORY if ri % 2 == 0 else FILL_WHITE
        vc.alignment = ALIGN_CENTER
        vc.border = THIN_BORDER
        rr += 1

    # ── Section 4: Vendor Payment Status ──
    rr += 1
    ws.merge_cells(f"E{rr}:F{rr}")
    ws.cell(row=rr, column=5, value="Vendor Payment Status").font = Font(
        name="Calibri", size=14, bold=True, color=WHITE
    )
    ws.cell(row=rr, column=5).fill = FILL_GOLD
    ws.cell(row=rr, column=5).alignment = ALIGN_CENTER
    ws.cell(row=rr, column=6).fill = FILL_GOLD
    rr += 1

    vendor_stats = [
        ("Contracts Signed:",  '=COUNTIF(\'Vendor Tracker\'!H5:H24,"Yes")'),
        ("Deposits Paid:",     '=COUNTIF(\'Vendor Tracker\'!J5:J24,"Yes")'),
        ("Fully Paid:",        '=COUNTIF(\'Vendor Tracker\'!N5:N24,"Paid")'),
        ("Total Vendor Cost:", "=SUM('Vendor Tracker'!K5:K24)"),
    ]
    for vi, (lbl, val) in enumerate(vendor_stats):
        ws.cell(row=rr, column=5, value=lbl).font = FONT_LABEL
        ws.cell(row=rr, column=5).fill = FILL_IVORY if vi % 2 == 0 else FILL_WHITE
        ws.cell(row=rr, column=5).alignment = ALIGN_RIGHT
        ws.cell(row=rr, column=5).border = THIN_BORDER
        vc = ws.cell(row=rr, column=6, value=val)
        vc.font = Font(name="Calibri", size=13, bold=True, color=DEEP_PLUM)
        vc.fill = FILL_IVORY if vi % 2 == 0 else FILL_WHITE
        vc.alignment = ALIGN_CENTER
        vc.border = THIN_BORDER
        if vi == 3:
            vc.number_format = MONEY_FMT
        rr += 1

    # ── Section 5: Key Contacts ──
    cr = r
    ws.merge_cells(f"B{cr}:F{cr}")
    ws.cell(row=cr, column=2, value="Key Contacts").font = Font(
        name="Calibri", size=14, bold=True, color=WHITE
    )
    ws.cell(row=cr, column=2).fill = FILL_ROSE_GOLD
    ws.cell(row=cr, column=2).alignment = ALIGN_CENTER
    for cc in range(3, 7):
        ws.cell(row=cr, column=cc).fill = FILL_ROSE_GOLD
    cr += 1

    contact_headers = ["Role", "Name", "Phone", "Email", "Notes"]
    write_header_row(ws, cr, contact_headers, fill=FILL_DEEP_PLUM)
    # Adjust to 5 cols (B-F mapped to 2-6, but write_header_row writes from col 1)
    # Let's do it manually
    for ci, h in enumerate(contact_headers):
        c = ws.cell(row=cr, column=ci + 2, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_DEEP_PLUM
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    cr += 1

    contact_roles = [
        "Wedding Coordinator", "Venue Contact", "Caterer", "Photographer",
        "DJ / Band", "Florist", "Officiant", "Hair & Makeup",
    ]
    for ki, role in enumerate(contact_roles):
        ws.cell(row=cr, column=2, value=role).font = FONT_BODY
        for cc in range(2, 7):
            c = ws.cell(row=cr, column=cc)
            c.fill = FILL_IVORY if ki % 2 == 0 else FILL_WHITE
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER
            c.font = FONT_BODY
        cr += 1

    # ── Section 6: Quick Countdown Checklist ──
    cr += 1
    ws.merge_cells(f"H{next_row+1}:I{next_row+1}")
    ws.cell(row=next_row+1, column=8, value="Quick To-Do Checklist").font = Font(
        name="Calibri", size=14, bold=True, color=WHITE
    )
    ws.cell(row=next_row+1, column=8).fill = FILL_MAUVE
    ws.cell(row=next_row+1, column=8).alignment = ALIGN_CENTER
    ws.cell(row=next_row+1, column=9).fill = FILL_MAUVE

    quick_tasks = [
        ("Book venue", "High"),
        ("Hire photographer", "High"),
        ("Send save-the-dates", "High"),
        ("Choose wedding party", "Medium"),
        ("Book caterer", "High"),
        ("Order invitations", "Medium"),
        ("Schedule tastings", "Medium"),
        ("Plan honeymoon", "Medium"),
        ("Buy wedding rings", "High"),
        ("Final dress fitting", "High"),
        ("Confirm all vendors", "High"),
        ("Write vows", "Medium"),
    ]
    tr = next_row + 2
    for ci, hdr in enumerate(["Task", "Priority"]):
        c = ws.cell(row=tr, column=8 + ci, value=hdr)
        c.font = FONT_HEADER
        c.fill = FILL_DEEP_PLUM
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    tr += 1

    for ti, (task, pri) in enumerate(quick_tasks):
        ws.cell(row=tr, column=8, value=task).font = FONT_BODY
        ws.cell(row=tr, column=8).fill = FILL_IVORY if ti % 2 == 0 else FILL_WHITE
        ws.cell(row=tr, column=8).border = THIN_BORDER
        ws.cell(row=tr, column=8).alignment = ALIGN_LEFT
        pc = ws.cell(row=tr, column=9, value=pri)
        pc.font = FONT_BODY
        pc.fill = FILL_IVORY if ti % 2 == 0 else FILL_WHITE
        pc.border = THIN_BORDER
        pc.alignment = ALIGN_CENTER
        tr += 1

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_budget_tracker(wb):
    ws = wb.create_sheet("Budget Tracker")
    ws.sheet_properties.tabColor = GOLD

    next_row = write_title_block(
        ws,
        "Wedding Budget Tracker",
        "Track Every Dollar  |  Estimated vs. Actual  |  Auto-Calculated Totals",
        merge_end_col="H",
    )

    set_col_widths(ws, {
        "A": 4, "B": 28, "C": 18, "D": 18,
        "E": 18, "F": 18, "G": 16, "H": 30,
    })

    headers = [
        "Category", "Estimated Cost", "Actual Cost",
        "Deposit Paid", "Balance Due", "Paid In Full", "Notes",
    ]
    hr = next_row
    # Write headers starting at col B but we map to 1-based col index properly
    # Actually let's use cols A=spacer. Let's keep it simple: data starts col A.
    # Re-do: col A is spacer implicitly with width 4. But headers col B..H = 2..8
    # For simplicity, let's use columns B through H (2-8).
    for ci, h in enumerate(headers):
        c = ws.cell(row=hr, column=ci + 1, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_DEEP_PLUM
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[hr].height = 26

    # Re-map widths for 1-based
    set_col_widths(ws, {
        "A": 30, "B": 18, "C": 18, "D": 18,
        "E": 18, "F": 16, "G": 30, "H": 4,
    })

    categories = [
        "Venue", "Catering", "Photography", "Videography",
        "Flowers / Florist", "Music / DJ", "Wedding Dress",
        "Groom's Attire", "Bridesmaids", "Groomsmen",
        "Hair & Makeup", "Invitations / Stationery", "Wedding Cake",
        "Favors", "Transportation", "Officiant", "Wedding Rings",
        "Decorations", "Rentals", "Coordinator / Planner",
        "Honeymoon", "Gifts", "Miscellaneous",
    ]

    dr = hr + 1
    for ri, cat in enumerate(categories):
        row_num = dr + ri
        ws.cell(row=row_num, column=1, value=cat)
        # Balance Due formula = Actual Cost - Deposit Paid
        ws.cell(row=row_num, column=5,
                value=f"=IF(C{row_num}=\"\",\"\",C{row_num}-D{row_num})")
        style_data_row(ws, row_num, 7, is_even=(ri % 2 == 0))
        ws.cell(row=row_num, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row_num, column=7).alignment = ALIGN_LEFT

    end_data = dr + len(categories) - 1

    # Money formatting
    for mc in [2, 3, 4, 5]:
        format_money_col(ws, mc, dr, end_data)

    # Paid In Full dropdown
    add_dropdown(ws, ["Yes", "No"], f"F{dr}:F{end_data}")

    # TOTAL row
    tr = end_data + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(
        name="Calibri", size=12, bold=True, color=WHITE
    )
    for tc in range(1, 8):
        ws.cell(row=tr, column=tc).fill = FILL_ROSE_GOLD
        ws.cell(row=tr, column=tc).border = THIN_BORDER
        ws.cell(row=tr, column=tc).font = Font(
            name="Calibri", size=12, bold=True, color=WHITE
        )
        ws.cell(row=tr, column=tc).alignment = ALIGN_CENTER

    for mc in [2, 3, 4, 5]:
        col_l = get_column_letter(mc)
        ws.cell(row=tr, column=mc,
                value=f"=SUM({col_l}{dr}:{col_l}{end_data})")
        ws.cell(row=tr, column=mc).number_format = MONEY_FMT

    ws.row_dimensions[tr].height = 28

    # ── Pie Chart ──
    chart = PieChart()
    chart.title = "Budget Breakdown by Category"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    labels = Reference(ws, min_col=1, min_row=dr, max_row=end_data)
    data = Reference(ws, min_col=2, min_row=dr, max_row=end_data)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False

    # Color slices with palette colors rotating
    palette = [BLUSH, ROSE_GOLD, CHAMPAGNE, SAGE_GREEN, DUSTY_MAUVE,
               GOLD, DEEP_PLUM, "D4A5A5", "8FB996", "E6D5C3",
               "7D5A5A", "A3C1AD", "C9B1D0", "F0C987", "B5838D",
               "6D6875", "FFCDB2", "B5E48C", "E5989B", "FFB4A2",
               "CDB4DB", "A2D2FF", "BDE0FE"]
    for i in range(len(categories)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = palette[i % len(palette)]
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, f"A{tr + 2}")

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_guest_list(wb):
    ws = wb.create_sheet("Guest List & RSVPs")
    ws.sheet_properties.tabColor = BLUSH

    next_row = write_title_block(
        ws,
        "Guest List & RSVPs",
        "Track Every Guest  |  RSVPs  |  Dietary Needs  |  Thank You Notes",
        merge_end_col="N",
    )

    headers = [
        "Guest Name", "Email", "Phone", "Address",
        "Side", "Relationship", "Plus One", "Plus One Name",
        "Dietary Needs", "RSVP Status", "Table #",
        "Gift Received", "Thank You Sent", "Notes",
    ]

    set_col_widths(ws, {
        "A": 22, "B": 24, "C": 16, "D": 28,
        "E": 12, "F": 16, "G": 12, "H": 20,
        "I": 18, "J": 16, "K": 10,
        "L": 16, "M": 16, "N": 22,
    })

    hr = next_row
    write_header_row(ws, hr, headers, fill=FILL_DEEP_PLUM)

    dr_start = hr + 1
    dr_end = dr_start + 149  # 150 rows

    for ri in range(dr_start, dr_end + 1):
        style_data_row(ws, ri, len(headers), is_even=((ri - dr_start) % 2 == 0))
        ws.cell(row=ri, column=1).alignment = ALIGN_LEFT
        ws.cell(row=ri, column=4).alignment = ALIGN_LEFT
        ws.cell(row=ri, column=14).alignment = ALIGN_LEFT

    # Dropdowns
    add_dropdown(ws, ["Bride", "Groom", "Both"], f"E{dr_start}:E{dr_end}")
    add_dropdown(ws, ["Yes", "No"], f"G{dr_start}:G{dr_end}")
    add_dropdown(ws, ["Invited", "Attending", "Declined", "Pending"],
                 f"J{dr_start}:J{dr_end}")
    add_dropdown(ws, ["Yes", "No"], f"M{dr_start}:M{dr_end}")

    # Conditional formatting for RSVP status
    rng = f"J{dr_start}:J{dr_end}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Attending"'],
                   fill=PatternFill("solid", fgColor=SAGE_GREEN),
                   font=Font(color=WHITE, bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Declined"'],
                   fill=PatternFill("solid", fgColor=DUSTY_MAUVE),
                   font=Font(color=WHITE, bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Pending"'],
                   fill=PatternFill("solid", fgColor=GOLD),
                   font=Font(color=DEEP_PLUM, bold=True)),
    )

    # TOTALS
    tr = dr_end + 2
    ws.merge_cells(f"A{tr}:B{tr}")
    ws.cell(row=tr, column=1, value="RSVP SUMMARY").font = Font(
        name="Calibri", size=13, bold=True, color=WHITE
    )
    ws.cell(row=tr, column=1).fill = FILL_ROSE_GOLD
    ws.cell(row=tr, column=1).alignment = ALIGN_CENTER
    ws.cell(row=tr, column=2).fill = FILL_ROSE_GOLD
    tr += 1

    summaries = [
        ("Total Invited:", f"=COUNTA(A{dr_start}:A{dr_end})"),
        ("Attending:", f'=COUNTIF(J{dr_start}:J{dr_end},"Attending")'),
        ("Declined:", f'=COUNTIF(J{dr_start}:J{dr_end},"Declined")'),
        ("Pending:", f'=COUNTIF(J{dr_start}:J{dr_end},"Pending")'),
        ("Plus Ones:", f'=COUNTIF(G{dr_start}:G{dr_end},"Yes")'),
        ("Thank Yous Sent:", f'=COUNTIF(M{dr_start}:M{dr_end},"Yes")'),
    ]
    for si, (lbl, formula) in enumerate(summaries):
        ws.cell(row=tr, column=1, value=lbl).font = FONT_LABEL
        ws.cell(row=tr, column=1).alignment = ALIGN_RIGHT
        ws.cell(row=tr, column=1).border = THIN_BORDER
        ws.cell(row=tr, column=1).fill = FILL_IVORY if si % 2 == 0 else FILL_WHITE
        vc = ws.cell(row=tr, column=2, value=formula)
        vc.font = Font(name="Calibri", size=13, bold=True, color=DEEP_PLUM)
        vc.alignment = ALIGN_CENTER
        vc.border = THIN_BORDER
        vc.fill = FILL_IVORY if si % 2 == 0 else FILL_WHITE
        tr += 1

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_vendor_tracker(wb):
    ws = wb.create_sheet("Vendor Tracker")
    ws.sheet_properties.tabColor = SAGE_GREEN

    next_row = write_title_block(
        ws,
        "Vendor Tracker",
        "Contacts  |  Contracts  |  Payments  |  All Vendor Details",
        merge_end_col="O",
    )

    headers = [
        "Vendor Category", "Vendor Name", "Contact Name", "Phone",
        "Email", "Website", "Contract Signed", "Deposit Amount",
        "Deposit Paid", "Total Cost", "Payment Due Date",
        "Balance Due", "Payment Status", "Notes",
    ]

    set_col_widths(ws, {
        "A": 22, "B": 22, "C": 20, "D": 16,
        "E": 24, "F": 24, "G": 16, "H": 16,
        "I": 14, "J": 16, "K": 18,
        "L": 16, "M": 16, "N": 26, "O": 4,
    })

    hr = next_row
    write_header_row(ws, hr, headers, fill=FILL_DEEP_PLUM)

    vendor_cats = [
        "Venue", "Caterer", "Photographer", "Videographer",
        "Florist", "DJ / Band", "Wedding Planner", "Officiant",
        "Baker (Cake)", "Hair Stylist", "Makeup Artist",
        "Dress / Bridal Shop", "Tuxedo / Suit Rental", "Invitations / Stationer",
        "Transportation / Limo", "Rentals (Tables/Chairs)", "Lighting / Decor",
        "Photo Booth", "Jeweler (Rings)", "Hotel / Accommodations",
    ]

    dr = hr + 1
    for ri, cat in enumerate(vendor_cats):
        row_num = dr + ri
        ws.cell(row=row_num, column=1, value=cat)
        # Balance Due = Total Cost - Deposit Amount
        ws.cell(row=row_num, column=12,
                value=f"=IF(J{row_num}=\"\",\"\",J{row_num}-H{row_num})")
        style_data_row(ws, row_num, len(headers), is_even=(ri % 2 == 0))
        ws.cell(row=row_num, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row_num, column=14).alignment = ALIGN_LEFT

    end_data = dr + len(vendor_cats) - 1

    # Money formatting
    for mc in [8, 10, 12]:
        format_money_col(ws, mc, dr, end_data)

    # Date formatting
    format_date_col(ws, 11, dr, end_data)

    # Dropdowns
    add_dropdown(ws, ["Yes", "No"], f"G{dr}:G{end_data}")
    add_dropdown(ws, ["Yes", "No"], f"I{dr}:I{end_data}")
    add_dropdown(ws, ["Paid", "Partial", "Unpaid", "N/A"],
                 f"M{dr}:M{end_data}")

    # Conditional formatting for Payment Status
    rng = f"M{dr}:M{end_data}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Paid"'],
                   fill=PatternFill("solid", fgColor=SAGE_GREEN),
                   font=Font(color=WHITE, bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Unpaid"'],
                   fill=PatternFill("solid", fgColor="D9534F"),
                   font=Font(color=WHITE, bold=True)),
    )

    # TOTAL row
    tr = end_data + 1
    ws.cell(row=tr, column=1, value="TOTALS").font = Font(
        name="Calibri", size=12, bold=True, color=WHITE
    )
    for tc in range(1, len(headers) + 1):
        ws.cell(row=tr, column=tc).fill = FILL_ROSE_GOLD
        ws.cell(row=tr, column=tc).border = THIN_BORDER
        ws.cell(row=tr, column=tc).font = Font(
            name="Calibri", size=11, bold=True, color=WHITE
        )
    for mc in [8, 10, 12]:
        col_l = get_column_letter(mc)
        ws.cell(row=tr, column=mc,
                value=f"=SUM({col_l}{dr}:{col_l}{end_data})")
        ws.cell(row=tr, column=mc).number_format = MONEY_FMT

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_timeline_checklist(wb):
    ws = wb.create_sheet("Timeline & Checklist")
    ws.sheet_properties.tabColor = DUSTY_MAUVE

    next_row = write_title_block(
        ws,
        "Wedding Timeline & Master Checklist",
        "12 Months to Wedding Day  |  Every Task Organized by Timeframe",
        merge_end_col="H",
    )

    headers = [
        "Timeframe", "Task", "Category", "Assigned To",
        "Due Date", "Status", "Priority", "Notes",
    ]

    set_col_widths(ws, {
        "A": 20, "B": 40, "C": 18, "D": 18,
        "E": 16, "F": 16, "G": 14, "H": 28,
    })

    hr = next_row
    write_header_row(ws, hr, headers, fill=FILL_DEEP_PLUM)

    # Tasks organized by timeframe
    tasks_by_time = {
        "12 Months Before": [
            ("Set budget", "Budget", "Couple"),
            ("Choose wedding date", "Planning", "Couple"),
            ("Book venue", "Venue", "Couple"),
            ("Hire wedding planner", "Planning", "Couple"),
            ("Start guest list draft", "Guest List", "Couple"),
            ("Research photographers", "Photography", "Bride"),
            ("Begin dress shopping", "Attire", "Bride"),
            ("Book officiant", "Ceremony", "Couple"),
        ],
        "9 Months Before": [
            ("Book photographer", "Photography", "Couple"),
            ("Book videographer", "Photography", "Couple"),
            ("Book caterer / tasting", "Catering", "Couple"),
            ("Choose wedding party", "Planning", "Couple"),
            ("Book florist", "Flowers", "Bride"),
            ("Book DJ or band", "Music", "Couple"),
            ("Register for gifts", "Gifts", "Couple"),
            ("Book hotel room blocks", "Accommodations", "Couple"),
        ],
        "6 Months Before": [
            ("Order invitations", "Stationery", "Bride"),
            ("Book transportation", "Logistics", "Groom"),
            ("Plan honeymoon", "Honeymoon", "Couple"),
            ("Choose wedding cake / baker", "Catering", "Couple"),
            ("Shop for bridesmaids dresses", "Attire", "Bride"),
            ("Book hair & makeup", "Beauty", "Bride"),
            ("Order wedding rings", "Rings", "Couple"),
        ],
        "4 Months Before": [
            ("Send save-the-dates", "Stationery", "Bride"),
            ("Finalize ceremony details", "Ceremony", "Couple"),
            ("Plan rehearsal dinner", "Events", "Groom"),
            ("Book rentals (tables, chairs)", "Rentals", "Planner"),
            ("Choose readings & music", "Ceremony", "Couple"),
            ("Schedule dress fittings", "Attire", "Bride"),
        ],
        "2 Months Before": [
            ("Mail invitations", "Stationery", "Bride"),
            ("Finalize menu", "Catering", "Couple"),
            ("Order favors", "Favors", "Bride"),
            ("Plan seating chart", "Guest List", "Couple"),
            ("Buy gifts for wedding party", "Gifts", "Couple"),
            ("Confirm honeymoon bookings", "Honeymoon", "Couple"),
        ],
        "1 Month Before": [
            ("Final dress fitting", "Attire", "Bride"),
            ("Get marriage license", "Legal", "Couple"),
            ("Confirm all vendor details", "Vendors", "Planner"),
            ("Create day-of timeline", "Planning", "Planner"),
            ("Finalize seating chart", "Guest List", "Couple"),
            ("Break in wedding shoes", "Attire", "Bride"),
            ("Write vows", "Ceremony", "Couple"),
        ],
        "2 Weeks Before": [
            ("Confirm final guest count", "Guest List", "Couple"),
            ("Final vendor payments", "Budget", "Couple"),
            ("Prepare tips/envelopes for vendors", "Budget", "Couple"),
            ("Confirm rehearsal dinner details", "Events", "Groom"),
            ("Pack for honeymoon", "Honeymoon", "Couple"),
        ],
        "1 Week Before": [
            ("Confirm transportation", "Logistics", "Groom"),
            ("Prepare emergency kit", "Planning", "MOH"),
            ("Delegate day-of tasks", "Planning", "Planner"),
            ("Steam/press wedding attire", "Attire", "Bride"),
            ("Finalize playlist with DJ", "Music", "Couple"),
        ],
        "Day Before": [
            ("Rehearsal", "Ceremony", "Couple"),
            ("Rehearsal dinner", "Events", "Couple"),
            ("Deliver items to venue", "Logistics", "Planner"),
            ("Get a good night's sleep!", "Self-Care", "Couple"),
            ("Prepare day-of bag", "Planning", "Bride"),
        ],
        "Wedding Day": [
            ("Hair & makeup", "Beauty", "Bride"),
            ("Photographer arrival", "Photography", "Photographer"),
            ("Get dressed", "Attire", "Couple"),
            ("First look (optional)", "Photography", "Couple"),
            ("Ceremony", "Ceremony", "All"),
            ("Cocktail hour", "Reception", "Guests"),
            ("Reception & dancing", "Reception", "All"),
            ("Send-off / exit", "Reception", "All"),
        ],
    }

    dr = hr + 1
    row_idx = 0
    for timeframe, tasks in tasks_by_time.items():
        for task_name, category, assigned in tasks:
            r = dr + row_idx
            ws.cell(row=r, column=1, value=timeframe)
            ws.cell(row=r, column=2, value=task_name)
            ws.cell(row=r, column=3, value=category)
            ws.cell(row=r, column=4, value=assigned)
            style_data_row(ws, r, len(headers), is_even=(row_idx % 2 == 0))
            ws.cell(row=r, column=2).alignment = ALIGN_LEFT
            ws.cell(row=r, column=8).alignment = ALIGN_LEFT
            ws.cell(row=r, column=5).number_format = DATE_FMT
            row_idx += 1

    end_data = dr + row_idx - 1

    # Dropdowns
    add_dropdown(ws, ["Not Started", "In Progress", "Done"],
                 f"F{dr}:F{end_data}")
    add_dropdown(ws, ["Low", "Medium", "High"],
                 f"G{dr}:G{end_data}")

    # Conditional formatting
    apply_conditional_status(ws, "F", dr, end_data)

    # Priority conditional formatting
    rng_p = f"G{dr}:G{end_data}"
    ws.conditional_formatting.add(
        rng_p,
        CellIsRule(operator="equal", formula=['"High"'],
                   fill=PatternFill("solid", fgColor="D9534F"),
                   font=Font(color=WHITE, bold=True)),
    )
    ws.conditional_formatting.add(
        rng_p,
        CellIsRule(operator="equal", formula=['"Medium"'],
                   fill=PatternFill("solid", fgColor=GOLD),
                   font=Font(color=DEEP_PLUM, bold=True)),
    )
    ws.conditional_formatting.add(
        rng_p,
        CellIsRule(operator="equal", formula=['"Low"'],
                   fill=PatternFill("solid", fgColor=SAGE_GREEN),
                   font=Font(color=WHITE, bold=True)),
    )

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_seating_chart(wb):
    ws = wb.create_sheet("Seating Chart")
    ws.sheet_properties.tabColor = CHAMPAGNE

    next_row = write_title_block(
        ws,
        "Seating Chart",
        "Table Assignments  |  20 Tables x 8 Guests  |  Dietary & Special Needs",
        merge_end_col="L",
    )

    headers = [
        "Table #", "Table Name / Theme", "Capacity",
        "Guest 1", "Guest 2", "Guest 3", "Guest 4",
        "Guest 5", "Guest 6", "Guest 7", "Guest 8",
        "Dietary Notes", "Special Needs",
    ]

    # Adjusted - 13 cols
    set_col_widths(ws, {
        "A": 10, "B": 22, "C": 12,
        "D": 18, "E": 18, "F": 18, "G": 18,
        "H": 18, "I": 18, "J": 18, "K": 18,
        "L": 24, "M": 24,
    })

    hr = next_row
    write_header_row(ws, hr, headers, fill=FILL_DEEP_PLUM)

    dr = hr + 1
    for ti in range(20):
        r = dr + ti
        ws.cell(row=r, column=1, value=ti + 1)
        ws.cell(row=r, column=3, value=8)
        style_data_row(ws, r, len(headers), is_even=(ti % 2 == 0))
        ws.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.cell(row=r, column=12).alignment = ALIGN_LEFT
        ws.cell(row=r, column=13).alignment = ALIGN_LEFT
        for gc in range(4, 12):
            ws.cell(row=r, column=gc).alignment = ALIGN_LEFT

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_day_of_timeline(wb):
    ws = wb.create_sheet("Day-Of Timeline")
    ws.sheet_properties.tabColor = ROSE_GOLD

    next_row = write_title_block(
        ws,
        "Day-Of Timeline",
        "Minute-by-Minute Wedding Day Schedule  |  8 AM to 11 PM",
        merge_end_col="F",
    )

    headers = [
        "Time", "Event / Activity", "Location",
        "Person Responsible", "Notes / Details", "Status",
    ]

    set_col_widths(ws, {
        "A": 14, "B": 34, "C": 22,
        "D": 22, "E": 32, "F": 16,
    })

    hr = next_row
    write_header_row(ws, hr, headers, fill=FILL_DEEP_PLUM)

    timeline_events = [
        ("8:00 AM", "Bridal party hair & makeup begins", "Bridal Suite", "Hair/Makeup Team", "Allow 3-4 hours for full party"),
        ("8:30 AM", "Breakfast for bridal party", "Bridal Suite", "MOH / Coordinator", "Light, non-messy foods"),
        ("9:00 AM", "Groom & groomsmen breakfast", "Groom's Room", "Best Man", "Keep it casual and fun"),
        ("10:00 AM", "Photographer arrives", "Bridal Suite", "Photographer", "Detail shots: dress, rings, shoes, invites"),
        ("10:30 AM", "Videographer arrives", "Venue", "Videographer", "Establish shots of venue"),
        ("11:00 AM", "Bride gets dressed", "Bridal Suite", "MOH / Mother", "Schedule 30 min for photos"),
        ("11:30 AM", "Groom gets dressed", "Groom's Room", "Best Man", "Boutonniere pinning"),
        ("12:00 PM", "First Look (optional)", "Garden / Venue", "Photographer", "Private moment, 15-20 min"),
        ("12:30 PM", "Wedding party photos", "Venue Grounds", "Photographer", "Group and individual shots"),
        ("1:00 PM", "Family formal photos", "Venue", "Photographer", "Pre-arrange family groupings"),
        ("1:30 PM", "Bridal party hidden / guests arrive", "Ceremony Area", "Coordinator", "Ushers seat guests"),
        ("2:00 PM", "Musicians / DJ begin prelude", "Ceremony Area", "DJ / Musicians", "30 min of ambient music"),
        ("2:25 PM", "Seat immediate family", "Ceremony Area", "Ushers", "Mothers seated last"),
        ("2:30 PM", "Ceremony begins", "Ceremony Area", "Officiant", "Processional order confirmed"),
        ("2:35 PM", "Wedding party processional", "Ceremony Area", "Coordinator", "Cue music changes"),
        ("2:40 PM", "Bride's entrance", "Ceremony Area", "Father / Escort", "Cue bridal march"),
        ("2:45 PM", "Readings & vows", "Ceremony Area", "Officiant", "Mic check beforehand"),
        ("3:00 PM", "Ring exchange & kiss", "Ceremony Area", "Best Man / Ring Bearer", "Have backup ring holder"),
        ("3:05 PM", "Recessional", "Ceremony Area", "All", "Upbeat exit music"),
        ("3:15 PM", "Receiving line (optional)", "Ceremony Exit", "Couple & Parents", "Keep it moving, 2-3 min each"),
        ("3:30 PM", "Cocktail hour begins", "Cocktail Area", "Caterer / Bar", "Passed apps + signature cocktails"),
        ("3:30 PM", "Couple photos (golden hour)", "Venue Grounds", "Photographer", "30-45 min, best light"),
        ("4:30 PM", "Guests move to reception", "Reception Hall", "Coordinator / DJ", "Announce transition"),
        ("4:45 PM", "Grand entrance", "Reception Hall", "DJ", "Introduce wedding party + couple"),
        ("5:00 PM", "First dance", "Dance Floor", "DJ / Band", "Song cued and tested"),
        ("5:10 PM", "Parent dances", "Dance Floor", "DJ", "Father-daughter, mother-son"),
        ("5:20 PM", "Welcome toast", "Reception Hall", "Couple / Host", "Keep it short and sweet"),
        ("5:30 PM", "Dinner service begins", "Reception Hall", "Caterer", "Table-by-table service"),
        ("6:00 PM", "Best Man toast", "Reception Hall", "Best Man", "Mic provided"),
        ("6:10 PM", "Maid of Honor toast", "Reception Hall", "MOH", "Coordinate order with DJ"),
        ("6:20 PM", "Additional toasts (optional)", "Reception Hall", "Family / Friends", "Limit to 2-3 min each"),
        ("6:30 PM", "Cake cutting", "Cake Table", "Couple", "Photographer ready"),
        ("6:45 PM", "Bouquet toss", "Dance Floor", "Bride", "DJ announces"),
        ("6:50 PM", "Garter toss (optional)", "Dance Floor", "Groom", "Fun music cue"),
        ("7:00 PM", "Open dancing begins", "Dance Floor", "DJ / Band", "Mix of genres and eras"),
        ("8:00 PM", "Special dances / group dances", "Dance Floor", "DJ", "Anniversary dance, hora, etc."),
        ("9:00 PM", "Late-night snacks served", "Food Station", "Caterer", "Pizza, sliders, dessert bar"),
        ("9:30 PM", "Last call for bar", "Bar", "Bartender", "30-min warning"),
        ("10:00 PM", "Last dance", "Dance Floor", "DJ / Couple", "Meaningful slow song"),
        ("10:15 PM", "Sparkler exit / send-off", "Venue Exit", "Coordinator", "Safety briefing for sparklers"),
        ("10:30 PM", "Couple departs", "Venue Exit", "Transportation", "Getaway car ready"),
        ("11:00 PM", "Venue cleanup / breakdown", "Venue", "Coordinator / Rentals", "Collect personal items & gifts"),
    ]

    dr = hr + 1
    for ri, (time, event, location, person, notes) in enumerate(timeline_events):
        r = dr + ri
        ws.cell(row=r, column=1, value=time)
        ws.cell(row=r, column=2, value=event)
        ws.cell(row=r, column=3, value=location)
        ws.cell(row=r, column=4, value=person)
        ws.cell(row=r, column=5, value=notes)
        style_data_row(ws, r, len(headers), is_even=(ri % 2 == 0))
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color=DEEP_PLUM)
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.cell(row=r, column=5).alignment = ALIGN_LEFT

    end_data = dr + len(timeline_events) - 1

    add_dropdown(ws, ["Pending", "In Progress", "Done"],
                 f"F{dr}:F{end_data}")

    apply_conditional_status(ws, "F", dr, end_data)

    # Also make "Pending" have conditional formatting
    rng = f"F{dr}:F{end_data}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Pending"'],
                   fill=PatternFill("solid", fgColor=DUSTY_MAUVE),
                   font=Font(color=WHITE, bold=True)),
    )

    ws.sheet_view.showGridLines = False


# ───────────────────────────────────────────────────────────────────────────

def build_gift_tracker(wb):
    ws = wb.create_sheet("Gift & Thank You Tracker")
    ws.sheet_properties.tabColor = GOLD

    next_row = write_title_block(
        ws,
        "Gift & Thank You Tracker",
        "Log Every Gift  |  Track Thank You Notes  |  Never Miss a Thank You",
        merge_end_col="G",
    )

    headers = [
        "Guest Name", "Gift Description", "Estimated Value",
        "Date Received", "Thank You Written", "Date Sent", "Notes",
    ]

    set_col_widths(ws, {
        "A": 24, "B": 32, "C": 18,
        "D": 16, "E": 18, "F": 16, "G": 28,
    })

    hr = next_row
    write_header_row(ws, hr, headers, fill=FILL_DEEP_PLUM)

    dr = hr + 1
    dr_end = dr + 99  # 100 rows

    for ri in range(dr, dr_end + 1):
        style_data_row(ws, ri, len(headers), is_even=((ri - dr) % 2 == 0))
        ws.cell(row=ri, column=1).alignment = ALIGN_LEFT
        ws.cell(row=ri, column=2).alignment = ALIGN_LEFT
        ws.cell(row=ri, column=7).alignment = ALIGN_LEFT

    format_money_col(ws, 3, dr, dr_end)
    format_date_col(ws, 4, dr, dr_end)
    format_date_col(ws, 6, dr, dr_end)

    add_dropdown(ws, ["Yes", "No"], f"E{dr}:E{dr_end}")

    # Conditional formatting for Thank You Written
    rng = f"E{dr}:E{dr_end}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill("solid", fgColor=SAGE_GREEN),
                   font=Font(color=WHITE, bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"No"'],
                   fill=PatternFill("solid", fgColor=DUSTY_MAUVE),
                   font=Font(color=WHITE, bold=True)),
    )

    # Summary
    tr = dr_end + 2
    ws.merge_cells(f"A{tr}:B{tr}")
    ws.cell(row=tr, column=1, value="GIFT SUMMARY").font = Font(
        name="Calibri", size=13, bold=True, color=WHITE
    )
    ws.cell(row=tr, column=1).fill = FILL_ROSE_GOLD
    ws.cell(row=tr, column=1).alignment = ALIGN_CENTER
    ws.cell(row=tr, column=2).fill = FILL_ROSE_GOLD
    tr += 1

    summaries = [
        ("Total Gifts Received:", f"=COUNTA(A{dr}:A{dr_end})"),
        ("Total Estimated Value:", f"=SUM(C{dr}:C{dr_end})"),
        ("Thank Yous Sent:", f'=COUNTIF(E{dr}:E{dr_end},"Yes")'),
        ("Thank Yous Remaining:", f'=COUNTIF(E{dr}:E{dr_end},"No")'),
    ]
    for si, (lbl, formula) in enumerate(summaries):
        ws.cell(row=tr, column=1, value=lbl).font = FONT_LABEL
        ws.cell(row=tr, column=1).alignment = ALIGN_RIGHT
        ws.cell(row=tr, column=1).border = THIN_BORDER
        ws.cell(row=tr, column=1).fill = FILL_IVORY if si % 2 == 0 else FILL_WHITE
        vc = ws.cell(row=tr, column=2, value=formula)
        vc.font = Font(name="Calibri", size=13, bold=True, color=DEEP_PLUM)
        vc.alignment = ALIGN_CENTER
        vc.border = THIN_BORDER
        vc.fill = FILL_IVORY if si % 2 == 0 else FILL_WHITE
        if si == 1:
            vc.number_format = MONEY_FMT
        tr += 1

    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    # Build all sheets
    build_how_to_use(wb)        # Sheet 1 (uses wb.active)
    build_dashboard(wb)         # Sheet 2
    build_budget_tracker(wb)    # Sheet 3
    build_guest_list(wb)        # Sheet 4
    build_vendor_tracker(wb)    # Sheet 5
    build_timeline_checklist(wb)  # Sheet 6
    build_seating_chart(wb)     # Sheet 7
    build_day_of_timeline(wb)   # Sheet 8
    build_gift_tracker(wb)      # Sheet 9

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "Wedding_Planner_Bundle_2026.xlsx")
    wb.save(output_path)

    print(f"Workbook saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Total sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
