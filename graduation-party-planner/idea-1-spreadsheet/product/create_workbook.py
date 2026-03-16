#!/usr/bin/env python3
"""
Graduation Party Planner — Excel Workbook Generator
Creates a polished, structured party planning tool with:
  - Dashboard (summary cards)
  - Guest List (RSVP tracking)
  - Budget Tracker (estimated vs actual)
  - Decoration Checklist
  - Food Planner
  - Party Timeline
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── COLOR SYSTEM ──────────────────────────────────────────────────────────────
COLORS = {
    "primary":        "fb5887",   # Pink — headers, accents
    "primary_light":  "fde8ef",   # Light pink — alternating rows
    "secondary":      "3ca4d7",   # Blue — section highlights
    "secondary_light":"dff0f9",   # Light blue — alt rows in some tabs
    "highlight":      "fe8c43",   # Orange — callouts
    "highlight_light":"fff0e6",   # Light orange
    "success":        "27AE60",   # Green — completed / yes
    "success_light":  "D5F5E3",   # Light green
    "warning":        "F39C12",   # Amber — maybe / pending
    "warning_light":  "FEF9E7",   # Light amber
    "danger":         "E74C3C",   # Red — no / alert
    "danger_light":   "FDEDEC",   # Light red
    "header_bg":      "2C3E50",   # Dark slate — main headers
    "header_text":    "FFFFFF",   # White text on headers
    "sub_header_bg":  "34495E",   # Slightly lighter slate
    "body_bg":        "FEFEFE",   # Off-white background
    "input_bg":       "FFFFF0",   # Ivory — editable input cells
    "border":         "BDC3C7",   # Soft border
    "white":          "FFFFFF",
    "dark_text":      "2C3E50",
    "gray_text":      "7F8C8D",
}

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def header_font(size=12, bold=True, color=None):
    return Font(name="Calibri", size=size, bold=bold, color=color or COLORS["header_text"])

def body_font(size=11, bold=False, color=None):
    return Font(name="Calibri", size=size, bold=bold, color=color or COLORS["dark_text"])

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False, indent=1):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap, indent=indent)

def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:          cell.font = font
    if fill:          cell.fill = fill
    if alignment:     cell.alignment = alignment
    if border:        cell.border = border
    if number_format: cell.number_format = number_format

def style_header_row(ws, row, col_start, col_end, text, fill_color=None, font_size=13):
    """Style a merged header row."""
    fill_color = fill_color or COLORS["header_bg"]
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    apply_style(cell,
                font=header_font(size=font_size),
                fill=make_fill(fill_color),
                alignment=center_align())

def style_col_headers(ws, row, headers, start_col=1, fill_color=None):
    """Apply column header styling to a row."""
    fill_color = fill_color or COLORS["primary"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        apply_style(cell,
                    font=header_font(size=11),
                    fill=make_fill(fill_color),
                    alignment=center_align(wrap=True),
                    border=thin_border)

def style_data_row(ws, row, num_cols, start_col=1, alt=False, fill_color=None):
    """Apply alternating row styling."""
    if fill_color:
        bg = fill_color
    else:
        bg = COLORS["primary_light"] if alt else COLORS["white"]
    for col in range(start_col, start_col + num_cols):
        cell = ws.cell(row=row, column=col)
        apply_style(cell,
                    font=body_font(),
                    fill=make_fill(bg),
                    alignment=left_align(),
                    border=thin_border)

def set_col_width(ws, col_widths):
    """Set column widths. col_widths is a dict {col_letter: width}."""
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width


# ── TAB 1: DASHBOARD ──────────────────────────────────────────────────────────
def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 15

    # Title banner
    ws.merge_cells("A2:H2")
    title_cell = ws["A2"]
    title_cell.value = "🎓  GRADUATION PARTY PLANNER"
    apply_style(title_cell,
                font=Font(name="Calibri", size=22, bold=True, color=COLORS["header_text"]),
                fill=make_fill(COLORS["header_bg"]),
                alignment=center_align())
    ws.row_dimensions[2].height = 42

    # Subtitle
    ws.merge_cells("A3:H3")
    sub = ws["A3"]
    sub.value = "Everything you need to plan the perfect graduation party"
    apply_style(sub,
                font=Font(name="Calibri", size=12, italic=True, color=COLORS["gray_text"]),
                fill=make_fill(COLORS["white"]),
                alignment=center_align())
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 12  # spacer

    # Party info input section
    style_header_row(ws, 5, 1, 8, "PARTY DETAILS", fill_color=COLORS["primary"], font_size=12)
    ws.row_dimensions[5].height = 28

    info_labels = [
        ("Graduate Name:", "A6", "B6"),
        ("School / Program:", "A7", "B7"),
        ("Party Date:", "D6", "E6"),
        ("Party Venue:", "D7", "E7"),
        ("Start Time:", "A8", "B8"),
        ("End Time:", "D8", "E8"),
    ]
    for label, lbl_cell, val_cell in info_labels:
        lc = ws[lbl_cell]
        lc.value = label
        apply_style(lc, font=body_font(bold=True), fill=make_fill(COLORS["secondary_light"]),
                    alignment=left_align(), border=thin_border)
        vc = ws[val_cell]
        apply_style(vc, font=body_font(), fill=make_fill(COLORS["input_bg"]),
                    alignment=left_align(), border=thin_border)
        ws.row_dimensions[int(lbl_cell[1])].height = 22

    ws.row_dimensions[9].height = 12  # spacer

    # Summary cards row label
    style_header_row(ws, 10, 1, 8, "PARTY SUMMARY", fill_color=COLORS["secondary"], font_size=12)
    ws.row_dimensions[10].height = 28

    # Card definitions: (label, formula_or_value, col_start, fill_color)
    cards = [
        ("TOTAL GUESTS", "=COUNTA('Guest List'!A3:A102)", 1, COLORS["primary"]),
        ("CONFIRMED", "=COUNTIF('Guest List'!B3:B102,\"Yes\")", 3, COLORS["success"]),
        ("TOTAL BUDGET", "=SUM('Budget Tracker'!C3:C52)", 5, COLORS["highlight"]),
        ("AMOUNT SPENT", "=SUM('Budget Tracker'!D3:D52)", 7, COLORS["secondary"]),
    ]

    for label, formula, col, color in cards:
        # Label cell
        lbl = ws.cell(row=11, column=col, value=label)
        apply_style(lbl, font=header_font(size=10), fill=make_fill(color),
                    alignment=center_align(), border=thin_border)
        ws.merge_cells(start_row=11, start_column=col, end_row=11, end_column=col + 1)

        # Value cell
        val = ws.cell(row=12, column=col, value=formula)
        apply_style(val, font=Font(name="Calibri", size=20, bold=True, color=color),
                    fill=make_fill(COLORS["white"]),
                    alignment=center_align(), border=thin_border)
        ws.merge_cells(start_row=12, start_column=col, end_row=12, end_column=col + 1)
        if "BUDGET" in label or "SPENT" in label:
            val.number_format = '"$"#,##0.00'
        ws.row_dimensions[11].height = 28
        ws.row_dimensions[12].height = 48

    ws.row_dimensions[13].height = 12  # spacer

    # Quick links / navigation guide
    style_header_row(ws, 14, 1, 8, "QUICK NAVIGATION", fill_color=COLORS["header_bg"], font_size=12)
    ws.row_dimensions[14].height = 28

    nav_items = [
        ("Guest List", "Track RSVPs, meal choices, and gifts"),
        ("Budget Tracker", "Monitor estimated vs actual costs"),
        ("Decoration Checklist", "Check off items as you purchase them"),
        ("Food Planner", "Plan your menu and who's bringing what"),
        ("Party Timeline", "Keep tasks on schedule leading up to the party"),
    ]
    for i, (tab, desc) in enumerate(nav_items):
        row = 15 + i
        tab_cell = ws.cell(row=row, column=1, value=tab)
        apply_style(tab_cell,
                    font=Font(name="Calibri", size=11, bold=True, color=COLORS["primary"]),
                    fill=make_fill(COLORS["primary_light"]),
                    alignment=left_align(), border=thin_border)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        desc_cell = ws.cell(row=row, column=3, value=desc)
        apply_style(desc_cell,
                    font=body_font(color=COLORS["gray_text"]),
                    fill=make_fill(COLORS["white"]),
                    alignment=left_align(), border=thin_border)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 22

    # Column widths
    set_col_width(ws, {
        "A": 20, "B": 22, "C": 4, "D": 20, "E": 22, "F": 4, "G": 20, "H": 22,
    })


# ── TAB 2: GUEST LIST ─────────────────────────────────────────────────────────
def create_guest_list(wb):
    ws = wb.create_sheet("Guest List")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    style_header_row(ws, 1, 1, 6, "GUEST LIST", fill_color=COLORS["primary"], font_size=14)
    ws.row_dimensions[1].height = 36

    headers = ["Guest Name", "RSVP", "Meal Choice", "Gift", "Notes"]
    style_col_headers(ws, 2, headers, fill_color=COLORS["header_bg"])
    ws.row_dimensions[2].height = 24

    # RSVP dropdown
    rsvp_dv = DataValidation(type="list", formula1='"Yes,No,Maybe,Pending"', allow_blank=True)
    rsvp_dv.sqref = "B3:B102"
    ws.add_data_validation(rsvp_dv)

    # Meal choice dropdown
    meal_dv = DataValidation(type="list", formula1='"Standard,Vegetarian,Vegan,Gluten-Free,Kids Meal"', allow_blank=True)
    meal_dv.sqref = "C3:C102"
    ws.add_data_validation(meal_dv)

    for i in range(100):
        row = 3 + i
        alt = i % 2 == 1
        style_data_row(ws, row, 5, alt=alt)
        # Input bg for editable columns
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, fill=make_fill(COLORS["input_bg"] if not alt else COLORS["primary_light"]))
        ws.row_dimensions[row].height = 20

    # Conditional formatting for RSVP
    green_fill = make_fill(COLORS["success_light"])
    red_fill = make_fill(COLORS["danger_light"])
    yellow_fill = make_fill(COLORS["warning_light"])

    ws.conditional_formatting.add("B3:B102",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=green_fill,
                   font=Font(color=COLORS["success"], bold=True)))
    ws.conditional_formatting.add("B3:B102",
        CellIsRule(operator="equal", formula=['"No"'], fill=red_fill,
                   font=Font(color=COLORS["danger"], bold=True)))
    ws.conditional_formatting.add("B3:B102",
        CellIsRule(operator="equal", formula=['"Maybe"'], fill=yellow_fill,
                   font=Font(color=COLORS["warning"], bold=True)))

    # Summary row
    summary_row = 103
    ws.row_dimensions[summary_row].height = 24
    ws.merge_cells(f"A{summary_row}:A{summary_row}")
    total_label = ws.cell(row=summary_row, column=1, value="TOTAL GUESTS:")
    apply_style(total_label, font=header_font(size=11, color=COLORS["dark_text"]),
                fill=make_fill(COLORS["secondary_light"]), alignment=left_align(), border=thin_border)
    total_val = ws.cell(row=summary_row, column=2, value="=COUNTA(A3:A102)")
    apply_style(total_val, font=header_font(size=11, color=COLORS["primary"]),
                fill=make_fill(COLORS["secondary_light"]), alignment=center_align(), border=thin_border)

    confirmed_label = ws.cell(row=summary_row, column=3, value="CONFIRMED:")
    apply_style(confirmed_label, font=header_font(size=11, color=COLORS["dark_text"]),
                fill=make_fill(COLORS["success_light"]), alignment=left_align(), border=thin_border)
    confirmed_val = ws.cell(row=summary_row, column=4, value='=COUNTIF(B3:B102,"Yes")')
    apply_style(confirmed_val, font=header_font(size=11, color=COLORS["success"]),
                fill=make_fill(COLORS["success_light"]), alignment=center_align(), border=thin_border)

    set_col_width(ws, {"A": 30, "B": 14, "C": 18, "D": 20, "E": 35})


# ── TAB 3: BUDGET TRACKER ─────────────────────────────────────────────────────
def create_budget_tracker(wb):
    ws = wb.create_sheet("Budget Tracker")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    style_header_row(ws, 1, 1, 6, "BUDGET TRACKER", fill_color=COLORS["highlight"], font_size=14)
    ws.row_dimensions[1].height = 36

    headers = ["Item", "Category", "Estimated Cost", "Actual Cost", "Paid", "Notes"]
    style_col_headers(ws, 2, headers, fill_color=COLORS["header_bg"])
    ws.row_dimensions[2].height = 24

    # Category dropdown
    cat_dv = DataValidation(
        type="list",
        formula1='"Venue,Catering,Decorations,Invitations,Entertainment,Photography,Favors,Cake & Desserts,Beverages,Flowers,Transportation,Other"',
        allow_blank=True
    )
    cat_dv.sqref = "B3:B52"
    ws.add_data_validation(cat_dv)

    # Paid dropdown
    paid_dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    paid_dv.sqref = "E3:E52"
    ws.add_data_validation(paid_dv)

    for i in range(50):
        row = 3 + i
        alt = i % 2 == 1
        style_data_row(ws, row, 6, alt=alt)
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            bg = COLORS["primary_light"] if alt else COLORS["input_bg"]
            apply_style(cell, fill=make_fill(bg))
        # Currency format
        for col in [3, 4]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 20

    # Conditional formatting for Paid
    ws.conditional_formatting.add("E3:E52",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=make_fill(COLORS["success_light"]),
                   font=Font(color=COLORS["success"], bold=True)))
    ws.conditional_formatting.add("E3:E52",
        CellIsRule(operator="equal", formula=['"No"'], fill=make_fill(COLORS["danger_light"]),
                   font=Font(color=COLORS["danger"], bold=True)))
    ws.conditional_formatting.add("E3:E52",
        CellIsRule(operator="equal", formula=['"Partial"'], fill=make_fill(COLORS["warning_light"]),
                   font=Font(color=COLORS["warning"], bold=True)))

    # Totals row
    totals_row = 53
    ws.row_dimensions[totals_row].height = 28
    total_label = ws.cell(row=totals_row, column=1, value="TOTALS")
    apply_style(total_label, font=header_font(size=12, color=COLORS["header_text"]),
                fill=make_fill(COLORS["header_bg"]), alignment=center_align(), border=thin_border)
    ws.merge_cells(f"A{totals_row}:B{totals_row}")

    estimated_total = ws.cell(row=totals_row, column=3, value="=SUM(C3:C52)")
    apply_style(estimated_total, font=header_font(size=12, color=COLORS["highlight"]),
                fill=make_fill(COLORS["highlight_light"]), alignment=center_align(), border=thin_border)
    estimated_total.number_format = '"$"#,##0.00'

    actual_total = ws.cell(row=totals_row, column=4, value="=SUM(D3:D52)")
    apply_style(actual_total, font=header_font(size=12, color=COLORS["primary"]),
                fill=make_fill(COLORS["primary_light"]), alignment=center_align(), border=thin_border)
    actual_total.number_format = '"$"#,##0.00'

    remaining_label = ws.cell(row=totals_row, column=5, value="REMAINING:")
    apply_style(remaining_label, font=header_font(size=11, color=COLORS["dark_text"]),
                fill=make_fill(COLORS["secondary_light"]), alignment=center_align(), border=thin_border)
    remaining_val = ws.cell(row=totals_row, column=6, value=f"=C{totals_row}-D{totals_row}")
    apply_style(remaining_val, font=header_font(size=12, color=COLORS["secondary"]),
                fill=make_fill(COLORS["secondary_light"]), alignment=center_align(), border=thin_border)
    remaining_val.number_format = '"$"#,##0.00'

    set_col_width(ws, {"A": 30, "B": 18, "C": 16, "D": 16, "E": 12, "F": 30})


# ── TAB 4: DECORATION CHECKLIST ──────────────────────────────────────────────
def create_decoration_checklist(wb):
    ws = wb.create_sheet("Decoration Checklist")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    style_header_row(ws, 1, 1, 5, "DECORATION CHECKLIST", fill_color=COLORS["secondary"], font_size=14)
    ws.row_dimensions[1].height = 36

    headers = ["Decoration Item", "Quantity", "Purchased", "Store / Source", "Notes"]
    style_col_headers(ws, 2, headers, fill_color=COLORS["header_bg"])
    ws.row_dimensions[2].height = 24

    purchased_dv = DataValidation(type="list", formula1='"Yes,No,Ordered"', allow_blank=True)
    purchased_dv.sqref = "C3:C62"
    ws.add_data_validation(purchased_dv)

    # Pre-fill common decoration items
    default_items = [
        "Graduation banners & signs", "Balloons (school colors)", "Table centerpieces",
        "Photo booth backdrop", "Tablecloths", "Plates & napkins (themed)",
        "Cups & straws", "Cutlery", "Guest book / memory display",
        "Photo display (grad photos)", "Fairy lights / string lights",
        "Graduation cap decorations", "Confetti", "Welcome sign",
        "Flower arrangements",
    ]

    for i in range(60):
        row = 3 + i
        alt = i % 2 == 1
        style_data_row(ws, row, 5, alt=alt)
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            bg = COLORS["secondary_light"] if alt else COLORS["input_bg"]
            apply_style(cell, fill=make_fill(bg))
        if i < len(default_items):
            ws.cell(row=row, column=1).value = default_items[i]
        ws.row_dimensions[row].height = 20

    ws.conditional_formatting.add("C3:C62",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=make_fill(COLORS["success_light"]),
                   font=Font(color=COLORS["success"], bold=True)))
    ws.conditional_formatting.add("C3:C62",
        CellIsRule(operator="equal", formula=['"No"'], fill=make_fill(COLORS["danger_light"]),
                   font=Font(color=COLORS["danger"], bold=True)))

    # Progress summary
    summary_row = 63
    ws.row_dimensions[summary_row].height = 24
    prog_label = ws.cell(row=summary_row, column=1, value="PURCHASED:")
    apply_style(prog_label, font=header_font(size=11, color=COLORS["dark_text"]),
                fill=make_fill(COLORS["success_light"]), alignment=left_align(), border=thin_border)
    prog_val = ws.cell(row=summary_row, column=2, value='=COUNTIF(C3:C62,"Yes")&" of "&COUNTA(A3:A62)')
    apply_style(prog_val, font=header_font(size=11, color=COLORS["success"]),
                fill=make_fill(COLORS["success_light"]), alignment=center_align(), border=thin_border)
    ws.merge_cells(f"B{summary_row}:E{summary_row}")

    set_col_width(ws, {"A": 32, "B": 12, "C": 14, "D": 22, "E": 30})


# ── TAB 5: FOOD PLANNER ──────────────────────────────────────────────────────
def create_food_planner(wb):
    ws = wb.create_sheet("Food Planner")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    style_header_row(ws, 1, 1, 6, "FOOD PLANNER", fill_color=COLORS["highlight"], font_size=14)
    ws.row_dimensions[1].height = 36

    headers = ["Food Item", "Serves", "Dietary Notes", "Who's Bringing", "Done", "Notes"]
    style_col_headers(ws, 2, headers, fill_color=COLORS["header_bg"])
    ws.row_dimensions[2].height = 24

    course_dv = DataValidation(
        type="list",
        formula1='"Appetizer,Main Course,Side Dish,Salad,Dessert,Cake,Beverages,Snacks,Other"',
        allow_blank=True
    )
    course_dv.sqref = "C3:C52"
    ws.add_data_validation(course_dv)

    done_dv = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=True)
    done_dv.sqref = "E3:E52"
    ws.add_data_validation(done_dv)

    for i in range(50):
        row = 3 + i
        alt = i % 2 == 1
        style_data_row(ws, row, 6, alt=alt)
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            bg = COLORS["highlight_light"] if alt else COLORS["input_bg"]
            apply_style(cell, fill=make_fill(bg))
        ws.row_dimensions[row].height = 20

    ws.conditional_formatting.add("E3:E52",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=make_fill(COLORS["success_light"]),
                   font=Font(color=COLORS["success"], bold=True)))
    ws.conditional_formatting.add("E3:E52",
        CellIsRule(operator="equal", formula=['"No"'], fill=make_fill(COLORS["danger_light"]),
                   font=Font(color=COLORS["danger"], bold=True)))

    # Total servings
    summary_row = 53
    ws.row_dimensions[summary_row].height = 24
    ts_label = ws.cell(row=summary_row, column=1, value="TOTAL ITEMS PLANNED:")
    apply_style(ts_label, font=header_font(size=11, color=COLORS["dark_text"]),
                fill=make_fill(COLORS["highlight_light"]), alignment=left_align(), border=thin_border)
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ts_val = ws.cell(row=summary_row, column=4, value="=COUNTA(A3:A52)")
    apply_style(ts_val, font=header_font(size=12, color=COLORS["highlight"]),
                fill=make_fill(COLORS["highlight_light"]), alignment=center_align(), border=thin_border)
    ws.merge_cells(f"D{summary_row}:F{summary_row}")

    set_col_width(ws, {"A": 30, "B": 10, "C": 18, "D": 22, "E": 14, "F": 28})


# ── TAB 6: PARTY TIMELINE ────────────────────────────────────────────────────
def create_party_timeline(wb):
    ws = wb.create_sheet("Party Timeline")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    style_header_row(ws, 1, 1, 5, "PARTY TIMELINE", fill_color=COLORS["secondary"], font_size=14)
    ws.row_dimensions[1].height = 36

    headers = ["Time / Week", "Task", "Assigned To", "Status", "Notes"]
    style_col_headers(ws, 2, headers, fill_color=COLORS["header_bg"])
    ws.row_dimensions[2].height = 24

    status_dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Done,Skipped"',
        allow_blank=True
    )
    status_dv.sqref = "D3:D72"
    ws.add_data_validation(status_dv)

    # Pre-fill timeline items grouped by timeframe
    timeline_items = [
        # 4 Weeks Before
        ("4 Weeks Before", "Choose and book the venue", "Host", "Not Started", ""),
        ("4 Weeks Before", "Set guest list and send invitations", "Host", "Not Started", ""),
        ("4 Weeks Before", "Book catering or plan the menu", "Host", "Not Started", ""),
        ("4 Weeks Before", "Order graduation cake", "Host", "Not Started", ""),
        ("4 Weeks Before", "Plan decorations theme", "Host", "Not Started", ""),
        # 3 Weeks Before
        ("3 Weeks Before", "Purchase decorations", "Host", "Not Started", ""),
        ("3 Weeks Before", "Arrange entertainment / music playlist", "Host", "Not Started", ""),
        ("3 Weeks Before", "Confirm RSVPs", "Host", "Not Started", ""),
        ("3 Weeks Before", "Arrange accommodations if needed", "Host", "Not Started", ""),
        # 2 Weeks Before
        ("2 Weeks Before", "Finalize headcount for catering", "Host", "Not Started", ""),
        ("2 Weeks Before", "Prepare party favors", "Host", "Not Started", ""),
        ("2 Weeks Before", "Create photo display", "Host", "Not Started", ""),
        ("2 Weeks Before", "Set up photo booth area", "Host", "Not Started", ""),
        # 1 Week Before
        ("1 Week Before", "Confirm all vendor bookings", "Host", "Not Started", ""),
        ("1 Week Before", "Prepare graduation slideshow", "Host", "Not Started", ""),
        ("1 Week Before", "Buy any remaining groceries / drinks", "Host", "Not Started", ""),
        ("1 Week Before", "Charge cameras and prepare tech", "Host", "Not Started", ""),
        # Day Before
        ("Day Before", "Set up venue decorations", "Host", "Not Started", ""),
        ("Day Before", "Prepare food items that can be made ahead", "Host", "Not Started", ""),
        ("Day Before", "Chill beverages", "Host", "Not Started", ""),
        ("Day Before", "Lay out serving dishes and utensils", "Host", "Not Started", ""),
        # Day Of
        ("Day Of — Morning", "Final venue setup and decorations", "Host", "Not Started", ""),
        ("Day Of — Morning", "Prepare remaining food items", "Host", "Not Started", ""),
        ("Day Of — 1hr Before", "Final walk-through of venue", "Host", "Not Started", ""),
        ("Day Of — 1hr Before", "Set out guest book and pens", "Host", "Not Started", ""),
        ("Party Start", "Welcome guests as they arrive", "Host", "Not Started", ""),
        ("Party End", "Thank guests and hand out favors", "Host", "Not Started", ""),
    ]

    for i, (time, task, assigned, status, notes) in enumerate(timeline_items):
        row = 3 + i
        alt = i % 2 == 1
        bg = COLORS["secondary_light"] if alt else COLORS["input_bg"]
        data = [time, task, assigned, status, notes]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, font=body_font(), fill=make_fill(bg),
                        alignment=left_align(wrap=True), border=thin_border)
        ws.row_dimensions[row].height = 22

    # Remaining blank rows
    for i in range(len(timeline_items), 70):
        row = 3 + i
        alt = i % 2 == 1
        style_data_row(ws, row, 5, alt=alt)
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col),
                        fill=make_fill(COLORS["secondary_light"] if alt else COLORS["input_bg"]))
        ws.row_dimensions[row].height = 20

    ws.conditional_formatting.add("D3:D72",
        CellIsRule(operator="equal", formula=['"Done"'], fill=make_fill(COLORS["success_light"]),
                   font=Font(color=COLORS["success"], bold=True)))
    ws.conditional_formatting.add("D3:D72",
        CellIsRule(operator="equal", formula=['"In Progress"'], fill=make_fill(COLORS["warning_light"]),
                   font=Font(color=COLORS["warning"], bold=True)))
    ws.conditional_formatting.add("D3:D72",
        CellIsRule(operator="equal", formula=['"Not Started"'], fill=make_fill(COLORS["danger_light"]),
                   font=Font(color=COLORS["danger"])))

    # Progress summary
    summary_row = 73
    ws.row_dimensions[summary_row].height = 26
    prog_label = ws.cell(row=summary_row, column=1, value="TASKS COMPLETED:")
    apply_style(prog_label, font=header_font(size=11, color=COLORS["dark_text"]),
                fill=make_fill(COLORS["success_light"]), alignment=left_align(), border=thin_border)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    prog_val = ws.cell(row=summary_row, column=3, value='=COUNTIF(D3:D72,"Done")&" of "&COUNTA(B3:B72)')
    apply_style(prog_val, font=header_font(size=12, color=COLORS["success"]),
                fill=make_fill(COLORS["success_light"]), alignment=center_align(), border=thin_border)
    ws.merge_cells(f"C{summary_row}:E{summary_row}")

    set_col_width(ws, {"A": 20, "B": 38, "C": 18, "D": 16, "E": 28})


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_dashboard(wb)
    create_guest_list(wb)
    create_budget_tracker(wb)
    create_decoration_checklist(wb)
    create_food_planner(wb)
    create_party_timeline(wb)

    output_path = "Graduation_Party_Planner.xlsx"
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")
    print(f"  Tabs: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
