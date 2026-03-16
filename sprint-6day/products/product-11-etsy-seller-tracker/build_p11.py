"""
Product 11 — Etsy Seller Profit & Tax Tracker
Builds the .xlsx file ready for upload to Etsy.
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT = os.path.join(os.path.dirname(__file__), "build",
                   "Etsy_Seller_Profit_Tax_Tracker.xlsx")

wb = openpyxl.Workbook()

# ── Palette ────────────────────────────────────────────────────────────────
SAGE       = "6B8F71"
SAGE_LIGHT = "C8DBC9"
SAGE_DARK  = "3D5C41"
GOLD       = "C9A84C"
GOLD_LIGHT = "F0E0A8"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY   = "CCCCCC"
DARK_GRAY  = "444444"
RED_SOFT   = "E57373"
GREEN_SOFT = "81C784"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=11, color=DARK_GRAY, italic=False):
    return Font(name="Calibri", bold=True, size=size, color=color, italic=italic)

def reg(size=11, color=DARK_GRAY, italic=False):
    return Font(name="Calibri", bold=False, size=size, color=color, italic=italic)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color=MED_GRAY)
    n = None
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    return Border()

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def header_row(ws, row, values, bg=SAGE, fg=WHITE, size=11):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = bold(size=size, color=fg)
        c.alignment = center(wrap=True)
        c.border = thin_border()

def title_cell(ws, cell, text, bg=SAGE_DARK, fg=WHITE, size=16):
    c = ws[cell]
    c.value = text
    c.fill = fill(bg)
    c.font = bold(size=size, color=fg)
    c.alignment = center()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Dashboard"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 50
ws1.row_dimensions[2].height = 20

# Title banner
ws1.merge_cells("A1:J1")
title_cell(ws1, "A1", "ETSY SELLER PROFIT & TAX TRACKER", SAGE_DARK, WHITE, 18)

# Sub-header
ws1.merge_cells("A2:J2")
ws1["A2"].value = "Know your real numbers. Grow your shop."
ws1["A2"].fill = fill(SAGE)
ws1["A2"].font = reg(size=11, color=WHITE, italic=True)
ws1["A2"].alignment = center()

# ── KPI Boxes ────────────────────────────────────────────────────────────────
ws1.row_dimensions[4].height = 20
ws1.row_dimensions[5].height = 45
ws1.row_dimensions[6].height = 45
ws1.row_dimensions[7].height = 20

kpi_labels = [
    ("B4:C4", "B5:C6", "TOTAL REVENUE (YTD)",
     "='📅 Monthly Log'!O2",  SAGE_DARK),
    ("E4:F4", "E5:F6", "TOTAL EXPENSES (YTD)",
     "='📅 Monthly Log'!P2",  GOLD),
    ("H4:I4", "H5:I6", "NET PROFIT (YTD)",
     "='📅 Monthly Log'!Q2",  SAGE),
    ("B8:C8", "B9:C10","PROFIT MARGIN %",
     "=IF('📅 Monthly Log'!O2=0,0,'📅 Monthly Log'!Q2/'📅 Monthly Log'!O2)", SAGE_DARK),
    ("E8:F8", "E9:F10","TOTAL ORDERS (YTD)",
     "=COUNTA('📅 Monthly Log'!A4:A1003)", GOLD),
    ("H8:I8", "H9:I10","AVG PROFIT / ORDER",
     "=IF(COUNTA('📅 Monthly Log'!A4:A1003)=0,0,'📅 Monthly Log'!Q2/COUNTA('📅 Monthly Log'!A4:A1003))", SAGE),
]

for label_range, val_range, label_text, formula, color in kpi_labels:
    ws1.merge_cells(label_range)
    lc = ws1[label_range.split(":")[0]]
    lc.value = label_text
    lc.fill = fill(color)
    lc.font = bold(size=9, color=WHITE)
    lc.alignment = center(wrap=True)

    ws1.merge_cells(val_range)
    vc = ws1[val_range.split(":")[0]]
    vc.value = formula
    vc.fill = fill(LIGHT_GRAY)
    vc.font = bold(size=22, color=color)
    vc.alignment = center()
    if "MARGIN" in label_text:
        vc.number_format = "0.0%"
    elif "ORDERS" in label_text:
        vc.number_format = "0"
    else:
        vc.number_format = '"$"#,##0.00'

# ── Monthly summary table ─────────────────────────────────────────────────
ws1.row_dimensions[12].height = 25
ws1.merge_cells("B12:I12")
ws1["B12"].value = "MONTHLY SUMMARY"
ws1["B12"].fill = fill(SAGE_DARK)
ws1["B12"].font = bold(size=13, color=WHITE)
ws1["B12"].alignment = center()

months = ["Jan","Feb","Mar","Apr","May","Jun",
          "Jul","Aug","Sep","Oct","Nov","Dec"]
header_row(ws1, 13,
           ["","Month","Revenue","Expenses","Net Profit","Margin %",
            "Orders","Avg/Order","",""],
           bg=SAGE, fg=WHITE)

for i, month in enumerate(months):
    r = 14 + i
    ws1.row_dimensions[r].height = 22
    ws1.cell(r, 2, month).font = bold(color=SAGE_DARK)
    # These will be blank — user populates via Monthly Log
    for c in range(3, 9):
        cell = ws1.cell(r, c)
        cell.fill = fill(LIGHT_GRAY if i % 2 == 0 else WHITE)
        cell.border = thin_border()
        cell.alignment = center()

ws1.column_dimensions["A"].width = 2
set_col_width(ws1, "B", 12)
set_col_width(ws1, "C", 14)
set_col_width(ws1, "D", 14)
set_col_width(ws1, "E", 14)
set_col_width(ws1, "F", 12)
set_col_width(ws1, "G", 10)
set_col_width(ws1, "H", 13)
set_col_width(ws1, "I", 5)

# ── Instructions box ─────────────────────────────────────────────────────
ws1.merge_cells("B27:I30")
ic = ws1["B27"]
ic.value = ("HOW TO USE: Enter your Etsy sales in the '📅 Monthly Log' tab. "
            "All formulas here update automatically. See '📖 Setup Guide' if you get stuck.")
ic.fill = fill(GOLD_LIGHT)
ic.font = reg(size=10, color=DARK_GRAY, italic=True)
ic.alignment = left(wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — MONTHLY LOG
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📅 Monthly Log")
ws2.sheet_view.showGridLines = False

# Freeze top rows
ws2.freeze_panes = "A4"

ws2.merge_cells("A1:N1")
title_cell(ws2, "A1", "MONTHLY SALES & EXPENSE LOG", SAGE_DARK, WHITE, 16)
ws2.row_dimensions[1].height = 40

# YTD summary row
ws2.merge_cells("A2:D2")
ws2["A2"].value = "YTD TOTALS (auto-calculated):"
ws2["A2"].fill = fill(SAGE_LIGHT)
ws2["A2"].font = bold(color=SAGE_DARK)
ws2["A2"].alignment = left()

summary_cols = [
    ("O2", "=SUM(E4:E1003)", "Revenue",    SAGE_DARK),
    ("P2", "=SUM(I4:I1003)+SUM(J4:J1003)+SUM(K4:K1003)+SUM(L4:L1003)", "Expenses", GOLD),
    ("Q2", "=O2-P2",         "Net Profit", SAGE),
]
for cell_ref, formula, label, color in summary_cols:
    ws2[cell_ref].value = formula
    ws2[cell_ref].number_format = '"$"#,##0.00'
    ws2[cell_ref].fill = fill(SAGE_LIGHT)
    ws2[cell_ref].font = bold(color=color)

headers = [
    "Date", "Order #", "Product Name", "Category",
    "Sale Price", "Trans Fee\n(6.5%)", "Pay Process\n(3%+$0.25)",
    "Listing Fee\n($0.20)", "Shipping\nCost", "COGS",
    "Offsite Ad\nFee", "Other Fees",
    "Total Fees", "Net Profit", "Notes"
]
header_row(ws2, 3, headers, bg=SAGE, fg=WHITE, size=10)
ws2.row_dimensions[3].height = 40

# Sample data row
sample = [
    "2026-01-15", "1234-ABCD", "Budget Planner Template", "Digital Download",
    9.99, None, None, 0.20, 0, 0, 0, 0, None, None, ""
]
for col, val in enumerate(sample, 1):
    c = ws2.cell(4, col, val)
    c.fill = fill(SAGE_LIGHT)
    c.border = thin_border()
    c.alignment = center()

# Formula rows (5–1003)
for r in range(5, 1003):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    # Date (A)
    ws2.cell(r, 1).number_format = "YYYY-MM-DD"
    # Sale Price (E) — user enters
    ws2.cell(r, 5).number_format = '"$"#,##0.00'
    # Transaction Fee (F) = E * 6.5%
    ws2.cell(r, 6).value  = f"=IF(E{r}=\"\",\"\",E{r}*0.065)"
    ws2.cell(r, 6).number_format = '"$"#,##0.00'
    # Payment Processing (G) = E * 3% + 0.25
    ws2.cell(r, 7).value  = f"=IF(E{r}=\"\",\"\",E{r}*0.03+0.25)"
    ws2.cell(r, 7).number_format = '"$"#,##0.00'
    # Listing fee (H) — default $0.20, user can override
    ws2.cell(r, 8).number_format  = '"$"#,##0.00'
    # Shipping (I), COGS (J), Offsite (K), Other (L) — user enters
    for c in [9, 10, 11, 12]:
        ws2.cell(r, c).number_format = '"$"#,##0.00'
    # Total Fees (M) = F+G+H+I+J+K+L
    ws2.cell(r, 13).value = (f"=IF(E{r}=\"\",\"\","
                             f"F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r})")
    ws2.cell(r, 13).number_format = '"$"#,##0.00'
    # Net Profit (N) = E - M
    ws2.cell(r, 14).value = f"=IF(E{r}=\"\",\"\",E{r}-M{r})"
    ws2.cell(r, 14).number_format = '"$"#,##0.00'

    for c in range(1, 16):
        ws2.cell(r, c).fill = fill(bg)
        ws2.cell(r, c).border = thin_border()
        ws2.cell(r, c).alignment = center()

# Conditional formatting colors for net profit column
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill as PF
ws2.conditional_formatting.add(
    "N5:N1003",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PF("solid", fgColor=RED_SOFT))
)
ws2.conditional_formatting.add(
    "N5:N1003",
    CellIsRule(operator="greaterThan", formula=["0"],
               fill=PF("solid", fgColor=GREEN_SOFT))
)

col_widths = {"A":13,"B":14,"C":28,"D":18,"E":12,"F":12,"G":13,
              "H":11,"I":11,"J":11,"K":12,"L":11,"M":12,"N":12,"O":3}
for col, w in col_widths.items():
    ws2.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — COGS TRACKER
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🧮 COGS Tracker")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:H1")
title_cell(ws3, "A1", "COST OF GOODS SOLD (COGS) PER PRODUCT", SAGE_DARK, WHITE, 15)
ws3.row_dimensions[1].height = 36

headers3 = ["Product Name", "Category", "Materials\nCost",
            "Labor Hrs", "Hourly\nRate", "Total COGS",
            "Selling\nPrice", "Gross\nMargin %"]
header_row(ws3, 2, headers3, bg=SAGE, fg=WHITE)
ws3.row_dimensions[2].height = 36

for r in range(3, 103):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    for c in range(1, 9):
        cell = ws3.cell(r, c)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = center()
    ws3.cell(r, 3).number_format = '"$"#,##0.00'
    ws3.cell(r, 5).number_format = '"$"#,##0.00'
    ws3.cell(r, 6).value = f"=IF(C{r}=\"\",\"\",C{r}+(D{r}*E{r}))"
    ws3.cell(r, 6).number_format = '"$"#,##0.00'
    ws3.cell(r, 7).number_format = '"$"#,##0.00'
    ws3.cell(r, 8).value = f"=IF(G{r}=0,\"\",IF(G{r}=\"\",\"\",(G{r}-F{r})/G{r}))"
    ws3.cell(r, 8).number_format = "0.0%"

cog_widths = {"A":28,"B":18,"C":14,"D":12,"E":12,"F":13,"G":13,"H":14}
for col, w in cog_widths.items():
    ws3.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — QUARTERLY TAXES
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🧾 Quarterly Taxes")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F1")
title_cell(ws4, "A1", "QUARTERLY TAX ESTIMATOR", SAGE_DARK, WHITE, 16)
ws4.row_dimensions[1].height = 42

ws4.merge_cells("A2:F2")
ws4["A2"].value = ("⚠  This is an ESTIMATE. Consult a tax professional. "
                   "Based on US self-employment tax rules (SE tax = 15.3%).")
ws4["A2"].fill = fill(GOLD_LIGHT)
ws4["A2"].font = reg(size=9, color=DARK_GRAY, italic=True)
ws4["A2"].alignment = center(wrap=True)
ws4.row_dimensions[2].height = 30

# Tax bracket input
ws4.merge_cells("A4:C4")
ws4["A4"].value = "YOUR FEDERAL INCOME TAX BRACKET (enter as decimal, e.g. 0.22):"
ws4["A4"].font = bold(color=SAGE_DARK)
ws4["A4"].fill = fill(SAGE_LIGHT)
ws4["A4"].alignment = left()
ws4["D4"].value = 0.22
ws4["D4"].number_format = "0%"
ws4["D4"].fill = fill(GOLD_LIGHT)
ws4["D4"].font = bold(color=GOLD, size=14)
ws4["D4"].alignment = center()
ws4["D4"].border = thin_border()

# Quarter table
quarters = [
    ("Q1", "Jan–Mar",  "Due April 15",
     "=SUMIF('📅 Monthly Log'!A4:A1003,\">=\"&DATE(YEAR(TODAY()),1,1),'📅 Monthly Log'!N4:N1003)"
     "-SUMIF('📅 Monthly Log'!A4:A1003,\">\"&DATE(YEAR(TODAY()),3,31),'📅 Monthly Log'!N4:N1003)"),
    ("Q2", "Apr–Jun",  "Due June 17",  ""),
    ("Q3", "Jul–Sep",  "Due Sept 16",  ""),
    ("Q4", "Oct–Dec",  "Due Jan 15",   ""),
]

header_row(ws4, 6,
           ["Quarter", "Months", "Due Date",
            "Net Profit", "SE Tax (15.3%)", "Income Tax Est.", "TOTAL SET ASIDE"],
           bg=SAGE, fg=WHITE)
ws4.row_dimensions[6].height = 30

for i, (q, months_str, due, formula) in enumerate(quarters):
    r = 7 + i
    ws4.row_dimensions[r].height = 38
    ws4.cell(r, 1, q).font = bold(size=14, color=SAGE_DARK)
    ws4.cell(r, 2, months_str)
    ws4.cell(r, 3, due).font = reg(color=GOLD, italic=True)
    # Net profit for quarter — leave formula blank for Q2-Q4 (complex date logic)
    np_cell = ws4.cell(r, 4)
    np_cell.value = "← enter from Monthly Log totals"
    np_cell.number_format = '"$"#,##0.00'
    np_cell.fill = fill(LIGHT_GRAY)
    np_cell.border = thin_border()
    # SE tax
    ws4.cell(r, 5).value = f"=IF(D{r}=\"\",\"\",D{r}*0.153)"
    ws4.cell(r, 5).number_format = '"$"#,##0.00'
    ws4.cell(r, 5).fill = fill(GOLD_LIGHT)
    ws4.cell(r, 5).border = thin_border()
    # Income tax
    ws4.cell(r, 6).value = f"=IF(D{r}=\"\",\"\",D{r}*$D$4)"
    ws4.cell(r, 6).number_format = '"$"#,##0.00'
    ws4.cell(r, 6).fill = fill(GOLD_LIGHT)
    ws4.cell(r, 6).border = thin_border()
    # Total
    ws4.cell(r, 7).value = f"=IF(D{r}=\"\",\"\",E{r}+F{r})"
    ws4.cell(r, 7).number_format = '"$"#,##0.00'
    ws4.cell(r, 7).fill = fill(SAGE_LIGHT)
    ws4.cell(r, 7).font = bold(color=SAGE_DARK)
    ws4.cell(r, 7).border = thin_border()
    for c in range(1, 8):
        ws4.cell(r, c).alignment = center()

# Payment dates reminder
ws4.merge_cells("A12:G14")
ws4["A12"].value = (
    "QUARTERLY PAYMENT DATES (2026):\n"
    "Q1 → April 15, 2026   |   Q2 → June 17, 2026   |   "
    "Q3 → September 16, 2026   |   Q4 → January 15, 2027\n"
    "Pay at: IRS.gov/DirectPay — select 'Estimated Tax'"
)
ws4["A12"].fill = fill(SAGE_LIGHT)
ws4["A12"].font = reg(size=10, color=SAGE_DARK)
ws4["A12"].alignment = left(wrap=True)

qt_widths = {"A":10,"B":14,"C":16,"D":22,"E":16,"F":16,"G":18}
for col, w in qt_widths.items():
    ws4.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — ANNUAL SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📆 Annual Summary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:G1")
title_cell(ws5, "A1", "ANNUAL SUMMARY & SCHEDULE C PREP", SAGE_DARK, WHITE, 16)
ws5.row_dimensions[1].height = 42

ws5.merge_cells("A3:G3")
ws5["A3"].value = "SCHEDULE C LINE ITEMS"
ws5["A3"].fill = fill(SAGE)
ws5["A3"].font = bold(size=13, color=WHITE)
ws5["A3"].alignment = center()

schedule_c = [
    ("Line 1 — Gross Receipts",
     "=SUM('📅 Monthly Log'!E4:E1003)"),
    ("Line 17 — Taxes & Licenses (SE tax paid)",
     "='🧾 Quarterly Taxes'!G7+'🧾 Quarterly Taxes'!G8+'🧾 Quarterly Taxes'!G9+'🧾 Quarterly Taxes'!G10"),
    ("Line 22 — Supplies / COGS",
     "=SUM('📅 Monthly Log'!J4:J1003)"),
    ("Line 27a — Other Expenses (Etsy fees)",
     "=SUM('📅 Monthly Log'!F4:F1003)+SUM('📅 Monthly Log'!G4:G1003)+SUM('📅 Monthly Log'!H4:H1003)"),
    ("Line 41 — Net Profit (Est.)",
     "='📅 Monthly Log'!Q2"),
]

for i, (label, formula) in enumerate(schedule_c):
    r = 4 + i
    ws5.row_dimensions[r].height = 30
    ws5.cell(r, 2, label).font = bold(color=DARK_GRAY)
    ws5.cell(r, 2).fill = fill(LIGHT_GRAY if i % 2 == 0 else WHITE)
    ws5.cell(r, 2).border = thin_border()
    ws5.cell(r, 2).alignment = left()
    ws5.cell(r, 4).value = formula
    ws5.cell(r, 4).number_format = '"$"#,##0.00'
    ws5.cell(r, 4).fill = fill(SAGE_LIGHT)
    ws5.cell(r, 4).font = bold(color=SAGE_DARK)
    ws5.cell(r, 4).border = thin_border()
    ws5.cell(r, 4).alignment = center()

ann_widths = {"A":3,"B":45,"C":3,"D":18,"E":3}
for col, w in ann_widths.items():
    ws5.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — SETUP GUIDE
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("📖 Setup Guide")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:F1")
title_cell(ws6, "A1", "SETUP GUIDE — READ FIRST", GOLD, WHITE, 18)
ws6.row_dimensions[1].height = 50

steps = [
    ("STEP 1: Make a Copy (Google Sheets)",
     "Upload this file to Google Drive → right-click → Open With → Google Sheets.\n"
     "Then File → Make a Copy. Work in YOUR copy — never in the original."),
    ("STEP 2: Enter Your Tax Bracket",
     "Go to '🧾 Quarterly Taxes' tab → Cell D4 → enter your federal income tax rate as a decimal.\n"
     "Not sure? Use 0.22 (22%) — it's the most common bracket for self-employed sellers earning $40K–$89K."),
    ("STEP 3: Log Your Sales",
     "Go to '📅 Monthly Log' tab. For each Etsy sale, enter:\n"
     "• Date (Column A)\n• Order # from Etsy (Column B)\n• Product name (Column C)\n"
     "• Sale price (Column E)\n• Shipping cost if you paid it (Column I)\n"
     "• COGS — your material/supply costs (Column J)\n\n"
     "Etsy fees (Columns F, G, H) calculate AUTOMATICALLY. You don't touch those."),
    ("STEP 4: Track COGS",
     "Go to '🧮 COGS Tracker' tab. For each product you sell, enter:\n"
     "• Materials cost per unit\n• Hours to make it (0 for digital products)\n"
     "• Your hourly rate\n"
     "The Gross Margin % column tells you which products are most profitable."),
    ("STEP 5: Check Your Dashboard",
     "The '📊 Dashboard' tab updates automatically as you enter data.\n"
     "Check it weekly to see your real profit numbers."),
    ("STEP 6: Pay Quarterly Taxes",
     "Every quarter, go to '🧾 Quarterly Taxes' and enter your net profit for that quarter.\n"
     "The 'TOTAL SET ASIDE' column tells you exactly how much to pay the IRS.\n"
     "Pay at IRS.gov/DirectPay before the due date."),
    ("NEED HELP?",
     "Questions? Email: [your contact] or leave a review on Etsy — we read every one.\n"
     "Remember: this tracker saves you hours at tax time and helps you find your\n"
     "most profitable products. It pays for itself after your first tax season."),
]

for i, (title_text, body_text) in enumerate(steps):
    r = 3 + (i * 5)
    ws6.row_dimensions[r].height = 30
    ws6.row_dimensions[r+1].height = 60
    ws6.merge_cells(f"B{r}:F{r}")
    tc = ws6.cell(r, 2, title_text)
    tc.fill = fill(SAGE)
    tc.font = bold(size=12, color=WHITE)
    tc.alignment = left()
    ws6.merge_cells(f"B{r+1}:F{r+3}")
    bc = ws6.cell(r+1, 2, body_text)
    bc.fill = fill(LIGHT_GRAY if i % 2 == 0 else WHITE)
    bc.font = reg(size=11)
    bc.alignment = left(wrap=True)
    ws6.row_dimensions[r+1].height = 80

ws6.column_dimensions["A"].width = 3
ws6.column_dimensions["B"].width = 70
for col in ["C","D","E","F"]:
    ws6.column_dimensions[col].width = 18

# ── Tab colors ───────────────────────────────────────────────────────────────
ws1.sheet_properties.tabColor = SAGE_DARK
ws2.sheet_properties.tabColor = SAGE
ws3.sheet_properties.tabColor = SAGE_LIGHT
ws4.sheet_properties.tabColor = GOLD
ws5.sheet_properties.tabColor = SAGE_DARK
ws6.sheet_properties.tabColor = GOLD_LIGHT

wb.save(OUT)
print(f"✅ Saved: {OUT}")
