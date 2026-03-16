"""
Product 14 — Real Estate Investor Deal Analyzer
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT = os.path.join(os.path.dirname(__file__), "build",
                   "Real_Estate_Deal_Analyzer.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

FOREST  = "1A4731"
FOR_LT  = "D1E7DD"
GOLD    = "B8860B"
GOLD_LT = "FFF3CD"
WHITE   = "FFFFFF"
LT_GRY  = "F8F9FA"
MED_GRY = "CCCCCC"
DARK    = "212529"
RED     = "DC3545"
GREEN   = "198754"
BLUE    = "0D6EFD"

def fill(h): return PatternFill("solid", fgColor=h)
def bold(sz=11, col=DARK): return Font(name="Calibri", bold=True, size=sz, color=col)
def reg(sz=11, col=DARK, it=False): return Font(name="Calibri", size=sz, color=col, italic=it)
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=False): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
thin = Side(style="thin", color=MED_GRY)
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def input_cell(ws, row, col, label, value="", fmt=None, hint=""):
    lc = ws.cell(row, col, label)
    lc.font = reg(10, DARK); lc.fill = fill(LT_GRY)
    lc.alignment = lft(); lc.border = bdr()
    vc = ws.cell(row, col+1, value)
    vc.fill = fill(GOLD_LT); vc.font = bold(12, GOLD)
    vc.alignment = ctr(); vc.border = bdr()
    if fmt: vc.number_format = fmt
    if hint:
        hc = ws.cell(row, col+2, hint)
        hc.font = reg(9, "888888", True); hc.alignment = lft()

def output_cell(ws, row, col, label, formula, fmt=None, color=FOREST):
    lc = ws.cell(row, col, label)
    lc.font = reg(10, DARK); lc.fill = fill(LT_GRY)
    lc.alignment = lft(); lc.border = bdr()
    vc = ws.cell(row, col+1, formula)
    vc.fill = fill(FOR_LT); vc.font = bold(12, color)
    vc.alignment = ctr(); vc.border = bdr()
    if fmt: vc.number_format = fmt

def section_hdr(ws, row, cols, text, bg=FOREST):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.fill = fill(bg); c.font = bold(12, WHITE); c.alignment = lft()
    ws.row_dimensions[row].height = 28

from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# TAB 1: DEAL ANALYZER (main)
# ══════════════════════════════════════════════════════════════
da = wb.active; da.title = "🏘 Deal Analyzer"
da.sheet_view.showGridLines = False

da.merge_cells("A1:H1")
da["A1"].value = "REAL ESTATE DEAL ANALYZER"
da["A1"].fill = fill(FOREST); da["A1"].font = bold(20, WHITE)
da["A1"].alignment = ctr(); da.row_dimensions[1].height = 50

da.merge_cells("A2:H2")
da["A2"].value = "Enter yellow cells. All results calculate automatically."
da["A2"].fill = fill(GOLD_LT); da["A2"].font = reg(10, GOLD, True)
da["A2"].alignment = ctr(); da.row_dimensions[2].height = 24

# ── SECTION: PROPERTY INFO ────────────────────────────────────
section_hdr(da, 4, 8, "  📍 PROPERTY INFORMATION")
input_cell(da, 5, 2, "Property Address", "123 Main St, City, ST", hint="Enter full address")
input_cell(da, 6, 2, "Property Type", "Single Family", hint="SFR / Duplex / Multi / Commercial")
input_cell(da, 7, 2, "Square Footage", 1500, "#,##0 \"sqft\"")
input_cell(da, 8, 2, "Year Built", 1995, "0")
input_cell(da, 9, 2, "Bedrooms / Baths", "3 / 2")
input_cell(da,10, 2, "Deal Type", "Buy & Hold", hint="Buy & Hold / Fix & Flip / BRRRR")

# ── SECTION: PURCHASE ─────────────────────────────────────────
section_hdr(da, 12, 8, "  💰 PURCHASE & ACQUISITION COSTS")
input_cell(da,13, 2, "Purchase Price",           250000, '"$"#,##0')
input_cell(da,14, 2, "Down Payment %",            0.20,  "0%",   "typical: 20-25%")
input_cell(da,15, 2, "Closing Costs",             4000,  '"$"#,##0', "est. 1.5-2%")
input_cell(da,16, 2, "Rehab / Repair Costs",      15000, '"$"#,##0')
input_cell(da,17, 2, "Other Acquisition Costs",   500,   '"$"#,##0', "inspection, etc.")

output_cell(da,18, 2, "Down Payment $",     "=C13*C14", '"$"#,##0')
output_cell(da,19, 2, "TOTAL CASH NEEDED",  "=C18+C15+C16+C17", '"$"#,##0', GOLD)
output_cell(da,20, 2, "Loan Amount",        "=C13-C18", '"$"#,##0')
output_cell(da,21, 2, "After Repair Value (ARV)", 295000, '"$"#,##0')  # user input

# ── SECTION: FINANCING ────────────────────────────────────────
section_hdr(da, 23, 8, "  🏦 FINANCING")
input_cell(da,24, 2, "Loan Type",           "Conventional")
input_cell(da,25, 2, "Interest Rate",        0.075, "0.000%", "e.g. 0.075 = 7.5%")
input_cell(da,26, 2, "Loan Term (years)",    30,    "0")
input_cell(da,27, 2, "Points Paid",          0,     "0.0", "1 point = 1% of loan")

output_cell(da,28, 2, "Monthly P&I Payment",
    '=IFERROR(PMT(C25/12,C26*12,-C20),0)', '"$"#,##0.00', GOLD)

# ── SECTION: RENTAL INCOME ────────────────────────────────────
section_hdr(da, 30, 8, "  🏠 RENTAL INCOME (Buy & Hold)")
input_cell(da,31, 2, "Monthly Rent (Market)",  2000,  '"$"#,##0')
input_cell(da,32, 2, "Vacancy Rate %",          0.05, "0%",  "typical: 5-8%")
input_cell(da,33, 2, "Other Monthly Income",    0,    '"$"#,##0', "laundry, storage, etc.")

output_cell(da,34, 2, "Effective Gross Income (mo.)",
    "=C31*(1-C32)+C33", '"$"#,##0')
output_cell(da,35, 2, "Annual Gross Income",
    "=C34*12", '"$"#,##0')

# ── SECTION: OPERATING EXPENSES ───────────────────────────────
section_hdr(da, 37, 8, "  📋 MONTHLY OPERATING EXPENSES")
expenses = [
    ("Property Tax (mo.)",    250,  '"$"#,##0'),
    ("Insurance (mo.)",       120,  '"$"#,##0'),
    ("Property Mgmt (mo.)",   0,    '"$"#,##0', "8-10% of rent if managed"),
    ("Repairs Reserve (mo.)", 100,  '"$"#,##0', "1% of value/yr ÷ 12"),
    ("CapEx Reserve (mo.)",   100,  '"$"#,##0', "roof, HVAC, appliances"),
    ("Utilities (mo.)",       0,    '"$"#,##0', "if landlord pays any"),
    ("HOA (mo.)",             0,    '"$"#,##0'),
    ("Other Expenses (mo.)",  50,   '"$"#,##0'),
]
for i, exp in enumerate(expenses):
    r = 38 + i
    if len(exp) == 3:
        input_cell(da, r, 2, exp[0], exp[1], exp[2])
    else:
        input_cell(da, r, 2, exp[0], exp[1], exp[2], exp[3])

output_cell(da,46, 2, "Total Operating Expenses (mo.)",
    "=SUM(C38:C45)", '"$"#,##0', GOLD)
output_cell(da,47, 2, "Total Operating Expenses (yr)",
    "=C46*12", '"$"#,##0')

# ── SECTION: RETURNS ──────────────────────────────────────────
section_hdr(da, 49, 8, "  📈 RETURNS & METRICS", FOREST)
da.row_dimensions[49].height = 30

metrics = [
    ("Monthly NOI",            "=C34-C46",        '"$"#,##0',   "Net Operating Income"),
    ("Annual NOI",             "=C50*12",         '"$"#,##0',   ""),
    ("Monthly Cash Flow",      "=C50-C28",        '"$"#,##0',   "NOI minus mortgage"),
    ("Annual Cash Flow",       "=C52*12",         '"$"#,##0',   ""),
    ("Cash-on-Cash Return",    "=IF(C19=0,0,C53/C19)", "0.00%", "annual CF / cash invested"),
    ("Cap Rate",               "=IF(C13=0,0,C51/C13)", "0.00%", "NOI / purchase price"),
    ("Gross Rent Multiplier",  "=IF(C35=0,0,C13/C35)", "0.00",  "purchase price / annual rent"),
    ("Price per Sq Ft",        "=IF(C7=0,0,C13/C7)",   '"$"#,##0.00', ""),
    ("Equity at Purchase",     "=C13-C20",        '"$"#,##0',   ""),
    ("ARV Equity",             "=C21-C20+C16",    '"$"#,##0',   "after rehab equity"),
    ("70% ARV Rule",           "=C21*0.70-C16",   '"$"#,##0',   "max offer for flip"),
]
for i, (lbl, frm, fmt, hint) in enumerate(metrics):
    r = 50 + i
    da.row_dimensions[r].height = 26
    lc = da.cell(r,2,lbl); lc.font = reg(10,DARK)
    lc.fill = fill(LT_GRY if i%2==0 else WHITE); lc.border = bdr(); lc.alignment = lft()
    vc = da.cell(r,3,frm); vc.number_format = fmt
    vc.fill = fill(FOR_LT); vc.font = bold(13, FOREST); vc.border = bdr(); vc.alignment = ctr()
    if hint:
        hc = da.cell(r,4,hint); hc.font = reg(9,"888888",True); hc.alignment = lft()

# Deal score
da.merge_cells("B61:D61"); da["B61"].value = "DEAL SCORE"
da["B61"].fill = fill(FOREST); da["B61"].font = bold(12,WHITE); da["B61"].alignment = ctr()

da.merge_cells("B62:D64")
da["B62"].value = ('=IF(C54>=0.08, "🟢 STRONG DEAL — CoC return ≥ 8%", '
                   'IF(C54>=0.05, "🟡 MARGINAL — CoC 5-8%, negotiate harder", '
                   '"🔴 PASS — CoC below 5%, numbers don\'t work"))')
da["B62"].fill = fill(GOLD_LT); da["B62"].font = bold(14, GOLD)
da["B62"].alignment = ctr(True)

for col, w in {"A":2,"B":32,"C":16,"D":30,"E":3,"F":3,"G":3,"H":3}.items():
    da.column_dimensions[col].width = w
for r in range(5, 65): da.row_dimensions[r].height = 26

# ══════════════════════════════════════════════════════════════
# TAB 2: DEAL COMPARISON (compare up to 5 deals)
# ══════════════════════════════════════════════════════════════
dc = wb.create_sheet("🔍 Deal Comparison")
dc.sheet_view.showGridLines = False

dc.merge_cells("A1:F1")
dc["A1"].value = "DEAL COMPARISON — Compare Up to 5 Properties"
dc["A1"].fill = fill(FOREST); dc["A1"].font = bold(16,WHITE); dc["A1"].alignment = ctr()
dc.row_dimensions[1].height = 44

deal_rows = [
    "Property Address", "Purchase Price", "Total Cash Needed",
    "Monthly Cash Flow", "Annual Cash Flow", "Cash-on-Cash Return",
    "Cap Rate", "GRM", "Monthly NOI", "ARV", "70% ARV Max Offer",
    "Notes",
]
dc.cell(3,1,"METRIC").fill = fill(FOREST); dc.cell(3,1).font = bold(col=WHITE)
dc.cell(3,1).alignment = ctr(); dc.cell(3,1).border = bdr()

for ci, deal_name in enumerate(["Deal A","Deal B","Deal C","Deal D","Deal E"], 2):
    dc.cell(3,ci,deal_name).fill = fill(FOREST)
    dc.cell(3,ci,deal_name).font = bold(col=GOLD)
    dc.cell(3,ci,deal_name).alignment = ctr(); dc.cell(3,ci,deal_name).border = bdr()

for ri, row_label in enumerate(deal_rows):
    r = 4+ri; bg = LT_GRY if ri%2==0 else WHITE
    dc.cell(r,1,row_label).font = bold(9,DARK)
    dc.cell(r,1).fill = fill(bg); dc.cell(r,1).border = bdr(); dc.cell(r,1).alignment = lft()
    for c in range(2,7):
        cell = dc.cell(r,c); cell.fill = fill(GOLD_LT); cell.border = bdr()
        cell.alignment = ctr()
        if "Price" in row_label or "Cash" in row_label or "NOI" in row_label or "ARV" in row_label or "Offer" in row_label:
            cell.number_format = '"$"#,##0'
        if "Return" in row_label or "Rate" in row_label:
            cell.number_format = "0.00%"

# Best deal highlight note
dc.merge_cells("A19:F20")
dc["A19"].value = ("TIP: Highlight the column with the best Cash-on-Cash Return in green. "
                   "That's usually your best deal — unless it requires too much cash upfront.")
dc["A19"].fill = fill(FOR_LT); dc["A19"].font = reg(10,FOREST,True)
dc["A19"].alignment = lft(True)

for col, w in {"A":28,"B":18,"C":18,"D":18,"E":18,"F":18}.items():
    dc.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 3: REHAB ESTIMATOR
# ══════════════════════════════════════════════════════════════
re_ = wb.create_sheet("🔨 Rehab Estimator")
re_.sheet_view.showGridLines = False

re_.merge_cells("A1:F1")
re_["A1"].value = "REHAB COST ESTIMATOR"
re_["A1"].fill = fill(FOREST); re_["A1"].font = bold(16,WHITE); re_["A1"].alignment = ctr()
re_.row_dimensions[1].height = 44

hdr_vals = ["Category","Item","Low Estimate","High Estimate","Your Estimate","Notes"]
for ci, v in enumerate(hdr_vals, 1):
    c = re_.cell(3,ci,v); c.fill = fill(FOREST); c.font = bold(col=WHITE)
    c.alignment = ctr(True); c.border = bdr()
re_.row_dimensions[3].height = 32

rehab_items = [
    ("Kitchen","Full Kitchen Remodel",15000,45000),
    ("Kitchen","Cabinet Refacing",3000,8000),
    ("Kitchen","Countertops (granite)",2000,5000),
    ("Kitchen","Appliances",2000,6000),
    ("Bathroom","Full Bath Remodel",6000,20000),
    ("Bathroom","Vanity + Fixtures",500,2000),
    ("Bathroom","Tile Work",1500,5000),
    ("Flooring","Hardwood (per sqft x sqft)",3,8),
    ("Flooring","LVP / Laminate",1500,5000),
    ("Flooring","Carpet",1000,3500),
    ("Exterior","Roof Replacement",8000,18000),
    ("Exterior","Siding",6000,15000),
    ("Exterior","Landscaping / Curb Appeal",500,3000),
    ("HVAC","Full HVAC Replacement",5000,12000),
    ("HVAC","Water Heater",800,1800),
    ("Electrical","Panel Upgrade",2000,5000),
    ("Plumbing","Re-pipe",4000,12000),
    ("Interior","Paint (interior)",1500,4000),
    ("Interior","Paint (exterior)",2000,6000),
    ("Interior","Doors (interior set)",1000,3000),
    ("Windows","Window Replacement (ea.)",300,700),
    ("Foundation","Foundation Repair",3000,20000),
    ("Other","Permits",500,2000),
    ("Other","Dumpster",400,800),
    ("Other","Contingency (10-15%)",0,0),
]
for i, (cat, item, lo, hi) in enumerate(rehab_items):
    r = 4+i; bg = LT_GRY if i%2==0 else WHITE
    re_.cell(r,1,cat).font = bold(9,FOREST); re_.cell(r,1).fill = fill(bg)
    re_.cell(r,1).border = bdr(); re_.cell(r,1).alignment = ctr()
    re_.cell(r,2,item).fill = fill(bg); re_.cell(r,2).border = bdr(); re_.cell(r,2).alignment = lft()
    for c, val in [(3,lo),(4,hi)]:
        re_.cell(r,c,val).fill = fill(bg); re_.cell(r,c).number_format = '"$"#,##0'
        re_.cell(r,c).border = bdr(); re_.cell(r,c).alignment = ctr()
    re_.cell(r,5).fill = fill(GOLD_LT); re_.cell(r,5).number_format = '"$"#,##0'
    re_.cell(r,5).border = bdr(); re_.cell(r,5).alignment = ctr()
    re_.cell(r,6).fill = fill(bg); re_.cell(r,6).border = bdr()

# Totals
total_r = 4+len(rehab_items)
re_.cell(total_r,2,"TOTAL REHAB ESTIMATE →").font = bold(col=FOREST)
re_.cell(total_r,2).fill = fill(FOR_LT); re_.cell(total_r,2).alignment = lft()
re_.cell(total_r,3).value = f"=SUM(C4:C{total_r-1})"
re_.cell(total_r,4).value = f"=SUM(D4:D{total_r-1})"
re_.cell(total_r,5).value = f"=SUM(E4:E{total_r-1})"
for c in [3,4,5]:
    re_.cell(total_r,c).number_format = '"$"#,##0'
    re_.cell(total_r,c).fill = fill(FOR_LT); re_.cell(total_r,c).font = bold(col=FOREST)
    re_.cell(total_r,c).border = bdr(); re_.cell(total_r,c).alignment = ctr()

for col,w in {"A":14,"B":32,"C":16,"D":16,"E":16,"F":22}.items():
    re_.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 4: RENTAL PROJECTIONS
# ══════════════════════════════════════════════════════════════
rp = wb.create_sheet("📈 Rental Projections")
rp.sheet_view.showGridLines = False
rp.merge_cells("A1:H1")
rp["A1"].value = "10-YEAR RENTAL INCOME PROJECTION"
rp["A1"].fill = fill(FOREST); rp["A1"].font = bold(16,WHITE); rp["A1"].alignment = ctr()
rp.row_dimensions[1].height = 44

# Inputs
for r,lbl,val,fmt in [
    (3,"Starting Monthly Rent",2000,'"$"#,##0'),
    (4,"Annual Rent Increase %",0.03,"0.0%"),
    (5,"Starting Mortgage Payment",1350,'"$"#,##0'),
    (6,"Annual Expense Increase %",0.02,"0.0%"),
    (7,"Starting Operating Expenses",570,'"$"#,##0'),
]:
    rp.cell(r,2,lbl).font = reg(10,DARK); rp.cell(r,2).fill = fill(LT_GRY)
    rp.cell(r,2).border = bdr(); rp.cell(r,2).alignment = lft()
    rp.cell(r,3,val).font = bold(12,GOLD); rp.cell(r,3).number_format = fmt
    rp.cell(r,3).fill = fill(GOLD_LT); rp.cell(r,3).border = bdr(); rp.cell(r,3).alignment = ctr()

# Projection table
hdr_rp = ["Year","Monthly Rent","Annual Rent","Operating Exp (yr)",
          "Mortgage (yr)","NOI","Cash Flow","Cumulative CF"]
for ci,v in enumerate(hdr_rp,1):
    c = rp.cell(10,ci,v); c.fill = fill(FOREST); c.font = bold(col=WHITE)
    c.alignment = ctr(True); c.border = bdr()
rp.row_dimensions[10].height = 32

for yr in range(1,11):
    r = 10+yr; bg = LT_GRY if yr%2==0 else WHITE
    rp.cell(r,1,f"Year {yr}").font = bold(col=FOREST)
    rp.cell(r,1).fill = fill(bg); rp.cell(r,1).border = bdr(); rp.cell(r,1).alignment = ctr()
    if yr == 1:
        rent_mo = "=$C$3"
        op_exp  = "=$C$7*12"
    else:
        rent_mo = f"=B{r-1}*(1+$C$4)"
        op_exp  = f"=D{r-1}*(1+$C$6)"
    rp.cell(r,2).value = rent_mo; rp.cell(r,2).number_format = '"$"#,##0'
    rp.cell(r,3).value = f"=B{r}*12"; rp.cell(r,3).number_format = '"$"#,##0'
    rp.cell(r,4).value = op_exp; rp.cell(r,4).number_format = '"$"#,##0'
    rp.cell(r,5).value = "=$C$5*12"; rp.cell(r,5).number_format = '"$"#,##0'
    rp.cell(r,6).value = f"=C{r}-D{r}"; rp.cell(r,6).number_format = '"$"#,##0'
    rp.cell(r,7).value = f"=F{r}-E{r}"; rp.cell(r,7).number_format = '"$"#,##0'
    rp.cell(r,8).value = f"=IF({r-1}=10,G{r},H{r-1}+G{r})" if yr>1 else f"=G{r}"
    rp.cell(r,8).number_format = '"$"#,##0'
    for c in range(1,9):
        rp.cell(r,c).fill = fill(bg); rp.cell(r,c).border = bdr(); rp.cell(r,c).alignment = ctr()
    rp.cell(r,8).fill = fill(FOR_LT); rp.cell(r,8).font = bold(col=FOREST)

for col,w in {"A":10,"B":16,"C":16,"D":20,"E":16,"F":14,"G":14,"H":18}.items():
    rp.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 5: SETUP GUIDE
# ══════════════════════════════════════════════════════════════
sg = wb.create_sheet("📖 Setup Guide")
sg.sheet_view.showGridLines = False
sg.merge_cells("A1:E1")
sg["A1"].value = "SETUP GUIDE — HOW TO ANALYZE YOUR FIRST DEAL"
sg["A1"].fill = fill(FOREST); sg["A1"].font = bold(16,WHITE); sg["A1"].alignment = ctr()
sg.row_dimensions[1].height = 44

steps = [
    ("STEP 1 — Enter the Property Info",
     "Go to '🏘 Deal Analyzer'. Fill in all YELLOW cells.\n"
     "Start with: Purchase Price, Down Payment %, Rehab Costs, Monthly Rent, and operating expenses.\n"
     "All metrics (Cash-on-Cash Return, Cap Rate, NOI) calculate automatically."),
    ("STEP 2 — Check the Deal Score",
     "Scroll to the bottom of '🏘 Deal Analyzer'.\n"
     "🟢 Green = strong deal (CoC ≥ 8%)   🟡 Yellow = marginal   🔴 Red = pass.\n"
     "The 70% ARV Rule shows the maximum you should pay for a fix-and-flip."),
    ("STEP 3 — Estimate Rehab Costs",
     "Go to '🔨 Rehab Estimator'. Enter your estimate for each repair category.\n"
     "Low/High ranges are pre-filled as guides. Enter YOUR number in the yellow 'Your Estimate' column."),
    ("STEP 4 — Compare Multiple Deals",
     "Use '🔍 Deal Comparison' to compare up to 5 properties side-by-side.\n"
     "Copy key metrics from each deal analysis into the comparison table."),
    ("STEP 5 — Project Long-Term Returns",
     "Go to '📈 Rental Projections'. Set your starting rent, annual increase %, and expenses.\n"
     "The table projects 10 years of cash flow and cumulative returns."),
    ("KEY METRICS EXPLAINED",
     "Cash-on-Cash Return: Annual cash flow ÷ total cash invested. Target: 8%+\n"
     "Cap Rate: NOI ÷ purchase price. Good markets: 5-8%+\n"
     "GRM (Gross Rent Multiplier): Purchase price ÷ annual rent. Lower = better.\n"
     "70% ARV Rule: For flips — never pay more than 70% of ARV minus repair costs."),
]
for i,(h,b) in enumerate(steps):
    r = 3+i*4
    sg.merge_cells(f"B{r}:E{r}")
    sg.cell(r,2,h).fill = fill(FOREST); sg.cell(r,2,h).font = bold(12,WHITE); sg.cell(r,2,h).alignment = lft()
    sg.merge_cells(f"B{r+1}:E{r+3}")
    sg.cell(r+1,2,b).fill = fill(LT_GRY if i%2==0 else WHITE)
    sg.cell(r+1,2,b).font = reg(11); sg.cell(r+1,2,b).alignment = lft(True)
    sg.row_dimensions[r+1].height = 80

sg.column_dimensions["A"].width = 3; sg.column_dimensions["B"].width = 75

da.sheet_properties.tabColor  = FOREST; dc.sheet_properties.tabColor  = FOREST
re_.sheet_properties.tabColor = GOLD;   rp.sheet_properties.tabColor  = GOLD
sg.sheet_properties.tabColor  = "AAAAAA"

wb.save(OUT)
print(f"✅ Saved: {OUT}")
