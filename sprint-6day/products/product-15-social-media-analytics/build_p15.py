"""
Product 15 — Social Media Content Calendar & Analytics Tracker
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT = os.path.join(os.path.dirname(__file__), "build",
                   "Social_Media_Analytics_Tracker.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

PURPLE  = "5B2D8E"
PUR_LT  = "EDE7F6"
PINK    = "C2185B"
PINK_LT = "FCE4EC"
WHITE   = "FFFFFF"
LT_GRY  = "FAFAFA"
MED_GRY = "CCCCCC"
DARK    = "212121"
AMBER   = "FF8F00"
TEAL    = "00695C"
GOLD_LT = "FFF9C4"

def fill(h): return PatternFill("solid", fgColor=h)
def bold(sz=11, col=DARK): return Font(name="Calibri", bold=True, size=sz, color=col)
def reg(sz=11, col=DARK, it=False): return Font(name="Calibri", size=sz, color=col, italic=it)
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=False): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
thin = Side(style="thin", color=MED_GRY)
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, vals, bg=PURPLE, fg=WHITE, sz=10):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill(bg); cell.font = bold(sz, fg)
        cell.alignment = ctr(True); cell.border = bdr()
    ws.row_dimensions[row].height = 38

def title(ws, rng, txt, bg=PURPLE, fg=WHITE, sz=16):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = txt; c.fill = fill(bg)
    c.font = bold(sz, fg); c.alignment = ctr()

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# TAB 1: DASHBOARD
# ══════════════════════════════════════════════════════════════
dsh = wb.active; dsh.title = "📊 Dashboard"
dsh.sheet_view.showGridLines = False
dsh.row_dimensions[1].height = 50; dsh.row_dimensions[2].height = 22

title(dsh, "A1:J1", "SOCIAL MEDIA ANALYTICS TRACKER", PURPLE, WHITE, 18)
dsh.merge_cells("A2:J2")
dsh["A2"].value = "Track every platform. Find what actually grows your audience."
dsh["A2"].fill = fill(PUR_LT); dsh["A2"].font = reg(11, PURPLE, True); dsh["A2"].alignment = ctr()

kpis = [
    ("B4:C4","B5:C6","TOTAL POSTS (MTD)",   "=COUNTIFS('📅 Content Calendar'!C2:C1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📅 Content Calendar'!D2:D1001,\"Posted\")", PURPLE),
    ("E4:F4","E5:F6","AVG ENGAGEMENT RATE", "=IFERROR(AVERAGEIF('📈 Analytics Log'!I2:I1001,\">0\",'📈 Analytics Log'!I2:I1001),0)", PINK),
    ("H4:I4","H5:I6","BEST PLATFORM",       "=\"See Analytics tab\"", PURPLE),
    ("B8:C8","B9:C10","TOTAL FOLLOWERS",    "=SUM('📊 Follower Growth'!C2:C8)", PINK),
    ("E8:F8","E9:F10","FOLLOWER GROWTH MTD","=SUM('📊 Follower Growth'!D2:D8)", PURPLE),
    ("H8:I8","H9:I10","POSTS SCHEDULED",    "=COUNTIF('📅 Content Calendar'!D2:D1001,\"Scheduled\")", PINK),
]
for lr, vr, lbl, frm, col in kpis:
    dsh.merge_cells(lr); dsh.merge_cells(vr)
    lc = dsh[lr.split(":")[0]]
    lc.value = lbl; lc.fill = fill(col); lc.font = bold(9, WHITE); lc.alignment = ctr(True)
    vc = dsh[vr.split(":")[0]]
    vc.value = frm; vc.fill = fill(LT_GRY); vc.font = bold(20, col); vc.alignment = ctr()
    if "RATE" in lbl: vc.number_format = "0.00%"
    if "GROWTH" in lbl: vc.number_format = "+#,##0;-#,##0"

for col,w in {"A":2,"B":16,"C":16,"D":3,"E":20,"F":14,"G":3,"H":18,"I":14}.items():
    dsh.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 2: CONTENT CALENDAR (12-month)
# ══════════════════════════════════════════════════════════════
cc = wb.create_sheet("📅 Content Calendar")
cc.sheet_view.showGridLines = False; cc.freeze_panes = "A3"
title(cc, "A1:K1", "CONTENT CALENDAR", PURPLE, WHITE, 15)

hdr(cc, 2, ["#","Date","Platform","Status","Content Type",
            "Caption / Hook","Hashtags","Visual Description",
            "CTA","Link","Notes"])

platforms = DataValidation(type="list",
    formula1='"Instagram,TikTok,Pinterest,Facebook,Twitter/X,YouTube,LinkedIn,Threads"',
    allow_blank=True)
status_dv = DataValidation(type="list",
    formula1='"Idea,In Progress,Scheduled,Posted,Archived"', allow_blank=True)
ctype_dv = DataValidation(type="list",
    formula1='"Reel,Carousel,Static Post,Story,Video,Pin,Tweet,Article,Short"', allow_blank=True)
cc.add_data_validation(platforms)
cc.add_data_validation(status_dv)
cc.add_data_validation(ctype_dv)

for i in range(3, 1003):
    bg = LT_GRY if i % 2 == 0 else WHITE
    cc.cell(i,1,i-2).fill = fill(PUR_LT); cc.cell(i,1).font = bold(9,PURPLE)
    cc.cell(i,1).alignment = ctr(); cc.cell(i,1).border = bdr()
    for c in range(2,12):
        cell = cc.cell(i,c)
        cell.fill = fill(bg); cell.border = bdr()
        cell.alignment = ctr(True if c in [6,7,8] else False)
    cc.cell(i,2).number_format = "YYYY-MM-DD"
    cc.cell(i,3,"Instagram"); platforms.add(cc.cell(i,3))
    cc.cell(i,4,"Idea"); status_dv.add(cc.cell(i,4))
    cc.cell(i,5,"Reel"); ctype_dv.add(cc.cell(i,5))

for col,w in {"A":5,"B":13,"C":13,"D":13,"E":14,"F":40,"G":30,"H":28,"I":18,"J":22,"K":20}.items():
    cc.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 3: ANALYTICS LOG (per post performance)
# ══════════════════════════════════════════════════════════════
al = wb.create_sheet("📈 Analytics Log")
al.sheet_view.showGridLines = False; al.freeze_panes = "A3"
title(al, "A1:K1", "POST ANALYTICS LOG", PINK, WHITE, 15)

hdr(al, 2, ["Date Posted","Platform","Content Type","Caption Preview",
            "Likes","Comments","Shares/Saves","Reach","Engagement Rate %",
            "Link Clicks","Notes"], bg=PINK)

for i in range(3, 1003):
    bg = LT_GRY if i % 2 == 0 else WHITE
    for c in range(1,12):
        cell = al.cell(i,c)
        cell.fill = fill(bg); cell.border = bdr(); cell.alignment = ctr()
    al.cell(i,1).number_format = "YYYY-MM-DD"
    # Engagement Rate = (Likes + Comments + Shares) / Reach
    al.cell(i,9).value = f'=IF(OR(H{i}=0,H{i}=""),"", (E{i}+F{i}+G{i})/H{i})'
    al.cell(i,9).number_format = "0.00%"
    al.cell(i,4).alignment = lft(True)

for col,w in {"A":14,"B":14,"C":14,"D":30,"E":10,"F":12,"G":14,"H":12,"I":18,"J":13,"K":22}.items():
    al.column_dimensions[col].width = w

# Summary by platform
al.merge_cells("A1005:K1005")
al["A1005"].value = "PLATFORM AVERAGES (auto-calculated below)"
al["A1005"].fill = fill(PURPLE); al["A1005"].font = bold(col=WHITE); al["A1005"].alignment = ctr()

platforms_list = ["Instagram","TikTok","Pinterest","Facebook","Twitter/X","YouTube","LinkedIn"]
hdr(al, 1006, ["Platform","Avg Likes","Avg Comments","Avg Shares",
               "Avg Reach","Avg Engagement %","Total Posts","","","",""], bg=PURPLE)
for i, plat in enumerate(platforms_list):
    r = 1007+i
    al.cell(r,1,plat).font = bold(col=PURPLE); al.cell(r,1).fill = fill(PUR_LT)
    al.cell(r,1).alignment = ctr(); al.cell(r,1).border = bdr()
    for c, col_src, fmt in [
        (2, "E", "#,##0"), (3,"F","#,##0"), (4,"G","#,##0"),
        (5,"H","#,##0"), (6,"I","0.00%"), (7,"A","0")
    ]:
        if c == 7:
            al.cell(r,c).value = f'=COUNTIF(B3:B1002,"{plat}")'
        else:
            al.cell(r,c).value = f'=IFERROR(AVERAGEIF(B3:B1002,"{plat}",{col_src}3:{col_src}1002),0)'
        al.cell(r,c).number_format = fmt
        al.cell(r,c).fill = fill(LT_GRY if i%2==0 else WHITE)
        al.cell(r,c).font = bold(col=PURPLE); al.cell(r,c).border = bdr()
        al.cell(r,c).alignment = ctr()

# ══════════════════════════════════════════════════════════════
# TAB 4: FOLLOWER GROWTH TRACKER
# ══════════════════════════════════════════════════════════════
fg = wb.create_sheet("📊 Follower Growth")
fg.sheet_view.showGridLines = False; fg.freeze_panes = "A3"
title(fg, "A1:F1", "FOLLOWER GROWTH TRACKER", PURPLE, WHITE, 15)

hdr(fg, 2, ["Platform","Current Followers","This Month +/-","Goal","% to Goal","Notes"], bg=PURPLE)
fg.row_dimensions[2].height = 30

plat_data = [
    ("Instagram",   0), ("TikTok",      0), ("Pinterest",    0),
    ("Facebook",    0), ("Twitter/X",   0), ("YouTube",      0), ("LinkedIn",   0),
]
for i, (plat, start) in enumerate(plat_data):
    r = 3+i; bg = LT_GRY if i%2==0 else WHITE
    fg.cell(r,1,plat).font = bold(col=PURPLE); fg.cell(r,1).fill = fill(PUR_LT)
    fg.cell(r,1).alignment = ctr(); fg.cell(r,1).border = bdr()
    for c in range(2,7):
        cell = fg.cell(r,c); cell.fill = fill(GOLD_LT if c in [3,5] else bg)
        cell.border = bdr(); cell.alignment = ctr()
    fg.cell(r,2).number_format = "#,##0"; fg.cell(r,3).number_format = "+#,##0;-#,##0"
    fg.cell(r,4).number_format = "#,##0"
    fg.cell(r,5).value = f"=IF(D{r}=0,\"\",C{r}/D{r})"
    fg.cell(r,5).number_format = "0.0%"; fg.cell(r,5).font = bold(col=PURPLE)

# Monthly tracking table
fg.merge_cells("A12:F12")
fg["A12"].value = "MONTHLY FOLLOWER LOG — Update at month end"
fg["A12"].fill = fill(PURPLE); fg["A12"].font = bold(12,WHITE); fg["A12"].alignment = ctr()

hdr(fg, 13, ["Month","Instagram","TikTok","Pinterest","Facebook","Twitter/X"], bg=PINK)
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for i, mo in enumerate(months):
    r = 14+i; bg = LT_GRY if i%2==0 else WHITE
    fg.cell(r,1,mo).font = bold(col=PURPLE); fg.cell(r,1).fill = fill(bg)
    fg.cell(r,1).alignment = ctr(); fg.cell(r,1).border = bdr()
    for c in range(2,7):
        cell = fg.cell(r,c); cell.fill = fill(bg)
        cell.border = bdr(); cell.number_format = "#,##0"; cell.alignment = ctr()

for col,w in {"A":16,"B":18,"C":16,"D":14,"E":12,"F":18}.items():
    fg.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 5: HASHTAG BANK
# ══════════════════════════════════════════════════════════════
hb = wb.create_sheet("# Hashtag Bank")
hb.sheet_view.showGridLines = False
title(hb, "A1:F1", "HASHTAG BANK", PURPLE, WHITE, 15)

hdr(hb, 2, ["Category / Niche","Hashtag","Platform","Est. Size","Performance","Notes"], bg=PURPLE)

categories = [
    ("Small Business","#smallbusiness","Instagram","Large","High",""),
    ("Small Business","#shopsmall","Instagram","Medium","High",""),
    ("Small Business","#entrepreneurlife","Instagram","Large","Medium",""),
    ("Etsy Seller","#etsyseller","Instagram","Medium","High",""),
    ("Etsy Seller","#etsyshop","Instagram","Large","High",""),
    ("Etsy Seller","#etsydigitaldownload","Instagram","Small","Very High","Niche, low competition"),
    ("Digital Products","#digitaldownload","Instagram","Medium","High",""),
    ("Digital Products","#passiveincome","Instagram","Large","Medium",""),
    ("Spreadsheets","#googlesheets","All","Small","Very High","Low competition"),
    ("Spreadsheets","#exceltips","TikTok","Small","Very High","viral potential"),
    ("Spreadsheets","#spreadsheettemplate","Instagram","Small","Very High",""),
    ("Finance","#budgetingmom","Instagram","Small","Very High","super niche, loyal"),
    ("Finance","#debtfreejourney","Instagram","Medium","High",""),
    ("Side Hustle","#sidehustle","Instagram","Large","Medium",""),
    ("Side Hustle","#makemoneyonline","TikTok","Large","Medium","competitive"),
    ("Content Creator","#contentcreator","Instagram","Large","Low","oversaturated"),
    ("Content Creator","#creatortips","Instagram","Small","High",""),
]
for i,(cat,tag,plat,size,perf,note) in enumerate(categories):
    r = 3+i; bg = LT_GRY if i%2==0 else WHITE
    for c,val in enumerate([cat,tag,plat,size,perf,note],1):
        cell = hb.cell(r,c,val); cell.fill = fill(bg)
        cell.border = bdr(); cell.font = (bold(col=PURPLE) if c==2 else reg())
        cell.alignment = ctr()

for col,w in {"A":22,"B":28,"C":14,"D":12,"E":14,"F":28}.items():
    hb.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════
# TAB 6: SETUP GUIDE
# ══════════════════════════════════════════════════════════════
sg = wb.create_sheet("📖 Setup Guide")
sg.sheet_view.showGridLines = False
title(sg, "A1:E1", "SETUP GUIDE", PURPLE, WHITE, 16)
sg.row_dimensions[1].height = 44

steps = [
    ("STEP 1 — Set Your Platforms & Goals",
     "Go to '📊 Follower Growth'. Enter your current follower count for each platform.\n"
     "Set a monthly follower goal for each. The % to Goal column tracks your progress automatically."),
    ("STEP 2 — Plan Your Content",
     "Go to '📅 Content Calendar'. Add your content ideas.\n"
     "Use the Status dropdown: Idea → In Progress → Scheduled → Posted.\n"
     "Fill in platform, caption/hook, hashtags, and your CTA (call to action)."),
    ("STEP 3 — Log Analytics After Each Post",
     "Go to '📈 Analytics Log'. After 24-48 hours, log your likes, comments, shares, and reach.\n"
     "The Engagement Rate calculates automatically: (Likes + Comments + Shares) ÷ Reach."),
    ("STEP 4 — Find Your Best Platform",
     "Scroll to the bottom of '📈 Analytics Log' to see the Platform Averages table.\n"
     "This shows which platform gives you the best average engagement.\n"
     "Double down on that platform. Don't spread yourself thin."),
    ("STEP 5 — Build Your Hashtag Bank",
     "Go to '# Hashtag Bank'. Add hashtags that work in your niche.\n"
     "Rate performance as Very High / High / Medium / Low based on reach they drive.\n"
     "Focus on small-to-medium size hashtags — they're less competitive."),
    ("THE 80/20 RULE FOR CONTENT",
     "80% of your results come from 20% of your content.\n"
     "After 30 days, look at your top 5 posts by engagement rate.\n"
     "What did they have in common? Do MORE of that. Stop guessing."),
]
for i,(h,b) in enumerate(steps):
    r = 3+i*4
    sg.merge_cells(f"B{r}:E{r}")
    sg.cell(r,2,h).fill = fill(PURPLE); sg.cell(r,2,h).font = bold(12,WHITE); sg.cell(r,2,h).alignment = lft()
    sg.merge_cells(f"B{r+1}:E{r+3}")
    sg.cell(r+1,2,b).fill = fill(LT_GRY if i%2==0 else WHITE)
    sg.cell(r+1,2,b).font = reg(11); sg.cell(r+1,2,b).alignment = lft(True)
    sg.row_dimensions[r+1].height = 75

sg.column_dimensions["A"].width = 3; sg.column_dimensions["B"].width = 72

dsh.sheet_properties.tabColor = PURPLE; cc.sheet_properties.tabColor  = PURPLE
al.sheet_properties.tabColor  = PINK;   fg.sheet_properties.tabColor  = PINK
hb.sheet_properties.tabColor  = PURPLE; sg.sheet_properties.tabColor  = "AAAAAA"

wb.save(OUT)
print(f"✅ Saved: {OUT}")
