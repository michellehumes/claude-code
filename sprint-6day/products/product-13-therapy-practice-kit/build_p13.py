"""
Product 13 — Therapy Practice Management Kit
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT = os.path.join(os.path.dirname(__file__), "build",
                   "Therapy_Practice_Management_Kit.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

TEAL    = "2E7D82"
TEAL_LT = "B2DFDB"
PLUM    = "6A3D7A"
PLUM_LT = "E1D5E7"
WHITE   = "FFFFFF"
LT_GRAY = "F5F5F5"
MED_GRY = "CCCCCC"
DARK    = "2C2C2C"
SAND    = "FDF6EC"
AMBER   = "F59E0B"

def fill(h): return PatternFill("solid", fgColor=h)
def bold(sz=11, col=DARK): return Font(name="Calibri", bold=True, size=sz, color=col)
def reg(sz=11, col=DARK, it=False): return Font(name="Calibri", size=sz, color=col, italic=it)
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=False): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
thin = Side(style="thin", color=MED_GRY)
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, vals, bg=TEAL, fg=WHITE, sz=10):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill(bg); cell.font = bold(sz, fg)
        cell.alignment = ctr(True); cell.border = bdr()
    ws.row_dimensions[row].height = 38

def title(ws, rng, txt, bg=TEAL, fg=WHITE, sz=16):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = txt; c.fill = fill(bg)
    c.font = bold(sz, fg); c.alignment = ctr()

wb = openpyxl.Workbook()

# ── TAB 1: DASHBOARD ─────────────────────────────────────────────────────────
dsh = wb.active; dsh.title = "📊 Dashboard"
dsh.sheet_view.showGridLines = False
dsh.row_dimensions[1].height = 50; dsh.row_dimensions[2].height = 22
title(dsh, "A1:J1", "THERAPY PRACTICE MANAGEMENT KIT", TEAL, WHITE, 18)
dsh.merge_cells("A2:J2")
dsh["A2"].value = "Organized practice. More time for your clients."
dsh["A2"].fill = fill(TEAL_LT); dsh["A2"].font = reg(11, TEAL, True); dsh["A2"].alignment = ctr()

kpis = [
    ("B4:C4","B5:C6","ACTIVE CLIENTS",   "=COUNTIF('👥 Client Roster'!E2:E201,\"Active\")", TEAL),
    ("E4:F4","E5:F6","SESSIONS THIS MO.", "=COUNTIFS('📅 Session Log'!C2:C1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📅 Session Log'!C2:C1001,\"<=\"&EOMONTH(TODAY(),0))", PLUM),
    ("H4:I4","H5:I6","REVENUE THIS MO.", "=SUMIFS('💰 Billing'!E2:E1001,'💰 Billing'!B2:B1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1))", TEAL),
    ("B8:C8","B9:C10","OUTSTANDING BALANCE","=SUMIF('💰 Billing'!F2:F1001,\"Unpaid\",'💰 Billing'!E2:E1001)", PLUM),
    ("E8:F8","E9:F10","CANCELLATIONS MTD", "=COUNTIFS('📅 Session Log'!D2:D1001,\"Cancelled\",'📅 Session Log'!C2:C1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1))", TEAL),
    ("H8:I8","H9:I10","AVG SESSION RATE", "=IFERROR(AVERAGEIF('💰 Billing'!E2:E1001,\">0\",'💰 Billing'!E2:E1001),0)", PLUM),
]
for lr, vr, lbl, frm, col in kpis:
    dsh.merge_cells(lr); dsh.merge_cells(vr)
    lc = dsh[lr.split(":")[0]]
    lc.value = lbl; lc.fill = fill(col); lc.font = bold(9, WHITE); lc.alignment = ctr(True)
    vc = dsh[vr.split(":")[0]]
    vc.value = frm; vc.fill = fill(LT_GRAY); vc.font = bold(20, col); vc.alignment = ctr()
    if "REVENUE" in lbl or "BALANCE" in lbl or "RATE" in lbl:
        vc.number_format = '"$"#,##0.00'

# Privacy notice
dsh.merge_cells("B12:I13")
dsh["B12"].value = ("🔒  PRIVACY NOTE: This spreadsheet contains sensitive client information. "
                    "Store in a password-protected Google Drive folder. "
                    "Do not share the file directly — use 'share with specific people' only.")
dsh["B12"].fill = fill(PLUM_LT); dsh["B12"].font = reg(10, PLUM, True)
dsh["B12"].alignment = ctr(True)

for col, w in {"A":2,"B":16,"C":16,"D":3,"E":20,"F":16,"G":3,"H":18,"I":16}.items():
    dsh.column_dimensions[col].width = w

# ── TAB 2: CLIENT ROSTER ──────────────────────────────────────────────────────
cr = wb.create_sheet("👥 Client Roster")
cr.sheet_view.showGridLines = False; cr.freeze_panes = "A3"
title(cr, "A1:K1", "CLIENT ROSTER", TEAL, WHITE, 15)
cr.row_dimensions[1].height = 38

hdr(cr, 2, ["Client ID","Full Name","Phone","Email","Status",
            "Insurance","Diagnosis Code","Start Date","Session Rate",
            "Preferred Day/Time","Emergency Contact"])

statuses = DataValidation(type="list",
    formula1='"Active,Inactive,Waitlist,Discharged"', allow_blank=True)
cr.add_data_validation(statuses)

for i in range(3, 203):
    bg = LT_GRAY if i % 2 == 0 else WHITE
    cr.cell(i,1,f"C{i-2:03d}").fill = fill(TEAL_LT)
    cr.cell(i,1).font = bold(col=TEAL); cr.cell(i,1).alignment = ctr()
    for c in range(2,12):
        cell = cr.cell(i,c)
        cell.fill = fill(bg); cell.border = bdr(); cell.alignment = lft()
    cr.cell(i,5,"Active"); statuses.add(cr.cell(i,5))
    cr.cell(i,5).alignment = ctr()
    cr.cell(i,9).number_format = '"$"#,##0.00'; cr.cell(i,9).alignment = ctr()
    cr.cell(i,8).number_format = "YYYY-MM-DD"; cr.cell(i,8).alignment = ctr()

for col, w in {"A":9,"B":22,"C":16,"D":26,"E":12,"F":18,"G":16,
               "H":13,"I":14,"J":20,"K":22}.items():
    cr.column_dimensions[col].width = w

# ── TAB 3: SESSION LOG ────────────────────────────────────────────────────────
sl = wb.create_sheet("📅 Session Log")
sl.sheet_view.showGridLines = False; sl.freeze_panes = "A3"
title(sl, "A1:J1", "SESSION LOG", TEAL, WHITE, 15)

hdr(sl, 2, ["Session ID","Client ID","Date","Status","Session Type",
            "Duration (min)","Notes / Goals","Next Session","Homework Assigned","Billing #"])

statuses2 = DataValidation(type="list",
    formula1='"Completed,Cancelled,No-Show,Rescheduled"', allow_blank=True)
types = DataValidation(type="list",
    formula1='"Individual,Couples,Group,Initial Assessment,Crisis"', allow_blank=True)
sl.add_data_validation(statuses2); sl.add_data_validation(types)

for i in range(3, 1003):
    bg = LT_GRAY if i % 2 == 0 else WHITE
    sl.cell(i,1,f"S{i-2:04d}").fill = fill(TEAL_LT); sl.cell(i,1).font = bold(col=TEAL)
    sl.cell(i,1).alignment = ctr()
    for c in range(2,11):
        cell = sl.cell(i,c)
        cell.fill = fill(bg); cell.border = bdr(); cell.alignment = ctr(True if c==7 else False)
    sl.cell(i,3).number_format = "YYYY-MM-DD"
    sl.cell(i,4,"Completed"); statuses2.add(sl.cell(i,4))
    sl.cell(i,5,"Individual"); types.add(sl.cell(i,5))
    sl.cell(i,8).number_format = "YYYY-MM-DD"

for col, w in {"A":10,"B":10,"C":13,"D":14,"E":18,"F":15,"G":35,"H":15,"I":22,"J":12}.items():
    sl.column_dimensions[col].width = w

# ── TAB 4: BILLING ────────────────────────────────────────────────────────────
bl = wb.create_sheet("💰 Billing")
bl.sheet_view.showGridLines = False; bl.freeze_panes = "A3"
title(bl, "A1:J1", "BILLING & PAYMENT TRACKER", PLUM, WHITE, 15)

hdr(bl, 2, ["Invoice #","Date","Client ID","Service Description",
            "Amount","Status","Date Paid","Payment Method",
            "Insurance Claim #","Notes"], bg=PLUM)

pay_status = DataValidation(type="list",
    formula1='"Paid,Unpaid,Partial,Insurance Pending,Waived"', allow_blank=True)
pay_method = DataValidation(type="list",
    formula1='"Cash,Card,Check,Venmo,Insurance,Zelle,Other"', allow_blank=True)
bl.add_data_validation(pay_status); bl.add_data_validation(pay_method)

for i in range(3, 1003):
    bg = LT_GRAY if i % 2 == 0 else WHITE
    bl.cell(i,1,f"INV{i-2:04d}").fill = fill(PLUM_LT); bl.cell(i,1).font = bold(col=PLUM)
    bl.cell(i,1).alignment = ctr()
    for c in range(2,11):
        cell = bl.cell(i,c)
        cell.fill = fill(bg); cell.border = bdr(); cell.alignment = ctr()
    bl.cell(i,2).number_format = "YYYY-MM-DD"
    bl.cell(i,5).number_format = '"$"#,##0.00'
    bl.cell(i,6,"Unpaid"); pay_status.add(bl.cell(i,6))
    bl.cell(i,7).number_format = "YYYY-MM-DD"
    pay_method.add(bl.cell(i,8))

# Monthly revenue summary
bl.cell(1004,4,"TOTAL COLLECTED →").font = bold(col=PLUM)
bl.cell(1004,5).value = '=SUMIF(F3:F1002,"Paid",E3:E1002)'
bl.cell(1004,5).number_format = '"$"#,##0.00'
bl.cell(1004,5).fill = fill(PLUM_LT); bl.cell(1004,5).font = bold(col=PLUM)
bl.cell(1004,4).fill = fill(PLUM_LT); bl.cell(1004,4).alignment = lft()

for col, w in {"A":12,"B":13,"C":11,"D":28,"E":13,"F":18,"G":13,"H":17,"I":18,"J":22}.items():
    bl.column_dimensions[col].width = w

# ── TAB 5: CASELOAD VIEW ──────────────────────────────────────────────────────
cv = wb.create_sheet("📈 Caseload View")
cv.sheet_view.showGridLines = False
title(cv, "A1:G1", "CASELOAD OVERVIEW", TEAL, WHITE, 15)

cv.merge_cells("A3:G3"); cv["A3"].value = "CASELOAD SUMMARY"
cv["A3"].fill = fill(TEAL_LT); cv["A3"].font = bold(col=TEAL); cv["A3"].alignment = ctr()

summary_rows = [
    ("Total Active Clients",   "=COUNTIF('👥 Client Roster'!E2:E201,\"Active\")"),
    ("Total Inactive Clients", "=COUNTIF('👥 Client Roster'!E2:E201,\"Inactive\")"),
    ("Waitlist",               "=COUNTIF('👥 Client Roster'!E2:E201,\"Waitlist\")"),
    ("Sessions This Month",    "=COUNTIFS('📅 Session Log'!C2:C1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📅 Session Log'!D2:D1001,\"Completed\")"),
    ("No-Shows This Month",    "=COUNTIFS('📅 Session Log'!C2:C1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📅 Session Log'!D2:D1001,\"No-Show\")"),
    ("Revenue This Month",     "=SUMIFS('💰 Billing'!E2:E1001,'💰 Billing'!B2:B1001,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'💰 Billing'!F2:F1001,\"Paid\")"),
    ("Outstanding Balance",    "=SUMIF('💰 Billing'!F2:F1001,\"Unpaid\",'💰 Billing'!E2:E1001)"),
    ("YTD Revenue",            "=SUMIFS('💰 Billing'!E2:E1001,'💰 Billing'!B2:B1001,\">=\"&DATE(YEAR(TODAY()),1,1),'💰 Billing'!F2:F1001,\"Paid\")"),
]
for i, (lbl, frm) in enumerate(summary_rows):
    r = 4+i
    cv.row_dimensions[r].height = 28
    cv.cell(r,2,lbl).font = bold(col=DARK)
    cv.cell(r,2).fill = fill(LT_GRAY if i%2==0 else WHITE)
    cv.cell(r,2).border = bdr(); cv.cell(r,2).alignment = lft()
    cv.cell(r,4).value = frm
    cv.cell(r,4).fill = fill(TEAL_LT); cv.cell(r,4).font = bold(col=TEAL)
    cv.cell(r,4).border = bdr(); cv.cell(r,4).alignment = ctr()
    if "Revenue" in lbl or "Balance" in lbl:
        cv.cell(r,4).number_format = '"$"#,##0.00'

cv.column_dimensions["A"].width = 3; cv.column_dimensions["B"].width = 30
cv.column_dimensions["C"].width = 3; cv.column_dimensions["D"].width = 18

# ── TAB 6: SETUP GUIDE ────────────────────────────────────────────────────────
sg = wb.create_sheet("📖 Setup Guide")
sg.sheet_view.showGridLines = False
title(sg, "A1:E1", "SETUP GUIDE", TEAL, WHITE, 16)

steps = [
    ("PRIVACY FIRST",
     "Upload to Google Drive and set sharing to 'Only you' or specific trusted collaborators.\n"
     "Never share the link publicly. Client data is protected health information (PHI)."),
    ("STEP 1 — Add Your Clients",
     "Go to '👥 Client Roster'. Client IDs (C001, C002...) are pre-filled.\n"
     "Add each client's info. Set status to Active, Inactive, Waitlist, or Discharged."),
    ("STEP 2 — Log Sessions",
     "After each session, add a row to '📅 Session Log'.\n"
     "Use the same Client ID from the Roster. Status dropdown: Completed, Cancelled, No-Show."),
    ("STEP 3 — Track Billing",
     "Go to '💰 Billing'. Log each invoice. Mark as Paid when payment received.\n"
     "The Dashboard automatically calculates outstanding balances and monthly revenue."),
    ("STEP 4 — Review Caseload Monthly",
     "The '📈 Caseload View' tab gives you a one-page practice summary.\n"
     "Review at the start of each month to see revenue, no-show rates, and caseload size."),
]
for i, (h, b) in enumerate(steps):
    r = 3 + i*4
    sg.merge_cells(f"B{r}:E{r}")
    sg.cell(r,2,h).fill = fill(TEAL); sg.cell(r,2,h).font = bold(12,WHITE)
    sg.cell(r,2,h).alignment = lft()
    sg.merge_cells(f"B{r+1}:E{r+3}")
    sg.cell(r+1,2,b).fill = fill(LT_GRAY if i%2==0 else WHITE)
    sg.cell(r+1,2,b).font = reg(11); sg.cell(r+1,2,b).alignment = lft(True)
    sg.row_dimensions[r+1].height = 70

sg.column_dimensions["A"].width = 3; sg.column_dimensions["B"].width = 70

dsh.sheet_properties.tabColor = TEAL; cr.sheet_properties.tabColor  = TEAL
sl.sheet_properties.tabColor  = TEAL; bl.sheet_properties.tabColor  = PLUM
cv.sheet_properties.tabColor  = PLUM; sg.sheet_properties.tabColor  = "AAAAAA"

wb.save(OUT)
print(f"✅ Saved: {OUT}")
