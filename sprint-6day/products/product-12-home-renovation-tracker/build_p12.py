"""
Product 12 — Home Renovation Budget Tracker
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "build",
                   "Home_Renovation_Budget_Tracker.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# Palette
NAVY    = "1B3A5C"
NAVY_LT = "D6E4F0"
ORANGE  = "E07B39"
ORG_LT  = "FAE5D3"
WHITE   = "FFFFFF"
LT_GRAY = "F4F4F4"
MED_GRY = "CCCCCC"
DARK    = "333333"
GREEN   = "4CAF50"
RED     = "E53935"

def fill(h): return PatternFill("solid", fgColor=h)
def bold(sz=11, col=DARK): return Font(name="Calibri", bold=True, size=sz, color=col)
def reg(sz=11, col=DARK, it=False): return Font(name="Calibri", size=sz, color=col, italic=it)
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=False): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def rgt(): return Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color=MED_GRY)
def border(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, vals, bg=NAVY, fg=WHITE, sz=10):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill(bg); cell.font = bold(sz, fg)
        cell.alignment = ctr(True); cell.border = border()

def title(ws, rng, txt, bg=NAVY, fg=WHITE, sz=16):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = txt; c.fill = fill(bg)
    c.font = bold(sz, fg); c.alignment = ctr()

wb = openpyxl.Workbook()

# ── TAB 1: DASHBOARD ─────────────────────────────────────────────────────────
d = wb.active; d.title = "🏠 Dashboard"
d.sheet_view.showGridLines = False
d.row_dimensions[1].height = 48
title(d, "A1:I1", "HOME RENOVATION BUDGET TRACKER", NAVY, WHITE, 18)
d.merge_cells("A2:I2")
d["A2"].value = "Plan smart. Avoid surprises. Stay on budget."
d["A2"].fill = fill(ORANGE); d["A2"].font = reg(11, WHITE, True); d["A2"].alignment = ctr()

# KPI boxes row 4-6
kpis = [
    ("B4:C4","B5:C6","TOTAL BUDGET",    "='📋 Room Budgets'!L2", NAVY),
    ("E4:F4","E5:F6","TOTAL SPENT",     "='📋 Room Budgets'!L3", ORANGE),
    ("H4:I4","H5:I6","REMAINING",       "='📋 Room Budgets'!L4", NAVY),
    ("B8:C8","B9:C10","# ROOMS",        "=COUNTA('📋 Room Budgets'!A5:A30)", NAVY),
    ("E8:F8","E9:F10","CONTINGENCY %",  "='📋 Room Budgets'!L5", ORANGE),
    ("H8:I8","H9:I10","% COMPLETE",     "=IF('📋 Room Budgets'!L2=0,0,'📋 Room Budgets'!L3/'📋 Room Budgets'!L2)", NAVY),
]
for lr, vr, lbl, frm, col in kpis:
    d.merge_cells(lr); d.merge_cells(vr)
    lc = d[lr.split(":")[0]]
    lc.value = lbl; lc.fill = fill(col); lc.font = bold(9, WHITE); lc.alignment = ctr(True)
    vc = d[vr.split(":")[0]]
    vc.value = frm; vc.fill = fill(LT_GRAY); vc.font = bold(22, col); vc.alignment = ctr()
    vc.number_format = '"$"#,##0' if "%" not in lbl else '0.0%'
d["H9"].number_format = "0.0%"

# Over/under budget alert
d.merge_cells("B12:I13")
d["B12"].value = ('=IF(\'📋 Room Budgets\'!L4<0,'
                  '"⚠  OVER BUDGET BY $"&TEXT(ABS(\'📋 Room Budgets\'!L4),"#,##0")&"  — review your room budgets",'
                  '"✅  ON TRACK — $"&TEXT(\'📋 Room Budgets\'!L4,"#,##0")&" remaining")')
d["B12"].fill = fill(ORG_LT); d["B12"].font = bold(13, ORANGE)
d["B12"].alignment = ctr(True)

# Column widths
for col, w in {"A":2,"B":14,"C":14,"D":3,"E":14,"F":14,"G":3,"H":14,"I":14}.items():
    d.column_dimensions[col].width = w

# ── TAB 2: ROOM BUDGETS ───────────────────────────────────────────────────────
rb = wb.create_sheet("📋 Room Budgets")
rb.sheet_view.showGridLines = False; rb.freeze_panes = "A5"
rb.row_dimensions[1].height = 40; rb.row_dimensions[2].height = 22

title(rb, "A1:K1", "ROOM-BY-ROOM BUDGET PLANNER", NAVY, WHITE, 15)

# Summary totals (row 2, used by Dashboard)
rb.merge_cells("A2:D2"); rb["A2"].value = "TOTALS →"
rb["A2"].fill = fill(NAVY_LT); rb["A2"].font = bold(col=NAVY); rb["A2"].alignment = lft()
summaries = [
    ("L2","=SUM(G5:G30)","TOTAL BUDGET"),
    ("L3","=SUM(H5:H30)","TOTAL SPENT"),
    ("L4","=L2-L3",      "REMAINING"),
    ("L5","=IF(L2=0,0,M2/L2)","CONTINGENCY %"),
    ("M2","=L2*0.15",    "CONTINGENCY $"),
]
for ref, frm, lbl in summaries:
    rb[ref].value = frm
    rb[ref].number_format = '"$"#,##0' if "%" not in lbl else "0.0%"
    rb[ref].fill = fill(NAVY_LT); rb[ref].font = bold(col=NAVY)

hdr(rb, 4, ["Room / Area","Scope of Work","Contractor","Status",
            "Labor Est.","Materials Est.","TOTAL BUDGET",
            "Amount Paid","Amount Remaining","Start Date","End Date"])
rb.row_dimensions[4].height = 36

rooms = ["Kitchen","Master Bathroom","Guest Bathroom","Living Room","Master Bedroom",
         "Basement","Backyard / Deck","Roof","Windows","Flooring","HVAC","Other"]
for i, room in enumerate(rooms):
    r = 5 + i; bg = LT_GRAY if i % 2 == 0 else WHITE
    rb.cell(r,1,room).font = bold(col=NAVY)
    # Status dropdown
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1='"Not Started,In Progress,Complete,On Hold"', allow_blank=True)
    rb.add_data_validation(dv); dv.add(rb.cell(r,4))
    rb.cell(r,4,"Not Started")
    # Formulas
    rb.cell(r,7).value = f"=IF(E{r}=\"\",\"\",E{r}+F{r})"
    rb.cell(r,9).value = f"=IF(G{r}=\"\",\"\",G{r}-H{r})"
    for c in range(1,12):
        cell = rb.cell(r,c)
        cell.fill = fill(bg); cell.border = border(); cell.alignment = ctr(True if c in [2,3] else False)
    for c in [5,6,7,8,9]:
        rb.cell(r,c).number_format = '"$"#,##0'

for col, w in {"A":20,"B":28,"C":20,"D":13,"E":13,"F":15,"G":15,
               "H":14,"I":17,"J":13,"K":13,"L":1,"M":1}.items():
    rb.column_dimensions[col].width = w

# ── TAB 3: CONTRACTOR BIDS ────────────────────────────────────────────────────
cb = wb.create_sheet("🔨 Contractor Bids")
cb.sheet_view.showGridLines = False; cb.freeze_panes = "A4"
title(cb, "A1:J1", "CONTRACTOR BID COMPARISON", NAVY, WHITE, 15)
cb.row_dimensions[1].height = 40

hdr(cb, 3, ["Room / Project","Contractor Name","Phone / Email",
            "Bid Amount","Includes Materials?","Est. Duration",
            "License #","References?","Rating (1-5)","Notes"])
cb.row_dimensions[3].height = 36

for i in range(4, 34):
    bg = LT_GRAY if i % 2 == 0 else WHITE
    for c in range(1, 11):
        cell = cb.cell(i, c)
        cell.fill = fill(bg); cell.border = border(); cell.alignment = ctr(True if c in [2,3,10] else False)
    cb.cell(i,4).number_format = '"$"#,##0'

# Best bid formula row
cb.merge_cells("A35:C35"); cb["A35"].value = "LOWEST BID →"
cb["A35"].fill = fill(NAVY_LT); cb["A35"].font = bold(col=NAVY)
cb["D35"].value = "=MIN(D4:D33)"
cb["D35"].number_format = '"$"#,##0'
cb["D35"].fill = fill(ORG_LT); cb["D35"].font = bold(col=ORANGE)

for col, w in {"A":22,"B":22,"C":24,"D":14,"E":18,"F":15,
               "G":14,"H":13,"I":12,"J":28}.items():
    cb.column_dimensions[col].width = w

# ── TAB 4: COST TRACKER ───────────────────────────────────────────────────────
ct = wb.create_sheet("💳 Cost Tracker")
ct.sheet_view.showGridLines = False; ct.freeze_panes = "A4"
title(ct, "A1:I1", "DETAILED COST TRACKER", NAVY, WHITE, 15)
ct.row_dimensions[1].height = 40

hdr(ct, 3, ["Date","Room","Category","Vendor / Store",
            "Description","Amount","Payment Method","Receipt #","Notes"])
ct.row_dimensions[3].height = 36

cats = ["Labor","Materials","Fixtures","Appliances","Permits","Design","Other"]
dv_cat = openpyxl.worksheet.datavalidation.DataValidation(
    type="list", formula1='"%s"' % ",".join(cats), allow_blank=True)
ct.add_data_validation(dv_cat)

for i in range(4, 504):
    bg = LT_GRAY if i % 2 == 0 else WHITE
    for c in range(1, 10):
        cell = ct.cell(i, c)
        cell.fill = fill(bg); cell.border = border(); cell.alignment = ctr()
    ct.cell(i,1).number_format = "YYYY-MM-DD"
    ct.cell(i,6).number_format = '"$"#,##0.00'
    dv_cat.add(ct.cell(i,3))

# Totals
ct.cell(503,5,"TOTAL SPENT →").font = bold(col=NAVY)
ct.cell(503,6).value = "=SUM(F4:F502)"
ct.cell(503,6).number_format = '"$"#,##0.00'
ct.cell(503,6).fill = fill(ORG_LT); ct.cell(503,6).font = bold(col=ORANGE)

for col, w in {"A":13,"B":18,"C":14,"D":22,"E":30,"F":13,"G":18,"H":14,"I":22}.items():
    ct.column_dimensions[col].width = w

# ── TAB 5: SETUP GUIDE ────────────────────────────────────────────────────────
sg = wb.create_sheet("📖 Setup Guide")
sg.sheet_view.showGridLines = False
title(sg, "A1:E1", "HOW TO USE THIS TRACKER", ORANGE, WHITE, 16)
sg.row_dimensions[1].height = 44

steps = [
    ("STEP 1 — List Your Rooms",
     "Go to '📋 Room Budgets'. Add each room or area you're renovating in Column A.\n"
     "Fill in the scope of work, contractor name, and your budget for labor + materials."),
    ("STEP 2 — Compare Contractor Bids",
     "Go to '🔨 Contractor Bids'. Enter all bids you receive for each project.\n"
     "The bottom row automatically highlights the lowest bid for easy comparison."),
    ("STEP 3 — Log Every Purchase",
     "Every time you spend money — materials, labor, permits — log it in '💳 Cost Tracker'.\n"
     "This is what keeps you from going over budget."),
    ("STEP 4 — Check Your Dashboard",
     "The '🏠 Dashboard' tab shows your total budget, total spent, and remaining at a glance.\n"
     "The alert banner turns orange if you go over budget."),
    ("PRO TIP — The 15% Contingency Rule",
     "Renovations almost always run over. Budget an extra 15% for surprises.\n"
     "The dashboard shows your contingency amount automatically based on your total budget."),
]
for i, (h, b) in enumerate(steps):
    r = 3 + i*4
    sg.merge_cells(f"B{r}:E{r}")
    sg.cell(r,2,h).fill = fill(NAVY); sg.cell(r,2,h).font = bold(12,WHITE)
    sg.cell(r,2,h).alignment = lft()
    sg.merge_cells(f"B{r+1}:E{r+3}")
    sg.cell(r+1,2,b).fill = fill(LT_GRAY if i%2==0 else WHITE)
    sg.cell(r+1,2,b).font = reg(11); sg.cell(r+1,2,b).alignment = lft(True)
    sg.row_dimensions[r+1].height = 70

sg.column_dimensions["A"].width = 3
sg.column_dimensions["B"].width = 70
for col in ["C","D","E"]: sg.column_dimensions[col].width = 15

# Tab colors
d.sheet_properties.tabColor   = NAVY
rb.sheet_properties.tabColor  = NAVY
cb.sheet_properties.tabColor  = ORANGE
ct.sheet_properties.tabColor  = ORANGE
sg.sheet_properties.tabColor  = "AAAAAA"

wb.save(OUT)
print(f"✅ Saved: {OUT}")
