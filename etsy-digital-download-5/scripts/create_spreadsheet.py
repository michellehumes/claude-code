#!/usr/bin/env python3
"""Generate the Resume & Job Search Tracker 2026 Excel spreadsheet."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from copy import copy

OUTPUT = "/home/user/claude-code/etsy-digital-download-5/product/Resume_Job_Search_Tracker_2026.xlsx"

# ── Color palette ──────────────────────────────────────────────────
DEEP_BLUE   = "1A365D"
SLATE       = "4A5568"
TEAL        = "2B6CB0"
LIGHT_BLUE  = "BEE3F8"
ICE         = "EBF8FF"
CORAL       = "E53E3E"
GOLD        = "D69E2E"
WHITE       = "FFFFFF"

# ── Reusable style objects ─────────────────────────────────────────
header_font    = Font(name="Calibri", bold=True, color=WHITE, size=12)
header_fill    = PatternFill("solid", fgColor=DEEP_BLUE)
sub_font       = Font(name="Calibri", bold=True, color=DEEP_BLUE, size=11)
sub_fill       = PatternFill("solid", fgColor=LIGHT_BLUE)
normal_font    = Font(name="Calibri", size=11, color=SLATE)
alt_fill_a     = PatternFill("solid", fgColor=WHITE)
alt_fill_b     = PatternFill("solid", fgColor=ICE)
accent_fill    = PatternFill("solid", fgColor=TEAL)
accent_font    = Font(name="Calibri", bold=True, color=WHITE, size=11)
thin_border    = Border(
    left=Side(style="thin", color=LIGHT_BLUE),
    right=Side(style="thin", color=LIGHT_BLUE),
    top=Side(style="thin", color=LIGHT_BLUE),
    bottom=Side(style="thin", color=LIGHT_BLUE),
)
header_border  = Border(
    left=Side(style="thin", color=DEEP_BLUE),
    right=Side(style="thin", color=DEEP_BLUE),
    top=Side(style="thin", color=DEEP_BLUE),
    bottom=Side(style="medium", color=DEEP_BLUE),
)
title_font     = Font(name="Calibri", bold=True, color=WHITE, size=16)
title_fill     = PatternFill("solid", fgColor=DEEP_BLUE)


def style_header_row(ws, headers, row=1, start_col=1):
    """Apply header styling to a row."""
    for ci, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = header_border


def style_data_rows(ws, start_row, end_row, num_cols, start_col=1):
    """Apply alternating-row styling."""
    for r in range(start_row, end_row + 1):
        fill = alt_fill_b if (r - start_row) % 2 == 0 else alt_fill_a
        for ci in range(start_col, start_col + num_cols):
            c = ws.cell(row=r, column=ci)
            c.font = normal_font
            c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)


def add_title_banner(ws, title, num_cols):
    """Merge top row into a big banner."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40


def auto_col_width(ws, headers, start_col=1, header_row=2, min_w=14, max_w=35):
    """Set column widths based on header text."""
    for ci, h in enumerate(headers, start=start_col):
        w = max(min_w, min(max_w, len(str(h)) * 1.5 + 4))
        ws.column_dimensions[get_column_letter(ci)].width = w


# ════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── TAB 1: Job Application Tracker ────────────────────────────────
ws1 = wb.active
ws1.title = "Job Application Tracker"
ws1.sheet_properties.tabColor = DEEP_BLUE

headers1 = [
    "Company", "Position", "Date Applied", "Source",
    "Status", "Salary Range", "Contact Person",
    "Follow-up Date", "Notes"
]
add_title_banner(ws1, "JOB APPLICATION TRACKER 2026", len(headers1))
style_header_row(ws1, headers1, row=2)
auto_col_width(ws1, headers1, header_row=2)

DATA_START = 3
DATA_END = DATA_START + 49  # 50 rows

# Source dropdown
dv_source = DataValidation(
    type="list",
    formula1='"LinkedIn,Indeed,Glassdoor,Referral,Company Website,Recruiter,Job Fair,Other"',
    allow_blank=True,
)
dv_source.prompt = "Select the job source"
dv_source.promptTitle = "Source"
ws1.add_data_validation(dv_source)
dv_source.add(f"D{DATA_START}:D{DATA_END}")

# Status dropdown
dv_status = DataValidation(
    type="list",
    formula1='"Applied,Phone Screen,Interview,Technical,Final Round,Offer,Rejected,Ghosted,Withdrawn"',
    allow_blank=True,
)
dv_status.prompt = "Select application status"
ws1.add_data_validation(dv_status)
dv_status.add(f"E{DATA_START}:E{DATA_END}")

style_data_rows(ws1, DATA_START, DATA_END, len(headers1))

# Date format
for r in range(DATA_START, DATA_END + 1):
    ws1.cell(row=r, column=3).number_format = "YYYY-MM-DD"
    ws1.cell(row=r, column=8).number_format = "YYYY-MM-DD"

# Summary section
SUM_ROW = DATA_END + 2
ws1.merge_cells(start_row=SUM_ROW, start_column=1, end_row=SUM_ROW, end_column=len(headers1))
c = ws1.cell(row=SUM_ROW, column=1, value="APPLICATION SUMMARY")
c.font = accent_font; c.fill = accent_fill
c.alignment = Alignment(horizontal="center")

labels = [
    ("Total Applications", f'=COUNTA(A{DATA_START}:A{DATA_END})'),
    ("Phone Screens",      f'=COUNTIF(E{DATA_START}:E{DATA_END},"Phone Screen")'),
    ("Interviews",         f'=COUNTIF(E{DATA_START}:E{DATA_END},"Interview")+COUNTIF(E{DATA_START}:E{DATA_END},"Technical")+COUNTIF(E{DATA_START}:E{DATA_END},"Final Round")'),
    ("Offers",             f'=COUNTIF(E{DATA_START}:E{DATA_END},"Offer")'),
    ("Rejected",           f'=COUNTIF(E{DATA_START}:E{DATA_END},"Rejected")'),
    ("Ghosted",            f'=COUNTIF(E{DATA_START}:E{DATA_END},"Ghosted")'),
    ("Response Rate",      f'=IF(COUNTA(A{DATA_START}:A{DATA_END})>0,(COUNTA(A{DATA_START}:A{DATA_END})-COUNTIF(E{DATA_START}:E{DATA_END},"Applied")-COUNTIF(E{DATA_START}:E{DATA_END},"Ghosted"))/COUNTA(A{DATA_START}:A{DATA_END}),0)'),
]
for i, (lbl, formula) in enumerate(labels):
    r = SUM_ROW + 1 + i
    lc = ws1.cell(row=r, column=1, value=lbl)
    lc.font = sub_font; lc.fill = sub_fill; lc.border = thin_border
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    vc = ws1.cell(row=r, column=4, value=formula if formula.startswith("=") else formula)
    vc.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center")
    vc.border = thin_border
    if "Rate" in lbl:
        vc.number_format = "0.0%"

# Conditional formatting for status
from openpyxl.formatting.rule import CellIsRule
ws1.conditional_formatting.add(
    f"E{DATA_START}:E{DATA_END}",
    CellIsRule(operator="equal", formula=['"Offer"'],
               fill=PatternFill("solid", fgColor="C6F6D5"),
               font=Font(color="276749", bold=True))
)
ws1.conditional_formatting.add(
    f"E{DATA_START}:E{DATA_END}",
    CellIsRule(operator="equal", formula=['"Rejected"'],
               fill=PatternFill("solid", fgColor="FED7D7"),
               font=Font(color=CORAL, bold=True))
)
ws1.conditional_formatting.add(
    f"E{DATA_START}:E{DATA_END}",
    CellIsRule(operator="equal", formula=['"Interview"'],
               fill=PatternFill("solid", fgColor="FEFCBF"),
               font=Font(color=GOLD, bold=True))
)

ws1.freeze_panes = "A3"

# ── TAB 2: Resume Versions ────────────────────────────────────────
ws2 = wb.create_sheet("Resume Versions")
ws2.sheet_properties.tabColor = TEAL

headers2 = [
    "Version Name", "Target Role", "Date Created",
    "Keywords Included", "Where Submitted", "Success Rate (%)", "Notes"
]
add_title_banner(ws2, "RESUME VERSIONS TRACKER", len(headers2))
style_header_row(ws2, headers2, row=2)
auto_col_width(ws2, headers2, header_row=2)
style_data_rows(ws2, 3, 52, len(headers2))
for r in range(3, 53):
    ws2.cell(row=r, column=3).number_format = "YYYY-MM-DD"
    ws2.cell(row=r, column=6).number_format = '0.0"%"'
ws2.column_dimensions["D"].width = 40
ws2.column_dimensions["E"].width = 30
ws2.freeze_panes = "A3"

# Summary
sr = 54
ws2.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers2))
c = ws2.cell(row=sr, column=1, value="RESUME SUMMARY")
c.font = accent_font; c.fill = accent_fill; c.alignment = Alignment(horizontal="center")
for i, (lbl, frm) in enumerate([
    ("Total Versions", "=COUNTA(A3:A52)"),
    ("Avg Success Rate", "=IF(COUNTA(F3:F52)>0,AVERAGE(F3:F52),0)"),
]):
    r = sr + 1 + i
    lc = ws2.cell(row=r, column=1, value=lbl)
    lc.font = sub_font; lc.fill = sub_fill; lc.border = thin_border
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    vc = ws2.cell(row=r, column=4, value=frm)
    vc.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center"); vc.border = thin_border

# ── TAB 3: Interview Prep ─────────────────────────────────────────
ws3 = wb.create_sheet("Interview Prep")
ws3.sheet_properties.tabColor = GOLD

headers3 = [
    "Company", "Position", "Interview Date", "Type",
    "Interviewer Name", "Questions Asked", "My Answers",
    "Follow-Up Sent", "Result"
]
add_title_banner(ws3, "INTERVIEW PREPARATION TRACKER", len(headers3))
style_header_row(ws3, headers3, row=2)
auto_col_width(ws3, headers3, header_row=2)
style_data_rows(ws3, 3, 52, len(headers3))

dv_type = DataValidation(
    type="list",
    formula1='"Phone Screen,Video Call,In-Person,Technical,Panel,Case Study"',
    allow_blank=True,
)
ws3.add_data_validation(dv_type)
dv_type.add("D3:D52")

dv_fu = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
ws3.add_data_validation(dv_fu)
dv_fu.add("H3:H52")

dv_result = DataValidation(
    type="list",
    formula1='"Passed,Failed,Pending,Moved Forward,No Response"',
    allow_blank=True,
)
ws3.add_data_validation(dv_result)
dv_result.add("I3:I52")

for r in range(3, 53):
    ws3.cell(row=r, column=3).number_format = "YYYY-MM-DD"
ws3.column_dimensions["F"].width = 40
ws3.column_dimensions["G"].width = 40
ws3.freeze_panes = "A3"

# Summary
sr = 54
ws3.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers3))
c = ws3.cell(row=sr, column=1, value="INTERVIEW SUMMARY")
c.font = accent_font; c.fill = accent_fill; c.alignment = Alignment(horizontal="center")
for i, (lbl, frm) in enumerate([
    ("Total Interviews", "=COUNTA(A3:A52)"),
    ("Passed", '=COUNTIF(I3:I52,"Passed")+COUNTIF(I3:I52,"Moved Forward")'),
    ("Pending", '=COUNTIF(I3:I52,"Pending")'),
    ("Follow-Ups Sent", '=COUNTIF(H3:H52,"Yes")'),
]):
    r = sr + 1 + i
    lc = ws3.cell(row=r, column=1, value=lbl)
    lc.font = sub_font; lc.fill = sub_fill; lc.border = thin_border
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    vc = ws3.cell(row=r, column=4, value=frm)
    vc.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center"); vc.border = thin_border

# ── TAB 4: Networking Tracker ─────────────────────────────────────
ws4 = wb.create_sheet("Networking Tracker")
ws4.sheet_properties.tabColor = CORAL

headers4 = [
    "Contact Name", "Company", "Title", "How We Met",
    "Date", "Email", "Phone", "LinkedIn URL",
    "Last Contact", "Notes", "Follow-Up"
]
add_title_banner(ws4, "NETWORKING & CONTACTS TRACKER", len(headers4))
style_header_row(ws4, headers4, row=2)
auto_col_width(ws4, headers4, header_row=2)
style_data_rows(ws4, 3, 52, len(headers4))

dv_followup = DataValidation(
    type="list",
    formula1='"This Week,Next Week,This Month,Quarterly,As Needed,Done"',
    allow_blank=True,
)
ws4.add_data_validation(dv_followup)
dv_followup.add("K3:K52")

for r in range(3, 53):
    ws4.cell(row=r, column=5).number_format = "YYYY-MM-DD"
    ws4.cell(row=r, column=9).number_format = "YYYY-MM-DD"
ws4.freeze_panes = "A3"

sr = 54
ws4.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers4))
c = ws4.cell(row=sr, column=1, value="NETWORKING SUMMARY")
c.font = accent_font; c.fill = accent_fill; c.alignment = Alignment(horizontal="center")
for i, (lbl, frm) in enumerate([
    ("Total Contacts", "=COUNTA(A3:A52)"),
    ("Need Follow-Up This Week", '=COUNTIF(K3:K52,"This Week")'),
    ("Need Follow-Up This Month", '=COUNTIF(K3:K52,"This Month")'),
]):
    r = sr + 1 + i
    lc = ws4.cell(row=r, column=1, value=lbl)
    lc.font = sub_font; lc.fill = sub_fill; lc.border = thin_border
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    vc = ws4.cell(row=r, column=4, value=frm)
    vc.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center"); vc.border = thin_border

# ── TAB 5: Salary Research ────────────────────────────────────────
ws5 = wb.create_sheet("Salary Research")
ws5.sheet_properties.tabColor = GOLD

headers5 = [
    "Company", "Position", "Glassdoor Range", "LinkedIn Range",
    "Levels.fyi", "Avg Salary", "Benefits Notes", "Negotiation Target"
]
add_title_banner(ws5, "SALARY & COMPENSATION RESEARCH", len(headers5))
style_header_row(ws5, headers5, row=2)
auto_col_width(ws5, headers5, header_row=2)
style_data_rows(ws5, 3, 52, len(headers5))

for r in range(3, 53):
    for ci in [3, 4, 5, 6, 8]:
        ws5.cell(row=r, column=ci).number_format = '$#,##0'
ws5.column_dimensions["G"].width = 35
ws5.freeze_panes = "A3"

sr = 54
ws5.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers5))
c = ws5.cell(row=sr, column=1, value="SALARY SUMMARY")
c.font = accent_font; c.fill = accent_fill; c.alignment = Alignment(horizontal="center")
for i, (lbl, frm) in enumerate([
    ("Companies Researched", "=COUNTA(A3:A52)"),
    ("Avg Target Salary", "=IF(COUNTA(F3:F52)>0,AVERAGE(F3:F52),0)"),
    ("Max Negotiation Target", "=IF(COUNTA(H3:H52)>0,MAX(H3:H52),0)"),
]):
    r = sr + 1 + i
    lc = ws5.cell(row=r, column=1, value=lbl)
    lc.font = sub_font; lc.fill = sub_fill; lc.border = thin_border
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    vc = ws5.cell(row=r, column=4, value=frm)
    vc.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center"); vc.border = thin_border
    if "Salary" in lbl or "Target" in lbl:
        vc.number_format = '$#,##0'

# ── TAB 6: Skills & Certifications ────────────────────────────────
ws6 = wb.create_sheet("Skills & Certifications")
ws6.sheet_properties.tabColor = TEAL

headers6 = [
    "Skill Name", "Category", "Proficiency",
    "Certification", "Expiry Date", "Renewal Needed"
]
add_title_banner(ws6, "SKILLS & CERTIFICATIONS LOG", len(headers6))
style_header_row(ws6, headers6, row=2)
auto_col_width(ws6, headers6, header_row=2)
style_data_rows(ws6, 3, 52, len(headers6))

dv_cat = DataValidation(
    type="list",
    formula1='"Technical,Soft Skill,Language,Design,Management,Industry Knowledge"',
    allow_blank=True,
)
ws6.add_data_validation(dv_cat)
dv_cat.add("B3:B52")

dv_prof = DataValidation(
    type="list",
    formula1='"Beginner,Intermediate,Advanced,Expert"',
    allow_blank=True,
)
ws6.add_data_validation(dv_prof)
dv_prof.add("C3:C52")

dv_renew = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
ws6.add_data_validation(dv_renew)
dv_renew.add("F3:F52")

for r in range(3, 53):
    ws6.cell(row=r, column=5).number_format = "YYYY-MM-DD"
ws6.freeze_panes = "A3"

sr = 54
ws6.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers6))
c = ws6.cell(row=sr, column=1, value="SKILLS SUMMARY")
c.font = accent_font; c.fill = accent_fill; c.alignment = Alignment(horizontal="center")
for i, (lbl, frm) in enumerate([
    ("Total Skills", "=COUNTA(A3:A52)"),
    ("Technical", '=COUNTIF(B3:B52,"Technical")'),
    ("Soft Skills", '=COUNTIF(B3:B52,"Soft Skill")'),
    ("Expert Level", '=COUNTIF(C3:C52,"Expert")'),
    ("Renewals Due", '=COUNTIF(F3:F52,"Yes")'),
]):
    r = sr + 1 + i
    lc = ws6.cell(row=r, column=1, value=lbl)
    lc.font = sub_font; lc.fill = sub_fill; lc.border = thin_border
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    vc = ws6.cell(row=r, column=3, value=frm)
    vc.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center"); vc.border = thin_border

# ── TAB 7: Weekly Job Search Plan ─────────────────────────────────
ws7 = wb.create_sheet("Weekly Job Search Plan")
ws7.sheet_properties.tabColor = DEEP_BLUE

headers7 = [
    "Day", "Activities Planned", "Hours Spent",
    "Applications Sent", "Networking Done",
    "Skills Worked On", "Notes"
]
add_title_banner(ws7, "WEEKLY JOB SEARCH PLANNER", len(headers7))

# Create 8 weeks of planning space
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
current_row = 2
for week_num in range(1, 9):
    # Week header
    ws7.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers7))
    c = ws7.cell(row=current_row, column=1, value=f"WEEK {week_num}")
    c.font = accent_font; c.fill = accent_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[current_row].height = 28
    current_row += 1

    style_header_row(ws7, headers7, row=current_row)
    current_row += 1

    for di, day in enumerate(days):
        ws7.cell(row=current_row, column=1, value=day)
        fill = alt_fill_b if di % 2 == 0 else alt_fill_a
        for ci in range(1, len(headers7) + 1):
            c = ws7.cell(row=current_row, column=ci)
            c.font = normal_font; c.fill = fill; c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        current_row += 1

    # Weekly summary row
    sum_row = current_row
    ws7.cell(row=sum_row, column=1, value="WEEKLY TOTAL").font = sub_font
    ws7.cell(row=sum_row, column=1).fill = sub_fill
    ws7.cell(row=sum_row, column=1).border = thin_border
    
    data_start_row = sum_row - 7
    ws7.cell(row=sum_row, column=3, value=f"=SUM(C{data_start_row}:C{sum_row-1})")
    ws7.cell(row=sum_row, column=3).font = sub_font; ws7.cell(row=sum_row, column=3).fill = sub_fill
    ws7.cell(row=sum_row, column=3).border = thin_border
    ws7.cell(row=sum_row, column=4, value=f"=SUM(D{data_start_row}:D{sum_row-1})")
    ws7.cell(row=sum_row, column=4).font = sub_font; ws7.cell(row=sum_row, column=4).fill = sub_fill
    ws7.cell(row=sum_row, column=4).border = thin_border
    
    for ci in range(1, len(headers7) + 1):
        c = ws7.cell(row=sum_row, column=ci)
        if not c.font or not c.font.bold:
            c.font = sub_font
        c.fill = sub_fill; c.border = thin_border

    current_row += 2  # blank row between weeks

auto_col_width(ws7, headers7, header_row=3)
ws7.column_dimensions["B"].width = 35
ws7.column_dimensions["F"].width = 25

# ── TAB 8: How to Use ─────────────────────────────────────────────
ws8 = wb.create_sheet("How to Use")
ws8.sheet_properties.tabColor = SLATE

ws8.merge_cells("A1:F1")
c = ws8.cell(row=1, column=1, value="HOW TO USE YOUR RESUME & JOB SEARCH TRACKER")
c.font = Font(name="Calibri", bold=True, color=WHITE, size=18)
c.fill = title_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws8.row_dimensions[1].height = 50

instructions = [
    ("", ""),
    ("GETTING STARTED", ""),
    ("Welcome!", "Thank you for purchasing the Resume & Job Search Tracker 2026! This spreadsheet is designed to organize every aspect of your job search in one place. Below you'll find instructions for each tab."),
    ("", ""),
    ("TAB 1: JOB APPLICATION TRACKER", ""),
    ("Purpose:", "Track every job you apply to with dates, sources, and status updates."),
    ("How to use:", "Enter each application as a new row. Use the dropdown menus for Source and Status. The summary section at the bottom auto-calculates your response rate and totals."),
    ("Pro tip:", "Update the Status column regularly. Set Follow-up Dates 5-7 business days after applying."),
    ("", ""),
    ("TAB 2: RESUME VERSIONS", ""),
    ("Purpose:", "Keep track of different resume versions tailored to different roles."),
    ("How to use:", "Create a new row each time you customize your resume. Track which keywords you included and where you submitted it. Monitor success rates to refine your approach."),
    ("", ""),
    ("TAB 3: INTERVIEW PREP", ""),
    ("Purpose:", "Prepare for and track all your interviews."),
    ("How to use:", "Before each interview, fill in the company and interviewer details. After the interview, document questions asked and your answers for future reference. Mark follow-up status."),
    ("Pro tip:", "Review this tab before second-round interviews to recall what you discussed earlier."),
    ("", ""),
    ("TAB 4: NETWORKING TRACKER", ""),
    ("Purpose:", "Build and maintain your professional network during your search."),
    ("How to use:", "Add every meaningful professional contact. Note how you met and set follow-up reminders. Regular networking is the #1 way to land interviews."),
    ("", ""),
    ("TAB 5: SALARY RESEARCH", ""),
    ("Purpose:", "Research and compare compensation across companies."),
    ("How to use:", "Before applying or interviewing, research salary ranges from multiple sources. Use the Negotiation Target column to set your ask price."),
    ("", ""),
    ("TAB 6: SKILLS & CERTIFICATIONS", ""),
    ("Purpose:", "Inventory your skills and track certifications."),
    ("How to use:", "List all relevant skills and certifications. Track expiry dates and renewal needs. Identify skill gaps for roles you're targeting."),
    ("", ""),
    ("TAB 7: WEEKLY JOB SEARCH PLAN", ""),
    ("Purpose:", "Structure your weekly job search activities."),
    ("How to use:", "Plan each day's activities at the start of the week. Track hours spent and applications sent. Weekly totals help you stay accountable."),
    ("Pro tip:", "Aim for 3-5 quality applications per day rather than mass-applying."),
    ("", ""),
    ("GENERAL TIPS", ""),
    ("Consistency:", "Update this tracker daily for best results."),
    ("Backup:", "Save a copy to Google Drive or OneDrive for cloud backup."),
    ("Customize:", "Feel free to add columns or modify dropdowns to fit your needs."),
    ("", ""),
    ("Good luck with your job search! You've got this!", ""),
]

for ri, (col_a, col_b) in enumerate(instructions, start=2):
    ca = ws8.cell(row=ri, column=1, value=col_a)
    cb = ws8.cell(row=ri, column=2, value=col_b)

    if col_a and not col_b and col_a.isupper():
        # Section header
        ws8.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
        ca.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
        ca.fill = PatternFill("solid", fgColor=TEAL)
        ca.alignment = Alignment(horizontal="left", vertical="center")
        ws8.row_dimensions[ri].height = 30
    elif col_a.startswith("Pro tip") or col_a.startswith("Welcome"):
        ca.font = Font(name="Calibri", bold=True, color=GOLD, size=11)
        ws8.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=6)
        cb.font = normal_font
        cb.alignment = Alignment(wrap_text=True)
    elif col_a and col_b:
        ca.font = sub_font
        ws8.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=6)
        cb.font = normal_font
        cb.alignment = Alignment(wrap_text=True)
    elif col_a and not col_b and not col_a.isupper():
        ws8.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
        ca.font = Font(name="Calibri", bold=True, color=TEAL, size=12)
        ca.alignment = Alignment(horizontal="center")

ws8.column_dimensions["A"].width = 20
ws8.column_dimensions["B"].width = 80

# ── Save ───────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Tabs: {wb.sheetnames}")
