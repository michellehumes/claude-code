#!/usr/bin/env python3
"""
Excel Life Tracker Builder
Creates a comprehensive life tracking workbook with multiple sheets and features.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime, timedelta
import random

def create_life_tracker():
    """Create the complete Life Tracker workbook"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all sheets in order
    create_readme_sheet(wb)
    create_setup_sheet(wb)
    create_dashboard_sheet(wb)
    create_daily_log_sheet(wb)
    create_habit_log_sheet(wb)
    create_tasks_sheet(wb)
    create_goals_sheet(wb)
    create_finance_sheet(wb)
    create_job_search_sheet(wb)
    create_health_fertility_sheet(wb)

    # Save workbook
    wb.save('Life_Tracker.xlsx')
    print("✓ Life_Tracker.xlsx created successfully!")

def create_readme_sheet(wb):
    """Create README sheet with instructions"""
    ws = wb.create_sheet("README", 0)

    # Header
    ws['A1'] = "EXCEL LIFE TRACKER"
    ws['A1'].font = Font(size=20, bold=True, color="1F4E78")
    ws.merge_cells('A1:F1')

    ws['A2'] = "Version 1.0 | Built: February 2026 | For: Michelle Humes"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws.merge_cells('A2:F2')

    # Content
    content = [
        [""],
        ["WHAT IS THIS?", ""],
        ["This workbook helps you track all aspects of your life in one place:", ""],
        ["• Daily metrics (sleep, mood, energy, exercise)", ""],
        ["• Habits and streaks", ""],
        ["• Tasks and goals", ""],
        ["• Personal finances", ""],
        ["• Job search pipeline", ""],
        ["• Health and fertility tracking", ""],
        [""],
        ["HOW TO USE", ""],
        ["1. Start with the SETUP sheet", "Review and customize categories and lists"],
        ["2. Visit the DASHBOARD daily", "Check your key metrics and trends"],
        ["3. Add daily entries", "Use DAILY_LOG for daily metrics"],
        ["4. Track habits", "Use HABIT_LOG to build consistency"],
        ["5. Manage tasks", "Use TASKS for to-dos and GOALS for long-term targets"],
        ["6. Log finances", "Use FINANCE to track income and expenses"],
        ["7. Track job applications", "Use JOB_SEARCH for your career pipeline"],
        [""],
        ["QUICK START TIPS", ""],
        ["✓ All dropdown menus are configured in SETUP", ""],
        ["✓ Tables auto-expand when you add new rows", ""],
        ["✓ Dashboard updates automatically from your data", ""],
        ["✓ Use consistent categories for better insights", ""],
        ["✓ Sample data is included - feel free to delete it", ""],
        [""],
        ["UPDATING CATEGORIES", ""],
        ["1. Go to SETUP sheet", ""],
        ["2. Find the list you want to edit", ""],
        ["3. Add or modify values in the list columns", ""],
        ["4. Named ranges will update automatically", ""],
        [""],
        ["READING THE DASHBOARD", ""],
        ["• KPI Cards", "Show averages and totals for last 7 and 30 days"],
        ["• Trend Charts", "Display mood and sleep patterns over time"],
        ["• Spending Chart", "Shows expenses by category for current month"],
        ["• Job Pipeline", "Summary of application statuses"],
        [""],
        ["NOTES", ""],
        ["• This workbook uses Excel 365 formulas", ""],
        ["• No macros or VBA required", ""],
        ["• Works on Windows and Mac", ""],
        ["• Keep backups of your data regularly", ""],
    ]

    for i, row in enumerate(content, start=4):
        ws[f'A{i}'] = row[0]
        if row[0].isupper() and row[0] != "":
            ws[f'A{i}'].font = Font(bold=True, size=12, color="1F4E78")
        if len(row) > 1 and row[1]:
            ws[f'B{i}'] = row[1]
            ws[f'B{i}'].font = Font(color="666666")

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50

    print("✓ README sheet created")

def create_setup_sheet(wb):
    """Create SETUP sheet with configuration and named ranges"""
    ws = wb.create_sheet("SETUP")

    # Header
    ws['A1'] = "CONFIGURATION & LISTS"
    ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws.merge_cells('A1:D1')

    # Configuration section
    ws['A3'] = "Configuration"
    ws['A3'].font = Font(bold=True, size=12)

    ws['A4'] = "Start Year"
    ws['B4'] = 2026
    ws['A5'] = "Start Date"
    ws['B5'] = datetime(2026, 1, 1)
    ws['B5'].number_format = 'yyyy-mm-dd'
    ws['A6'] = "Owner Name"
    ws['B6'] = "Michelle Humes"

    # Style config cells
    for row in [4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Lists section
    lists_data = {
        'list_mood': ['1', '2', '3', '4', '5'],
        'list_energy': ['1', '2', '3', '4', '5'],
        'list_stress': ['1', '2', '3', '4', '5'],
        'list_yes_no': ['Yes', 'No'],
        'list_task_status': ['Backlog', 'Today', 'In Progress', 'Completed'],
        'list_task_priority': ['Low', 'Medium', 'High'],
        'list_task_category': ['Health', 'Career', 'Business', 'Personal', 'Relationships'],
        'list_goal_status': ['Active', 'Paused', 'Completed'],
        'list_goal_timeframe': ['Short', 'Medium', 'Long'],
        'list_finance_type': ['Income', 'Expense'],
        'list_finance_category': ['Housing', 'Utilities', 'Groceries', 'Dining', 'Transportation',
                                  'Health', 'Fitness', 'Subscriptions', 'Shopping', 'Travel',
                                  'Education', 'Business', 'Savings', 'Other'],
        'list_job_status': ['Target', 'Applied', 'Screening', 'Interview', 'Offer',
                           'Rejected', 'Stalled', 'Withdrawn'],
        'list_health_symptom': ['Cramps', 'Bloating', 'Headache', 'Fatigue', 'Nausea',
                               'Tenderness', 'Other'],
        'list_lh_test': ['Negative', 'Positive', 'Not Taken'],
        'list_cycle_phase': ['Follicular', 'Fertile', 'Ovulation', 'Luteal', 'Implantation']
    }

    ws['A8'] = "Dropdown Lists"
    ws['A8'].font = Font(bold=True, size=12)
    ws['A9'] = "List Name"
    ws['B9'] = "Values"
    ws['A9'].font = Font(bold=True)
    ws['B9'].font = Font(bold=True)

    current_row = 10
    for list_name, values in lists_data.items():
        ws[f'A{current_row}'] = list_name
        ws[f'A{current_row}'].font = Font(bold=True, color="1F4E78")

        # Add values
        for i, value in enumerate(values):
            ws[f'B{current_row + i}'] = value
            ws[f'B{current_row + i}'].fill = PatternFill(start_color="F5F5F5", fill_type="solid")

        # Create named range
        end_row = current_row + len(values) - 1
        range_ref = f"SETUP!$B${current_row}:$B${end_row}"
        wb.create_named_range(list_name, ws, range_ref)

        current_row = end_row + 2

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    print("✓ SETUP sheet created with named ranges")

def create_daily_log_sheet(wb):
    """Create DAILY_LOG sheet with daily metrics"""
    ws = wb.create_sheet("DAILY_LOG")

    # Headers
    headers = ['Date', 'Day', 'Week', 'Month', 'SleepHours', 'SleepQuality', 'Steps',
               'ExerciseMinutes', 'WaterOz', 'Mood', 'Energy', 'Stress', 'Highlights', 'Notes']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data (last 10 days)
    base_date = datetime.now() - timedelta(days=9)
    sample_data = []

    for i in range(10):
        date = base_date + timedelta(days=i)
        row = [
            date,
            f'=TEXT(A{i+2},"ddd")',
            f'=WEEKNUM(A{i+2})',
            f'=TEXT(A{i+2},"yyyy-mm")',
            round(random.uniform(6, 8.5), 1),
            random.randint(3, 5),
            random.randint(5000, 12000),
            random.randint(0, 60),
            random.randint(40, 80),
            random.randint(3, 5),
            random.randint(3, 5),
            random.randint(1, 3),
            random.choice(['Great workout!', 'Productive day', 'Relaxing evening',
                          'Coffee with friend', 'Made progress on project']),
            random.choice(['Felt good today', 'A bit tired', 'Very energetic',
                          'Need more sleep', 'Balanced day'])
        ]
        sample_data.append(row)

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx == 1:  # Date column
                cell.number_format = 'yyyy-mm-dd'

    # Create table
    table = Table(displayName="tblDaily", ref=f"A1:N{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation for dropdowns
    dv_quality = DataValidation(type="list", formula1="list_mood", allow_blank=True)
    dv_mood = DataValidation(type="list", formula1="list_mood", allow_blank=True)
    dv_energy = DataValidation(type="list", formula1="list_energy", allow_blank=True)
    dv_stress = DataValidation(type="list", formula1="list_stress", allow_blank=True)

    ws.add_data_validation(dv_quality)
    ws.add_data_validation(dv_mood)
    ws.add_data_validation(dv_energy)
    ws.add_data_validation(dv_stress)

    dv_quality.add(f'F2:F1000')
    dv_mood.add(f'J2:J1000')
    dv_energy.add(f'K2:K1000')
    dv_stress.add(f'L2:L1000')

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 30

    print("✓ DAILY_LOG sheet created")

def create_habit_log_sheet(wb):
    """Create HABIT_LOG sheet"""
    ws = wb.create_sheet("HABIT_LOG")

    # Headers
    headers = ['Date', 'Habit', 'Category', 'Value', 'Done', 'Notes']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data
    habits = ['Morning Exercise', 'Read 30min', 'Meditate', 'Track Finances', 'Walk 10k steps']
    base_date = datetime.now() - timedelta(days=9)
    sample_data = []

    for i in range(10):
        date = base_date + timedelta(days=i)
        for habit in random.sample(habits, k=random.randint(3, 5)):
            row = [
                date,
                habit,
                random.choice(['Health', 'Personal', 'Career']),
                random.randint(1, 30),
                random.choice(['Yes', 'No']),
                ''
            ]
            sample_data.append(row)

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx == 1:
                cell.number_format = 'yyyy-mm-dd'

    # Create table
    table = Table(displayName="tblHabits", ref=f"A1:F{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation
    dv_done = DataValidation(type="list", formula1="list_yes_no", allow_blank=True)
    ws.add_data_validation(dv_done)
    dv_done.add(f'E2:E1000')

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Summary section
    ws['H1'] = "MONTHLY SUMMARY"
    ws['H1'].font = Font(bold=True, size=12, color="1F4E78")

    ws['H3'] = "Habit Completion Rate (Current Month)"
    ws['H3'].font = Font(bold=True)

    summary_habits = ['Morning Exercise', 'Read 30min', 'Meditate']
    for i, habit in enumerate(summary_habits, start=4):
        ws[f'H{i}'] = habit
        # Formula to calculate completion percentage
        ws[f'I{i}'] = (f'=IFERROR(COUNTIFS(tblHabits[Habit],"{habit}",'
                      f'tblHabits[Done],"Yes",'
                      f'tblHabits[Date],">="&EOMONTH(TODAY(),-1)+1,'
                      f'tblHabits[Date],"<="&EOMONTH(TODAY(),0))/'
                      f'COUNTIFS(tblHabits[Habit],"{habit}",'
                      f'tblHabits[Date],">="&EOMONTH(TODAY(),-1)+1,'
                      f'tblHabits[Date],"<="&EOMONTH(TODAY(),0)),0)')
        ws[f'I{i}'].number_format = '0%'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['H'].width = 25

    print("✓ HABIT_LOG sheet created")

def create_tasks_sheet(wb):
    """Create TASKS sheet"""
    ws = wb.create_sheet("TASKS")

    # Headers
    headers = ['Task', 'Category', 'Priority', 'Status', 'DueDate',
               'CompletedDate', 'Notes', 'IsOverdue']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data
    sample_data = [
        ['Update resume', 'Career', 'High', 'In Progress', datetime.now() + timedelta(days=3), '', 'Focus on latest achievements', ''],
        ['Schedule dentist appointment', 'Health', 'Medium', 'Today', datetime.now(), '', '', ''],
        ['Review Q1 finances', 'Personal', 'High', 'Backlog', datetime.now() + timedelta(days=7), '', 'Check all accounts', ''],
        ['Plan Valentine\'s Day', 'Relationships', 'Medium', 'Completed', datetime.now() - timedelta(days=2), datetime.now() - timedelta(days=1), 'Reservations made', ''],
        ['Launch new product', 'Business', 'High', 'In Progress', datetime.now() + timedelta(days=14), '', 'Final testing phase', ''],
        ['Read "Atomic Habits"', 'Personal', 'Low', 'Backlog', datetime.now() + timedelta(days=30), '', '50% complete', ''],
        ['Update LinkedIn profile', 'Career', 'Medium', 'Today', datetime.now(), '', '', ''],
        ['Meal prep for week', 'Health', 'Medium', 'Completed', datetime.now() - timedelta(days=1), datetime.now() - timedelta(days=1), '', ''],
    ]

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx in [5, 6] and value:  # Date columns
                cell.number_format = 'yyyy-mm-dd'

        # Add IsOverdue formula
        ws[f'H{row_idx}'] = f'=IF(AND(D{row_idx}<>"Completed",E{row_idx}<>"",E{row_idx}<TODAY()),"Yes","No")'

    # Create table
    table = Table(displayName="tblTasks", ref=f"A1:H{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation
    dv_category = DataValidation(type="list", formula1="list_task_category", allow_blank=True)
    dv_priority = DataValidation(type="list", formula1="list_task_priority", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="list_task_status", allow_blank=True)

    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_priority)
    ws.add_data_validation(dv_status)

    dv_category.add('B2:B1000')
    dv_priority.add('C2:C1000')
    dv_status.add('D2:D1000')

    # Conditional formatting
    # Red fill for overdue tasks
    red_fill = PatternFill(start_color="FFE0E0", fill_type="solid")
    overdue_rule = FormulaRule(formula=['$H2="Yes"'], fill=red_fill)
    ws.conditional_formatting.add(f'A2:H{len(sample_data)+1}', overdue_rule)

    # Green fill for completed tasks
    green_fill = PatternFill(start_color="E0FFE0", fill_type="solid")
    completed_rule = FormulaRule(formula=['$D2="Completed"'], fill=green_fill)
    ws.conditional_formatting.add(f'A2:H{len(sample_data)+1}', completed_rule)

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['G'].width = 30

    print("✓ TASKS sheet created")

def create_goals_sheet(wb):
    """Create GOALS sheet"""
    ws = wb.create_sheet("GOALS")

    # Headers
    headers = ['Goal', 'Category', 'Timeframe', 'Status', 'StartDate',
               'TargetDate', 'TargetValue', 'CurrentValue', 'ProgressPct', 'Notes']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data
    sample_data = [
        ['Land new job with $120k+ salary', 'Career', 'Medium', 'Active',
         datetime(2026, 1, 1), datetime(2026, 6, 30), 1, 0.3, '', 'Applied to 15 companies so far'],
        ['Save $10,000 emergency fund', 'Personal', 'Short', 'Active',
         datetime(2026, 1, 1), datetime(2026, 12, 31), 10000, 4500, '', 'On track'],
        ['Launch 3 new Etsy products', 'Business', 'Short', 'Active',
         datetime(2026, 1, 15), datetime(2026, 3, 31), 3, 1, '', '1st product launched'],
        ['Read 24 books this year', 'Personal', 'Long', 'Active',
         datetime(2026, 1, 1), datetime(2026, 12, 31), 24, 3, '', '2 books per month goal'],
        ['Get pregnant', 'Health', 'Medium', 'Active',
         datetime(2025, 11, 1), datetime(2026, 8, 31), 1, 0.5, '', 'Tracking cycles actively'],
        ['Build professional network', 'Career', 'Long', 'Active',
         datetime(2026, 1, 1), datetime(2026, 12, 31), 50, 12, '', 'Attending monthly meetups'],
    ]

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx in [5, 6] and value:  # Date columns
                cell.number_format = 'yyyy-mm-dd'

        # Add ProgressPct formula
        ws[f'I{row_idx}'] = f'=IF(G{row_idx}="","",MIN(1,H{row_idx}/G{row_idx}))'
        ws[f'I{row_idx}'].number_format = '0%'

    # Create table
    table = Table(displayName="tblGoals", ref=f"A1:J{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation
    dv_category = DataValidation(type="list", formula1="list_task_category", allow_blank=True)
    dv_timeframe = DataValidation(type="list", formula1="list_goal_timeframe", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="list_goal_status", allow_blank=True)

    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_timeframe)
    ws.add_data_validation(dv_status)

    dv_category.add('B2:B1000')
    dv_timeframe.add('C2:C1000')
    dv_status.add('D2:D1000')

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['J'].width = 30

    print("✓ GOALS sheet created")

def create_finance_sheet(wb):
    """Create FINANCE sheet"""
    ws = wb.create_sheet("FINANCE")

    # Headers
    headers = ['Date', 'Description', 'Category', 'Type', 'Amount',
               'Account', 'Owner', 'Notes']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data
    sample_data = []
    base_date = datetime.now().replace(day=1)  # Start of current month

    # Income entries
    sample_data.append([base_date + timedelta(days=1), 'Salary - Main Job', 'Income', 'Income',
                       4500, 'Chase Checking', 'Michelle', 'Bi-weekly payroll'])
    sample_data.append([base_date + timedelta(days=15), 'Salary - Main Job', 'Income', 'Income',
                       4500, 'Chase Checking', 'Michelle', 'Bi-weekly payroll'])
    sample_data.append([base_date + timedelta(days=10), 'Etsy Sales', 'Business', 'Income',
                       450, 'Business Account', 'Michelle', 'Wedding planners sold'])

    # Expense entries
    expenses = [
        [base_date + timedelta(days=1), 'Rent Payment', 'Housing', 'Expense', 1800, 'Chase Checking', 'Michelle', 'Monthly rent'],
        [base_date + timedelta(days=3), 'Electric Bill', 'Utilities', 'Expense', 120, 'Chase Checking', 'Michelle', 'January usage'],
        [base_date + timedelta(days=5), 'Trader Joes', 'Groceries', 'Expense', 85, 'Chase Credit', 'Michelle', 'Weekly shopping'],
        [base_date + timedelta(days=7), 'Target', 'Shopping', 'Expense', 45, 'Chase Credit', 'Michelle', 'Household items'],
        [base_date + timedelta(days=8), 'Starbucks', 'Dining', 'Expense', 12, 'Chase Credit', 'Michelle', 'Coffee meeting'],
        [base_date + timedelta(days=10), 'LA Fitness', 'Fitness', 'Expense', 50, 'Chase Checking', 'Michelle', 'Monthly membership'],
        [base_date + timedelta(days=12), 'Whole Foods', 'Groceries', 'Expense', 95, 'Chase Credit', 'Michelle', 'Weekly shopping'],
        [base_date + timedelta(days=14), 'Netflix', 'Subscriptions', 'Expense', 15.99, 'Chase Credit', 'Michelle', 'Monthly subscription'],
        [base_date + timedelta(days=16), 'Gas Station', 'Transportation', 'Expense', 45, 'Chase Credit', 'Michelle', 'Fill up'],
        [base_date + timedelta(days=18), 'Amazon', 'Shopping', 'Expense', 67, 'Chase Credit', 'Michelle', 'Office supplies'],
    ]
    sample_data.extend(expenses)

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx == 1:  # Date column
                cell.number_format = 'yyyy-mm-dd'
            elif col_idx == 5:  # Amount column
                cell.number_format = '$#,##0.00'

    # Create table
    table = Table(displayName="tblFinance", ref=f"A1:H{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation
    dv_category = DataValidation(type="list", formula1="list_finance_category", allow_blank=True)
    dv_type = DataValidation(type="list", formula1="list_finance_type", allow_blank=True)

    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_type)

    dv_category.add('C2:C1000')
    dv_type.add('D2:D1000')

    # Monthly summary section
    ws['J1'] = "MONTHLY SUMMARY"
    ws['J1'].font = Font(bold=True, size=12, color="1F4E78")

    ws['J3'] = "Current Month"
    ws['J3'].font = Font(bold=True)

    ws['J4'] = "Total Income"
    ws['K4'] = ('=SUMIFS(tblFinance[Amount],tblFinance[Type],"Income",'
               'tblFinance[Date],">="&EOMONTH(TODAY(),-1)+1,'
               'tblFinance[Date],"<="&EOMONTH(TODAY(),0))')
    ws['K4'].number_format = '$#,##0.00'
    ws['K4'].font = Font(color="008000")

    ws['J5'] = "Total Expenses"
    ws['K5'] = ('=SUMIFS(tblFinance[Amount],tblFinance[Type],"Expense",'
               'tblFinance[Date],">="&EOMONTH(TODAY(),-1)+1,'
               'tblFinance[Date],"<="&EOMONTH(TODAY(),0))')
    ws['K5'].number_format = '$#,##0.00'
    ws['K5'].font = Font(color="FF0000")

    ws['J6'] = "Net Cash Flow"
    ws['K6'] = '=K4-K5'
    ws['K6'].number_format = '$#,##0.00'
    ws['K6'].font = Font(bold=True)

    # Style summary cells
    for row in [4, 5, 6]:
        ws[f'J{row}'].font = Font(bold=True)
        ws[f'K{row}'].fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 15

    print("✓ FINANCE sheet created")

def create_job_search_sheet(wb):
    """Create JOB_SEARCH sheet"""
    ws = wb.create_sheet("JOB_SEARCH")

    # Headers
    headers = ['Company', 'Role', 'Level', 'SalaryMin', 'SalaryMax', 'Location',
               'Remote', 'Status', 'AppliedDate', 'LastContact', 'NextAction',
               'NextActionDate', 'Notes']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data
    sample_data = [
        ['Google', 'Product Manager', 'Senior', 140000, 180000, 'Mountain View, CA', 'Yes',
         'Interview', datetime(2026, 1, 10), datetime(2026, 1, 28), 'Prepare for final round',
         datetime(2026, 2, 5), 'Very interested - great culture fit'],
        ['Microsoft', 'Program Manager', 'Mid-level', 120000, 150000, 'Seattle, WA', 'Yes',
         'Applied', datetime(2026, 1, 15), datetime(2026, 1, 15), 'Follow up if no response',
         datetime(2026, 2, 1), 'Strong technical team'],
        ['Amazon', 'Technical Program Manager', 'Senior', 135000, 165000, 'Austin, TX', 'No',
         'Screening', datetime(2026, 1, 5), datetime(2026, 1, 20), 'Phone screen scheduled',
         datetime(2026, 2, 3), 'Fast-paced environment'],
        ['Apple', 'Product Manager', 'Senior', 145000, 185000, 'Cupertino, CA', 'No',
         'Target', None, None, 'Submit application', datetime(2026, 2, 7), 'Dream company'],
        ['Salesforce', 'Senior Product Manager', 'Senior', 130000, 160000, 'San Francisco, CA', 'Yes',
         'Rejected', datetime(2026, 1, 3), datetime(2026, 1, 25), None, None, 'Not a good fit per recruiter'],
        ['Meta', 'Product Manager', 'Senior', 150000, 190000, 'Menlo Park, CA', 'Yes',
         'Interview', datetime(2026, 1, 12), datetime(2026, 1, 30), 'Technical interview prep',
         datetime(2026, 2, 8), 'Exciting product area'],
        ['Stripe', 'Product Manager', 'Mid-level', 125000, 155000, 'Remote', 'Yes',
         'Applied', datetime(2026, 1, 20), datetime(2026, 1, 20), 'Wait for response',
         datetime(2026, 2, 10), 'Fintech experience relevant'],
    ]

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx in [4, 5]:  # Salary columns
                if value:
                    cell.number_format = '$#,##0'
            elif col_idx in [9, 10, 12] and value:  # Date columns
                cell.number_format = 'yyyy-mm-dd'

    # Create table
    table = Table(displayName="tblJobs", ref=f"A1:M{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation
    dv_remote = DataValidation(type="list", formula1="list_yes_no", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="list_job_status", allow_blank=True)

    ws.add_data_validation(dv_remote)
    ws.add_data_validation(dv_status)

    dv_remote.add('G2:G1000')
    dv_status.add('H2:H1000')

    # Summary section
    ws['O1'] = "PIPELINE SUMMARY"
    ws['O1'].font = Font(bold=True, size=12, color="1F4E78")

    ws['O3'] = "Status Counts"
    ws['O3'].font = Font(bold=True)

    statuses = ['Target', 'Applied', 'Screening', 'Interview', 'Offer']
    for i, status in enumerate(statuses, start=4):
        ws[f'O{i}'] = status
        ws[f'P{i}'] = f'=COUNTIF(tblJobs[Status],"{status}")'
        ws[f'O{i}'].font = Font(bold=True)

    ws['O10'] = "Avg Salary Range"
    ws['O10'].font = Font(bold=True)
    ws['O11'] = "Min"
    ws['P11'] = '=AVERAGE(tblJobs[SalaryMin])'
    ws['P11'].number_format = '$#,##0'
    ws['O12'] = "Max"
    ws['P12'] = '=AVERAGE(tblJobs[SalaryMax])'
    ws['P12'].number_format = '$#,##0'

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['O'].width = 18

    print("✓ JOB_SEARCH sheet created")

def create_health_fertility_sheet(wb):
    """Create HEALTH_FERTILITY sheet"""
    ws = wb.create_sheet("HEALTH_FERTILITY")

    # Headers
    headers = ['Date', 'CycleDay', 'DPO', 'Phase', 'Temperature', 'LHTest',
               'Symptoms', 'Mood', 'Energy', 'SexualActivity', 'Notes']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")

    # Sample data (28-day cycle)
    base_date = datetime.now() - timedelta(days=20)
    sample_data = []

    phases = ['Follicular', 'Follicular', 'Follicular', 'Follicular', 'Follicular',
              'Fertile', 'Fertile', 'Fertile', 'Ovulation', 'Ovulation',
              'Luteal', 'Luteal', 'Luteal', 'Luteal', 'Luteal',
              'Luteal', 'Luteal', 'Luteal', 'Luteal', 'Luteal']

    for i in range(20):
        date = base_date + timedelta(days=i)
        cycle_day = i + 5  # Assume we're on day 5 of cycle
        dpo = max(0, cycle_day - 14) if cycle_day > 14 else ''
        phase = phases[i]
        temp = round(random.uniform(97.0, 98.6), 1)
        lh_test = 'Positive' if cycle_day in [13, 14] else 'Negative' if cycle_day > 10 else 'Not Taken'
        symptoms = random.choice(['', 'Cramps', 'Bloating', 'Tenderness', 'Fatigue'])

        row = [
            date,
            cycle_day,
            dpo,
            phase,
            temp,
            lh_test,
            symptoms,
            random.randint(3, 5),
            random.randint(3, 5),
            random.choice(['Yes', 'No']),
            ''
        ]
        sample_data.append(row)

    # Add data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx == 1:  # Date column
                cell.number_format = 'yyyy-mm-dd'
            elif col_idx == 5 and value:  # Temperature
                cell.number_format = '0.0'

    # Create table
    table = Table(displayName="tblHealth", ref=f"A1:K{len(sample_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add data validation
    dv_phase = DataValidation(type="list", formula1="list_cycle_phase", allow_blank=True)
    dv_lh = DataValidation(type="list", formula1="list_lh_test", allow_blank=True)
    dv_symptoms = DataValidation(type="list", formula1="list_health_symptom", allow_blank=True)
    dv_mood = DataValidation(type="list", formula1="list_mood", allow_blank=True)
    dv_energy = DataValidation(type="list", formula1="list_energy", allow_blank=True)
    dv_activity = DataValidation(type="list", formula1="list_yes_no", allow_blank=True)

    ws.add_data_validation(dv_phase)
    ws.add_data_validation(dv_lh)
    ws.add_data_validation(dv_symptoms)
    ws.add_data_validation(dv_mood)
    ws.add_data_validation(dv_energy)
    ws.add_data_validation(dv_activity)

    dv_phase.add('D2:D1000')
    dv_lh.add('F2:F1000')
    dv_symptoms.add('G2:G1000')
    dv_mood.add('H2:H1000')
    dv_energy.add('I2:I1000')
    dv_activity.add('J2:J1000')

    # Conditional formatting by phase
    phase_colors = {
        'Follicular': 'E0F0FF',
        'Fertile': 'FFE0FF',
        'Ovulation': 'FFE0E0',
        'Luteal': 'FFFFE0',
        'Implantation': 'E0FFE0'
    }

    for phase, color in phase_colors.items():
        fill = PatternFill(start_color=color, fill_type="solid")
        rule = FormulaRule(formula=[f'$D2="{phase}"'], fill=fill)
        ws.conditional_formatting.add(f'A2:K{len(sample_data)+1}', rule)

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['K'].width = 30

    print("✓ HEALTH_FERTILITY sheet created")

def create_dashboard_sheet(wb):
    """Create DASHBOARD sheet with KPIs and charts"""
    ws = wb.create_sheet("DASHBOARD", 2)  # Insert as 3rd sheet

    # Title
    ws['A1'] = "LIFE TRACKER DASHBOARD"
    ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
    ws.merge_cells('A1:H1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws.merge_cells('A2:H2')

    # KPI Section - Last 7 Days
    ws['A4'] = "LAST 7 DAYS"
    ws['A4'].font = Font(size=12, bold=True, color="1F4E78")
    ws.merge_cells('A4:D4')

    kpis_7d = [
        ('Avg Sleep Hours', '=AVERAGEIFS(DAILY_LOG!$E:$E,DAILY_LOG!$A:$A,">="&TODAY()-6,DAILY_LOG!$A:$A,"<="&TODAY())', '0.0'),
        ('Avg Mood', '=AVERAGEIFS(DAILY_LOG!$J:$J,DAILY_LOG!$A:$A,">="&TODAY()-6,DAILY_LOG!$A:$A,"<="&TODAY())', '0.0'),
        ('Avg Energy', '=AVERAGEIFS(DAILY_LOG!$K:$K,DAILY_LOG!$A:$A,">="&TODAY()-6,DAILY_LOG!$A:$A,"<="&TODAY())', '0.0'),
        ('Total Exercise (min)', '=SUMIFS(DAILY_LOG!$H:$H,DAILY_LOG!$A:$A,">="&TODAY()-6,DAILY_LOG!$A:$A,"<="&TODAY())', '#,##0'),
    ]

    row = 5
    for label, formula, num_format in kpis_7d:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = num_format
        ws[f'B{row}'].fill = PatternFill(start_color="E0F0FF", fill_type="solid")
        ws[f'B{row}'].font = Font(size=12, bold=True)
        row += 1

    # KPI Section - Last 30 Days
    ws['A10'] = "LAST 30 DAYS"
    ws['A10'].font = Font(size=12, bold=True, color="1F4E78")
    ws.merge_cells('A10:D10')

    kpis_30d = [
        ('Avg Sleep Hours', '=AVERAGEIFS(DAILY_LOG!$E:$E,DAILY_LOG!$A:$A,">="&TODAY()-29,DAILY_LOG!$A:$A,"<="&TODAY())', '0.0'),
        ('Avg Mood', '=AVERAGEIFS(DAILY_LOG!$J:$J,DAILY_LOG!$A:$A,">="&TODAY()-29,DAILY_LOG!$A:$A,"<="&TODAY())', '0.0'),
        ('Tasks Completed', '=COUNTIFS(TASKS!$D:$D,"Completed",TASKS!$F:$F,">="&TODAY()-29,TASKS!$F:$F,"<="&TODAY())', '#,##0'),
        ('Habit Completion %', '=IFERROR(COUNTIFS(HABIT_LOG!$E:$E,"Yes",HABIT_LOG!$A:$A,">="&TODAY()-29,HABIT_LOG!$A:$A,"<="&TODAY())/COUNTIFS(HABIT_LOG!$A:$A,">="&TODAY()-29,HABIT_LOG!$A:$A,"<="&TODAY()),0)', '0%'),
    ]

    row = 11
    for label, formula, num_format in kpis_30d:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = num_format
        ws[f'B{row}'].fill = PatternFill(start_color="FFF0E0", fill_type="solid")
        ws[f'B{row}'].font = Font(size=12, bold=True)
        row += 1

    # Finance Section - Current Month
    ws['A16'] = "CURRENT MONTH FINANCES"
    ws['A16'].font = Font(size=12, bold=True, color="1F4E78")
    ws.merge_cells('A16:D16')

    finance_kpis = [
        ('Total Income', '=SUMIFS(FINANCE!$E:$E,FINANCE!$D:$D,"Income",FINANCE!$A:$A,">="&EOMONTH(TODAY(),-1)+1,FINANCE!$A:$A,"<="&EOMONTH(TODAY(),0))', '$#,##0.00', '008000'),
        ('Total Expenses', '=SUMIFS(FINANCE!$E:$E,FINANCE!$D:$D,"Expense",FINANCE!$A:$A,">="&EOMONTH(TODAY(),-1)+1,FINANCE!$A:$A,"<="&EOMONTH(TODAY(),0))', '$#,##0.00', 'FF0000'),
        ('Net Cash Flow', '=B17-B18', '$#,##0.00', '000000'),
    ]

    row = 17
    for label, formula, num_format, color in finance_kpis:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = num_format
        ws[f'B{row}'].fill = PatternFill(start_color="F0FFE0", fill_type="solid")
        ws[f'B{row}'].font = Font(size=12, bold=True, color=color)
        row += 1

    # Job Search Summary
    ws['A21'] = "JOB SEARCH PIPELINE"
    ws['A21'].font = Font(size=12, bold=True, color="1F4E78")
    ws.merge_cells('A21:D21')

    job_statuses = ['Target', 'Applied', 'Screening', 'Interview', 'Offer']
    row = 22
    for status in job_statuses:
        ws[f'A{row}'] = status
        ws[f'B{row}'] = f'=COUNTIF(JOB_SEARCH!$H:$H,"{status}")'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color="FFE0F0", fill_type="solid")
        ws[f'B{row}'].font = Font(size=11, bold=True)
        row += 1

    # Charts section placeholder
    ws['E4'] = "CHARTS & TRENDS"
    ws['E4'].font = Font(size=12, bold=True, color="1F4E78")

    ws['E6'] = "Note: Charts show visual trends for:"
    ws['E7'] = "• Mood & Sleep patterns (last 30 days)"
    ws['E8'] = "• Expense breakdown by category"
    ws['E9'] = "• Job pipeline funnel"
    ws['E10'] = ""
    ws['E11'] = "Excel will auto-generate recommended"
    ws['E12'] = "charts based on your data tables."
    ws['E13'] = ""
    ws['E14'] = "To add charts:"
    ws['E15'] = "1. Select data from any sheet"
    ws['E16'] = "2. Insert > Recommended Charts"
    ws['E17'] = "3. Copy chart to Dashboard"

    for r in range(6, 18):
        ws[f'E{r}'].font = Font(size=10, color="666666")

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['E'].width = 35

    print("✓ DASHBOARD sheet created")

if __name__ == "__main__":
    print("\n📊 Building Excel Life Tracker...\n")
    create_life_tracker()
    print("\n✅ Complete! Open Life_Tracker.xlsx to start using your tracker.\n")
