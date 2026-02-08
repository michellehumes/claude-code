"""
Create a Wellness & Habit Tracker Spreadsheet Bundle
for Etsy Digital Download - Product #3
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.worksheet.datavalidation import DataValidation

# ── Color Palette (Warm, calming, wellness-inspired) ──
PLUM = "5B2C6F"
DUSTY_ROSE = "C77D98"
SOFT_PINK = "F2D7D5"
LAVENDER = "D7BDE2"
PALE_LILAC = "F0E6F6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "E0E0E0"
DARK_TEXT = "2C2C2C"
SAGE = "82A88C"
WARM_BEIGE = "F5E6D3"
DUSTY_TEAL = "6B9F9E"
SOFT_GOLD = "D4A853"
CORAL_SOFT = "E8A598"

def header_font(size=12, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=11, bold=False, color=DARK_TEXT):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    side = Side(style="thin", color=MED_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, bg_color=PLUM):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = fill(bg_color)
        cell.alignment = center_align()
        cell.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    bg = LIGHT_GRAY if alt else WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.fill = fill(bg)
        cell.alignment = center_align()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title_block(ws, title, subtitle, merge_end_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=PLUM)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = fill(PALE_LILAC)
    ws.row_dimensions[1].height = 45

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end_col)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color=DUSTY_ROSE)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.fill = fill(PALE_LILAC)
    ws.row_dimensions[2].height = 25


# ═══════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────
#  SHEET 1: DAILY HABIT TRACKER (30-day grid)
# ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Habit Tracker"
ws1.sheet_properties.tabColor = PLUM

add_title_block(ws1, "DAILY HABIT TRACKER", "Check off your habits each day — consistency creates transformation", 34)

# Month & Year inputs
ws1.cell(row=3, column=1, value="Month:").font = body_font(11, True)
ws1.cell(row=3, column=2, value="January").font = body_font(11, False, PLUM)
ws1.cell(row=3, column=3, value="Year:").font = body_font(11, True)
ws1.cell(row=3, column=4, value="2026").font = body_font(11, False, PLUM)

row = 4
# Headers: Habit | Day 1-31 | Total | %
ws1.cell(row=row, column=1, value="Habit")
ws1.cell(row=row, column=2, value="Category")
for d in range(1, 32):
    ws1.cell(row=row, column=d + 2, value=d)
ws1.cell(row=row, column=34, value="Total")
ws1.cell(row=row, column=35, value="%")
style_header_row(ws1, row, 35)

# Pre-fill sample habits
sample_habits = [
    ("Drink 8 glasses of water", "Health"),
    ("Exercise 30+ minutes", "Fitness"),
    ("Meditate / Mindfulness", "Mental Health"),
    ("Read 20+ minutes", "Growth"),
    ("Healthy meal prep", "Nutrition"),
    ("8 hours of sleep", "Health"),
    ("Gratitude journaling", "Mental Health"),
    ("No social media before 9am", "Mindfulness"),
    ("Take vitamins/supplements", "Health"),
    ("Stretch / Yoga", "Fitness"),
    ("Spend time outdoors", "Wellness"),
    ("Practice a hobby", "Joy"),
    ("Skincare routine (AM/PM)", "Self-Care"),
    ("Connect with a friend/family", "Social"),
    ("Tidy/organize for 10 min", "Home"),
]

yn_dv = DataValidation(type="list", formula1='"✓,X,"', allow_blank=True)
ws1.add_data_validation(yn_dv)

cat_dv = DataValidation(type="list", formula1='"Health,Fitness,Mental Health,Nutrition,Growth,Mindfulness,Wellness,Joy,Self-Care,Social,Home,Financial,Creative,Other"', allow_blank=True)
ws1.add_data_validation(cat_dv)

for idx, (habit, cat) in enumerate(sample_habits):
    r = 5 + idx
    ws1.cell(row=r, column=1, value=habit).font = body_font(10)
    ws1.cell(row=r, column=1).alignment = left_align()
    ws1.cell(row=r, column=2, value=cat).font = body_font(10)
    cat_dv.add(f"B{r}")

    for d in range(1, 32):
        yn_dv.add(f"{get_column_letter(d+2)}{r}")

    # Total formula: count checkmarks
    ws1.cell(row=r, column=34, value=f'=COUNTIF(C{r}:AG{r},"✓")')
    ws1.cell(row=r, column=34).font = body_font(11, True, PLUM)
    ws1.cell(row=r, column=35, value=f'=IF(AH{r}=0,0,AH{r}/31)').number_format = '0%'
    ws1.cell(row=r, column=35).font = body_font(11, True, PLUM)
    style_data_row(ws1, r, 35, idx % 2 == 1)

# Add 5 blank rows for custom habits
for idx in range(5):
    r = 5 + len(sample_habits) + idx
    cat_dv.add(f"B{r}")
    for d in range(1, 32):
        yn_dv.add(f"{get_column_letter(d+2)}{r}")
    ws1.cell(row=r, column=34, value=f'=COUNTIF(C{r}:AG{r},"✓")')
    ws1.cell(row=r, column=35, value=f'=IF(AH{r}=0,0,AH{r}/31)').number_format = '0%'
    style_data_row(ws1, r, 35, (len(sample_habits) + idx) % 2 == 1)

# Daily completion rate row
rate_row = 5 + len(sample_habits) + 5
ws1.cell(row=rate_row, column=1, value="DAILY COMPLETION").font = header_font(10, True, PLUM)
ws1.cell(row=rate_row, column=1).fill = fill(LAVENDER)
for d in range(1, 32):
    col = d + 2
    col_letter = get_column_letter(col)
    ws1.cell(row=rate_row, column=col,
             value=f'=IF(COUNTA({col_letter}5:{col_letter}{rate_row-1})=0,"",COUNTIF({col_letter}5:{col_letter}{rate_row-1},"✓")/COUNTA({col_letter}5:{col_letter}{rate_row-1}))')
    ws1.cell(row=rate_row, column=col).number_format = '0%'
    ws1.cell(row=rate_row, column=col).font = body_font(9, True, PLUM)
    ws1.cell(row=rate_row, column=col).fill = fill(LAVENDER)
    ws1.cell(row=rate_row, column=col).border = thin_border()
    ws1.cell(row=rate_row, column=col).alignment = center_align()

set_col_widths(ws1, [30, 14] + [4.5]*31 + [8, 8])

# ──────────────────────────────────────────────────
#  SHEET 2: MOOD & MENTAL HEALTH TRACKER
# ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Mood Tracker")
ws2.sheet_properties.tabColor = DUSTY_ROSE

add_title_block(ws2, "MOOD & MENTAL HEALTH TRACKER", "Notice patterns, honor your feelings, and track what lifts you up", 9)

row = 4
headers2 = ["Date", "Overall Mood", "Energy Level", "Anxiety Level", "Sleep Quality",
            "Exercise?", "Highlight of the Day", "Challenge/Stressor", "Self-Care Done"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, 9)

mood_dv = DataValidation(type="list", formula1='"😊 Great,🙂 Good,😐 Okay,😔 Low,😢 Rough"', allow_blank=True)
ws2.add_data_validation(mood_dv)

energy_dv = DataValidation(type="list", formula1='"⚡ High,🔋 Medium,🪫 Low,💤 Exhausted"', allow_blank=True)
ws2.add_data_validation(energy_dv)

anxiety_dv = DataValidation(type="list", formula1='"1 - Calm,2 - Mild,3 - Moderate,4 - High,5 - Severe"', allow_blank=True)
ws2.add_data_validation(anxiety_dv)

sleep_dv = DataValidation(type="list", formula1='"Excellent,Good,Fair,Poor,Terrible"', allow_blank=True)
ws2.add_data_validation(sleep_dv)

exercise_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws2.add_data_validation(exercise_dv)

for r in range(5, 36):  # 31 days
    ws2.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    mood_dv.add(f"B{r}")
    energy_dv.add(f"C{r}")
    anxiety_dv.add(f"D{r}")
    sleep_dv.add(f"E{r}")
    exercise_dv.add(f"F{r}")
    style_data_row(ws2, r, 9, (r - 5) % 2 == 1)

# Monthly reflection section
ref_row = 38
ws2.cell(row=ref_row, column=1, value="MONTHLY REFLECTION").font = header_font(14, True, PLUM)
ws2.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=9)
ws2.cell(row=ref_row, column=1).fill = fill(LAVENDER)
ws2.cell(row=ref_row, column=1).alignment = center_align()

reflections = [
    "What patterns did I notice in my mood this month?",
    "What activities or people lifted my mood the most?",
    "What were my biggest stressors?",
    "What self-care practices helped me the most?",
    "One thing I want to do differently next month:",
    "I'm proud of myself for:",
    "My intention for next month:",
]
for idx, prompt in enumerate(reflections):
    r = ref_row + 1 + idx
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws2.cell(row=r, column=1, value=prompt).font = body_font(11, True)
    ws2.cell(row=r, column=1).fill = fill(PALE_LILAC)
    ws2.cell(row=r, column=1).border = thin_border()
    ws2.cell(row=r, column=1).alignment = left_align()
    ws2.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    ws2.cell(row=r, column=4).border = thin_border()
    ws2.row_dimensions[r].height = 35

set_col_widths(ws2, [14, 16, 16, 16, 16, 12, 30, 30, 25])

# ──────────────────────────────────────────────────
#  SHEET 3: FITNESS & WORKOUT LOG
# ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Fitness Log")
ws3.sheet_properties.tabColor = SAGE

add_title_block(ws3, "FITNESS & WORKOUT LOG", "Track your workouts, celebrate your progress, and stay consistent", 9)

row = 4
headers3 = ["Date", "Workout Type", "Duration (min)", "Intensity", "Exercises / Details",
            "Calories Burned", "How I Felt", "Steps", "Notes"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 9, SAGE)

workout_dv = DataValidation(type="list", formula1='"Strength Training,Cardio,HIIT,Yoga,Pilates,Walking,Running,Swimming,Cycling,Dance,Sports,Stretching,Home Workout,Gym,Other"', allow_blank=True)
ws3.add_data_validation(workout_dv)

intensity_dv = DataValidation(type="list", formula1='"Light,Moderate,Intense,Maximum"', allow_blank=True)
ws3.add_data_validation(intensity_dv)

felt_dv = DataValidation(type="list", formula1='"Amazing,Good,Okay,Tired,Struggled"', allow_blank=True)
ws3.add_data_validation(felt_dv)

for r in range(5, 55):
    ws3.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    workout_dv.add(f"B{r}")
    intensity_dv.add(f"D{r}")
    felt_dv.add(f"G{r}")
    style_data_row(ws3, r, 9, (r - 5) % 2 == 1)

# Monthly fitness summary
sum_row = 56
ws3.cell(row=sum_row, column=1, value="MONTHLY SUMMARY").font = header_font(14, True, PLUM)
ws3.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=9)
ws3.cell(row=sum_row, column=1).fill = fill(LAVENDER)
ws3.cell(row=sum_row, column=1).alignment = center_align()

summary_items = [
    ("Total Workouts", "=COUNTA(B5:B54)"),
    ("Total Minutes", "=SUM(C5:C54)"),
    ("Average Duration", "=IF(COUNTA(B5:B54)=0,0,SUM(C5:C54)/COUNTA(B5:B54))"),
    ("Total Calories Burned", "=SUM(F5:F54)"),
    ("Total Steps", "=SUM(H5:H54)"),
]
for idx, (label, formula) in enumerate(summary_items):
    r = sum_row + 1 + idx
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws3.cell(row=r, column=1, value=label).font = body_font(11, True)
    ws3.cell(row=r, column=1).fill = fill(PALE_LILAC)
    ws3.cell(row=r, column=1).border = thin_border()
    ws3.cell(row=r, column=4, value=formula).font = body_font(12, True, PLUM)
    ws3.cell(row=r, column=4).border = thin_border()
    ws3.cell(row=r, column=4).alignment = center_align()

set_col_widths(ws3, [14, 20, 16, 14, 35, 16, 14, 14, 25])

# ──────────────────────────────────────────────────
#  SHEET 4: MEAL PLANNER & GROCERY LIST
# ──────────────────────────────────────────────────
ws4 = wb.create_sheet("Meal Planner")
ws4.sheet_properties.tabColor = CORAL_SOFT

add_title_block(ws4, "WEEKLY MEAL PLANNER & GROCERY LIST", "Plan your meals, save money, and eat better — one week at a time", 9)

# Weekly meal plan
row = 4
ws4.cell(row=row, column=1, value="WEEKLY MEAL PLAN").font = header_font(14, True, PLUM)
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
ws4.cell(row=row, column=1).fill = fill(LAVENDER)
ws4.cell(row=row, column=1).alignment = center_align()

row = 5
meal_headers = ["Day", "Breakfast", "Snack AM", "Lunch", "Snack PM", "Dinner", "Water (cups)", "Calories", "Notes"]
for i, h in enumerate(meal_headers, 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 9, DUSTY_ROSE)

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
for idx, day in enumerate(days):
    r = 6 + idx
    ws4.cell(row=r, column=1, value=day).font = body_font(11, True)
    style_data_row(ws4, r, 9, idx % 2 == 1)

# Weekly totals
wt_row = 13
ws4.cell(row=wt_row, column=1, value="WEEKLY TOTALS")
ws4.cell(row=wt_row, column=7, value="=SUM(G6:G12)")
ws4.cell(row=wt_row, column=8, value="=SUM(H6:H12)")
style_header_row(ws4, wt_row, 9, PLUM)

# Grocery List section
gl_row = 15
ws4.cell(row=gl_row, column=1, value="GROCERY LIST").font = header_font(14, True, PLUM)
ws4.merge_cells(start_row=gl_row, start_column=1, end_row=gl_row, end_column=9)
ws4.cell(row=gl_row, column=1).fill = fill(LAVENDER)
ws4.cell(row=gl_row, column=1).alignment = center_align()

gl_header_row = gl_row + 1
gl_headers = ["Item", "Category", "Quantity", "Unit", "Est. Price", "Got It?", "", "", ""]
for i, h in enumerate(gl_headers, 1):
    if h:
        ws4.cell(row=gl_header_row, column=i, value=h)
style_header_row(ws4, gl_header_row, 6, DUSTY_ROSE)

grocery_cats = DataValidation(type="list", formula1='"Produce,Dairy,Meat/Protein,Grains/Bread,Frozen,Canned/Dry,Snacks,Beverages,Condiments,Household,Other"', allow_blank=True)
ws4.add_data_validation(grocery_cats)

unit_dv = DataValidation(type="list", formula1='"each,lb,oz,bag,box,can,bottle,bunch,pack,dozen"', allow_blank=True)
ws4.add_data_validation(unit_dv)

got_dv = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
ws4.add_data_validation(got_dv)

for r in range(gl_header_row + 1, gl_header_row + 31):
    ws4.cell(row=r, column=5).number_format = '"$"#,##0.00'
    grocery_cats.add(f"B{r}")
    unit_dv.add(f"D{r}")
    got_dv.add(f"F{r}")
    style_data_row(ws4, r, 6, (r - gl_header_row - 1) % 2 == 1)

# Grocery total
gt_row = gl_header_row + 31
ws4.cell(row=gt_row, column=1, value="ESTIMATED TOTAL")
ws4.cell(row=gt_row, column=5, value=f"=SUM(E{gl_header_row+1}:E{gt_row-1})").number_format = '"$"#,##0.00'
style_header_row(ws4, gt_row, 6, PLUM)

# Meal prep tips
tips_row = gt_row + 2
ws4.cell(row=tips_row, column=1, value="MEAL PREP TIPS").font = header_font(12, True, PLUM)
ws4.merge_cells(start_row=tips_row, start_column=1, end_row=tips_row, end_column=6)
ws4.cell(row=tips_row, column=1).fill = fill(LAVENDER)
ws4.cell(row=tips_row, column=1).alignment = center_align()

tips = [
    "Plan meals around what's on sale to save money.",
    "Prep ingredients on Sunday for the whole week.",
    "Cook grains and proteins in bulk — they keep well for 4-5 days.",
    "Freeze extra portions for busy days.",
    "Keep a running list of staples you always need.",
]
for idx, tip in enumerate(tips):
    r = tips_row + 1 + idx
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws4.cell(row=r, column=1, value=f"  {idx+1}. {tip}").font = body_font(10)
    ws4.cell(row=r, column=1).alignment = left_align()

set_col_widths(ws4, [14, 22, 14, 22, 14, 22, 14, 14, 20])

# ──────────────────────────────────────────────────
#  SHEET 5: WATER & SLEEP TRACKER
# ──────────────────────────────────────────────────
ws5 = wb.create_sheet("Water & Sleep")
ws5.sheet_properties.tabColor = DUSTY_TEAL

add_title_block(ws5, "WATER INTAKE & SLEEP TRACKER", "Hydration and rest are the foundation of everything else", 10)

row = 4
headers5 = ["Date", "Water Goal", "Water Actual", "% of Goal",
            "Bedtime", "Wake Time", "Hours Slept", "Sleep Quality",
            "Dreams/Notes", "Energy Next Day"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, 10, DUSTY_TEAL)

sleep_qual_dv = DataValidation(type="list", formula1='"Excellent,Good,Fair,Poor,Terrible"', allow_blank=True)
ws5.add_data_validation(sleep_qual_dv)

energy_next_dv = DataValidation(type="list", formula1='"High,Medium,Low,Exhausted"', allow_blank=True)
ws5.add_data_validation(energy_next_dv)

for r in range(5, 36):  # 31 days
    ws5.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    ws5.cell(row=r, column=2, value=8)  # Default goal: 8 cups
    ws5.cell(row=r, column=4, value=f'=IF(C{r}="","",C{r}/B{r})').number_format = '0%'
    ws5.cell(row=r, column=5).number_format = "HH:MM"
    ws5.cell(row=r, column=6).number_format = "HH:MM"
    ws5.cell(row=r, column=7, value=f'=IF(OR(E{r}="",F{r}=""),"",IF(F{r}>E{r},F{r}-E{r},1-E{r}+F{r}))')
    ws5.cell(row=r, column=7).number_format = '[h]:mm'
    sleep_qual_dv.add(f"H{r}")
    energy_next_dv.add(f"J{r}")
    style_data_row(ws5, r, 10, (r - 5) % 2 == 1)

# Monthly averages
avg_row = 37
ws5.cell(row=avg_row, column=1, value="MONTHLY AVERAGES").font = header_font(12, True, PLUM)
ws5.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=10)
ws5.cell(row=avg_row, column=1).fill = fill(LAVENDER)
ws5.cell(row=avg_row, column=1).alignment = center_align()

avgs = [
    ("Avg Water (cups/day)", "=AVERAGE(C5:C35)"),
    ("Water Goal Hit Rate", '=COUNTIF(D5:D35,">="&1)/COUNTA(C5:C35)'),
    ("Avg Hours Slept", "=AVERAGE(G5:G35)"),
    ("Days with Good+ Sleep", '=COUNTIFS(H5:H35,"Excellent")+COUNTIFS(H5:H35,"Good")'),
]
for idx, (label, formula) in enumerate(avgs):
    r = avg_row + 1 + idx
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws5.cell(row=r, column=1, value=label).font = body_font(11, True)
    ws5.cell(row=r, column=1).fill = fill(PALE_LILAC)
    ws5.cell(row=r, column=1).border = thin_border()
    val = ws5.cell(row=r, column=4, value=formula)
    if idx == 1:
        val.number_format = '0%'
    elif idx == 2:
        val.number_format = '0.0'
    val.font = body_font(12, True, PLUM)
    val.border = thin_border()
    val.alignment = center_align()

set_col_widths(ws5, [14, 12, 12, 12, 12, 12, 12, 14, 25, 14])

# ──────────────────────────────────────────────────
#  SHEET 6: GRATITUDE & REFLECTION JOURNAL
# ──────────────────────────────────────────────────
ws6 = wb.create_sheet("Gratitude Journal")
ws6.sheet_properties.tabColor = SOFT_GOLD

add_title_block(ws6, "GRATITUDE & REFLECTION JOURNAL", "Start and end each day with intention — what you focus on grows", 7)

row = 4
headers6 = ["Date", "3 Things I'm Grateful For", "Today's Affirmation",
            "Best Part of Today", "Lesson Learned", "Tomorrow I Will...", "Mood"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 7, SOFT_GOLD)

grat_mood_dv = DataValidation(type="list", formula1='"Grateful,Peaceful,Happy,Content,Hopeful,Reflective,Mixed,Struggling"', allow_blank=True)
ws6.add_data_validation(grat_mood_dv)

for r in range(5, 36):
    ws6.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    grat_mood_dv.add(f"G{r}")
    style_data_row(ws6, r, 7, (r - 5) % 2 == 1)
    ws6.row_dimensions[r].height = 40

# Affirmation bank
aff_row = 38
ws6.cell(row=aff_row, column=1, value="AFFIRMATION IDEAS").font = header_font(14, True, PLUM)
ws6.merge_cells(start_row=aff_row, start_column=1, end_row=aff_row, end_column=7)
ws6.cell(row=aff_row, column=1).fill = fill(LAVENDER)
ws6.cell(row=aff_row, column=1).alignment = center_align()

affirmations = [
    "I am worthy of love, happiness, and success.",
    "I choose progress over perfection.",
    "I am resilient and can handle whatever comes my way.",
    "My potential is limitless, and I am growing every day.",
    "I release what I cannot control and focus on what I can.",
    "I deserve rest and give myself permission to slow down.",
    "I am creating a life that feels good on the inside.",
    "I attract positivity and abundance into my life.",
    "My feelings are valid and I honor them without judgment.",
    "Today I choose joy, kindness, and peace.",
    "I am enough, exactly as I am right now.",
    "I trust the timing of my life and embrace the journey.",
]
for idx, aff in enumerate(affirmations):
    r = aff_row + 1 + idx
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws6.cell(row=r, column=1, value=f'  "{aff}"').font = Font(name="Calibri", size=10, italic=True, color=PLUM)
    ws6.cell(row=r, column=1).alignment = left_align()
    ws6.cell(row=r, column=1).fill = fill(LIGHT_GRAY if idx % 2 == 0 else WHITE)

set_col_widths(ws6, [14, 35, 30, 30, 30, 30, 14])

# ──────────────────────────────────────────────────
#  SHEET 7: SELF-CARE ROUTINE CHECKLIST
# ──────────────────────────────────────────────────
ws7 = wb.create_sheet("Self-Care Checklist")
ws7.sheet_properties.tabColor = DUSTY_ROSE

add_title_block(ws7, "SELF-CARE ROUTINE CHECKLIST", "You can't pour from an empty cup — schedule time for YOU", 6)

# Categories with activities
categories = {
    "PHYSICAL SELF-CARE": [
        "Take a long bath or shower",
        "Go for a walk in nature",
        "Stretch or do yoga",
        "Get a full night's sleep",
        "Eat a nourishing meal",
        "Dance to your favorite music",
        "Book a massage or spa day",
        "Try a new healthy recipe",
    ],
    "MENTAL SELF-CARE": [
        "Read a book for pleasure",
        "Do a puzzle or brain game",
        "Learn something new",
        "Write in a journal",
        "Listen to a podcast",
        "Take a social media break",
        "Organize one small space",
        "Set healthy boundaries",
    ],
    "EMOTIONAL SELF-CARE": [
        "Practice gratitude",
        "Have a good cry if you need it",
        "Talk to someone you trust",
        "Write a letter to yourself",
        "Celebrate a small win",
        "Forgive yourself for something",
        "Say no to something draining",
        "Do something that makes you laugh",
    ],
    "SOCIAL SELF-CARE": [
        "Call or text a friend",
        "Have a meaningful conversation",
        "Plan a fun outing",
        "Join a community or group",
        "Compliment someone sincerely",
        "Ask for help when you need it",
        "Set a boundary with someone",
        "Volunteer or help someone",
    ],
    "SPIRITUAL SELF-CARE": [
        "Meditate for 10 minutes",
        "Practice deep breathing",
        "Spend time in nature",
        "Write down your values",
        "Practice mindfulness",
        "Listen to calming music",
        "Reflect on what gives you purpose",
        "Disconnect and be present",
    ],
}

check_dv = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
ws7.add_data_validation(check_dv)

current_row = 4
for cat_name, activities in categories.items():
    ws7.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws7.cell(row=current_row, column=1, value=cat_name).font = header_font(13, True, PLUM)
    ws7.cell(row=current_row, column=1).fill = fill(LAVENDER)
    ws7.cell(row=current_row, column=1).alignment = left_align()
    current_row += 1

    # Headers for this section
    sec_headers = ["Activity", "Week 1", "Week 2", "Week 3", "Week 4", "Notes"]
    for i, h in enumerate(sec_headers, 1):
        ws7.cell(row=current_row, column=i, value=h)
    style_header_row(ws7, current_row, 6, DUSTY_ROSE)
    current_row += 1

    for idx, activity in enumerate(activities):
        ws7.cell(row=current_row, column=1, value=activity).font = body_font(10)
        ws7.cell(row=current_row, column=1).alignment = left_align()
        for w in range(2, 6):
            check_dv.add(f"{get_column_letter(w)}{current_row}")
        style_data_row(ws7, current_row, 6, idx % 2 == 1)
        current_row += 1

    current_row += 1  # Blank row between categories

set_col_widths(ws7, [35, 10, 10, 10, 10, 25])

# ──────────────────────────────────────────────────
#  SHEET 8: GOALS & MONTHLY REVIEW
# ──────────────────────────────────────────────────
ws8 = wb.create_sheet("Goals & Review")
ws8.sheet_properties.tabColor = SAGE

add_title_block(ws8, "MONTHLY GOALS & WELLNESS REVIEW", "Set intentions, track progress, and reflect on your growth", 7)

# Monthly goals
row = 4
ws8.cell(row=row, column=1, value="THIS MONTH'S WELLNESS GOALS").font = header_font(14, True, PLUM)
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
ws8.cell(row=row, column=1).fill = fill(LAVENDER)
ws8.cell(row=row, column=1).alignment = center_align()

row = 5
goal_headers = ["Goal", "Category", "Why It Matters", "Weekly Check-in", "Target", "Actual", "Status"]
for i, h in enumerate(goal_headers, 1):
    ws8.cell(row=row, column=i, value=h)
style_header_row(ws8, row, 7, DUSTY_ROSE)

goal_cat_dv = DataValidation(type="list", formula1='"Physical Health,Mental Health,Fitness,Nutrition,Sleep,Self-Care,Relationships,Personal Growth,Mindfulness,Other"', allow_blank=True)
ws8.add_data_validation(goal_cat_dv)

goal_status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Almost There,Completed,Adjusted"', allow_blank=True)
ws8.add_data_validation(goal_status_dv)

for r in range(6, 14):
    goal_cat_dv.add(f"B{r}")
    goal_status_dv.add(f"G{r}")
    style_data_row(ws8, r, 7, (r - 6) % 2 == 1)

# Monthly review section
rev_row = 16
ws8.cell(row=rev_row, column=1, value="END-OF-MONTH WELLNESS REVIEW").font = header_font(14, True, PLUM)
ws8.merge_cells(start_row=rev_row, start_column=1, end_row=rev_row, end_column=7)
ws8.cell(row=rev_row, column=1).fill = fill(LAVENDER)
ws8.cell(row=rev_row, column=1).alignment = center_align()

review_prompts = [
    ("Rate your overall wellness this month (1-10):", ""),
    ("What healthy habit stuck the best?", ""),
    ("What habit was hardest to maintain?", ""),
    ("How did your body feel this month?", ""),
    ("How did your mind feel this month?", ""),
    ("What brought you the most joy?", ""),
    ("What drained your energy the most?", ""),
    ("One thing you're proud of this month:", ""),
    ("One thing to improve next month:", ""),
    ("Word or intention for next month:", ""),
]

for idx, (prompt, _) in enumerate(review_prompts):
    r = rev_row + 1 + idx
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws8.cell(row=r, column=1, value=prompt).font = body_font(11, True)
    ws8.cell(row=r, column=1).fill = fill(PALE_LILAC)
    ws8.cell(row=r, column=1).border = thin_border()
    ws8.cell(row=r, column=1).alignment = left_align()
    ws8.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    ws8.cell(row=r, column=4).border = thin_border()
    ws8.row_dimensions[r].height = 35

# Annual wellness at a glance
annual_row = rev_row + 1 + len(review_prompts) + 2
ws8.cell(row=annual_row, column=1, value="ANNUAL WELLNESS AT A GLANCE").font = header_font(14, True, PLUM)
ws8.merge_cells(start_row=annual_row, start_column=1, end_row=annual_row, end_column=7)
ws8.cell(row=annual_row, column=1).fill = fill(LAVENDER)
ws8.cell(row=annual_row, column=1).alignment = center_align()

ann_headers = ["Month", "Wellness Rating", "Best Habit", "Biggest Win", "Focus Area", "Weight", "Notes"]
ann_h_row = annual_row + 1
for i, h in enumerate(ann_headers, 1):
    ws8.cell(row=ann_h_row, column=i, value=h)
style_header_row(ws8, ann_h_row, 7, SAGE)

months = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
for idx, month in enumerate(months):
    r = ann_h_row + 1 + idx
    ws8.cell(row=r, column=1, value=month)
    style_data_row(ws8, r, 7, idx % 2 == 1)

set_col_widths(ws8, [22, 18, 25, 20, 14, 14, 14])

# ──────────────────────────────────────────────────
#  SHEET 9: HOW TO USE
# ──────────────────────────────────────────────────
ws9 = wb.create_sheet("How to Use")
ws9.sheet_properties.tabColor = PALE_LILAC

add_title_block(ws9, "HOW TO USE YOUR WELLNESS TRACKER", "A gentle guide to getting the most from each tab", 6)

instructions = [
    ("GETTING STARTED", [
        "1. Save a copy of this file to your computer or Google Drive.",
        "2. Start with the Habit Tracker — pick 5-10 habits that matter to you.",
        "3. Don't try to be perfect. Tracking 3 habits consistently beats tracking 15 sporadically.",
        "4. Set a daily reminder (morning or evening) to update your tracker.",
    ]),
    ("HABIT TRACKER TAB", [
        "- Pre-loaded with 15 popular wellness habits — customize any of them!",
        "- Use ✓ for completed and X for missed (or leave blank for days off).",
        "- The Total and % columns auto-calculate your consistency.",
        "- The bottom row shows your daily completion rate across all habits.",
        "- TIP: Start with just 3-5 habits and add more once those are solid.",
    ]),
    ("MOOD TRACKER TAB", [
        "- Log your mood, energy, anxiety level, and sleep quality daily.",
        "- Use the dropdown menus for quick, consistent entries.",
        "- Write highlights and challenges — these reveal patterns over time.",
        "- Complete the Monthly Reflection section at month's end.",
        "- TIP: There's no 'right' answer — this is about self-awareness, not judgment.",
    ]),
    ("FITNESS LOG TAB", [
        "- Log every workout with type, duration, intensity, and how you felt.",
        "- Dropdown menus for workout type, intensity, and feeling.",
        "- Monthly summary auto-calculates total workouts, minutes, and calories.",
        "- TIP: Consistency matters more than intensity. A 20-min walk counts!",
    ]),
    ("MEAL PLANNER TAB", [
        "- Plan your meals for the week (Breakfast, Lunch, Dinner + Snacks).",
        "- Use the Grocery List section below to organize your shopping.",
        "- Category dropdown for groceries helps you shop by aisle.",
        "- TIP: Plan meals on Sunday, grocery shop Monday, prep Tuesday.",
    ]),
    ("WATER & SLEEP TRACKER TAB", [
        "- Default water goal is 8 cups/day — adjust to your needs.",
        "- Log bedtime and wake time — hours slept calculates automatically.",
        "- Monthly averages at the bottom show your trends.",
        "- TIP: Put a water bottle at your desk and refill it every 2 hours.",
    ]),
    ("GRATITUDE JOURNAL TAB", [
        "- Write 3 things you're grateful for each day (even small things!).",
        "- Pick a daily affirmation — we've included 12 to get you started.",
        "- Capture the best part of your day and one lesson learned.",
        "- TIP: Gratitude journaling for just 5 minutes boosts happiness measurably.",
    ]),
    ("SELF-CARE CHECKLIST TAB", [
        "- 40 self-care activities organized into 5 categories.",
        "- Check off activities each week as you do them.",
        "- No pressure to do everything — pick what resonates with you.",
        "- TIP: Schedule self-care like an appointment. You deserve it.",
    ]),
    ("GOALS & REVIEW TAB", [
        "- Set up to 8 wellness goals each month.",
        "- Complete the End-of-Month Review to reflect on your progress.",
        "- Use the Annual Wellness at a Glance to track your year.",
        "- TIP: Review goals weekly and adjust as needed. Flexibility is healthy!",
    ]),
    ("GENTLE REMINDERS", [
        "- Progress, not perfection. A 60% consistency rate is still amazing.",
        "- Missed a day (or a week)? Just start again. No guilt needed.",
        "- Your wellness journey is unique — don't compare to anyone else.",
        "- Rest IS productive. Recovery IS progress.",
        "- Be as kind to yourself as you would be to a friend.",
        "- You're already doing great by starting. Keep going.",
    ]),
]

current_row = 4
for section_title, items in instructions:
    ws9.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws9.cell(row=current_row, column=1, value=section_title).font = header_font(13, True, PLUM)
    ws9.cell(row=current_row, column=1).fill = fill(LAVENDER)
    ws9.cell(row=current_row, column=1).alignment = left_align()
    ws9.row_dimensions[current_row].height = 28
    current_row += 1
    for item in items:
        ws9.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws9.cell(row=current_row, column=1, value=f"  {item}").font = body_font(10)
        ws9.cell(row=current_row, column=1).alignment = left_align()
        current_row += 1
    current_row += 1

set_col_widths(ws9, [80, 15, 15, 15, 15, 15])

# ── Move How to Use to front ──
wb.move_sheet("How to Use", offset=-8)

# ── Save ──
output_path = "/home/user/claude-code/etsy-digital-download-3/product/Wellness_Habit_Tracker_Bundle_2026.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
