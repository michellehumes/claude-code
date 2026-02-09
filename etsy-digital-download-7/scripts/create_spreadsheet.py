#!/usr/bin/env python3
"""Create Teacher Classroom Planner 2026 Excel Spreadsheet with 8 professional tabs."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import datetime
import random

OUTPUT = "/home/user/claude-code/etsy-digital-download-7/product/Teacher_Classroom_Planner_2026.xlsx"

# Color palette
APPLE_RED = "C53030"
SUNSHINE = "D69E2E"
SKY_BLUE = "3182CE"
CHALK_GREEN = "38A169"
SOFT_YELLOW = "FEFCBF"
PALE_PEACH = "FFF5F5"
DARK_TEXT = "2D3748"
WHITE = "FFFFFF"

# Fills
fill_red = PatternFill(start_color=APPLE_RED, end_color=APPLE_RED, fill_type="solid")
fill_yellow = PatternFill(start_color=SUNSHINE, end_color=SUNSHINE, fill_type="solid")
fill_blue = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type="solid")
fill_green = PatternFill(start_color=CHALK_GREEN, end_color=CHALK_GREEN, fill_type="solid")
fill_soft_yellow = PatternFill(start_color=SOFT_YELLOW, end_color=SOFT_YELLOW, fill_type="solid")
fill_pale_peach = PatternFill(start_color=PALE_PEACH, end_color=PALE_PEACH, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_light_blue = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
fill_light_green = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
fill_light_red = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")

# Fonts
font_title = Font(name="Calibri", size=18, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=12, bold=True, color=DARK_TEXT)
font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_header_dark = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
font_body = Font(name="Calibri", size=10, color=DARK_TEXT)
font_body_bold = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_small = Font(name="Calibri", size=9, color="718096")
font_formula = Font(name="Calibri", size=10, bold=True, color=APPLE_RED)
font_instructions = Font(name="Calibri", size=11, color=DARK_TEXT)
font_instructions_bold = Font(name="Calibri", size=11, bold=True, color=SKY_BLUE)

# Borders
thin_border = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='thin', color='CBD5E0')
)
thick_bottom = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='medium', color=DARK_TEXT)
)

# Alignment
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')

# Sample data
FIRST_NAMES = ["Emma", "Liam", "Olivia", "Noah", "Ava", "Ethan", "Sophia", "Mason",
    "Isabella", "William", "Mia", "James", "Charlotte", "Benjamin", "Amelia",
    "Lucas", "Harper", "Henry", "Evelyn", "Alexander", "Abigail", "Daniel",
    "Emily", "Michael", "Ella", "Sebastian", "Avery", "Jack", "Scarlett",
    "Aiden", "Grace", "Owen", "Chloe", "Samuel", "Victoria", "Ryan",
    "Riley", "Nathan", "Aria", "Caleb"]
LAST_NAMES = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez", "Wilson",
    "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin", "Lee",
    "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark", "Ramirez",
    "Lewis", "Robinson", "Walker", "Young", "Allen", "King", "Wright",
    "Scott", "Torres", "Nguyen", "Hill", "Flores"]

random.seed(42)
STUDENTS = [f"{FIRST_NAMES[i]} {LAST_NAMES[i]}" for i in range(40)]
STUDENT_IDS = [f"STU-2026-{str(i+1).zfill(3)}" for i in range(40)]
GRADES = ["3rd", "4th", "5th"] * 14  # Mix of grades
GRADES = GRADES[:40]


def style_title_row(ws, row, max_col, title_text, fill_color):
    """Create a merged title row across all columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = font_title
    cell.fill = fill_color
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 45
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill_color


def style_header_row(ws, row, headers, fill_color, widths=None):
    """Style a header row with specific fill."""
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        cell.font = font_header
        cell.fill = fill_color
        cell.alignment = align_center
        cell.border = thin_border
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = 30


def style_data_cell(ws, row, col, value, is_alt=False):
    """Style a data cell with alternating row colors."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font_body
    cell.fill = fill_light_blue if is_alt else fill_white
    cell.alignment = align_center if col > 1 else align_left
    cell.border = thin_border
    return cell


def create_class_roster(wb):
    """Tab 1: Class Roster with 40 students."""
    ws = wb.active
    ws.title = "Class Roster"
    ws.sheet_properties.tabColor = APPLE_RED

    headers = ["Student Name", "Student ID", "Grade", "Parent/Guardian",
               "Parent Email", "Parent Phone", "Emergency Contact",
               "Allergies/Medical", "Birthday", "Notes"]
    widths = [22, 16, 8, 22, 28, 16, 22, 20, 14, 20]

    style_title_row(ws, 1, len(headers), "CLASS ROSTER  |  Teacher Classroom Planner 2026", fill_red)
    
    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub = ws.cell(row=2, column=1, value="Track student information, emergency contacts, and important details all in one place")
    sub.font = font_subtitle
    sub.fill = fill_pale_peach
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, len(headers)+1):
        ws.cell(row=2, column=c).fill = fill_pale_peach

    style_header_row(ws, 3, headers, fill_red, widths)

    # Sample data
    parent_names = [f"Mr./Mrs. {LAST_NAMES[i]}" for i in range(40)]
    allergies = ["None", "Peanut allergy", "None", "Bee sting allergy", "None",
                 "Gluten intolerance", "None", "None", "Asthma", "None"] * 4

    for i in range(40):
        row = i + 4
        is_alt = i % 2 == 1
        bday = datetime.date(2016 + random.randint(0, 2), random.randint(1, 12), random.randint(1, 28))
        
        data = [
            STUDENTS[i], STUDENT_IDS[i], GRADES[i], parent_names[i],
            f"{FIRST_NAMES[i].lower()}.{LAST_NAMES[i].lower()}@email.com",
            f"(555) {random.randint(100,999)}-{random.randint(1000,9999)}",
            f"Grandparent: (555) {random.randint(100,999)}-{random.randint(1000,9999)}",
            allergies[i], bday.strftime("%m/%d/%Y"), ""
        ]
        for j, val in enumerate(data):
            cell = style_data_cell(ws, row, j+1, val, is_alt)
            if j == 8:
                cell.number_format = 'MM/DD/YYYY'
        ws.row_dimensions[row].height = 22

    # Summary section
    summary_row = 45
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
    s = ws.cell(row=summary_row, column=1, value="SUMMARY")
    s.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    s.fill = fill_red
    s.alignment = align_center
    for c in range(1, 4):
        ws.cell(row=summary_row, column=c).fill = fill_red
        ws.cell(row=summary_row, column=c).border = thin_border

    labels = ["Total Students:", "Students with Allergies:", "Students with Medical Notes:"]
    formulas = [
        '=COUNTA(A4:A43)',
        '=COUNTIF(H4:H43,"<>None")-COUNTBLANK(H4:H43)',
        '=COUNTIF(H4:H43,"<>None")-COUNTBLANK(H4:H43)'
    ]
    for idx, (label, formula) in enumerate(zip(labels, formulas)):
        r = summary_row + 1 + idx
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font_body_bold
        lc.fill = fill_pale_peach
        lc.border = thin_border
        fc = ws.cell(row=r, column=2, value=formula if idx == 0 else None)
        if idx == 0:
            fc.font = font_formula
        fc.fill = fill_pale_peach
        fc.border = thin_border
        fc.alignment = align_center
        ws.cell(row=r, column=3).fill = fill_pale_peach
        ws.cell(row=r, column=3).border = thin_border

    ws.cell(row=summary_row+1, column=2).value = '=COUNTA(A4:A43)'
    ws.cell(row=summary_row+2, column=2).value = '=COUNTIF(H4:H43,"<>None")'
    ws.cell(row=summary_row+2, column=2).font = font_formula
    ws.cell(row=summary_row+3, column=2).value = '=COUNTIF(J4:J43,"<>")'
    ws.cell(row=summary_row+3, column=2).font = font_formula

    # Freeze panes
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J43"


def create_lesson_planner(wb):
    """Tab 2: Lesson Planner - 40 weeks."""
    ws = wb.create_sheet("Lesson Planner")
    ws.sheet_properties.tabColor = SKY_BLUE

    headers = ["Week #", "Date Range", "Subject", "Topic", "Learning Objectives",
               "Materials Needed", "Activities", "Assessment", "Homework",
               "Standards Aligned", "Notes"]
    widths = [8, 18, 14, 18, 24, 20, 22, 18, 18, 16, 18]

    style_title_row(ws, 1, len(headers), "LESSON PLANNER  |  Teacher Classroom Planner 2026", fill_blue)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub = ws.cell(row=2, column=1, value="Plan your entire school year week by week - 40 weeks of organized instruction")
    sub.font = font_subtitle
    sub.fill = fill_light_blue
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, len(headers)+1):
        ws.cell(row=2, column=c).fill = fill_light_blue

    style_header_row(ws, 3, headers, fill_blue, widths)

    subjects = ["Math", "ELA", "Science", "Social Studies", "Math", "ELA", "Science", "Art"]
    topics = {
        "Math": ["Place Value", "Addition & Subtraction", "Multiplication", "Division", "Fractions", "Geometry", "Measurement", "Word Problems"],
        "ELA": ["Reading Comprehension", "Narrative Writing", "Grammar & Punctuation", "Vocabulary", "Poetry", "Informational Text", "Persuasive Writing", "Book Reports"],
        "Science": ["Scientific Method", "Life Cycles", "Earth & Space", "Matter & Energy", "Weather", "Ecosystems", "Simple Machines", "Human Body"],
        "Social Studies": ["Community", "Geography", "Government", "Economics", "History", "Culture", "Maps & Globes", "Current Events"],
        "Art": ["Color Theory", "Drawing Basics", "Sculpture", "Art History", "Mixed Media", "Printmaking", "Photography", "Crafts"]
    }
    
    start_date = datetime.date(2026, 8, 17)  # School year start
    for i in range(40):
        row = i + 4
        is_alt = i % 2 == 1
        week_start = start_date + datetime.timedelta(weeks=i)
        week_end = week_start + datetime.timedelta(days=4)
        subject = subjects[i % len(subjects)]
        topic = topics[subject][i % len(topics[subject])]
        
        data = [
            i + 1,
            f"{week_start.strftime('%m/%d')} - {week_end.strftime('%m/%d/%Y')}",
            subject, topic,
            f"Students will be able to understand {topic.lower()}",
            "Textbook, worksheets, manipulatives",
            f"Guided practice, group work on {topic.lower()}",
            "Exit ticket + worksheet",
            f"Pg. {random.randint(10,200)} problems {random.randint(1,10)}-{random.randint(15,30)}",
            f"CCSS.{subject[:2].upper()}.{random.randint(1,5)}.{random.randint(1,9)}",
            ""
        ]
        for j, val in enumerate(data):
            style_data_cell(ws, row, j+1, val, is_alt)
        ws.row_dimensions[row].height = 28

    # Subject dropdown
    subj_dv = DataValidation(type="list", formula1='"Math,ELA,Science,Social Studies,Art,Music,PE,Library"', allow_blank=True)
    subj_dv.error = "Please select a valid subject"
    subj_dv.prompt = "Choose a subject"
    ws.add_data_validation(subj_dv)
    subj_dv.add(f"C4:C43")

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K43"


def create_grade_tracker(wb):
    """Tab 3: Grade Tracker with auto-average and letter grades."""
    ws = wb.create_sheet("Grade Tracker")
    ws.sheet_properties.tabColor = CHALK_GREEN

    # Headers: Student Name + 20 assignments + Average + Letter Grade
    headers = ["Student Name"]
    for i in range(1, 21):
        headers.append(f"Assgn {i}")
    headers.extend(["Average", "Letter Grade"])
    
    total_cols = len(headers)
    widths = [22] + [10]*20 + [12, 12]

    style_title_row(ws, 1, total_cols, "GRADE TRACKER  |  Teacher Classroom Planner 2026", fill_green)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub = ws.cell(row=2, column=1, value="Track 20 assignments per student with automatic averages and letter grades")
    sub.font = font_subtitle
    sub.fill = fill_light_green
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, total_cols+1):
        ws.cell(row=2, column=c).fill = fill_light_green

    style_header_row(ws, 3, headers, fill_green, widths)

    for i in range(40):
        row = i + 4
        is_alt = i % 2 == 1
        # Student name
        style_data_cell(ws, row, 1, STUDENTS[i], is_alt)
        
        # 20 assignment grades (sample data)
        for j in range(1, 21):
            grade = random.randint(55, 100)
            cell = style_data_cell(ws, row, j+1, grade, is_alt)
            cell.alignment = align_center
            # Conditional color
            if grade >= 90:
                cell.font = Font(name="Calibri", size=10, color=CHALK_GREEN)
            elif grade < 70:
                cell.font = Font(name="Calibri", size=10, color=APPLE_RED)

        # Average formula
        avg_col = get_column_letter(22)
        first_grade = get_column_letter(2)
        last_grade = get_column_letter(21)
        avg_formula = f'=AVERAGE({first_grade}{row}:{last_grade}{row})'
        avg_cell = style_data_cell(ws, row, 22, avg_formula, is_alt)
        avg_cell.font = font_body_bold
        avg_cell.number_format = '0.0'
        avg_cell.alignment = align_center

        # Letter grade formula
        letter_formula = f'=IF({avg_col}{row}>=90,"A",IF({avg_col}{row}>=80,"B",IF({avg_col}{row}>=70,"C",IF({avg_col}{row}>=60,"D","F"))))'
        letter_cell = style_data_cell(ws, row, 23, letter_formula, is_alt)
        letter_cell.font = font_body_bold
        letter_cell.alignment = align_center

        ws.row_dimensions[row].height = 22

    # Class averages row
    avg_row = 45
    ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=1)
    lbl = ws.cell(row=avg_row, column=1, value="CLASS AVERAGE")
    lbl.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    lbl.fill = fill_green
    lbl.alignment = align_center
    lbl.border = thick_bottom

    for j in range(2, 23):
        col_letter = get_column_letter(j)
        formula = f'=AVERAGE({col_letter}4:{col_letter}43)'
        cell = ws.cell(row=avg_row, column=j, value=formula)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = fill_green
        cell.alignment = align_center
        cell.border = thick_bottom
        if j < 22:
            cell.number_format = '0.0'
        elif j == 22:
            cell.number_format = '0.0'

    # Conditional formatting for letter grades
    ws.conditional_formatting.add(f'W4:W43',
        CellIsRule(operator='equal', formula=['"A"'], fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")))
    ws.conditional_formatting.add(f'W4:W43',
        CellIsRule(operator='equal', formula=['"F"'], fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")))

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:W43"


def create_attendance(wb):
    """Tab 4: Attendance tracker with P/A/T/E dropdowns."""
    ws = wb.create_sheet("Attendance")
    ws.sheet_properties.tabColor = SUNSHINE

    headers = ["Student Name"]
    for d in range(1, 32):
        headers.append(f"Day {d}")
    headers.append("Attendance %")
    
    total_cols = len(headers)
    widths = [22] + [6]*31 + [14]

    style_title_row(ws, 1, total_cols, "ATTENDANCE TRACKER  |  Teacher Classroom Planner 2026", fill_yellow)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub = ws.cell(row=2, column=1, value="P = Present  |  A = Absent  |  T = Tardy  |  E = Excused  — Monthly tracking for up to 31 days")
    sub.font = font_subtitle
    sub.fill = fill_soft_yellow
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, total_cols+1):
        ws.cell(row=2, column=c).fill = fill_soft_yellow

    style_header_row(ws, 3, headers, fill_yellow, widths)
    # Make header text dark for yellow
    for c in range(1, total_cols+1):
        ws.cell(row=3, column=c).font = Font(name="Calibri", size=9, bold=True, color=DARK_TEXT)

    # Dropdown for P/A/T/E
    att_dv = DataValidation(type="list", formula1='"P,A,T,E"', allow_blank=True)
    att_dv.error = "Use P, A, T, or E"
    att_dv.prompt = "P=Present, A=Absent, T=Tardy, E=Excused"
    ws.add_data_validation(att_dv)

    attendance_marks = ["P", "P", "P", "P", "P", "P", "P", "P", "P", "A", "P", "P", "T", "P", "P", "P", "P", "E", "P", "P"]

    for i in range(40):
        row = i + 4
        is_alt = i % 2 == 1
        style_data_cell(ws, row, 1, STUDENTS[i], is_alt)
        
        for d in range(1, 32):
            mark = random.choice(attendance_marks)
            cell = style_data_cell(ws, row, d+1, mark, is_alt)
            cell.alignment = align_center
            cell.font = Font(name="Calibri", size=9, color=DARK_TEXT)
            if mark == "A":
                cell.font = Font(name="Calibri", size=9, bold=True, color=APPLE_RED)
            elif mark == "T":
                cell.font = Font(name="Calibri", size=9, color=SUNSHINE)
        
        # Attendance % formula
        first_col = get_column_letter(2)
        last_col = get_column_letter(32)
        pct_formula = f'=COUNTIF({first_col}{row}:{last_col}{row},"P")/COUNTA({first_col}{row}:{last_col}{row})'
        pct_cell = style_data_cell(ws, row, 33, pct_formula, is_alt)
        pct_cell.number_format = '0.0%'
        pct_cell.font = font_body_bold
        pct_cell.alignment = align_center
        
        ws.row_dimensions[row].height = 20

    att_dv.add(f"B4:AF43")

    # Class attendance rate
    rate_row = 45
    lbl = ws.cell(row=rate_row, column=1, value="CLASS RATE")
    lbl.font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
    lbl.fill = fill_yellow
    lbl.alignment = align_center
    lbl.border = thick_bottom
    
    rate_cell = ws.cell(row=rate_row, column=33, value='=AVERAGE(AG4:AG43)')
    rate_cell.font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
    rate_cell.fill = fill_yellow
    rate_cell.number_format = '0.0%'
    rate_cell.alignment = align_center
    rate_cell.border = thick_bottom
    
    for c in range(2, 33):
        ws.cell(row=rate_row, column=c).fill = fill_yellow
        ws.cell(row=rate_row, column=c).border = thick_bottom

    ws.freeze_panes = "B4"


def create_classroom_budget(wb):
    """Tab 5: Classroom Budget tracker."""
    ws = wb.create_sheet("Classroom Budget")
    ws.sheet_properties.tabColor = CHALK_GREEN

    headers = ["Category", "Item", "Qty", "Cost Per", "Total Cost",
               "Source", "Reimbursed", "Date", "Notes"]
    widths = [18, 24, 8, 12, 14, 16, 14, 14, 22]

    style_title_row(ws, 1, len(headers), "CLASSROOM BUDGET  |  Teacher Classroom Planner 2026", fill_green)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub = ws.cell(row=2, column=1, value="Track every classroom expense, funding source, and reimbursement status")
    sub.font = font_subtitle
    sub.fill = fill_light_green
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, len(headers)+1):
        ws.cell(row=2, column=c).fill = fill_light_green

    style_header_row(ws, 3, headers, fill_green, widths)

    # Category dropdown
    cat_dv = DataValidation(type="list", formula1='"Supplies,Books,Technology,Decorations,Crafts,Field Trips,Snacks,Other"', allow_blank=True)
    ws.add_data_validation(cat_dv)
    cat_dv.add("A4:A53")

    # Source dropdown
    src_dv = DataValidation(type="list", formula1='"School,Personal,Donation,Grant"', allow_blank=True)
    ws.add_data_validation(src_dv)
    src_dv.add("F4:F53")

    # Reimbursed dropdown
    reimb_dv = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
    ws.add_data_validation(reimb_dv)
    reimb_dv.add("G4:G53")

    # Sample budget items
    items = [
        ("Supplies", "Dry erase markers (24-pack)", 3, 8.99, "School", "N/A"),
        ("Supplies", "Construction paper (500 sheets)", 2, 12.49, "School", "N/A"),
        ("Books", "Class novel set - Charlotte's Web", 30, 6.99, "School", "N/A"),
        ("Technology", "iPad stylus pens", 10, 9.99, "Grant", "N/A"),
        ("Decorations", "Bulletin board borders", 5, 4.99, "Personal", "Pending"),
        ("Crafts", "Glue sticks (bulk 100)", 1, 15.99, "School", "N/A"),
        ("Supplies", "Pencils (gross box)", 2, 11.99, "Donation", "N/A"),
        ("Snacks", "Goldfish crackers (30-ct)", 4, 8.49, "Personal", "No"),
        ("Field Trips", "Science museum bus", 1, 350.00, "School", "N/A"),
        ("Books", "Math workbooks", 30, 8.50, "School", "N/A"),
        ("Technology", "Replacement charger cables", 5, 7.99, "Personal", "Yes"),
        ("Decorations", "Classroom calendar set", 1, 19.99, "Personal", "Pending"),
        ("Supplies", "Whiteboard cleaner spray", 3, 5.49, "School", "N/A"),
        ("Crafts", "Pipe cleaners (500-ct)", 2, 6.99, "Donation", "N/A"),
        ("Snacks", "Animal crackers (individual)", 30, 0.50, "Personal", "No"),
    ]

    for i, (cat, item, qty, cost, source, reimb) in enumerate(items):
        row = i + 4
        is_alt = i % 2 == 1
        style_data_cell(ws, row, 1, cat, is_alt)
        style_data_cell(ws, row, 2, item, is_alt)
        cell_qty = style_data_cell(ws, row, 3, qty, is_alt)
        cell_qty.alignment = align_center
        cell_cost = style_data_cell(ws, row, 4, cost, is_alt)
        cell_cost.number_format = '$#,##0.00'
        cell_cost.alignment = align_right
        # Total formula
        total_formula = f'=C{row}*D{row}'
        cell_total = style_data_cell(ws, row, 5, total_formula, is_alt)
        cell_total.number_format = '$#,##0.00'
        cell_total.font = font_body_bold
        cell_total.alignment = align_right
        style_data_cell(ws, row, 6, source, is_alt)
        style_data_cell(ws, row, 7, reimb, is_alt)
        style_data_cell(ws, row, 8, datetime.date(2026, random.randint(8,12), random.randint(1,28)).strftime("%m/%d/%Y"), is_alt)
        style_data_cell(ws, row, 9, "", is_alt)
        ws.row_dimensions[row].height = 22

    # Empty rows for more entries
    for i in range(15, 50):
        row = i + 4
        is_alt = i % 2 == 1
        for j in range(1, 10):
            cell = style_data_cell(ws, row, j, "", is_alt)
        # Pre-fill total formula
        total_formula = f'=IF(C{row}*D{row}=0,"",C{row}*D{row})'
        cell_total = ws.cell(row=row, column=5, value=total_formula)
        cell_total.number_format = '$#,##0.00'
        cell_total.border = thin_border
        cell_total.fill = fill_light_blue if is_alt else fill_white

    # Totals row
    totals_row = 55
    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=4)
    t = ws.cell(row=totals_row, column=1, value="GRAND TOTAL")
    t.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    t.fill = fill_green
    t.alignment = align_center
    for c in range(1, 5):
        ws.cell(row=totals_row, column=c).fill = fill_green
        ws.cell(row=totals_row, column=c).border = thick_bottom
    
    total_cell = ws.cell(row=totals_row, column=5, value='=SUM(E4:E53)')
    total_cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    total_cell.fill = fill_green
    total_cell.number_format = '$#,##0.00'
    total_cell.alignment = align_center
    total_cell.border = thick_bottom

    # Summary by source
    sum_row = totals_row + 2
    ws.cell(row=sum_row, column=1, value="BY SOURCE:").font = font_body_bold
    sources = ["School", "Personal", "Donation", "Grant"]
    for idx, src in enumerate(sources):
        r = sum_row + 1 + idx
        ws.cell(row=r, column=1, value=src).font = font_body
        ws.cell(row=r, column=1).border = thin_border
        f = ws.cell(row=r, column=2, value=f'=SUMIF(F4:F53,"{src}",E4:E53)')
        f.number_format = '$#,##0.00'
        f.font = font_formula
        f.border = thin_border

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I53"


def create_parent_communication(wb):
    """Tab 6: Parent Communication Log."""
    ws = wb.create_sheet("Parent Communication")
    ws.sheet_properties.tabColor = SKY_BLUE

    headers = ["Date", "Student", "Parent Name", "Method", "Topic",
               "Action Items", "Follow-Up Needed", "Follow-Up Date", "Resolved"]
    widths = [14, 20, 20, 14, 24, 24, 16, 14, 12]

    style_title_row(ws, 1, len(headers), "PARENT COMMUNICATION LOG  |  Teacher Classroom Planner 2026", fill_blue)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub = ws.cell(row=2, column=1, value="Document every parent interaction for organized, professional communication tracking")
    sub.font = font_subtitle
    sub.fill = fill_light_blue
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, len(headers)+1):
        ws.cell(row=2, column=c).fill = fill_light_blue

    style_header_row(ws, 3, headers, fill_blue, widths)

    # Method dropdown
    method_dv = DataValidation(type="list", formula1='"Email,Phone,Conference,Note Home,App"', allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add("D4:D53")

    # Follow-up dropdown
    fu_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(fu_dv)
    fu_dv.add("G4:G53")

    # Resolved dropdown
    res_dv = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=True)
    ws.add_data_validation(res_dv)
    res_dv.add("I4:I53")

    methods = ["Email", "Phone", "Conference", "Note Home", "App"]
    topics = ["Behavior update", "Academic progress", "Missing homework", "Field trip permission",
              "IEP meeting schedule", "Positive report", "Attendance concern", "Supply request",
              "Birthday celebration", "Schedule change"]

    for i in range(15):
        row = i + 4
        is_alt = i % 2 == 1
        d = datetime.date(2026, random.randint(8, 12), random.randint(1, 28))
        student = STUDENTS[i]
        parent = f"Mr./Mrs. {LAST_NAMES[i]}"
        method = random.choice(methods)
        topic = random.choice(topics)
        
        data = [
            d.strftime("%m/%d/%Y"), student, parent, method, topic,
            "Review and follow up" if random.random() > 0.5 else "No action needed",
            "Yes" if random.random() > 0.6 else "No",
            (d + datetime.timedelta(days=7)).strftime("%m/%d/%Y") if random.random() > 0.5 else "",
            random.choice(["Yes", "No", "In Progress"])
        ]
        for j, val in enumerate(data):
            style_data_cell(ws, row, j+1, val, is_alt)
        ws.row_dimensions[row].height = 24

    # Empty rows
    for i in range(15, 50):
        row = i + 4
        is_alt = i % 2 == 1
        for j in range(1, 10):
            style_data_cell(ws, row, j, "", is_alt)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I53"


def create_iep_tracker(wb):
    """Tab 7: IEP/504 Tracker."""
    ws = wb.create_sheet("IEP-504 Tracker")
    ws.sheet_properties.tabColor = APPLE_RED

    headers = ["Student Name", "Plan Type", "Accommodation Summary",
               "Review Date", "Goals", "Progress Notes",
               "Next Meeting Date", "Case Manager"]
    widths = [22, 14, 30, 14, 28, 28, 16, 20]

    style_title_row(ws, 1, len(headers), "IEP / 504 TRACKER  |  Teacher Classroom Planner 2026", fill_red)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub = ws.cell(row=2, column=1, value="Confidential — Track accommodations, goals, progress, and meeting dates for IEP/504 students")
    sub.font = font_subtitle
    sub.fill = fill_pale_peach
    sub.alignment = align_center
    ws.row_dimensions[2].height = 25
    for c in range(1, len(headers)+1):
        ws.cell(row=2, column=c).fill = fill_pale_peach

    style_header_row(ws, 3, headers, fill_red, widths)

    # Plan type dropdown
    plan_dv = DataValidation(type="list", formula1='"IEP,504"', allow_blank=True)
    ws.add_data_validation(plan_dv)
    plan_dv.add("B4:B23")

    accommodations = [
        "Extended time on tests (1.5x)", "Preferential seating near teacher",
        "Reduced homework load", "Audio books provided",
        "Frequent breaks (every 20 min)", "Written instructions supplement verbal",
        "Modified assignments", "Use of calculator allowed",
        "Behavioral support plan", "Speech therapy 2x/week"
    ]
    case_managers = ["Ms. Rivera", "Mr. Chen", "Dr. Patel", "Mrs. Okafor", "Mr. Torres"]

    for i in range(10):
        row = i + 4
        is_alt = i % 2 == 1
        plan = "IEP" if i < 6 else "504"
        review = datetime.date(2026, random.randint(9, 12), random.randint(1, 28))
        next_meeting = review + datetime.timedelta(days=90)
        
        data = [
            STUDENTS[i*3], plan, accommodations[i],
            review.strftime("%m/%d/%Y"),
            f"Achieve grade-level proficiency in {'reading' if i % 2 == 0 else 'math'}",
            f"{'On track' if random.random() > 0.3 else 'Needs additional support'}",
            next_meeting.strftime("%m/%d/%Y"),
            random.choice(case_managers)
        ]
        for j, val in enumerate(data):
            cell = style_data_cell(ws, row, j+1, val, is_alt)
        ws.row_dimensions[row].height = 28

    # Empty rows
    for i in range(10, 20):
        row = i + 4
        is_alt = i % 2 == 1
        for j in range(1, 9):
            style_data_cell(ws, row, j, "", is_alt)

    ws.freeze_panes = "A4"


def create_how_to_use(wb):
    """Tab 8: How to Use instructions."""
    ws = wb.create_sheet("How to Use")
    ws.sheet_properties.tabColor = SUNSHINE

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value="HOW TO USE YOUR TEACHER CLASSROOM PLANNER 2026")
    title.font = Font(name="Calibri", size=20, bold=True, color=WHITE)
    title.fill = fill_blue
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 55
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = fill_blue

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 55

    instructions = [
        ("", ""),
        ("GETTING STARTED", ""),
        ("", ""),
        ("Step 1:", 'Save a backup copy of this file before making any changes.'),
        ("Step 2:", 'Start with the Class Roster tab - enter all student information first.'),
        ("Step 3:", 'Customize the Lesson Planner with your curriculum and schedule.'),
        ("Step 4:", 'Use the Grade Tracker throughout the year to monitor performance.'),
        ("Step 5:", 'Track attendance daily - the formulas calculate rates automatically.'),
        ("", ""),
        ("TAB-BY-TAB GUIDE", ""),
        ("", ""),
        ("Class Roster", "Your central hub for student information. Fill in all fields for each student. Use the Notes column for any special information. Emergency contacts are critical - fill these in first!"),
        ("Lesson Planner", "Plan 40 weeks of instruction. Use the Subject dropdown to categorize lessons. Align each lesson to standards. The Topic column helps you search for specific content later."),
        ("Grade Tracker", "Enter scores for each assignment (0-100 scale). Averages and letter grades calculate automatically. You can modify the letter grade thresholds in the formula if your school uses a different scale."),
        ("Attendance", "Use the dropdown: P=Present, A=Absent, T=Tardy, E=Excused. The percentage column shows each student's attendance rate. The class rate at the bottom shows overall attendance."),
        ("Classroom Budget", "Track every purchase. Use the Category and Source dropdowns. Total Cost calculates automatically (Qty x Cost Per). The summary section at the bottom shows totals by funding source."),
        ("Parent Communication", "Document EVERY parent interaction. This protects you professionally and ensures follow-through. Use the Follow-Up columns to stay on top of action items."),
        ("IEP/504 Tracker", "CONFIDENTIAL tab. Track accommodations, meeting dates, and progress. Set reminders for review dates. Keep this information secure and up to date."),
        ("", ""),
        ("PRO TIPS", ""),
        ("", ""),
        ("Tip 1:", "Use Ctrl+F (Cmd+F on Mac) to quickly search for any student across all tabs."),
        ("Tip 2:", "Right-click a tab to change the tab color and make it easier to navigate."),
        ("Tip 3:", "Print individual tabs by selecting the tab first, then Print > Active Sheet."),
        ("Tip 4:", "Back up your file weekly to cloud storage (Google Drive, OneDrive, etc.)."),
        ("Tip 5:", "Protect individual tabs with a password if sharing the file with others."),
        ("Tip 6:", "Use conditional formatting to highlight students who need extra attention."),
        ("", ""),
        ("SUPPORT", ""),
        ("", ""),
        ("", "Thank you for purchasing the Teacher Classroom Planner 2026!"),
        ("", "If you love this planner, please leave a review on our Etsy shop."),
        ("", "For questions or customization requests, message us through Etsy."),
        ("", ""),
        ("", "© 2026 Teacher Classroom Planner. All rights reserved."),
    ]

    for i, (label, text) in enumerate(instructions):
        row = i + 3
        if label in ["GETTING STARTED", "TAB-BY-TAB GUIDE", "PRO TIPS", "SUPPORT"]:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=2, value=label)
            cell.font = Font(name="Calibri", size=14, bold=True, color=SKY_BLUE)
            cell.fill = fill_light_blue
            cell.border = Border(bottom=Side(style='medium', color=SKY_BLUE))
            for c in range(2, 5):
                ws.cell(row=row, column=c).fill = fill_light_blue
            ws.row_dimensions[row].height = 35
        elif label.startswith("Step") or label.startswith("Tip"):
            cell_l = ws.cell(row=row, column=2, value=label)
            cell_l.font = font_instructions_bold
            cell_t = ws.cell(row=row, column=3)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            cell_t = ws.cell(row=row, column=3, value=text)
            cell_t.font = font_instructions
            cell_t.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 28
        elif label and text:
            cell_l = ws.cell(row=row, column=2, value=f"📋 {label}")
            cell_l.font = Font(name="Calibri", size=11, bold=True, color=CHALK_GREEN)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            cell_t = ws.cell(row=row, column=3, value=text)
            cell_t.font = font_instructions
            cell_t.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 50
        elif text:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=2, value=text)
            cell.font = font_instructions
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 22


def main():
    wb = openpyxl.Workbook()
    
    print("Creating Class Roster...")
    create_class_roster(wb)
    
    print("Creating Lesson Planner...")
    create_lesson_planner(wb)
    
    print("Creating Grade Tracker...")
    create_grade_tracker(wb)
    
    print("Creating Attendance...")
    create_attendance(wb)
    
    print("Creating Classroom Budget...")
    create_classroom_budget(wb)
    
    print("Creating Parent Communication...")
    create_parent_communication(wb)
    
    print("Creating IEP/504 Tracker...")
    create_iep_tracker(wb)
    
    print("Creating How to Use...")
    create_how_to_use(wb)
    
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Tabs: {wb.sheetnames}")
    print(f"Total tabs: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
