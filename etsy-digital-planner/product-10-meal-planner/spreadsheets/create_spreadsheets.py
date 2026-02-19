#!/usr/bin/env python3
"""
Create Meal Planner & Grocery List Bundle Spreadsheets for 2026.
Generates 6 professional Excel workbooks with formatting, formulas, and instructions.
"""

import os
import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference

# ── Color Palette ──
TERRACOTTA = "C4713B"
SAGE = "8FAE8B"
CREAM = "FFF8F0"
SOFT_ORANGE = "F4A460"
WHITE = "FFFFFF"
CHARCOAL = "444444"
SOFT_RED = "D4726A"
LIGHT_SAGE = "D4E8D1"
LIGHT_TERRA = "F0C8A8"

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Shared style helpers ──
header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
subheader_font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
body_font = Font(name="Calibri", size=11, color=CHARCOAL)
title_font = Font(name="Calibri", bold=True, size=16, color=TERRACOTTA)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=TERRACOTTA)

terra_fill = PatternFill(start_color=TERRACOTTA, end_color=TERRACOTTA, fill_type="solid")
sage_fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
cream_fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type="solid")
light_sage_fill = PatternFill(start_color=LIGHT_SAGE, end_color=LIGHT_SAGE, fill_type="solid")
light_terra_fill = PatternFill(start_color=LIGHT_TERRA, end_color=LIGHT_TERRA, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
orange_fill = PatternFill(start_color=SOFT_ORANGE, end_color=SOFT_ORANGE, fill_type="solid")
red_fill = PatternFill(start_color=SOFT_RED, end_color=SOFT_RED, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=SAGE),
    right=Side(style="thin", color=SAGE),
    top=Side(style="thin", color=SAGE),
    bottom=Side(style="thin", color=SAGE),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
currency_format = '"$"#,##0.00'


def style_header_row(ws, row, max_col, fill=None):
    """Apply header styling to a row."""
    if fill is None:
        fill = terra_fill
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border


def style_cell(cell, font=None, fill=None, alignment=None, border=None, num_fmt=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt


def auto_col_width(ws, min_width=12, max_width=30):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 4, max_width))
        ws.column_dimensions[col_letter].width = max_len


def add_instructions_tab(wb, title, instructions_list):
    """Add an Instructions tab to a workbook."""
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = TERRACOTTA

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="Calibri", bold=True, size=20, color=TERRACOTTA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:G2")
    ws["A2"].value = "2026 Meal Planner & Grocery List Bundle"
    ws["A2"].font = Font(name="Calibri", size=13, color=SAGE, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    for item in instructions_list:
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1, value=item)
        if item.startswith("##"):
            cell.value = item.replace("## ", "")
            cell.font = Font(name="Calibri", bold=True, size=13, color=TERRACOTTA)
            ws.row_dimensions[row].height = 30
        elif item.startswith("- "):
            cell.font = Font(name="Calibri", size=11, color=CHARCOAL)
            cell.alignment = Alignment(indent=2, wrap_text=True)
        else:
            cell.font = body_font
            cell.alignment = left_align
        row += 1

    ws.column_dimensions["A"].width = 15
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 15


# ═══════════════════════════════════════════════════════════
# 1. WEEKLY MEAL PLANNER 2026
# ═══════════════════════════════════════════════════════════
def create_weekly_meal_planner():
    print("  Creating Weekly Meal Planner 2026...")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = [
        "## How to Use This Weekly Meal Planner",
        "- Each tab represents one week of 2026 (Week 1 through Week 52).",
        "- Fill in your meals for Breakfast, Lunch, Dinner, and Snacks each day.",
        "- Add main ingredients, prep time, and optional calorie counts.",
        "- Use the Grocery List section at the bottom to plan your shopping.",
        "- The weekly budget auto-calculates from the Estimated Cost column.",
        "",
        "## Tips for Meal Planning",
        "- Plan meals on the weekend for the upcoming week.",
        "- Check your pantry inventory before adding grocery items.",
        "- Batch-cook proteins and grains on prep day.",
        "- Keep a rotating list of 15-20 family favorites.",
        "- Use theme nights (Taco Tuesday, Stir-Fry Friday) for easy planning.",
        "",
        "## Column Guide",
        "- Meal Name: What you plan to eat (e.g., 'Chicken Stir Fry').",
        "- Main Ingredients: Key items needed (e.g., 'chicken, broccoli, soy sauce').",
        "- Prep Time: Estimated preparation time in minutes.",
        "- Calories: Optional field for tracking nutrition.",
        "",
        "## Grocery List Categories",
        "- Produce, Dairy, Meat, Pantry, Frozen, Other.",
        "- The total at the bottom auto-sums your estimated costs.",
    ]
    add_instructions_tab(wb, "Weekly Meal Planner 2026", instructions)

    # Find the Monday of the first week of 2026
    jan1 = datetime(2026, 1, 1)
    # ISO week: start from the Monday of ISO week 1
    # The first Monday on or before Jan 4
    jan4 = datetime(2026, 1, 4)
    first_monday = jan4 - timedelta(days=jan4.weekday())

    for week_num in range(1, 53):
        ws = wb.create_sheet(f"Week {week_num}")
        ws.sheet_properties.tabColor = SAGE if week_num % 2 == 0 else TERRACOTTA

        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Title
        ws.merge_cells("A1:R1")
        cell = ws["A1"]
        cell.value = f"Week {week_num} - {week_start.strftime('%b %d')} to {week_end.strftime('%b %d, %Y')}"
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Day headers
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        meals = ["Breakfast", "Lunch", "Dinner", "Snacks"]
        col_headers = ["Meal Name", "Main Ingredients", "Prep Time", "Calories"]

        row = 3
        for d_idx, day in enumerate(days):
            day_date = week_start + timedelta(days=d_idx)

            # Day header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=1, value=f"{day} - {day_date.strftime('%b %d')}")
            cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
            day_fill = terra_fill if d_idx % 2 == 0 else sage_fill
            cell.fill = day_fill
            cell.alignment = center_align
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = day_fill
                ws.cell(row=row, column=c).border = thin_border
            ws.row_dimensions[row].height = 28
            row += 1

            # Column sub-headers
            sub_headers = ["Meal", "Meal Name", "Main Ingredients", "Prep Time (min)", "Calories"]
            for c_idx, h in enumerate(sub_headers):
                cell = ws.cell(row=row, column=c_idx + 1, value=h)
                cell.font = Font(name="Calibri", bold=True, size=10, color=CHARCOAL)
                cell.fill = cream_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

            # Meal rows
            for meal in meals:
                ws.cell(row=row, column=1, value=meal).font = Font(name="Calibri", bold=True, size=10, color=TERRACOTTA)
                ws.cell(row=row, column=1).fill = light_terra_fill if meal == "Dinner" else white_fill
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=1).alignment = left_align
                for c in range(2, 6):
                    cell = ws.cell(row=row, column=c)
                    cell.border = thin_border
                    cell.font = body_font
                    cell.alignment = left_align
                row += 1
            row += 1  # blank row between days

        # ── Grocery List Section ──
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value="WEEKLY GROCERY LIST")
        cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        cell.fill = terra_fill
        cell.alignment = center_align
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = terra_fill
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        grocery_headers = ["Item", "Quantity", "Category", "Estimated Cost", "Purchased?", "Notes"]
        for c_idx, h in enumerate(grocery_headers):
            cell = ws.cell(row=row, column=c_idx + 1, value=h)
            cell.font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
            cell.fill = light_sage_fill
            cell.alignment = center_align
            cell.border = thin_border
        grocery_header_row = row
        row += 1

        grocery_start = row
        categories = ["Produce", "Dairy", "Meat", "Pantry", "Frozen", "Other"]
        for i in range(20):
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.border = thin_border
                cell.font = body_font
                cell.alignment = left_align
            # Add data validation hint for category
            ws.cell(row=row, column=3).value = None
            ws.cell(row=row, column=4).number_format = currency_format
            row += 1
        grocery_end = row - 1

        # Total row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value="WEEKLY GROCERY BUDGET TOTAL:")
        cell.font = Font(name="Calibri", bold=True, size=12, color=TERRACOTTA)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell = ws.cell(row=row, column=4)
        total_cell.value = f"=SUM(D{grocery_start}:D{grocery_end})"
        total_cell.font = Font(name="Calibri", bold=True, size=12, color=TERRACOTTA)
        total_cell.number_format = currency_format
        total_cell.border = Border(bottom=Side(style="double", color=TERRACOTTA))

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 20

        # Freeze panes
        ws.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "Weekly_Meal_Planner_2026.xlsx")
    wb.save(path)
    print(f"    Saved: {path}")


# ═══════════════════════════════════════════════════════════
# 2. MONTHLY MEAL CALENDAR 2026
# ═══════════════════════════════════════════════════════════
def create_monthly_meal_calendar():
    print("  Creating Monthly Meal Calendar 2026...")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = [
        "## How to Use This Monthly Meal Calendar",
        "- Each tab shows one month of 2026 in calendar format.",
        "- Write your planned dinners in each day's cell.",
        "- Use the side panel for meal prep notes and theme nights.",
        "- Track your monthly grocery budget at the bottom.",
        "",
        "## Theme Night Ideas",
        "- Monday: Meatless Monday",
        "- Tuesday: Taco Tuesday",
        "- Wednesday: Pasta Night",
        "- Thursday: Slow Cooker Thursday",
        "- Friday: Pizza / Takeout Friday",
        "- Saturday: Grill Night / Try Something New",
        "- Sunday: Soup & Salad / Meal Prep Day",
        "",
        "## Monthly Budget Tracker",
        "- Enter your grocery spending as you shop.",
        "- The total auto-calculates at the bottom.",
    ]
    add_instructions_tab(wb, "Monthly Meal Calendar 2026", instructions)

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    for month_idx in range(1, 13):
        ws = wb.create_sheet(month_names[month_idx - 1])
        ws.sheet_properties.tabColor = SAGE if month_idx % 2 == 0 else TERRACOTTA

        # Title
        ws.merge_cells("A1:G1")
        cell = ws["A1"]
        cell.value = f"{month_names[month_idx - 1]} 2026 - Dinner Calendar"
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 45

        # Day-of-week headers
        dow = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for c, d in enumerate(dow, 1):
            cell = ws.cell(row=3, column=c, value=d)
            cell.font = header_font
            cell.fill = terra_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(c)].width = 18

        # Calendar grid
        cal = calendar.monthcalendar(2026, month_idx)
        row = 4
        for week in cal:
            # Date row
            for c, day in enumerate(week, 1):
                cell = ws.cell(row=row, column=c)
                if day != 0:
                    cell.value = day
                    cell.font = Font(name="Calibri", bold=True, size=11, color=TERRACOTTA)
                cell.fill = cream_fill
                cell.alignment = Alignment(horizontal="left", vertical="top")
                cell.border = thin_border
            ws.row_dimensions[row].height = 18
            row += 1

            # Dinner plan row (taller for writing)
            for c, day in enumerate(week, 1):
                cell = ws.cell(row=row, column=c)
                cell.border = thin_border
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if day == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            ws.row_dimensions[row].height = 50
            row += 1

        # ── Side Panel: starts at column I ──
        panel_col = 9
        ws.column_dimensions[get_column_letter(panel_col)].width = 30
        ws.column_dimensions[get_column_letter(panel_col + 1)].width = 25

        ws.merge_cells(start_row=3, start_column=panel_col, end_row=3, end_column=panel_col + 1)
        cell = ws.cell(row=3, column=panel_col, value="MEAL PREP NOTES")
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = sage_fill
        cell.alignment = center_align

        for r in range(4, 10):
            for c in [panel_col, panel_col + 1]:
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = body_font

        ws.merge_cells(start_row=11, start_column=panel_col, end_row=11, end_column=panel_col + 1)
        cell = ws.cell(row=11, column=panel_col, value="THEME NIGHTS")
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = terra_fill
        cell.alignment = center_align

        themes = [
            ("Monday", "Meatless Monday"),
            ("Tuesday", "Taco Tuesday"),
            ("Wednesday", "Pasta Night"),
            ("Thursday", "Slow Cooker"),
            ("Friday", "Pizza / Takeout"),
            ("Saturday", "Grill Night"),
            ("Sunday", "Meal Prep Day"),
        ]
        for i, (day, theme) in enumerate(themes):
            r = 12 + i
            ws.cell(row=r, column=panel_col, value=day).font = Font(name="Calibri", bold=True, size=10, color=CHARCOAL)
            ws.cell(row=r, column=panel_col).border = thin_border
            ws.cell(row=r, column=panel_col + 1, value=theme).font = body_font
            ws.cell(row=r, column=panel_col + 1).border = thin_border

        # Special occasions
        ws.merge_cells(start_row=20, start_column=panel_col, end_row=20, end_column=panel_col + 1)
        cell = ws.cell(row=20, column=panel_col, value="SPECIAL OCCASIONS")
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = orange_fill
        cell.alignment = center_align

        for r in range(21, 26):
            ws.cell(row=r, column=panel_col, value="Date:").font = body_font
            ws.cell(row=r, column=panel_col).border = thin_border
            ws.cell(row=r, column=panel_col + 1).border = thin_border

        # ── Monthly Grocery Budget Tracker ──
        budget_row = row + 2
        ws.merge_cells(start_row=budget_row, start_column=1, end_row=budget_row, end_column=5)
        cell = ws.cell(row=budget_row, column=1, value="MONTHLY GROCERY BUDGET")
        cell.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        cell.fill = terra_fill
        cell.alignment = center_align
        for c in range(1, 6):
            ws.cell(row=budget_row, column=c).fill = terra_fill
        budget_row += 1

        budget_headers = ["Week", "Budget", "Actual", "Difference", "Notes"]
        for c, h in enumerate(budget_headers, 1):
            cell = ws.cell(row=budget_row, column=c, value=h)
            cell.font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
            cell.fill = light_sage_fill
            cell.alignment = center_align
            cell.border = thin_border
        budget_row += 1

        budget_start = budget_row
        for w in range(1, 6):
            ws.cell(row=budget_row, column=1, value=f"Week {w}").font = body_font
            ws.cell(row=budget_row, column=1).border = thin_border
            for c in [2, 3]:
                ws.cell(row=budget_row, column=c).number_format = currency_format
                ws.cell(row=budget_row, column=c).border = thin_border
            ws.cell(row=budget_row, column=4).border = thin_border
            ws.cell(row=budget_row, column=4).number_format = currency_format
            ws.cell(row=budget_row, column=4, value=f"=B{budget_row}-C{budget_row}")
            ws.cell(row=budget_row, column=5).border = thin_border
            budget_row += 1
        budget_end = budget_row - 1

        # Monthly total
        ws.cell(row=budget_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11, color=TERRACOTTA)
        ws.cell(row=budget_row, column=1).border = thin_border
        for c in [2, 3, 4]:
            cell = ws.cell(row=budget_row, column=c)
            cell.value = f"=SUM({get_column_letter(c)}{budget_start}:{get_column_letter(c)}{budget_end})"
            cell.font = Font(name="Calibri", bold=True, size=11, color=TERRACOTTA)
            cell.number_format = currency_format
            cell.border = Border(bottom=Side(style="double", color=TERRACOTTA))

        ws.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "Monthly_Meal_Calendar_2026.xlsx")
    wb.save(path)
    print(f"    Saved: {path}")


# ═══════════════════════════════════════════════════════════
# 3. RECIPE COLLECTION TRACKER
# ═══════════════════════════════════════════════════════════
def create_recipe_collection_tracker():
    print("  Creating Recipe Collection Tracker...")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = [
        "## How to Use This Recipe Collection Tracker",
        "- Add your favorite recipes to the 'All Recipes' tab.",
        "- Fill in category, cuisine, times, servings, difficulty, and rating.",
        "- Use the 'Category Summary' tab to see counts by category.",
        "- Mark favorites with a 5-star rating to easily find them.",
        "",
        "## Categories",
        "- Breakfast, Lunch, Dinner, Dessert, Snack",
        "",
        "## Difficulty Levels",
        "- Easy: Under 30 min, simple techniques",
        "- Medium: 30-60 min, some cooking skills needed",
        "- Hard: Over 60 min or advanced techniques",
        "",
        "## Rating System",
        "- 1 = Won't make again",
        "- 2 = Okay, needs changes",
        "- 3 = Good, would make again",
        "- 4 = Great, family favorite",
        "- 5 = Outstanding, must-make",
    ]
    add_instructions_tab(wb, "Recipe Collection Tracker", instructions)

    # ── All Recipes tab ──
    ws = wb.create_sheet("All Recipes")
    ws.sheet_properties.tabColor = TERRACOTTA

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Recipe Collection Tracker"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = [
        "Recipe Name", "Category", "Cuisine Type", "Prep Time (min)",
        "Cook Time (min)", "Servings", "Difficulty", "Rating (1-5)",
        "Source / URL", "Notes"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = terra_fill
        cell.alignment = center_align
        cell.border = thin_border
    style_header_row(ws, 3, len(headers))

    # 100 pre-formatted rows
    for r in range(4, 104):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = left_align
            if r % 2 == 0:
                cell.fill = cream_fill

    widths = [30, 14, 18, 15, 15, 12, 14, 14, 30, 25]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A4"

    # ── Category Summary tab ──
    ws2 = wb.create_sheet("Category Summary")
    ws2.sheet_properties.tabColor = SAGE

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Category Summary"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    summary_headers = ["Category", "Count", "Avg Rating", "Avg Prep+Cook Time"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        style_cell(cell, font=header_font, fill=sage_fill, alignment=center_align, border=thin_border)

    categories = ["Breakfast", "Lunch", "Dinner", "Dessert", "Snack"]
    for i, cat in enumerate(categories):
        r = 4 + i
        ws2.cell(row=r, column=1, value=cat).font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=1).fill = light_sage_fill

        # COUNTIF for category
        ws2.cell(row=r, column=2, value=f'=COUNTIF(\'All Recipes\'!B4:B103,"{cat}")')
        ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=2).alignment = center_align

        # AVERAGEIF for rating
        ws2.cell(row=r, column=3, value=f'=IFERROR(AVERAGEIF(\'All Recipes\'!B4:B103,"{cat}",\'All Recipes\'!H4:H103),"N/A")')
        ws2.cell(row=r, column=3).border = thin_border
        ws2.cell(row=r, column=3).alignment = center_align
        ws2.cell(row=r, column=3).number_format = "0.0"

        # Average total time
        ws2.cell(row=r, column=4, value=f'=IFERROR(AVERAGEIF(\'All Recipes\'!B4:B103,"{cat}",\'All Recipes\'!D4:D103)+AVERAGEIF(\'All Recipes\'!B4:B103,"{cat}",\'All Recipes\'!E4:E103),"N/A")')
        ws2.cell(row=r, column=4).border = thin_border
        ws2.cell(row=r, column=4).alignment = center_align

    # Totals row
    r = 4 + len(categories)
    ws2.cell(row=r, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11, color=TERRACOTTA)
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})")
    ws2.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=11, color=TERRACOTTA)
    ws2.cell(row=r, column=2).border = Border(bottom=Side(style="double", color=TERRACOTTA))
    ws2.cell(row=r, column=2).alignment = center_align

    for c in range(1, 5):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    # ── Favorites tab ──
    ws3 = wb.create_sheet("Favorites (5 Stars)")
    ws3.sheet_properties.tabColor = SOFT_ORANGE

    ws3.merge_cells("A1:E1")
    ws3["A1"].value = "Favorite Recipes (Rating 5)"
    ws3["A1"].font = title_font
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 40

    ws3.merge_cells("A3:E3")
    ws3["A3"].value = "Tip: Filter the 'All Recipes' tab by Rating = 5 to find your favorites. Copy them here for quick reference!"
    ws3["A3"].font = Font(name="Calibri", italic=True, size=11, color=SAGE)
    ws3["A3"].alignment = left_align

    fav_headers = ["Recipe Name", "Category", "Cuisine", "Total Time", "Notes"]
    for c, h in enumerate(fav_headers, 1):
        cell = ws3.cell(row=5, column=c, value=h)
        style_cell(cell, font=header_font, fill=orange_fill, alignment=center_align, border=thin_border)

    for r in range(6, 31):
        for c in range(1, 6):
            cell = ws3.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            if r % 2 == 0:
                cell.fill = cream_fill

    for c in range(1, 6):
        ws3.column_dimensions[get_column_letter(c)].width = 24

    ws3.freeze_panes = "A6"

    path = os.path.join(OUTPUT_DIR, "Recipe_Collection_Tracker.xlsx")
    wb.save(path)
    print(f"    Saved: {path}")


# ═══════════════════════════════════════════════════════════
# 4. GROCERY BUDGET TRACKER 2026
# ═══════════════════════════════════════════════════════════
def create_grocery_budget_tracker():
    print("  Creating Grocery Budget Tracker 2026...")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = [
        "## How to Use This Grocery Budget Tracker",
        "- Each month has its own tab for recording grocery purchases.",
        "- Log every trip: date, store, items, category, amount, payment method.",
        "- Monthly totals and category breakdowns auto-calculate.",
        "- The 'Annual Dashboard' tab shows yearly trends and comparisons.",
        "",
        "## Spending Categories",
        "- Produce, Dairy, Meat/Protein, Bakery, Pantry/Dry Goods",
        "- Frozen, Beverages, Snacks, Household, Other",
        "",
        "## Tips for Budget Tracking",
        "- Record purchases the same day you shop.",
        "- Review weekly to stay on track.",
        "- Set a monthly budget goal and compare against actual spending.",
        "- Look for trends in the annual dashboard to find savings opportunities.",
    ]
    add_instructions_tab(wb, "Grocery Budget Tracker 2026", instructions)

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    categories = [
        "Produce", "Dairy", "Meat/Protein", "Bakery", "Pantry/Dry Goods",
        "Frozen", "Beverages", "Snacks", "Household", "Other"
    ]

    for m_idx, month in enumerate(month_names, 1):
        ws = wb.create_sheet(month)
        ws.sheet_properties.tabColor = SAGE if m_idx % 2 == 0 else TERRACOTTA

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"{month} 2026 - Grocery Budget Tracker"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Purchase log headers
        headers = ["Date", "Store", "Items Description", "Category", "Amount", "Payment Method"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            style_cell(cell, font=header_font, fill=terra_fill, alignment=center_align, border=thin_border)

        # 35 rows for entries
        for r in range(4, 39):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = body_font
                cell.alignment = left_align
                if r % 2 == 0:
                    cell.fill = cream_fill
            ws.cell(row=r, column=1).number_format = "MM/DD/YYYY"
            ws.cell(row=r, column=5).number_format = currency_format

        # Monthly totals
        r = 40
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r, column=1, value="MONTHLY TOTAL:").font = Font(name="Calibri", bold=True, size=12, color=TERRACOTTA)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=5, value="=SUM(E4:E38)").font = Font(name="Calibri", bold=True, size=12, color=TERRACOTTA)
        ws.cell(row=r, column=5).number_format = currency_format
        ws.cell(row=r, column=5).border = Border(bottom=Side(style="double", color=TERRACOTTA))

        r += 1
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r, column=1, value="AVERAGE PER TRIP:").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=5, value="=IFERROR(E40/COUNTA(E4:E38),0)").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
        ws.cell(row=r, column=5).number_format = currency_format

        r += 1
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r, column=1, value="NUMBER OF TRIPS:").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=5, value="=COUNTA(E4:E38)").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)

        # Category breakdown
        r = 45
        ws.merge_cells(f"A{r}:C{r}")
        cell = ws.cell(row=r, column=1, value="CATEGORY BREAKDOWN")
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = sage_fill
        cell.alignment = center_align
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = sage_fill
        r += 1

        cat_headers = ["Category", "Total Spent", "% of Total"]
        for c, h in enumerate(cat_headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            style_cell(cell, font=Font(name="Calibri", bold=True, size=11, color=CHARCOAL),
                       fill=light_sage_fill, alignment=center_align, border=thin_border)
        r += 1

        cat_start = r
        for cat in categories:
            ws.cell(row=r, column=1, value=cat).font = body_font
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2, value=f'=SUMIF(D4:D38,"{cat}",E4:E38)')
            ws.cell(row=r, column=2).number_format = currency_format
            ws.cell(row=r, column=2).border = thin_border
            ws.cell(row=r, column=3, value=f"=IFERROR(B{r}/E40,0)")
            ws.cell(row=r, column=3).number_format = "0.0%"
            ws.cell(row=r, column=3).border = thin_border
            r += 1

        widths = [16, 22, 30, 20, 16, 18]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        ws.freeze_panes = "A4"

    # ── Annual Dashboard ──
    ws_dash = wb.create_sheet("Annual Dashboard")
    ws_dash.sheet_properties.tabColor = SOFT_ORANGE

    ws_dash.merge_cells("A1:N1")
    ws_dash["A1"].value = "2026 Annual Grocery Budget Dashboard"
    ws_dash["A1"].font = Font(name="Calibri", bold=True, size=18, color=TERRACOTTA)
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 50

    # Monthly comparison
    ws_dash.merge_cells("A3:C3")
    ws_dash["A3"].value = "MONTHLY SPENDING COMPARISON"
    ws_dash["A3"].font = subtitle_font
    ws_dash["A3"].fill = light_terra_fill
    ws_dash["A3"].alignment = center_align
    for c in range(1, 4):
        ws_dash.cell(row=3, column=c).fill = light_terra_fill

    comp_headers = ["Month", "Total Spent", "# Trips"]
    for c, h in enumerate(comp_headers, 1):
        cell = ws_dash.cell(row=4, column=c, value=h)
        style_cell(cell, font=header_font, fill=terra_fill, alignment=center_align, border=thin_border)

    for i, month in enumerate(month_names):
        r = 5 + i
        ws_dash.cell(row=r, column=1, value=month).font = body_font
        ws_dash.cell(row=r, column=1).border = thin_border
        ws_dash.cell(row=r, column=2, value=f"='{month}'!E40").number_format = currency_format
        ws_dash.cell(row=r, column=2).border = thin_border
        ws_dash.cell(row=r, column=3, value=f"='{month}'!E42").border = thin_border
        ws_dash.cell(row=r, column=3).alignment = center_align

    # Annual totals
    r = 17
    ws_dash.cell(row=r, column=1, value="ANNUAL TOTAL").font = Font(name="Calibri", bold=True, size=12, color=TERRACOTTA)
    ws_dash.cell(row=r, column=1).border = Border(bottom=Side(style="double", color=TERRACOTTA))
    ws_dash.cell(row=r, column=2, value="=SUM(B5:B16)").font = Font(name="Calibri", bold=True, size=12, color=TERRACOTTA)
    ws_dash.cell(row=r, column=2).number_format = currency_format
    ws_dash.cell(row=r, column=2).border = Border(bottom=Side(style="double", color=TERRACOTTA))

    r = 18
    ws_dash.cell(row=r, column=1, value="MONTHLY AVERAGE").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
    ws_dash.cell(row=r, column=2, value="=B17/12").number_format = currency_format
    ws_dash.cell(row=r, column=2).font = body_font

    r = 19
    ws_dash.cell(row=r, column=1, value="HIGHEST MONTH").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
    ws_dash.cell(row=r, column=2, value="=MAX(B5:B16)").number_format = currency_format

    r = 20
    ws_dash.cell(row=r, column=1, value="LOWEST MONTH").font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
    ws_dash.cell(row=r, column=2, value="=MIN(B5:B16)").number_format = currency_format

    # Category annual breakdown
    r = 22
    ws_dash.merge_cells(f"A{r}:C{r}")
    ws_dash.cell(row=r, column=1, value="ANNUAL CATEGORY BREAKDOWN").font = subtitle_font
    ws_dash.cell(row=r, column=1).fill = light_sage_fill
    for c in range(1, 4):
        ws_dash.cell(row=r, column=c).fill = light_sage_fill

    r = 23
    for c, h in enumerate(["Category", "Annual Total", "% of Total"], 1):
        cell = ws_dash.cell(row=r, column=c, value=h)
        style_cell(cell, font=header_font, fill=sage_fill, alignment=center_align, border=thin_border)

    r = 24
    for cat in categories:
        ws_dash.cell(row=r, column=1, value=cat).font = body_font
        ws_dash.cell(row=r, column=1).border = thin_border
        # Sum across all months for this category
        formula_parts = []
        for m_idx, month in enumerate(month_names):
            cat_row = 47 + categories.index(cat)  # row in monthly tab
            formula_parts.append(f"'{month}'!B{cat_row}")
        ws_dash.cell(row=r, column=2, value="=" + "+".join(formula_parts))
        ws_dash.cell(row=r, column=2).number_format = currency_format
        ws_dash.cell(row=r, column=2).border = thin_border
        ws_dash.cell(row=r, column=3, value=f"=IFERROR(B{r}/B17,0)")
        ws_dash.cell(row=r, column=3).number_format = "0.0%"
        ws_dash.cell(row=r, column=3).border = thin_border
        r += 1

    # Bar chart for monthly spending
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Grocery Spending 2026"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 25
    chart.height = 15

    data = Reference(ws_dash, min_col=2, min_row=4, max_row=16)
    cats_ref = Reference(ws_dash, min_col=1, min_row=5, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = TERRACOTTA
    ws_dash.add_chart(chart, "E3")

    for c in range(1, 4):
        ws_dash.column_dimensions[get_column_letter(c)].width = 22

    ws_dash.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "Grocery_Budget_Tracker_2026.xlsx")
    wb.save(path)
    print(f"    Saved: {path}")


# ═══════════════════════════════════════════════════════════
# 5. PANTRY INVENTORY TRACKER
# ═══════════════════════════════════════════════════════════
def create_pantry_inventory_tracker():
    print("  Creating Pantry Inventory Tracker...")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = [
        "## How to Use This Pantry Inventory Tracker",
        "- List all items in your pantry, fridge, and freezer.",
        "- Update quantities as you use items and add new ones.",
        "- Check expiration dates regularly - items expiring soon highlight automatically.",
        "- Set Reorder to 'Y' for items that need restocking.",
        "- Use the Shopping List Generator tab to see all items needing reorder.",
        "",
        "## Location Guide",
        "- Pantry: Dry goods, canned items, spices, oils",
        "- Fridge: Fresh produce, dairy, leftovers, condiments",
        "- Freezer: Frozen meals, meats, vegetables, ice cream",
        "",
        "## Tips",
        "- Do a full inventory check monthly.",
        "- Use the FIFO method (First In, First Out) for rotation.",
        "- Keep a running list of items to reorder.",
        "- Group items by location for easier tracking.",
    ]
    add_instructions_tab(wb, "Pantry Inventory Tracker", instructions)

    # ── Inventory tab ──
    ws = wb.create_sheet("Inventory")
    ws.sheet_properties.tabColor = TERRACOTTA

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Pantry Inventory Tracker"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["Item Name", "Category", "Quantity", "Unit", "Expiration Date", "Location", "Reorder (Y/N)", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        style_cell(cell, font=header_font, fill=terra_fill, alignment=center_align, border=thin_border)

    # Pre-populated pantry items
    pantry_items = [
        # Pantry staples
        ("All-Purpose Flour", "Baking", 1, "bag", "", "Pantry", "N", "5 lb bag"),
        ("Sugar (White)", "Baking", 1, "bag", "", "Pantry", "N", ""),
        ("Brown Sugar", "Baking", 1, "bag", "", "Pantry", "N", ""),
        ("Baking Soda", "Baking", 1, "box", "", "Pantry", "N", ""),
        ("Baking Powder", "Baking", 1, "can", "", "Pantry", "N", ""),
        ("Vanilla Extract", "Baking", 1, "bottle", "", "Pantry", "N", ""),
        ("Cocoa Powder", "Baking", 1, "container", "", "Pantry", "N", ""),
        # Oils & Vinegars
        ("Olive Oil", "Oils & Vinegar", 1, "bottle", "", "Pantry", "N", "Extra virgin"),
        ("Vegetable Oil", "Oils & Vinegar", 1, "bottle", "", "Pantry", "N", ""),
        ("Apple Cider Vinegar", "Oils & Vinegar", 1, "bottle", "", "Pantry", "N", ""),
        ("Balsamic Vinegar", "Oils & Vinegar", 1, "bottle", "", "Pantry", "N", ""),
        # Grains & Pasta
        ("White Rice", "Grains", 2, "bag", "", "Pantry", "N", "Long grain"),
        ("Brown Rice", "Grains", 1, "bag", "", "Pantry", "N", ""),
        ("Spaghetti", "Pasta", 2, "box", "", "Pantry", "N", ""),
        ("Penne", "Pasta", 1, "box", "", "Pantry", "N", ""),
        ("Macaroni", "Pasta", 1, "box", "", "Pantry", "N", ""),
        ("Quinoa", "Grains", 1, "bag", "", "Pantry", "N", ""),
        ("Oats (Rolled)", "Grains", 1, "container", "", "Pantry", "N", ""),
        ("Bread Crumbs", "Grains", 1, "container", "", "Pantry", "N", ""),
        # Canned Goods
        ("Canned Tomatoes (Diced)", "Canned", 3, "can", "", "Pantry", "N", "14.5 oz"),
        ("Tomato Paste", "Canned", 2, "can", "", "Pantry", "N", "6 oz"),
        ("Tomato Sauce", "Canned", 2, "can", "", "Pantry", "N", "8 oz"),
        ("Canned Black Beans", "Canned", 2, "can", "", "Pantry", "N", ""),
        ("Canned Kidney Beans", "Canned", 2, "can", "", "Pantry", "N", ""),
        ("Canned Chickpeas", "Canned", 2, "can", "", "Pantry", "N", ""),
        ("Canned Corn", "Canned", 2, "can", "", "Pantry", "N", ""),
        ("Canned Tuna", "Canned", 3, "can", "", "Pantry", "N", ""),
        ("Chicken Broth", "Canned", 2, "carton", "", "Pantry", "N", "32 oz"),
        ("Coconut Milk", "Canned", 1, "can", "", "Pantry", "N", ""),
        # Spices
        ("Salt", "Spices", 1, "container", "", "Pantry", "N", ""),
        ("Black Pepper", "Spices", 1, "container", "", "Pantry", "N", ""),
        ("Garlic Powder", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Onion Powder", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Paprika", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Cumin", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Chili Powder", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Italian Seasoning", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Cinnamon", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Bay Leaves", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Red Pepper Flakes", "Spices", 1, "jar", "", "Pantry", "N", ""),
        ("Oregano", "Spices", 1, "jar", "", "Pantry", "N", ""),
        # Condiments
        ("Soy Sauce", "Condiments", 1, "bottle", "", "Pantry", "N", ""),
        ("Hot Sauce", "Condiments", 1, "bottle", "", "Pantry", "N", ""),
        ("Ketchup", "Condiments", 1, "bottle", "", "Fridge", "N", ""),
        ("Mustard", "Condiments", 1, "bottle", "", "Fridge", "N", ""),
        ("Mayonnaise", "Condiments", 1, "jar", "", "Fridge", "N", ""),
        ("Worcestershire Sauce", "Condiments", 1, "bottle", "", "Pantry", "N", ""),
        ("Honey", "Condiments", 1, "bottle", "", "Pantry", "N", ""),
        ("Maple Syrup", "Condiments", 1, "bottle", "", "Pantry", "N", ""),
        ("Peanut Butter", "Condiments", 1, "jar", "", "Pantry", "N", ""),
        ("Jam/Jelly", "Condiments", 1, "jar", "", "Fridge", "N", ""),
        # Fridge staples
        ("Butter", "Dairy", 1, "box", "", "Fridge", "N", "Unsalted"),
        ("Eggs", "Dairy", 1, "dozen", "", "Fridge", "Y", "Large"),
        ("Milk", "Dairy", 1, "gallon", "", "Fridge", "Y", "Whole or 2%"),
        ("Cheddar Cheese", "Dairy", 1, "block", "", "Fridge", "N", ""),
        ("Parmesan Cheese", "Dairy", 1, "container", "", "Fridge", "N", "Grated"),
        ("Cream Cheese", "Dairy", 1, "package", "", "Fridge", "N", ""),
        ("Greek Yogurt", "Dairy", 2, "container", "", "Fridge", "Y", "Plain"),
        ("Sour Cream", "Dairy", 1, "container", "", "Fridge", "N", ""),
        ("Heavy Cream", "Dairy", 1, "pint", "", "Fridge", "N", ""),
        # Produce
        ("Garlic", "Produce", 1, "head", "", "Pantry", "Y", ""),
        ("Onions (Yellow)", "Produce", 3, "each", "", "Pantry", "Y", ""),
        ("Potatoes", "Produce", 5, "each", "", "Pantry", "N", ""),
        ("Lemons", "Produce", 3, "each", "", "Fridge", "N", ""),
        ("Carrots", "Produce", 1, "bag", "", "Fridge", "N", ""),
        ("Celery", "Produce", 1, "bunch", "", "Fridge", "N", ""),
        ("Bell Peppers", "Produce", 2, "each", "", "Fridge", "N", ""),
        ("Tomatoes", "Produce", 3, "each", "", "Counter", "Y", ""),
        ("Lettuce/Salad Mix", "Produce", 1, "bag", "", "Fridge", "Y", ""),
        ("Bananas", "Produce", 1, "bunch", "", "Counter", "Y", ""),
        ("Apples", "Produce", 4, "each", "", "Fridge", "N", ""),
        # Freezer
        ("Frozen Chicken Breasts", "Protein", 1, "bag", "", "Freezer", "N", ""),
        ("Ground Beef", "Protein", 2, "lb", "", "Freezer", "N", ""),
        ("Frozen Vegetables (Mixed)", "Frozen", 2, "bag", "", "Freezer", "N", ""),
        ("Frozen Broccoli", "Frozen", 1, "bag", "", "Freezer", "N", ""),
        ("Frozen Berries", "Frozen", 1, "bag", "", "Freezer", "N", ""),
        ("Ice Cream", "Frozen", 1, "container", "", "Freezer", "N", ""),
        ("Frozen Pizza", "Frozen", 1, "box", "", "Freezer", "N", ""),
        # Snacks & Misc
        ("Tortilla Chips", "Snacks", 1, "bag", "", "Pantry", "N", ""),
        ("Crackers", "Snacks", 1, "box", "", "Pantry", "N", ""),
        ("Nuts (Mixed)", "Snacks", 1, "container", "", "Pantry", "N", ""),
        ("Granola Bars", "Snacks", 1, "box", "", "Pantry", "N", ""),
        ("Coffee", "Beverages", 1, "bag", "", "Pantry", "Y", "Ground or beans"),
        ("Tea Bags", "Beverages", 1, "box", "", "Pantry", "N", "Assorted"),
    ]

    for i, item in enumerate(pantry_items):
        r = 4 + i
        for c, val in enumerate(item, 1):
            cell = ws.cell(row=r, column=c, value=val if val != "" else None)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = left_align
            if i % 2 == 0:
                cell.fill = cream_fill
        if item[4]:  # If expiration date exists
            ws.cell(row=r, column=5).number_format = "MM/DD/YYYY"

    # Additional blank rows after pre-populated items
    for r in range(4 + len(pantry_items), 4 + len(pantry_items) + 30):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            if (r - 4) % 2 == 0:
                cell.fill = cream_fill

    widths = [28, 16, 12, 14, 16, 14, 14, 25]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A4"

    # Conditional formatting for expiring items (highlight yellow if within 7 days)
    expire_col_letter = "E"
    data_end = 4 + len(pantry_items) + 29
    yellow_fill_cf = PatternFill(start_color="FFFF00", end_color="FFFF00")
    red_fill_cf = PatternFill(start_color="FFB3B3", end_color="FFB3B3")

    ws.conditional_formatting.add(
        f"E4:E{data_end}",
        FormulaRule(
            formula=[f'AND(E4<>"",E4<TODAY()+7,E4>=TODAY())'],
            fill=PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"E4:E{data_end}",
        FormulaRule(
            formula=[f'AND(E4<>"",E4<TODAY())'],
            fill=PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
        )
    )

    # ── Shopping List Generator tab ──
    ws2 = wb.create_sheet("Shopping List Generator")
    ws2.sheet_properties.tabColor = SAGE

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Shopping List Generator"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    ws2.merge_cells("A2:E2")
    ws2["A2"].value = "Items below are those marked 'Y' in the Reorder column on the Inventory tab."
    ws2["A2"].font = Font(name="Calibri", italic=True, size=11, color=SAGE)

    shop_headers = ["Item Name", "Category", "Current Qty", "Location", "Notes"]
    for c, h in enumerate(shop_headers, 1):
        cell = ws2.cell(row=4, column=c, value=h)
        style_cell(cell, font=header_font, fill=sage_fill, alignment=center_align, border=thin_border)

    # Add formulas that reference inventory items where Reorder = Y
    # Since we can't dynamically filter in Excel without VBA easily, we pre-populate
    reorder_items = [(i, item) for i, item in enumerate(pantry_items) if item[6] == "Y"]
    for idx, (orig_idx, item) in enumerate(reorder_items):
        r = 5 + idx
        ws2.cell(row=r, column=1, value=item[0]).font = body_font
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=item[1]).font = body_font
        ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=3, value=item[2]).font = body_font
        ws2.cell(row=r, column=3).border = thin_border
        ws2.cell(row=r, column=4, value=item[5]).font = body_font
        ws2.cell(row=r, column=4).border = thin_border
        ws2.cell(row=r, column=5, value=item[7]).font = body_font
        ws2.cell(row=r, column=5).border = thin_border
        if idx % 2 == 0:
            for c in range(1, 6):
                ws2.cell(row=r, column=c).fill = cream_fill

    # Add blank rows for manual additions
    for r in range(5 + len(reorder_items), 5 + len(reorder_items) + 15):
        for c in range(1, 6):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font

    for c in range(1, 6):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    ws2.freeze_panes = "A5"

    path = os.path.join(OUTPUT_DIR, "Pantry_Inventory_Tracker.xlsx")
    wb.save(path)
    print(f"    Saved: {path}")


# ═══════════════════════════════════════════════════════════
# 6. MEAL PREP PLANNER
# ═══════════════════════════════════════════════════════════
def create_meal_prep_planner():
    print("  Creating Meal Prep Planner...")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = [
        "## How to Use This Meal Prep Planner",
        "- Use the 'Prep Day Schedule' tab to plan your cooking day hour by hour.",
        "- Track batch cooking in the 'Batch Cooking Tracker' tab.",
        "- Print container labels from the 'Container Labels' tab.",
        "- Check off items on the 'Weekly Prep Checklist' each week.",
        "",
        "## Meal Prep Tips",
        "- Start with proteins that take longest to cook.",
        "- Prep vegetables while proteins are cooking.",
        "- Let everything cool completely before storing.",
        "- Label all containers with meal name, date, and reheating instructions.",
        "- Most prepped meals last 3-4 days in the fridge, 2-3 months in the freezer.",
        "",
        "## Time-Saving Strategies",
        "- Cook grains in bulk (rice, quinoa, pasta).",
        "- Wash and chop all produce at once.",
        "- Marinate proteins overnight for next-day cooking.",
        "- Use sheet pan meals for easy cleanup.",
        "- Double recipes and freeze half.",
    ]
    add_instructions_tab(wb, "Meal Prep Planner", instructions)

    # ── Prep Day Schedule tab ──
    ws = wb.create_sheet("Prep Day Schedule")
    ws.sheet_properties.tabColor = TERRACOTTA

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Prep Day Schedule"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    ws["A2"].value = "Plan your meal prep day with time blocks for maximum efficiency"
    ws["A2"].font = Font(name="Calibri", italic=True, size=11, color=SAGE)

    headers = ["Time Block", "Task", "Recipe", "Duration (min)", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        style_cell(cell, font=header_font, fill=terra_fill, alignment=center_align, border=thin_border)

    # Pre-populate time blocks
    time_blocks = [
        ("8:00 AM", "Preheat oven, gather ingredients", "", "15", "Get everything out and ready"),
        ("8:15 AM", "Start marinating proteins", "", "10", ""),
        ("8:30 AM", "Wash and chop all vegetables", "", "30", "Batch process all veggies"),
        ("9:00 AM", "Start cooking grains (rice, quinoa)", "", "20", "Set timers"),
        ("9:20 AM", "Cook proteins (oven/stovetop)", "", "30", ""),
        ("9:50 AM", "Prepare sauces and dressings", "", "15", ""),
        ("10:05 AM", "Assemble salads and cold items", "", "20", ""),
        ("10:25 AM", "Cook second batch of proteins", "", "30", ""),
        ("10:55 AM", "Portion meals into containers", "", "20", ""),
        ("11:15 AM", "Label all containers", "", "10", "Use Container Labels tab"),
        ("11:25 AM", "Clean up kitchen", "", "20", ""),
        ("11:45 AM", "Store meals in fridge/freezer", "", "10", ""),
        ("12:00 PM", "DONE! Enjoy the rest of your day", "", "", ""),
    ]

    for i, (time, task, recipe, duration, notes) in enumerate(time_blocks):
        r = 5 + i
        vals = [time, task, recipe, duration, notes]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = left_align
            if i % 2 == 0:
                cell.fill = cream_fill
        ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11, color=TERRACOTTA)

    # Blank rows
    for r in range(5 + len(time_blocks), 5 + len(time_blocks) + 10):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font

    widths = [14, 35, 25, 16, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A5"

    # ── Batch Cooking Tracker ──
    ws2 = wb.create_sheet("Batch Cooking Tracker")
    ws2.sheet_properties.tabColor = SAGE

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "Batch Cooking Tracker"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    batch_headers = ["Recipe", "Servings Made", "Storage Method", "Date Made", "Use By Date", "Notes"]
    for c, h in enumerate(batch_headers, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        style_cell(cell, font=header_font, fill=sage_fill, alignment=center_align, border=thin_border)

    for r in range(4, 34):
        for c in range(1, 7):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = left_align
            if r % 2 == 0:
                cell.fill = cream_fill
        ws2.cell(row=r, column=4).number_format = "MM/DD/YYYY"
        ws2.cell(row=r, column=5).number_format = "MM/DD/YYYY"

    widths2 = [30, 16, 18, 14, 14, 25]
    for i, w in enumerate(widths2):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    ws2.freeze_panes = "A4"

    # ── Container Labels ──
    ws3 = wb.create_sheet("Container Labels")
    ws3.sheet_properties.tabColor = SOFT_ORANGE

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "Container Labels"
    ws3["A1"].font = title_font
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 40

    ws3.merge_cells("A2:D2")
    ws3["A2"].value = "Fill in and print to label your meal prep containers"
    ws3["A2"].font = Font(name="Calibri", italic=True, size=11, color=SAGE)

    label_headers = ["Meal Name", "Date Prepared", "Reheating Instructions", "Servings"]
    for c, h in enumerate(label_headers, 1):
        cell = ws3.cell(row=4, column=c, value=h)
        style_cell(cell, font=header_font, fill=orange_fill, alignment=center_align, border=thin_border)

    for r in range(5, 30):
        for c in range(1, 5):
            cell = ws3.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = left_align
            if r % 2 == 0:
                cell.fill = cream_fill
        ws3.cell(row=r, column=2).number_format = "MM/DD/YYYY"

    widths3 = [28, 16, 35, 12]
    for i, w in enumerate(widths3):
        ws3.column_dimensions[get_column_letter(i + 1)].width = w

    ws3.freeze_panes = "A5"

    # ── Weekly Prep Checklist ──
    ws4 = wb.create_sheet("Weekly Prep Checklist")
    ws4.sheet_properties.tabColor = SAGE

    ws4.merge_cells("A1:D1")
    ws4["A1"].value = "Weekly Prep Checklist"
    ws4["A1"].font = title_font
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 40

    checklist_headers = ["Done?", "Task", "Day", "Notes"]
    for c, h in enumerate(checklist_headers, 1):
        cell = ws4.cell(row=3, column=c, value=h)
        style_cell(cell, font=header_font, fill=terra_fill, alignment=center_align, border=thin_border)

    checklist_items = [
        ("", "Review weekly meal plan", "Saturday", "Check what needs to be prepped"),
        ("", "Check pantry inventory", "Saturday", "Update Pantry Tracker"),
        ("", "Create grocery list", "Saturday", "Use Weekly Planner grocery section"),
        ("", "Go grocery shopping", "Saturday/Sunday", "Buy fresh produce last"),
        ("", "Wash all produce", "Sunday", "Spin dry leafy greens"),
        ("", "Chop vegetables", "Sunday", "Store in airtight containers"),
        ("", "Marinate proteins", "Sunday AM", "Overnight for best flavor"),
        ("", "Cook grains in bulk", "Sunday", "Rice, quinoa, pasta"),
        ("", "Cook proteins", "Sunday", "Bake, grill, or slow cook"),
        ("", "Prepare sauces/dressings", "Sunday", "Store in glass jars"),
        ("", "Assemble meals", "Sunday", "Use Container Labels tab"),
        ("", "Label all containers", "Sunday", "Include date and reheating info"),
        ("", "Store meals properly", "Sunday", "Fridge: 3-4 days, Freezer: 2-3 months"),
        ("", "Prepare overnight oats/breakfasts", "Sunday", "Mason jars for easy grab-and-go"),
        ("", "Portion snacks", "Sunday", "Nuts, fruits, veggies"),
        ("", "Clean and organize kitchen", "Sunday", "Reset for the week"),
        ("", "Take inventory of leftover supplies", "Sunday", "Note anything to buy next week"),
    ]

    for i, (done, task, day, notes) in enumerate(checklist_items):
        r = 4 + i
        ws4.cell(row=r, column=1, value=done).border = thin_border
        ws4.cell(row=r, column=1).alignment = center_align
        ws4.cell(row=r, column=2, value=task).font = body_font
        ws4.cell(row=r, column=2).border = thin_border
        ws4.cell(row=r, column=3, value=day).font = body_font
        ws4.cell(row=r, column=3).border = thin_border
        ws4.cell(row=r, column=3).alignment = center_align
        ws4.cell(row=r, column=4, value=notes).font = body_font
        ws4.cell(row=r, column=4).border = thin_border
        if i % 2 == 0:
            for c in range(1, 5):
                ws4.cell(row=r, column=c).fill = cream_fill

    widths4 = [10, 38, 18, 35]
    for i, w in enumerate(widths4):
        ws4.column_dimensions[get_column_letter(i + 1)].width = w

    ws4.freeze_panes = "A4"

    path = os.path.join(OUTPUT_DIR, "Meal_Prep_Planner.xlsx")
    wb.save(path)
    print(f"    Saved: {path}")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("Generating 2026 Meal Planner & Grocery List Bundle")
    print("=" * 60)

    create_weekly_meal_planner()
    create_monthly_meal_calendar()
    create_recipe_collection_tracker()
    create_grocery_budget_tracker()
    create_pantry_inventory_tracker()
    create_meal_prep_planner()

    print("\n" + "=" * 60)
    print("All 6 spreadsheets created successfully!")
    print("=" * 60)
