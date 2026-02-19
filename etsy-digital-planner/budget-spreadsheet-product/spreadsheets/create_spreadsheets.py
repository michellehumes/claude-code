#!/usr/bin/env python3
"""
Create 2026 Budget & Finance Spreadsheet Bundle
6 professionally formatted Excel spreadsheets with formulas
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
import os

# Color palette - Modern sage/money green theme
SAGE_GREEN = "7B9E87"
DEEP_GREEN = "4A6B55"
SOFT_MINT = "D4E6D9"
WARM_CREAM = "FDF8F0"
CHARCOAL = "3D3D3D"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
CORAL_ACCENT = "E8917A"
GOLD_ACCENT = "D4A853"
WHITE = "FFFFFF"

# Reusable styles
header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=DEEP_GREEN, end_color=DEEP_GREEN, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
subheader_fill = PatternFill(start_color=SOFT_MINT, end_color=SOFT_MINT, fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=18, color=DEEP_GREEN)
subtitle_font = Font(name="Calibri", size=11, color=SAGE_GREEN)
body_font = Font(name="Calibri", size=11, color=CHARCOAL)
currency_format = '"$"#,##0.00'
percent_format = '0.0%'
thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)
accent_fill = PatternFill(start_color=WARM_CREAM, end_color=WARM_CREAM, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
sage_fill = PatternFill(start_color=SAGE_GREEN, end_color=SAGE_GREEN, fill_type="solid")
mint_fill = PatternFill(start_color=SOFT_MINT, end_color=SOFT_MINT, fill_type="solid")
coral_fill = PatternFill(start_color=CORAL_ACCENT, end_color=CORAL_ACCENT, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD_ACCENT, end_color=GOLD_ACCENT, fill_type="solid")

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_data_cell(ws, row, col, is_currency=False, is_percent=False):
    cell = ws.cell(row=row, column=col)
    cell.font = body_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_currency:
        cell.number_format = currency_format
    if is_percent:
        cell.number_format = percent_format
    if row % 2 == 0:
        cell.fill = PatternFill(start_color="F9FAF9", end_color="F9FAF9", fill_type="solid")

def add_title(ws, title, subtitle=""):
    ws.merge_cells("A1:H1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"].value = subtitle
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 25


def create_monthly_budget():
    """Sheet 1: Monthly Budget Tracker with categories and charts"""
    wb = openpyxl.Workbook()

    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]

    # Instructions sheet
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.sheet_properties.tabColor = SAGE_GREEN
    add_title(ws_inst, "2026 Monthly Budget Tracker", "Your complete monthly spending & income tracker")

    instructions = [
        "",
        "HOW TO USE THIS SPREADSHEET:",
        "",
        "1. Start on the 'January' tab (or your current month)",
        "2. Enter your income sources at the top",
        "3. Fill in your budgeted amounts for each category",
        "4. As you spend, enter actual amounts",
        "5. The spreadsheet will auto-calculate differences",
        "6. Check the Dashboard tab for yearly overview",
        "",
        "TIPS:",
        "- Green cells = under budget (good!)",
        "- Red cells = over budget (review spending)",
        "- Adjust categories to fit YOUR life",
        "- Review monthly to stay on track",
        "",
        "CATEGORIES INCLUDED:",
        "Housing, Utilities, Transportation, Groceries,",
        "Dining Out, Entertainment, Shopping, Health,",
        "Insurance, Debt Payments, Savings, Subscriptions,",
        "Personal Care, Education, Gifts, Miscellaneous",
    ]
    for i, line in enumerate(instructions, 4):
        ws_inst.cell(row=i, column=1, value=line).font = body_font
    ws_inst.column_dimensions["A"].width = 60

    categories = [
        ("HOUSING", ["Rent/Mortgage", "Property Tax", "HOA Fees", "Home Insurance", "Maintenance"]),
        ("UTILITIES", ["Electric", "Gas/Heat", "Water", "Internet", "Phone", "Trash/Recycling"]),
        ("TRANSPORTATION", ["Car Payment", "Gas/Fuel", "Car Insurance", "Parking", "Public Transit", "Maintenance"]),
        ("FOOD", ["Groceries", "Dining Out", "Coffee Shops", "Meal Delivery"]),
        ("HEALTH", ["Health Insurance", "Medications", "Doctor Visits", "Gym/Fitness", "Mental Health"]),
        ("ENTERTAINMENT", ["Streaming Services", "Movies/Events", "Hobbies", "Books/Media", "Vacations"]),
        ("FINANCIAL", ["Savings", "Investments", "Emergency Fund", "Debt Extra Payments"]),
        ("PERSONAL", ["Clothing", "Personal Care", "Gifts", "Education", "Subscriptions", "Miscellaneous"]),
    ]

    # Create monthly sheets
    for month in months:
        ws = wb.create_sheet(title=month[:3])
        ws.sheet_properties.tabColor = SAGE_GREEN

        add_title(ws, f"{month} 2026 Budget", "Track your income and expenses")

        # Income section
        row = 4
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1, value="INCOME").font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        ws.cell(row=row, column=1).fill = PatternFill(start_color=SAGE_GREEN, end_color=SAGE_GREEN, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=SAGE_GREEN, end_color=SAGE_GREEN, fill_type="solid")

        row = 5
        headers = ["Source", "Expected", "Actual", "Difference", "Notes"]
        for col, h in enumerate(headers, 1):
            style_header_row(ws, row, 5)
            ws.cell(row=row, column=col, value=h)

        income_sources = ["Primary Job", "Side Income", "Freelance", "Investments", "Other"]
        for i, src in enumerate(income_sources):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=src).font = body_font
            ws.cell(row=r, column=1).border = thin_border
            for c in range(2, 4):
                style_data_cell(ws, r, c, is_currency=True)
            ws.cell(row=r, column=4).value = f"=C{r}-B{r}"
            style_data_cell(ws, r, 4, is_currency=True)
            style_data_cell(ws, r, 5)

        total_row = row + len(income_sources) + 1
        ws.cell(row=total_row, column=1, value="TOTAL INCOME").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_row, column=1).border = thin_border
        ws.cell(row=total_row, column=2).value = f"=SUM(B{row+1}:B{total_row-1})"
        ws.cell(row=total_row, column=2).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_row, column=2).number_format = currency_format
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3).value = f"=SUM(C{row+1}:C{total_row-1})"
        ws.cell(row=total_row, column=3).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_row, column=3).number_format = currency_format
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=4).value = f"=C{total_row}-B{total_row}"
        ws.cell(row=total_row, column=4).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_row, column=4).number_format = currency_format
        ws.cell(row=total_row, column=4).border = thin_border

        income_total_row = total_row

        # Expense section
        row = total_row + 2
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1, value="EXPENSES").font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=CORAL_ACCENT, end_color=CORAL_ACCENT, fill_type="solid")

        row += 1
        headers = ["Category", "Budgeted", "Actual", "Difference", "Notes"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header_row(ws, row, 5)

        expense_start = row + 1
        current_row = expense_start

        for cat_name, items in categories:
            # Category header
            ws.cell(row=current_row, column=1, value=cat_name).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
            ws.cell(row=current_row, column=1).fill = mint_fill
            for c in range(1, 6):
                ws.cell(row=current_row, column=c).fill = mint_fill
                ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

            for item in items:
                ws.cell(row=current_row, column=1, value=item).font = body_font
                ws.cell(row=current_row, column=1).border = thin_border
                for c in range(2, 4):
                    style_data_cell(ws, current_row, c, is_currency=True)
                ws.cell(row=current_row, column=4).value = f"=B{current_row}-C{current_row}"
                style_data_cell(ws, current_row, 4, is_currency=True)
                style_data_cell(ws, current_row, 5)
                current_row += 1

        expense_end = current_row - 1

        # Total expenses
        ws.cell(row=current_row, column=1, value="TOTAL EXPENSES").font = Font(name="Calibri", bold=True, size=11, color=CORAL_ACCENT)
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2).value = f"=SUMPRODUCT((LEN(B{expense_start}:B{expense_end})>0)*B{expense_start}:B{expense_end})"
        ws.cell(row=current_row, column=2).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=current_row, column=2).number_format = currency_format
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).value = f"=SUMPRODUCT((LEN(C{expense_start}:C{expense_end})>0)*C{expense_start}:C{expense_end})"
        ws.cell(row=current_row, column=3).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=current_row, column=3).number_format = currency_format
        ws.cell(row=current_row, column=3).border = thin_border

        expense_total_row = current_row

        # Net summary
        current_row += 2
        ws.cell(row=current_row, column=1, value="NET INCOME (Income - Expenses)").font = Font(name="Calibri", bold=True, size=13, color=DEEP_GREEN)
        ws.cell(row=current_row, column=2).value = f"=B{income_total_row}-B{expense_total_row}"
        ws.cell(row=current_row, column=2).font = Font(name="Calibri", bold=True, size=13, color=DEEP_GREEN)
        ws.cell(row=current_row, column=2).number_format = currency_format
        ws.cell(row=current_row, column=3).value = f"=C{income_total_row}-C{expense_total_row}"
        ws.cell(row=current_row, column=3).font = Font(name="Calibri", bold=True, size=13, color=DEEP_GREEN)
        ws.cell(row=current_row, column=3).number_format = currency_format

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 25

    # Dashboard sheet
    ws_dash = wb.create_sheet(title="Dashboard", index=1)
    ws_dash.sheet_properties.tabColor = DEEP_GREEN
    add_title(ws_dash, "2026 Annual Budget Dashboard", "Your financial year at a glance")

    row = 4
    headers_dash = ["Month", "Total Income", "Total Expenses", "Net Savings", "Savings Rate"]
    for col, h in enumerate(headers_dash, 1):
        ws_dash.cell(row=row, column=col, value=h)
    style_header_row(ws_dash, row, 5)

    for i, month in enumerate(months):
        r = row + 1 + i
        ws_dash.cell(row=r, column=1, value=month).font = body_font
        ws_dash.cell(row=r, column=1).border = thin_border
        for c in range(2, 5):
            style_data_cell(ws_dash, r, c, is_currency=True)
        style_data_cell(ws_dash, r, 5, is_percent=True)
        ws_dash.cell(row=r, column=4).value = f"=B{r}-C{r}"
        ws_dash.cell(row=r, column=5).value = f'=IF(B{r}>0,D{r}/B{r},0)'

    total_r = row + 13
    ws_dash.cell(row=total_r, column=1, value="YEARLY TOTAL").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws_dash.cell(row=total_r, column=1).border = thin_border
    for c in range(2, 5):
        ws_dash.cell(row=total_r, column=c).value = f"=SUM({get_column_letter(c)}{row+1}:{get_column_letter(c)}{total_r-1})"
        ws_dash.cell(row=total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws_dash.cell(row=total_r, column=c).number_format = currency_format
        ws_dash.cell(row=total_r, column=c).border = thin_border
    ws_dash.cell(row=total_r, column=5).value = f"=IF(B{total_r}>0,D{total_r}/B{total_r},0)"
    ws_dash.cell(row=total_r, column=5).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws_dash.cell(row=total_r, column=5).number_format = percent_format
    ws_dash.cell(row=total_r, column=5).border = thin_border

    ws_dash.column_dimensions["A"].width = 18
    for c in "BCDE":
        ws_dash.column_dimensions[c].width = 18

    path = os.path.join(OUT_DIR, "01_Monthly_Budget_Tracker_2026.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def create_annual_overview():
    """Sheet 2: Annual Budget Overview with yearly planning"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Overview"
    ws.sheet_properties.tabColor = SAGE_GREEN

    add_title(ws, "2026 Annual Budget Overview", "Plan your entire financial year")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Headers
    row = 4
    ws.cell(row=row, column=1, value="Category")
    for i, m in enumerate(months):
        ws.cell(row=row, column=i+2, value=m)
    ws.cell(row=row, column=14, value="Annual Total")
    ws.cell(row=row, column=15, value="Monthly Avg")
    style_header_row(ws, row, 15)

    # Income categories
    row = 5
    ws.cell(row=row, column=1, value="INCOME").font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    for c in range(1, 16):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=SAGE_GREEN, end_color=SAGE_GREEN, fill_type="solid")
        ws.cell(row=row, column=c).border = thin_border

    income_items = ["Primary Salary", "Side Income", "Freelance/Contract", "Investment Returns", "Other Income"]
    for i, item in enumerate(income_items):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=item).font = body_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 14):
            style_data_cell(ws, r, c, is_currency=True)
        ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
        style_data_cell(ws, r, 14, is_currency=True)
        ws.cell(row=r, column=14).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=r, column=15).value = f"=N{r}/12"
        style_data_cell(ws, r, 15, is_currency=True)

    income_total_r = row + len(income_items) + 1
    ws.cell(row=income_total_r, column=1, value="TOTAL INCOME").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws.cell(row=income_total_r, column=1).fill = mint_fill
    ws.cell(row=income_total_r, column=1).border = thin_border
    for c in range(2, 16):
        col_letter = get_column_letter(c)
        ws.cell(row=income_total_r, column=c).value = f"=SUM({col_letter}{row+1}:{col_letter}{income_total_r-1})"
        ws.cell(row=income_total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=income_total_r, column=c).number_format = currency_format
        ws.cell(row=income_total_r, column=c).fill = mint_fill
        ws.cell(row=income_total_r, column=c).border = thin_border

    # Expense categories
    exp_start = income_total_r + 2
    ws.cell(row=exp_start, column=1, value="EXPENSES").font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    for c in range(1, 16):
        ws.cell(row=exp_start, column=c).fill = PatternFill(start_color=CORAL_ACCENT, end_color=CORAL_ACCENT, fill_type="solid")
        ws.cell(row=exp_start, column=c).border = thin_border

    expense_items = [
        "Housing/Rent", "Utilities", "Transportation", "Groceries", "Dining Out",
        "Health & Medical", "Insurance", "Debt Payments", "Savings & Investments",
        "Entertainment", "Clothing", "Personal Care", "Education",
        "Subscriptions", "Gifts & Donations", "Miscellaneous"
    ]

    for i, item in enumerate(expense_items):
        r = exp_start + 1 + i
        ws.cell(row=r, column=1, value=item).font = body_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 14):
            style_data_cell(ws, r, c, is_currency=True)
        ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
        style_data_cell(ws, r, 14, is_currency=True)
        ws.cell(row=r, column=14).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=r, column=15).value = f"=N{r}/12"
        style_data_cell(ws, r, 15, is_currency=True)

    exp_total_r = exp_start + len(expense_items) + 1
    ws.cell(row=exp_total_r, column=1, value="TOTAL EXPENSES").font = Font(name="Calibri", bold=True, size=11, color=CORAL_ACCENT)
    ws.cell(row=exp_total_r, column=1).fill = PatternFill(start_color="FCE8E3", end_color="FCE8E3", fill_type="solid")
    ws.cell(row=exp_total_r, column=1).border = thin_border
    for c in range(2, 16):
        col_letter = get_column_letter(c)
        ws.cell(row=exp_total_r, column=c).value = f"=SUM({col_letter}{exp_start+1}:{col_letter}{exp_total_r-1})"
        ws.cell(row=exp_total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=CORAL_ACCENT)
        ws.cell(row=exp_total_r, column=c).number_format = currency_format
        ws.cell(row=exp_total_r, column=c).fill = PatternFill(start_color="FCE8E3", end_color="FCE8E3", fill_type="solid")
        ws.cell(row=exp_total_r, column=c).border = thin_border

    # Net row
    net_r = exp_total_r + 2
    ws.cell(row=net_r, column=1, value="NET SAVINGS").font = Font(name="Calibri", bold=True, size=13, color=DEEP_GREEN)
    ws.cell(row=net_r, column=1).border = thin_border
    for c in range(2, 16):
        col_letter = get_column_letter(c)
        ws.cell(row=net_r, column=c).value = f"={col_letter}{income_total_r}-{col_letter}{exp_total_r}"
        ws.cell(row=net_r, column=c).font = Font(name="Calibri", bold=True, size=13, color=DEEP_GREEN)
        ws.cell(row=net_r, column=c).number_format = currency_format
        ws.cell(row=net_r, column=c).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 24
    for i in range(2, 16):
        ws.column_dimensions[get_column_letter(i)].width = 14

    path = os.path.join(OUT_DIR, "02_Annual_Budget_Overview_2026.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def create_debt_payoff():
    """Sheet 3: Debt Payoff Tracker"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Debt Tracker"
    ws.sheet_properties.tabColor = CORAL_ACCENT

    add_title(ws, "Debt Payoff Tracker 2026", "Track your journey to debt freedom")

    # Debt summary section
    row = 4
    headers = ["Debt Name", "Type", "Original Balance", "Current Balance", "Interest Rate",
               "Min Payment", "Extra Payment", "Total Payment", "Payoff Date", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 10)

    debt_types = [
        ("Credit Card 1", "Credit Card"),
        ("Credit Card 2", "Credit Card"),
        ("Student Loan", "Student Loan"),
        ("Car Loan", "Auto"),
        ("Personal Loan", "Personal"),
        ("Medical Debt", "Medical"),
        ("Other Debt 1", "Other"),
        ("Other Debt 2", "Other"),
    ]

    for i, (name, dtype) in enumerate(debt_types):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=name).font = body_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=dtype).font = body_font
        ws.cell(row=r, column=2).border = thin_border
        for c in [3, 4, 6, 7]:
            style_data_cell(ws, r, c, is_currency=True)
        style_data_cell(ws, r, 5, is_percent=True)
        ws.cell(row=r, column=8).value = f"=F{r}+G{r}"
        style_data_cell(ws, r, 8, is_currency=True)
        ws.cell(row=r, column=8).font = Font(name="Calibri", bold=True, size=11)
        style_data_cell(ws, r, 9)
        style_data_cell(ws, r, 10)

    total_r = row + len(debt_types) + 1
    ws.cell(row=total_r, column=1, value="TOTALS").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws.cell(row=total_r, column=1).border = thin_border
    for c in [3, 4, 6, 7, 8]:
        col_letter = get_column_letter(c)
        ws.cell(row=total_r, column=c).value = f"=SUM({col_letter}{row+1}:{col_letter}{total_r-1})"
        ws.cell(row=total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_r, column=c).number_format = currency_format
        ws.cell(row=total_r, column=c).border = thin_border

    # Payment log section
    pay_row = total_r + 3
    ws.merge_cells(f"A{pay_row}:G{pay_row}")
    ws.cell(row=pay_row, column=1, value="PAYMENT LOG").font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    for c in range(1, 8):
        ws.cell(row=pay_row, column=c).fill = PatternFill(start_color=DEEP_GREEN, end_color=DEEP_GREEN, fill_type="solid")

    pay_row += 1
    pay_headers = ["Date", "Debt Name", "Payment Amount", "Extra Amount", "New Balance", "Interest Paid", "Notes"]
    for col, h in enumerate(pay_headers, 1):
        ws.cell(row=pay_row, column=col, value=h)
    style_header_row(ws, pay_row, 7)

    for i in range(20):
        r = pay_row + 1 + i
        for c in range(1, 8):
            style_data_cell(ws, r, c, is_currency=(c in [3, 4, 5, 6]))

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14

    # Snowball vs Avalanche sheet
    ws2 = wb.create_sheet(title="Payoff Strategy")
    ws2.sheet_properties.tabColor = GOLD_ACCENT
    add_title(ws2, "Debt Payoff Strategy Comparison", "Snowball vs Avalanche Method")

    row = 4
    ws2.cell(row=row, column=1, value="SNOWBALL METHOD (Smallest Balance First)").font = Font(name="Calibri", bold=True, size=12, color=DEEP_GREEN)
    ws2.cell(row=row, column=1).fill = mint_fill
    ws2.merge_cells(f"A{row}:F{row}")

    row = 5
    s_headers = ["Priority", "Debt Name", "Balance", "Min Payment", "Extra Payment", "Total"]
    for col, h in enumerate(s_headers, 1):
        ws2.cell(row=row, column=col, value=h)
    style_header_row(ws2, row, 6)

    for i in range(8):
        r = row + 1 + i
        ws2.cell(row=r, column=1, value=i+1).font = body_font
        ws2.cell(row=r, column=1).border = thin_border
        for c in range(2, 7):
            style_data_cell(ws2, r, c, is_currency=(c >= 3))

    aval_row = row + 11
    ws2.cell(row=aval_row, column=1, value="AVALANCHE METHOD (Highest Interest First)").font = Font(name="Calibri", bold=True, size=12, color=CORAL_ACCENT)
    ws2.merge_cells(f"A{aval_row}:F{aval_row}")

    aval_row += 1
    for col, h in enumerate(s_headers, 1):
        ws2.cell(row=aval_row, column=col, value=h)
    style_header_row(ws2, aval_row, 6)

    for i in range(8):
        r = aval_row + 1 + i
        ws2.cell(row=r, column=1, value=i+1).font = body_font
        ws2.cell(row=r, column=1).border = thin_border
        for c in range(2, 7):
            style_data_cell(ws2, r, c, is_currency=(c >= 3))

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    for c in "CDEF":
        ws2.column_dimensions[c].width = 16

    path = os.path.join(OUT_DIR, "03_Debt_Payoff_Tracker_2026.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def create_savings_tracker():
    """Sheet 4: Savings Goal Tracker"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savings Goals"
    ws.sheet_properties.tabColor = GOLD_ACCENT

    add_title(ws, "Savings Goal Tracker 2026", "Visualize your progress toward every goal")

    # Goals summary
    row = 4
    headers = ["Goal Name", "Target Amount", "Current Saved", "Remaining", "% Complete",
               "Monthly Target", "Deadline", "Priority", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 9)

    goals = [
        "Emergency Fund (3-6 months)",
        "Vacation Fund",
        "New Car Down Payment",
        "Home Down Payment",
        "Wedding Fund",
        "Education/Course",
        "Holiday Gift Fund",
        "Tech/Gadget Fund",
        "Home Renovation",
        "Investment Seed Money",
    ]

    for i, goal in enumerate(goals):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=goal).font = body_font
        ws.cell(row=r, column=1).border = thin_border
        for c in [2, 3, 6]:
            style_data_cell(ws, r, c, is_currency=True)
        ws.cell(row=r, column=4).value = f"=B{r}-C{r}"
        style_data_cell(ws, r, 4, is_currency=True)
        ws.cell(row=r, column=5).value = f"=IF(B{r}>0,C{r}/B{r},0)"
        style_data_cell(ws, r, 5, is_percent=True)
        style_data_cell(ws, r, 7)
        style_data_cell(ws, r, 8)
        style_data_cell(ws, r, 9)

    total_r = row + len(goals) + 1
    ws.cell(row=total_r, column=1, value="TOTALS").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws.cell(row=total_r, column=1).border = thin_border
    for c in [2, 3, 4, 6]:
        col_letter = get_column_letter(c)
        ws.cell(row=total_r, column=c).value = f"=SUM({col_letter}{row+1}:{col_letter}{total_r-1})"
        ws.cell(row=total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_r, column=c).number_format = currency_format
        ws.cell(row=total_r, column=c).border = thin_border

    # Monthly savings log
    log_row = total_r + 3
    ws.merge_cells(f"A{log_row}:F{log_row}")
    ws.cell(row=log_row, column=1, value="MONTHLY SAVINGS LOG").font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    for c in range(1, 7):
        ws.cell(row=log_row, column=c).fill = PatternFill(start_color=GOLD_ACCENT, end_color=GOLD_ACCENT, fill_type="solid")

    log_row += 1
    log_headers = ["Date", "Goal", "Amount Saved", "Running Total", "Source", "Notes"]
    for col, h in enumerate(log_headers, 1):
        ws.cell(row=log_row, column=col, value=h)
    style_header_row(ws, log_row, 6)

    for i in range(25):
        r = log_row + 1 + i
        for c in range(1, 7):
            style_data_cell(ws, r, c, is_currency=(c in [3, 4]))

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12

    path = os.path.join(OUT_DIR, "04_Savings_Goal_Tracker_2026.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def create_expense_tracker():
    """Sheet 5: Daily Expense Tracker"""
    wb = openpyxl.Workbook()

    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]

    # Instructions
    ws_inst = wb.active
    ws_inst.title = "How To Use"
    ws_inst.sheet_properties.tabColor = SAGE_GREEN
    add_title(ws_inst, "Daily Expense Tracker 2026", "Log every purchase and watch your habits change")

    instructions = [
        "",
        "HOW TO USE:",
        "",
        "1. Go to the tab for the current month",
        "2. Each day, log your expenses as they happen",
        "3. Enter the date, description, category, amount, and payment method",
        "4. The spreadsheet auto-calculates daily and monthly totals",
        "5. Review the Summary tab to see spending patterns",
        "",
        "CATEGORIES:",
        "Food & Drink | Transportation | Shopping | Entertainment",
        "Bills & Utilities | Health | Education | Personal",
        "Gifts | Travel | Home | Other",
        "",
        "PAYMENT METHODS:",
        "Cash | Debit Card | Credit Card | Venmo/Zelle | Other",
    ]
    for i, line in enumerate(instructions, 4):
        ws_inst.cell(row=i, column=1, value=line).font = body_font
    ws_inst.column_dimensions["A"].width = 60

    categories = ["Food & Drink", "Transportation", "Shopping", "Entertainment",
                  "Bills & Utilities", "Health", "Education", "Personal",
                  "Gifts", "Travel", "Home", "Other"]

    for month in months:
        ws = wb.create_sheet(title=month[:3])
        ws.sheet_properties.tabColor = SAGE_GREEN

        add_title(ws, f"{month} 2026 - Daily Expenses", "Log every purchase below")

        row = 4
        headers = ["Date", "Description", "Category", "Amount", "Payment Method", "Notes", "Running Total"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header_row(ws, row, 7)

        # 35 rows for daily entries
        for i in range(35):
            r = row + 1 + i
            for c in range(1, 8):
                style_data_cell(ws, r, c, is_currency=(c == 4))
            if i == 0:
                ws.cell(row=r, column=7).value = f"=D{r}"
            else:
                ws.cell(row=r, column=7).value = f"=G{r-1}+D{r}"
            ws.cell(row=r, column=7).number_format = currency_format

        # Monthly summary at bottom
        summary_row = row + 37
        ws.merge_cells(f"A{summary_row}:D{summary_row}")
        ws.cell(row=summary_row, column=1, value="MONTHLY SUMMARY BY CATEGORY").font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        for c in range(1, 5):
            ws.cell(row=summary_row, column=c).fill = header_fill

        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Category").font = subheader_font
        ws.cell(row=summary_row, column=2, value="Total Spent").font = subheader_font
        ws.cell(row=summary_row, column=3, value="# of Transactions").font = subheader_font
        ws.cell(row=summary_row, column=4, value="% of Total").font = subheader_font
        for c in range(1, 5):
            ws.cell(row=summary_row, column=c).fill = subheader_fill
            ws.cell(row=summary_row, column=c).border = thin_border

        for i, cat in enumerate(categories):
            r = summary_row + 1 + i
            ws.cell(row=r, column=1, value=cat).font = body_font
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2).value = f'=SUMIF(C{row+1}:C{row+35},A{r},D{row+1}:D{row+35})'
            style_data_cell(ws, r, 2, is_currency=True)
            ws.cell(row=r, column=3).value = f'=COUNTIF(C{row+1}:C{row+35},A{r})'
            style_data_cell(ws, r, 3)
            grand_total_r = summary_row + len(categories) + 1
            ws.cell(row=r, column=4).value = f'=IF(B{grand_total_r}>0,B{r}/B{grand_total_r},0)'
            style_data_cell(ws, r, 4, is_percent=True)

        ws.cell(row=grand_total_r, column=1, value="GRAND TOTAL").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=grand_total_r, column=1).border = thin_border
        ws.cell(row=grand_total_r, column=2).value = f"=SUM(B{summary_row+1}:B{grand_total_r-1})"
        ws.cell(row=grand_total_r, column=2).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=grand_total_r, column=2).number_format = currency_format
        ws.cell(row=grand_total_r, column=2).border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 16

    path = os.path.join(OUT_DIR, "05_Daily_Expense_Tracker_2026.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def create_bill_calendar():
    """Sheet 6: Bill Payment Calendar & Subscription Tracker"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Calendar"
    ws.sheet_properties.tabColor = DEEP_GREEN

    add_title(ws, "Bill Payment Calendar 2026", "Never miss a payment again")

    # Bills overview
    row = 4
    headers = ["Bill Name", "Amount", "Due Date", "Frequency", "Auto-Pay?",
               "Account", "Category", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Annual Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 20)

    bills = [
        "Rent/Mortgage", "Electric", "Gas/Heat", "Water/Sewer",
        "Internet", "Phone", "Car Insurance", "Health Insurance",
        "Netflix", "Spotify", "Gym Membership", "Cloud Storage",
        "Car Payment", "Student Loan", "Credit Card Min",
        "Trash/Recycling", "HOA", "Pet Insurance", "Life Insurance",
        "Other Subscription"
    ]

    for i, bill in enumerate(bills):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=bill).font = body_font
        ws.cell(row=r, column=1).border = thin_border
        style_data_cell(ws, r, 2, is_currency=True)
        for c in range(3, 8):
            style_data_cell(ws, r, c)
        for c in range(8, 20):
            style_data_cell(ws, r, c, is_currency=True)
        ws.cell(row=r, column=20).value = f"=SUM(H{r}:S{r})"
        style_data_cell(ws, r, 20, is_currency=True)
        ws.cell(row=r, column=20).font = Font(name="Calibri", bold=True, size=11)

    total_r = row + len(bills) + 1
    ws.cell(row=total_r, column=1, value="MONTHLY TOTALS").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws.cell(row=total_r, column=1).border = thin_border
    for c in range(8, 21):
        col_letter = get_column_letter(c)
        ws.cell(row=total_r, column=c).value = f"=SUM({col_letter}{row+1}:{col_letter}{total_r-1})"
        ws.cell(row=total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws.cell(row=total_r, column=c).number_format = currency_format
        ws.cell(row=total_r, column=c).border = thin_border

    # Subscription audit sheet
    ws2 = wb.create_sheet(title="Subscription Audit")
    ws2.sheet_properties.tabColor = CORAL_ACCENT
    add_title(ws2, "Subscription Audit 2026", "Review and optimize your recurring charges")

    row = 4
    audit_headers = ["Service", "Monthly Cost", "Annual Cost", "Category",
                     "Last Used", "Usage Level", "Keep?", "Notes"]
    for col, h in enumerate(audit_headers, 1):
        ws2.cell(row=row, column=col, value=h)
    style_header_row(ws2, row, 8)

    subs = [
        "Netflix", "Spotify", "Amazon Prime", "Disney+", "Hulu",
        "YouTube Premium", "Apple iCloud", "Google One", "Adobe CC",
        "Gym/Fitness", "Meal Kit", "News Subscription", "VPN",
        "Password Manager", "Dropbox", "Microsoft 365",
        "Other 1", "Other 2", "Other 3", "Other 4"
    ]

    for i, sub in enumerate(subs):
        r = row + 1 + i
        ws2.cell(row=r, column=1, value=sub).font = body_font
        ws2.cell(row=r, column=1).border = thin_border
        style_data_cell(ws2, r, 2, is_currency=True)
        ws2.cell(row=r, column=3).value = f"=B{r}*12"
        style_data_cell(ws2, r, 3, is_currency=True)
        for c in range(4, 9):
            style_data_cell(ws2, r, c)

    total_r = row + len(subs) + 1
    ws2.cell(row=total_r, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
    ws2.cell(row=total_r, column=1).border = thin_border
    for c in [2, 3]:
        col_letter = get_column_letter(c)
        ws2.cell(row=total_r, column=c).value = f"=SUM({col_letter}{row+1}:{col_letter}{total_r-1})"
        ws2.cell(row=total_r, column=c).font = Font(name="Calibri", bold=True, size=11, color=DEEP_GREEN)
        ws2.cell(row=total_r, column=c).number_format = currency_format
        ws2.cell(row=total_r, column=c).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for c_letter in "CDEFG":
        ws.column_dimensions[c_letter].width = 14
    for i in range(8, 21):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    for c_letter in "DEFGH":
        ws2.column_dimensions[c_letter].width = 16

    path = os.path.join(OUT_DIR, "06_Bill_Payment_Calendar_2026.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    print("Creating 2026 Budget & Finance Spreadsheet Bundle...")
    print("=" * 50)
    create_monthly_budget()
    create_annual_overview()
    create_debt_payoff()
    create_savings_tracker()
    create_expense_tracker()
    create_bill_calendar()
    print("=" * 50)
    print("All 6 spreadsheets created successfully!")
