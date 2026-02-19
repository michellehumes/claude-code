#!/usr/bin/env python3
"""
Small Business Planner & Finance Tracker Bundle
Generates 8 professionally formatted Excel spreadsheets for entrepreneurs.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# Color palette
ROYAL_BLUE = "2C5F8A"
STEEL_BLUE = "4682B4"
LIGHT_BLUE = "D6E8F5"
WHITE = "FFFFFF"
CHARCOAL = "333333"
AMBER = "E8943A"
RED_ACCENT = "D44638"

# Reusable styles
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color=WHITE)
HEADER_FILL = PatternFill(start_color=ROYAL_BLUE, end_color=ROYAL_BLUE, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=ROYAL_BLUE)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=13, color=STEEL_BLUE)
NORMAL_FONT = Font(name="Calibri", size=11, color=CHARCOAL)
BOLD_FONT = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
AMBER_FONT = Font(name="Calibri", bold=True, size=12, color=AMBER)
RED_FONT = Font(name="Calibri", bold=True, size=11, color=RED_ACCENT)
ALT_ROW_FILL = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
AMBER_FILL = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
STEEL_FILL = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def style_header_row(ws, row, max_col, font=None, fill=None):
    """Apply header styling to a row."""
    f = font or HEADER_FONT
    fl = fill or HEADER_FILL
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = f
        cell.fill = fl
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, max_col):
    """Apply alternating row colors and borders."""
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL
            else:
                cell.fill = WHITE_FILL


def set_col_widths(ws, widths):
    """Set column widths from a dict or list."""
    if isinstance(widths, dict):
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
    elif isinstance(widths, list):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def add_instructions_sheet(wb, title, instructions_list):
    """Add an Instructions/How-to-Use tab."""
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = AMBER

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"How to Use: {title}"
    cell.font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A3:H3")
    ws["A3"].value = "Getting Started"
    ws["A3"].font = SUBTITLE_FONT

    row = 5
    for i, instruction in enumerate(instructions_list, 1):
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = f"{i}. {instruction}"
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"].value = "Tips & Notes"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    tips = [
        "Use the dropdown menus for consistent data entry.",
        "Do not delete or move formula cells - they auto-calculate totals.",
        "Back up your file regularly.",
        "This spreadsheet works in Excel and Google Sheets.",
        "Customize categories and columns to fit your specific business needs.",
    ]
    for tip in tips:
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1).value = f"  - {tip}"
        ws.cell(row=row, column=1).font = NORMAL_FONT
        row += 1

    ws.column_dimensions["A"].width = 80
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 12


def create_income_tracker():
    """1. Business Income Tracker 2026"""
    wb = Workbook()
    wb.remove(wb.active)

    # Instructions
    add_instructions_sheet(wb, "Business Income Tracker 2026", [
        "Each monthly tab tracks income received from clients/customers.",
        "Enter the date, client name, product/service, invoice number, amount, payment method, and status.",
        "Use the Status dropdown to mark payments as Paid, Pending, or Overdue.",
        "Monthly totals auto-calculate at the bottom of each tab.",
        "The Dashboard tab shows your annual income summary and YTD running totals.",
        "Filter or sort columns to quickly find specific entries.",
        "Update the Status column regularly to keep cash flow visibility.",
    ])

    headers = ["Date", "Client/Customer", "Product/Service", "Invoice #",
               "Amount", "Payment Method", "Status"]
    col_widths = [14, 25, 25, 14, 15, 18, 14]
    payment_methods = '"Cash,Check,Bank Transfer,Credit Card,PayPal,Venmo,Zelle,Other"'
    statuses = '"Paid,Pending,Overdue"'

    for m_idx, month in enumerate(MONTHS):
        ws = wb.create_sheet(month)
        ws.sheet_properties.tabColor = ROYAL_BLUE

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"].value = f"Income Tracker - {month} 2026"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Headers
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        style_header_row(ws, 3, 7)
        set_col_widths(ws, col_widths)
        ws.freeze_panes = "A4"

        # Data rows (35 rows)
        num_rows = 35
        style_data_rows(ws, 4, 4 + num_rows - 1, 7)

        # Date format for column A
        for r in range(4, 4 + num_rows):
            ws.cell(row=r, column=1).number_format = "MM/DD/YYYY"
            ws.cell(row=r, column=5).number_format = '$#,##0.00'

        # Data validation
        dv_payment = DataValidation(type="list", formula1=payment_methods, allow_blank=True)
        dv_payment.error = "Please select a valid payment method"
        dv_payment.errorTitle = "Invalid Entry"
        ws.add_data_validation(dv_payment)
        dv_payment.add(f"F4:F{3 + num_rows}")

        dv_status = DataValidation(type="list", formula1=statuses, allow_blank=True)
        dv_status.error = "Please select Paid, Pending, or Overdue"
        ws.add_data_validation(dv_status)
        dv_status.add(f"G4:G{3 + num_rows}")

        # Monthly total
        total_row = 4 + num_rows + 1
        ws.cell(row=total_row, column=4, value="Monthly Total:").font = BOLD_FONT
        ws.cell(row=total_row, column=5).value = f"=SUM(E4:E{3 + num_rows})"
        ws.cell(row=total_row, column=5).font = BOLD_FONT
        ws.cell(row=total_row, column=5).number_format = '$#,##0.00'

        # Paid total
        ws.cell(row=total_row + 1, column=4, value="Total Paid:").font = BOLD_FONT
        ws.cell(row=total_row + 1, column=5).value = f'=SUMIF(G4:G{3 + num_rows},"Paid",E4:E{3 + num_rows})'
        ws.cell(row=total_row + 1, column=5).font = Font(name="Calibri", bold=True, size=11, color="228B22")
        ws.cell(row=total_row + 1, column=5).number_format = '$#,##0.00'

        # Pending total
        ws.cell(row=total_row + 2, column=4, value="Total Pending:").font = BOLD_FONT
        ws.cell(row=total_row + 2, column=5).value = f'=SUMIF(G4:G{3 + num_rows},"Pending",E4:E{3 + num_rows})'
        ws.cell(row=total_row + 2, column=5).font = Font(name="Calibri", bold=True, size=11, color=AMBER)
        ws.cell(row=total_row + 2, column=5).number_format = '$#,##0.00'

        # Overdue total
        ws.cell(row=total_row + 3, column=4, value="Total Overdue:").font = BOLD_FONT
        ws.cell(row=total_row + 3, column=5).value = f'=SUMIF(G4:G{3 + num_rows},"Overdue",E4:E{3 + num_rows})'
        ws.cell(row=total_row + 3, column=5).font = RED_FONT
        ws.cell(row=total_row + 3, column=5).number_format = '$#,##0.00'

    # Dashboard
    ws = wb.create_sheet("Dashboard", 1)
    ws.sheet_properties.tabColor = AMBER

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Annual Income Dashboard - 2026"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    dash_headers = ["Month", "Total Income", "Paid", "Pending", "Overdue", "YTD Running Total"]
    for i, h in enumerate(dash_headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 6)
    set_col_widths(ws, [16, 18, 18, 18, 18, 20])
    ws.freeze_panes = "A4"

    for m_idx, month in enumerate(MONTHS):
        row = 4 + m_idx
        ws.cell(row=row, column=1, value=month).font = BOLD_FONT
        total_row_ref = 4 + 35 + 1
        ws.cell(row=row, column=2).value = f"='{month}'!E{total_row_ref}"
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).value = f"='{month}'!E{total_row_ref + 1}"
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).value = f"='{month}'!E{total_row_ref + 2}"
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).value = f"='{month}'!E{total_row_ref + 3}"
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        if m_idx == 0:
            ws.cell(row=row, column=6).value = f"=B{row}"
        else:
            ws.cell(row=row, column=6).value = f"=F{row - 1}+B{row}"
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=6).font = BOLD_FONT

    style_data_rows(ws, 4, 15, 6)

    # Annual totals
    ws.cell(row=17, column=1, value="ANNUAL TOTAL").font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    ws.cell(row=17, column=1).fill = PatternFill(start_color=ROYAL_BLUE, end_color=ROYAL_BLUE, fill_type="solid")
    for c in range(2, 7):
        ws.cell(row=17, column=c).fill = PatternFill(start_color=ROYAL_BLUE, end_color=ROYAL_BLUE, fill_type="solid")
        ws.cell(row=17, column=c).font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        ws.cell(row=17, column=c).number_format = '$#,##0.00'
    ws.cell(row=17, column=2).value = "=SUM(B4:B15)"
    ws.cell(row=17, column=3).value = "=SUM(C4:C15)"
    ws.cell(row=17, column=4).value = "=SUM(D4:D15)"
    ws.cell(row=17, column=5).value = "=SUM(E4:E15)"
    ws.cell(row=17, column=6).value = "=F15"

    fp = os.path.join(OUTPUT_DIR, "Business_Income_Tracker_2026.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_expense_tracker():
    """2. Business Expense Tracker 2026"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Business Expense Tracker 2026", [
        "Each monthly tab tracks all business expenses.",
        "Enter date, vendor, category (use dropdown), description, amount, payment method, tax deductible status, and receipt number.",
        "Categories are pre-set: Supplies, Marketing, Software, Travel, Meals, Office, Professional Services, Insurance, Utilities, Other.",
        "Mark tax-deductible expenses with 'Y' for easy tax preparation.",
        "Monthly totals and category breakdowns auto-calculate.",
        "The Annual Summary tab aggregates data across all months.",
        "Keep receipt numbers consistent for easy cross-referencing.",
    ])

    headers = ["Date", "Vendor", "Category", "Description", "Amount",
               "Payment Method", "Tax Deductible", "Receipt #"]
    col_widths = [14, 22, 18, 28, 15, 18, 16, 14]
    categories = '"Supplies,Marketing,Software,Travel,Meals,Office,Professional Services,Insurance,Utilities,Other"'
    payment_methods = '"Cash,Check,Bank Transfer,Credit Card,PayPal,Venmo,Zelle,Other"'
    tax_ded = '"Y,N"'

    cat_list = ["Supplies", "Marketing", "Software", "Travel", "Meals",
                "Office", "Professional Services", "Insurance", "Utilities", "Other"]

    for m_idx, month in enumerate(MONTHS):
        ws = wb.create_sheet(month)
        ws.sheet_properties.tabColor = STEEL_BLUE

        ws.merge_cells("A1:H1")
        ws["A1"].value = f"Expense Tracker - {month} 2026"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        style_header_row(ws, 3, 8)
        set_col_widths(ws, col_widths)
        ws.freeze_panes = "A4"

        num_rows = 40
        style_data_rows(ws, 4, 4 + num_rows - 1, 8)

        for r in range(4, 4 + num_rows):
            ws.cell(row=r, column=1).number_format = "MM/DD/YYYY"
            ws.cell(row=r, column=5).number_format = '$#,##0.00'

        dv_cat = DataValidation(type="list", formula1=categories, allow_blank=True)
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"C4:C{3 + num_rows}")

        dv_pm = DataValidation(type="list", formula1=payment_methods, allow_blank=True)
        ws.add_data_validation(dv_pm)
        dv_pm.add(f"F4:F{3 + num_rows}")

        dv_td = DataValidation(type="list", formula1=tax_ded, allow_blank=True)
        ws.add_data_validation(dv_td)
        dv_td.add(f"G4:G{3 + num_rows}")

        total_row = 4 + num_rows + 1
        ws.cell(row=total_row, column=4, value="Monthly Total:").font = BOLD_FONT
        ws.cell(row=total_row, column=5).value = f"=SUM(E4:E{3 + num_rows})"
        ws.cell(row=total_row, column=5).font = BOLD_FONT
        ws.cell(row=total_row, column=5).number_format = '$#,##0.00'

        ws.cell(row=total_row + 1, column=4, value="Tax Deductible Total:").font = BOLD_FONT
        ws.cell(row=total_row + 1, column=5).value = f'=SUMIF(G4:G{3 + num_rows},"Y",E4:E{3 + num_rows})'
        ws.cell(row=total_row + 1, column=5).font = Font(name="Calibri", bold=True, size=11, color="228B22")
        ws.cell(row=total_row + 1, column=5).number_format = '$#,##0.00'

        # Category breakdown
        ws.cell(row=total_row + 3, column=4, value="Category Breakdown:").font = SUBTITLE_FONT
        for ci, cat in enumerate(cat_list):
            r = total_row + 4 + ci
            ws.cell(row=r, column=4, value=cat).font = NORMAL_FONT
            ws.cell(row=r, column=5).value = f'=SUMIF(C4:C{3 + num_rows},"{cat}",E4:E{3 + num_rows})'
            ws.cell(row=r, column=5).number_format = '$#,##0.00'

    # Annual Summary
    ws = wb.create_sheet("Annual Summary", 1)
    ws.sheet_properties.tabColor = AMBER

    ws.merge_cells("A1:N1")
    ws["A1"].value = "Annual Expense Summary - 2026"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    sum_headers = ["Month", "Total Expenses", "Tax Deductible"] + cat_list
    for i, h in enumerate(sum_headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(sum_headers))
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(sum_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A4"

    for m_idx, month in enumerate(MONTHS):
        row = 4 + m_idx
        total_ref = 4 + 40 + 1
        ws.cell(row=row, column=1, value=month).font = BOLD_FONT
        ws.cell(row=row, column=2).value = f"='{month}'!E{total_ref}"
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).value = f"='{month}'!E{total_ref + 1}"
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        for ci, cat in enumerate(cat_list):
            ws.cell(row=row, column=4 + ci).value = f"='{month}'!E{total_ref + 4 + ci}"
            ws.cell(row=row, column=4 + ci).number_format = '$#,##0.00'

    style_data_rows(ws, 4, 15, len(sum_headers))

    # Total row
    ws.cell(row=17, column=1, value="ANNUAL TOTAL").font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    for c in range(1, len(sum_headers) + 1):
        ws.cell(row=17, column=c).fill = PatternFill(start_color=ROYAL_BLUE, end_color=ROYAL_BLUE, fill_type="solid")
        ws.cell(row=17, column=c).font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        if c >= 2:
            ws.cell(row=17, column=c).value = f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}15)"
            ws.cell(row=17, column=c).number_format = '$#,##0.00'

    fp = os.path.join(OUTPUT_DIR, "Business_Expense_Tracker_2026.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_profit_loss():
    """3. Profit & Loss Statement 2026"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Profit & Loss Statement 2026", [
        "Enter your Revenue, Cost of Goods Sold (COGS), and Operating Expenses each month.",
        "Gross Profit, Operating Income, and Net Profit are auto-calculated.",
        "Margins and percentages update automatically based on your entries.",
        "The 12-Month Comparison view shows trends across the full year.",
        "Use this to understand your profitability month over month.",
        "Compare categories to identify areas for cost reduction.",
        "Share with your accountant for tax preparation.",
    ])

    expense_cats = [
        "Supplies & Materials", "Marketing & Advertising", "Software & Subscriptions",
        "Travel & Transportation", "Meals & Entertainment", "Office Expenses",
        "Professional Services", "Insurance", "Utilities & Rent", "Payroll & Contractors",
        "Depreciation", "Other Expenses"
    ]

    for m_idx, month in enumerate(MONTHS):
        ws = wb.create_sheet(month)
        ws.sheet_properties.tabColor = ROYAL_BLUE

        ws.merge_cells("A1:D1")
        ws["A1"].value = f"Profit & Loss Statement - {month} 2026"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        set_col_widths(ws, [35, 18, 18, 18])

        # Revenue Section
        row = 3
        ws.cell(row=row, column=1, value="REVENUE").font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=ROYAL_BLUE, end_color=ROYAL_BLUE, fill_type="solid")

        ws.cell(row=row, column=2, value="Amount").font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        ws.cell(row=row, column=3, value="% of Revenue").font = Font(name="Calibri", bold=True, size=11, color=WHITE)

        row = 4
        revenue_items = ["Product Sales", "Service Revenue", "Other Revenue"]
        for ri, item in enumerate(revenue_items):
            ws.cell(row=row + ri, column=1, value=f"  {item}").font = NORMAL_FONT
            ws.cell(row=row + ri, column=2).number_format = '$#,##0.00'
            ws.cell(row=row + ri, column=2).font = NORMAL_FONT
            ws.cell(row=row + ri, column=3).number_format = '0.0%'
            ws.cell(row=row + ri, column=3).font = NORMAL_FONT

        rev_total_row = row + len(revenue_items)
        ws.cell(row=rev_total_row, column=1, value="Total Revenue").font = BOLD_FONT
        ws.cell(row=rev_total_row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=rev_total_row, column=2).value = f"=SUM(B{row}:B{row + len(revenue_items) - 1})"
        ws.cell(row=rev_total_row, column=2).font = BOLD_FONT
        ws.cell(row=rev_total_row, column=2).fill = SUBHEADER_FILL
        ws.cell(row=rev_total_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=rev_total_row, column=3).value = 1
        ws.cell(row=rev_total_row, column=3).number_format = '0.0%'
        ws.cell(row=rev_total_row, column=3).fill = SUBHEADER_FILL

        # Percentage formulas
        for ri in range(len(revenue_items)):
            ws.cell(row=row + ri, column=3).value = f"=IF(B{rev_total_row}=0,0,B{row + ri}/B{rev_total_row})"

        # COGS
        cogs_header = rev_total_row + 2
        ws.cell(row=cogs_header, column=1, value="COST OF GOODS SOLD").font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        for c in range(1, 5):
            ws.cell(row=cogs_header, column=c).fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")

        cogs_items = ["Materials & Supplies", "Direct Labor", "Shipping & Fulfillment"]
        cogs_start = cogs_header + 1
        for ci, item in enumerate(cogs_items):
            ws.cell(row=cogs_start + ci, column=1, value=f"  {item}").font = NORMAL_FONT
            ws.cell(row=cogs_start + ci, column=2).number_format = '$#,##0.00'
            ws.cell(row=cogs_start + ci, column=3).value = f"=IF(B{rev_total_row}=0,0,B{cogs_start + ci}/B{rev_total_row})"
            ws.cell(row=cogs_start + ci, column=3).number_format = '0.0%'

        cogs_total_row = cogs_start + len(cogs_items)
        ws.cell(row=cogs_total_row, column=1, value="Total COGS").font = BOLD_FONT
        ws.cell(row=cogs_total_row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=cogs_total_row, column=2).value = f"=SUM(B{cogs_start}:B{cogs_start + len(cogs_items) - 1})"
        ws.cell(row=cogs_total_row, column=2).font = BOLD_FONT
        ws.cell(row=cogs_total_row, column=2).fill = SUBHEADER_FILL
        ws.cell(row=cogs_total_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=cogs_total_row, column=3).value = f"=IF(B{rev_total_row}=0,0,B{cogs_total_row}/B{rev_total_row})"
        ws.cell(row=cogs_total_row, column=3).number_format = '0.0%'
        ws.cell(row=cogs_total_row, column=3).fill = SUBHEADER_FILL

        # Gross Profit
        gp_row = cogs_total_row + 1
        ws.cell(row=gp_row, column=1, value="GROSS PROFIT").font = Font(name="Calibri", bold=True, size=12, color=ROYAL_BLUE)
        ws.cell(row=gp_row, column=2).value = f"=B{rev_total_row}-B{cogs_total_row}"
        ws.cell(row=gp_row, column=2).font = Font(name="Calibri", bold=True, size=12, color=ROYAL_BLUE)
        ws.cell(row=gp_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=gp_row, column=3).value = f"=IF(B{rev_total_row}=0,0,B{gp_row}/B{rev_total_row})"
        ws.cell(row=gp_row, column=3).number_format = '0.0%'
        ws.cell(row=gp_row, column=3).font = Font(name="Calibri", bold=True, size=12, color=ROYAL_BLUE)

        # Operating Expenses
        opex_header = gp_row + 2
        ws.cell(row=opex_header, column=1, value="OPERATING EXPENSES").font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        for c in range(1, 5):
            ws.cell(row=opex_header, column=c).fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")

        opex_start = opex_header + 1
        for ei, item in enumerate(expense_cats):
            ws.cell(row=opex_start + ei, column=1, value=f"  {item}").font = NORMAL_FONT
            ws.cell(row=opex_start + ei, column=2).number_format = '$#,##0.00'
            ws.cell(row=opex_start + ei, column=3).value = f"=IF(B{rev_total_row}=0,0,B{opex_start + ei}/B{rev_total_row})"
            ws.cell(row=opex_start + ei, column=3).number_format = '0.0%'

        opex_total_row = opex_start + len(expense_cats)
        ws.cell(row=opex_total_row, column=1, value="Total Operating Expenses").font = BOLD_FONT
        ws.cell(row=opex_total_row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=opex_total_row, column=2).value = f"=SUM(B{opex_start}:B{opex_start + len(expense_cats) - 1})"
        ws.cell(row=opex_total_row, column=2).font = BOLD_FONT
        ws.cell(row=opex_total_row, column=2).fill = SUBHEADER_FILL
        ws.cell(row=opex_total_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=opex_total_row, column=3).value = f"=IF(B{rev_total_row}=0,0,B{opex_total_row}/B{rev_total_row})"
        ws.cell(row=opex_total_row, column=3).number_format = '0.0%'
        ws.cell(row=opex_total_row, column=3).fill = SUBHEADER_FILL

        # Net Profit
        np_row = opex_total_row + 2
        ws.cell(row=np_row, column=1, value="NET PROFIT / (LOSS)").font = Font(name="Calibri", bold=True, size=14, color=ROYAL_BLUE)
        ws.cell(row=np_row, column=2).value = f"=B{gp_row}-B{opex_total_row}"
        ws.cell(row=np_row, column=2).font = Font(name="Calibri", bold=True, size=14, color=ROYAL_BLUE)
        ws.cell(row=np_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=np_row, column=3).value = f"=IF(B{rev_total_row}=0,0,B{np_row}/B{rev_total_row})"
        ws.cell(row=np_row, column=3).number_format = '0.0%'
        ws.cell(row=np_row, column=3).font = Font(name="Calibri", bold=True, size=14, color=ROYAL_BLUE)

        for c in range(1, 5):
            ws.cell(row=np_row, column=c).border = Border(
                top=Side(style="double", color=ROYAL_BLUE),
                bottom=Side(style="double", color=ROYAL_BLUE)
            )

    # 12-Month Comparison
    ws = wb.create_sheet("12-Month Comparison", 1)
    ws.sheet_properties.tabColor = AMBER

    ws.merge_cells("A1:N1")
    ws["A1"].value = "12-Month Profit & Loss Comparison - 2026"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    comp_headers = ["Category"] + MONTHS + ["Annual Total"]
    for i, h in enumerate(comp_headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(comp_headers))
    ws.column_dimensions["A"].width = 25
    for c in range(2, len(comp_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B4"

    # For each month, reference key rows from monthly tabs
    # Revenue row 7, COGS row 12, Gross Profit row 13, OpEx total row 28, Net Profit row 30
    key_rows = {
        "Total Revenue": 7,
        "Total COGS": 12,
        "Gross Profit": 13,
        "Total Operating Expenses": 28,
        "Net Profit / (Loss)": 30,
    }

    comp_row = 4
    for label, src_row in key_rows.items():
        ws.cell(row=comp_row, column=1, value=label).font = BOLD_FONT
        for mi, month in enumerate(MONTHS):
            ws.cell(row=comp_row, column=2 + mi).value = f"='{month}'!B{src_row}"
            ws.cell(row=comp_row, column=2 + mi).number_format = '$#,##0.00'
        # Annual total
        ws.cell(row=comp_row, column=14).value = f"=SUM(B{comp_row}:M{comp_row})"
        ws.cell(row=comp_row, column=14).number_format = '$#,##0.00'
        ws.cell(row=comp_row, column=14).font = BOLD_FONT
        comp_row += 1

    style_data_rows(ws, 4, comp_row - 1, len(comp_headers))

    fp = os.path.join(OUTPUT_DIR, "Profit_Loss_Statement_2026.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_invoice_tracker():
    """4. Invoice Tracker 2026"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Invoice Tracker 2026", [
        "Track all invoices sent to clients in one organized spreadsheet.",
        "Enter invoice details: number, date, client, description, amount, due date, status, and payment date.",
        "Days Overdue auto-calculates for unpaid invoices past their due date.",
        "Use the Status dropdown: Paid, Pending, Overdue, Cancelled.",
        "The Summary tab shows aging reports (30/60/90 day brackets).",
        "Filter by status to quickly see outstanding invoices.",
        "Update payment dates when invoices are settled.",
    ])

    # Main tracker
    ws = wb.create_sheet("Invoice Tracker")
    ws.sheet_properties.tabColor = ROYAL_BLUE

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Invoice Tracker - 2026"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    headers = ["Invoice #", "Date", "Client", "Description", "Amount",
               "Due Date", "Status", "Days Overdue", "Payment Date"]
    col_widths = [14, 14, 25, 30, 15, 14, 14, 14, 14]

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 9)
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A4"

    num_rows = 100
    style_data_rows(ws, 4, 4 + num_rows - 1, 9)

    statuses = '"Paid,Pending,Overdue,Cancelled"'
    dv_status = DataValidation(type="list", formula1=statuses, allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"G4:G{3 + num_rows}")

    for r in range(4, 4 + num_rows):
        ws.cell(row=r, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        ws.cell(row=r, column=6).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=9).number_format = "MM/DD/YYYY"
        # Days overdue formula
        ws.cell(row=r, column=8).value = f'=IF(AND(G{r}="Overdue",F{r}<>""),MAX(0,TODAY()-F{r}),IF(AND(G{r}="Pending",F{r}<>""),MAX(0,TODAY()-F{r}),0))'
        ws.cell(row=r, column=8).number_format = '0'

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = AMBER

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Invoice Summary & Aging Report"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 50

    set_col_widths(ws2, [30, 20, 20, 20])

    summary_items = [
        ("Total Invoiced", f"=SUM('Invoice Tracker'!E4:E{3 + num_rows})"),
        ("Total Paid", f'=SUMIF(\'Invoice Tracker\'!G4:G{3 + num_rows},"Paid",\'Invoice Tracker\'!E4:E{3 + num_rows})'),
        ("Total Pending", f'=SUMIF(\'Invoice Tracker\'!G4:G{3 + num_rows},"Pending",\'Invoice Tracker\'!E4:E{3 + num_rows})'),
        ("Total Overdue", f'=SUMIF(\'Invoice Tracker\'!G4:G{3 + num_rows},"Overdue",\'Invoice Tracker\'!E4:E{3 + num_rows})'),
        ("Total Cancelled", f'=SUMIF(\'Invoice Tracker\'!G4:G{3 + num_rows},"Cancelled",\'Invoice Tracker\'!E4:E{3 + num_rows})'),
    ]

    ws2.cell(row=3, column=1, value="Overview").font = SUBTITLE_FONT
    for i, (label, formula) in enumerate(summary_items):
        row = 4 + i
        ws2.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws2.cell(row=row, column=2).value = formula
        ws2.cell(row=row, column=2).number_format = '$#,##0.00'
        ws2.cell(row=row, column=2).font = BOLD_FONT

    # Aging report
    ws2.cell(row=11, column=1, value="Aging Report").font = SUBTITLE_FONT
    aging_labels = ["Current (0-30 days)", "31-60 days", "61-90 days", "Over 90 days"]
    ws2.cell(row=12, column=1, value="Category").font = HEADER_FONT
    ws2.cell(row=12, column=1).fill = HEADER_FILL
    ws2.cell(row=12, column=2, value="Amount").font = HEADER_FONT
    ws2.cell(row=12, column=2).fill = HEADER_FILL
    ws2.cell(row=12, column=3, value="Count").font = HEADER_FONT
    ws2.cell(row=12, column=3).fill = HEADER_FILL

    # Aging formulas using SUMPRODUCT
    inv_range = f"'Invoice Tracker'!E4:E{3 + num_rows}"
    days_range = f"'Invoice Tracker'!H4:H{3 + num_rows}"
    status_range = f"'Invoice Tracker'!G4:G{3 + num_rows}"

    aging_formulas = [
        f'=SUMPRODUCT(({days_range}>=0)*({days_range}<=30)*({status_range}<>"Paid")*({status_range}<>"")*({inv_range}))',
        f'=SUMPRODUCT(({days_range}>30)*({days_range}<=60)*({status_range}<>"Paid")*({status_range}<>"")*({inv_range}))',
        f'=SUMPRODUCT(({days_range}>60)*({days_range}<=90)*({status_range}<>"Paid")*({status_range}<>"")*({inv_range}))',
        f'=SUMPRODUCT(({days_range}>90)*({status_range}<>"Paid")*({status_range}<>"")*({inv_range}))',
    ]

    count_formulas = [
        f'=SUMPRODUCT(({days_range}>=0)*({days_range}<=30)*({status_range}<>"Paid")*({status_range}<>""))',
        f'=SUMPRODUCT(({days_range}>30)*({days_range}<=60)*({status_range}<>"Paid")*({status_range}<>""))',
        f'=SUMPRODUCT(({days_range}>60)*({days_range}<=90)*({status_range}<>"Paid")*({status_range}<>""))',
        f'=SUMPRODUCT(({days_range}>90)*({status_range}<>"Paid")*({status_range}<>""))',
    ]

    for i, label in enumerate(aging_labels):
        row = 13 + i
        ws2.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws2.cell(row=row, column=2).value = aging_formulas[i]
        ws2.cell(row=row, column=2).number_format = '$#,##0.00'
        ws2.cell(row=row, column=3).value = count_formulas[i]

    style_data_rows(ws2, 13, 16, 3)

    fp = os.path.join(OUTPUT_DIR, "Invoice_Tracker_2026.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_client_crm():
    """5. Client CRM Tracker"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Client CRM Tracker", [
        "Track all client/customer information in one centralized database.",
        "Enter client name, contact details, services purchased, and key dates.",
        "Total Revenue per client auto-calculates if you enter amounts in the Revenue column.",
        "Use Status dropdown to mark clients as Active or Inactive.",
        "Add notes for follow-ups, preferences, or special requests.",
        "The Summary tab shows key statistics about your client base.",
        "Sort by Last Purchase Date to identify clients who may need re-engagement.",
    ])

    ws = wb.create_sheet("Client Database")
    ws.sheet_properties.tabColor = ROYAL_BLUE

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Client CRM Tracker"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    headers = ["Client Name", "Email", "Phone", "Service/Product",
               "First Purchase", "Last Purchase", "Total Revenue",
               "Status", "Priority", "Notes"]
    col_widths = [22, 28, 16, 22, 16, 16, 16, 14, 12, 35]

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 10)
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A4"

    num_rows = 100
    style_data_rows(ws, 4, 4 + num_rows - 1, 10)

    statuses = '"Active,Inactive,Lead,Prospect"'
    dv_status = DataValidation(type="list", formula1=statuses, allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"H4:H{3 + num_rows}")

    priorities = '"High,Medium,Low"'
    dv_pri = DataValidation(type="list", formula1=priorities, allow_blank=True)
    ws.add_data_validation(dv_pri)
    dv_pri.add(f"I4:I{3 + num_rows}")

    for r in range(4, 4 + num_rows):
        ws.cell(row=r, column=5).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=6).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=7).number_format = '$#,##0.00'

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = AMBER

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Client Summary Statistics"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 50
    set_col_widths(ws2, [30, 20, 20, 20])

    stats = [
        ("Total Clients", f"=COUNTA('Client Database'!A4:A{3 + num_rows})"),
        ("Active Clients", f'=COUNTIF(\'Client Database\'!H4:H{3 + num_rows},"Active")'),
        ("Inactive Clients", f'=COUNTIF(\'Client Database\'!H4:H{3 + num_rows},"Inactive")'),
        ("Leads", f'=COUNTIF(\'Client Database\'!H4:H{3 + num_rows},"Lead")'),
        ("Prospects", f'=COUNTIF(\'Client Database\'!H4:H{3 + num_rows},"Prospect")'),
        ("Total Revenue (All Clients)", f"=SUM('Client Database'!G4:G{3 + num_rows})"),
        ("Average Revenue per Client", f"=IF(COUNTA('Client Database'!A4:A{3 + num_rows})=0,0,SUM('Client Database'!G4:G{3 + num_rows})/COUNTA('Client Database'!A4:A{3 + num_rows}))"),
        ("High Priority Clients", f'=COUNTIF(\'Client Database\'!I4:I{3 + num_rows},"High")'),
    ]

    ws2.cell(row=3, column=1, value="Metric").font = HEADER_FONT
    ws2.cell(row=3, column=1).fill = HEADER_FILL
    ws2.cell(row=3, column=2, value="Value").font = HEADER_FONT
    ws2.cell(row=3, column=2).fill = HEADER_FILL

    for i, (label, formula) in enumerate(stats):
        row = 4 + i
        ws2.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws2.cell(row=row, column=2).value = formula
        if "Revenue" in label or "Average" in label:
            ws2.cell(row=row, column=2).number_format = '$#,##0.00'
        ws2.cell(row=row, column=2).font = BOLD_FONT

    style_data_rows(ws2, 4, 4 + len(stats) - 1, 2)

    fp = os.path.join(OUTPUT_DIR, "Client_CRM_Tracker.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_social_media_calendar():
    """6. Social Media Content Calendar 2026"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Social Media Content Calendar 2026", [
        "Plan and track your social media content across all platforms.",
        "Each monthly tab has a weekly view with space for daily content planning.",
        "Use Platform dropdown: Instagram, TikTok, Pinterest, Facebook, Twitter/X, LinkedIn.",
        "Content Types include: Photo, Video, Reel, Story, Carousel, Blog Post, Pin, Live.",
        "Track status from Draft to Scheduled to Posted.",
        "Add engagement notes after posting to track performance.",
        "The Monthly Overview tab provides a high-level content calendar view.",
    ])

    headers = ["Date", "Day", "Platform", "Content Type", "Caption/Topic",
               "Hashtags", "Status", "Engagement Notes"]
    col_widths = [14, 12, 16, 16, 35, 30, 14, 30]
    platforms = '"Instagram,TikTok,Pinterest,Facebook,Twitter/X,LinkedIn,YouTube"'
    content_types = '"Photo,Video,Reel,Story,Carousel,Blog Post,Pin,Live,Thread"'
    statuses = '"Idea,Draft,In Progress,Scheduled,Posted,Repurpose"'

    import datetime
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for m_idx, month in enumerate(MONTHS):
        ws = wb.create_sheet(month)
        ws.sheet_properties.tabColor = STEEL_BLUE

        ws.merge_cells("A1:H1")
        ws["A1"].value = f"Content Calendar - {month} 2026"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        style_header_row(ws, 3, 8)
        set_col_widths(ws, col_widths)
        ws.freeze_panes = "A4"

        # Calculate days in month
        if m_idx == 11:
            next_month = datetime.date(2027, 1, 1)
        else:
            next_month = datetime.date(2026, m_idx + 2, 1)
        first_day = datetime.date(2026, m_idx + 1, 1)
        days_in_month = (next_month - first_day).days

        num_rows = days_in_month
        style_data_rows(ws, 4, 4 + num_rows - 1, 8)

        # Pre-fill dates and days
        for d in range(days_in_month):
            r = 4 + d
            dt = datetime.date(2026, m_idx + 1, d + 1)
            ws.cell(row=r, column=1, value=dt).number_format = "MM/DD/YYYY"
            ws.cell(row=r, column=2, value=days_of_week[dt.weekday()]).font = NORMAL_FONT

        dv_platform = DataValidation(type="list", formula1=platforms, allow_blank=True)
        ws.add_data_validation(dv_platform)
        dv_platform.add(f"C4:C{3 + num_rows}")

        dv_ct = DataValidation(type="list", formula1=content_types, allow_blank=True)
        ws.add_data_validation(dv_ct)
        dv_ct.add(f"D4:D{3 + num_rows}")

        dv_status = DataValidation(type="list", formula1=statuses, allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add(f"G4:G{3 + num_rows}")

        # Monthly summary at bottom
        sum_row = 4 + num_rows + 2
        ws.cell(row=sum_row, column=1, value="Monthly Summary").font = SUBTITLE_FONT
        ws.cell(row=sum_row + 1, column=1, value="Total Posts Planned:").font = BOLD_FONT
        ws.cell(row=sum_row + 1, column=2).value = f'=COUNTA(C4:C{3 + num_rows})'
        ws.cell(row=sum_row + 2, column=1, value="Posts Scheduled:").font = BOLD_FONT
        ws.cell(row=sum_row + 2, column=2).value = f'=COUNTIF(G4:G{3 + num_rows},"Scheduled")'
        ws.cell(row=sum_row + 3, column=1, value="Posts Published:").font = BOLD_FONT
        ws.cell(row=sum_row + 3, column=2).value = f'=COUNTIF(G4:G{3 + num_rows},"Posted")'

    fp = os.path.join(OUTPUT_DIR, "Social_Media_Content_Calendar_2026.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_inventory_tracker():
    """7. Inventory Tracker"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Inventory Tracker", [
        "Track all products, stock levels, and supplier information.",
        "Enter product details: name, SKU, category, cost price, selling price, stock levels, and supplier.",
        "Profit Margin auto-calculates based on Cost Price and Selling Price.",
        "The Status column auto-flags items as 'In Stock', 'Low Stock', or 'Out of Stock' based on current stock vs. reorder point.",
        "Set Reorder Points to get visual alerts when stock is running low.",
        "The Summary tab shows total inventory value and key metrics.",
        "Regularly update Current Stock after sales and restocking.",
    ])

    ws = wb.create_sheet("Inventory")
    ws.sheet_properties.tabColor = ROYAL_BLUE

    ws.merge_cells("A1:K1")
    ws["A1"].value = "Inventory Tracker"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    headers = ["Product Name", "SKU", "Category", "Cost Price", "Selling Price",
               "Margin %", "Current Stock", "Reorder Point", "Supplier", "Status", "Notes"]
    col_widths = [25, 14, 18, 14, 14, 12, 14, 14, 22, 14, 30]

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 11)
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A4"

    num_rows = 100
    style_data_rows(ws, 4, 4 + num_rows - 1, 11)

    categories = '"Finished Goods,Raw Materials,Packaging,Handmade,Digital,Supplies,Other"'
    dv_cat = DataValidation(type="list", formula1=categories, allow_blank=True)
    ws.add_data_validation(dv_cat)
    dv_cat.add(f"C4:C{3 + num_rows}")

    for r in range(4, 4 + num_rows):
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        # Margin formula
        ws.cell(row=r, column=6).value = f'=IF(E{r}=0,0,(E{r}-D{r})/E{r})'
        ws.cell(row=r, column=6).number_format = '0.0%'
        # Status formula (auto-calculated)
        ws.cell(row=r, column=10).value = (
            f'=IF(G{r}="","",IF(G{r}=0,"Out of Stock",IF(G{r}<=H{r},"Low Stock","In Stock")))'
        )

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = AMBER

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Inventory Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 50
    set_col_widths(ws2, [35, 20, 20, 20])

    stats = [
        ("Total Products", f"=COUNTA(Inventory!A4:A{3 + num_rows})"),
        ("Total Inventory Value (Cost)", f"=SUMPRODUCT(Inventory!D4:D{3 + num_rows},Inventory!G4:G{3 + num_rows})"),
        ("Total Inventory Value (Retail)", f"=SUMPRODUCT(Inventory!E4:E{3 + num_rows},Inventory!G4:G{3 + num_rows})"),
        ("Total Units in Stock", f"=SUM(Inventory!G4:G{3 + num_rows})"),
        ("Items In Stock", f'=COUNTIF(Inventory!J4:J{3 + num_rows},"In Stock")'),
        ("Items Low Stock", f'=COUNTIF(Inventory!J4:J{3 + num_rows},"Low Stock")'),
        ("Items Out of Stock", f'=COUNTIF(Inventory!J4:J{3 + num_rows},"Out of Stock")'),
        ("Average Margin", f"=AVERAGE(Inventory!F4:F{3 + num_rows})"),
    ]

    ws2.cell(row=3, column=1, value="Metric").font = HEADER_FONT
    ws2.cell(row=3, column=1).fill = HEADER_FILL
    ws2.cell(row=3, column=2, value="Value").font = HEADER_FONT
    ws2.cell(row=3, column=2).fill = HEADER_FILL

    for i, (label, formula) in enumerate(stats):
        row = 4 + i
        ws2.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws2.cell(row=row, column=2).value = formula
        if "Value" in label:
            ws2.cell(row=row, column=2).number_format = '$#,##0.00'
        elif "Margin" in label:
            ws2.cell(row=row, column=2).number_format = '0.0%'
        ws2.cell(row=row, column=2).font = BOLD_FONT

    style_data_rows(ws2, 4, 4 + len(stats) - 1, 2)

    fp = os.path.join(OUTPUT_DIR, "Inventory_Tracker.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def create_tax_deduction_tracker():
    """8. Tax Deduction Tracker 2026"""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, "Tax Deduction Tracker 2026", [
        "Track all tax-deductible business expenses throughout the year.",
        "Enter the date, category (use dropdown), description, amount, and whether you have a receipt.",
        "Categories include: Home Office, Vehicle, Supplies, Travel, Meals, Marketing, Professional Development, Insurance, Software/Subscriptions.",
        "Quarterly subtotals and annual totals auto-calculate.",
        "The estimated tax savings row helps you plan (adjust the tax rate on the Summary tab).",
        "Always keep receipts for deductions over $75 (IRS requirement).",
        "Review quarterly to ensure you are capturing all eligible deductions.",
    ])

    categories = '"Home Office,Vehicle,Supplies,Travel,Meals (50%),Marketing,Professional Development,Insurance,Software/Subscriptions,Bank/Payment Fees,Other"'
    cat_list = ["Home Office", "Vehicle", "Supplies", "Travel", "Meals (50%)",
                "Marketing", "Professional Development", "Insurance",
                "Software/Subscriptions", "Bank/Payment Fees", "Other"]

    # Quarterly tabs
    quarters = [
        ("Q1 (Jan-Mar)", [0, 1, 2]),
        ("Q2 (Apr-Jun)", [3, 4, 5]),
        ("Q3 (Jul-Sep)", [6, 7, 8]),
        ("Q4 (Oct-Dec)", [9, 10, 11]),
    ]

    headers = ["Date", "Category", "Description", "Amount", "Receipt (Y/N)", "Notes"]
    col_widths = [14, 24, 35, 15, 14, 30]

    for q_idx, (q_name, month_indices) in enumerate(quarters):
        ws = wb.create_sheet(q_name)
        ws.sheet_properties.tabColor = ROYAL_BLUE

        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Tax Deductions - {q_name} 2026"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        style_header_row(ws, 3, 6)
        set_col_widths(ws, col_widths)
        ws.freeze_panes = "A4"

        num_rows = 60
        style_data_rows(ws, 4, 4 + num_rows - 1, 6)

        dv_cat = DataValidation(type="list", formula1=categories, allow_blank=True)
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"B4:B{3 + num_rows}")

        dv_receipt = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
        ws.add_data_validation(dv_receipt)
        dv_receipt.add(f"E4:E{3 + num_rows}")

        for r in range(4, 4 + num_rows):
            ws.cell(row=r, column=1).number_format = "MM/DD/YYYY"
            ws.cell(row=r, column=4).number_format = '$#,##0.00'

        # Quarterly total
        total_row = 4 + num_rows + 1
        ws.cell(row=total_row, column=3, value="Quarterly Total:").font = BOLD_FONT
        ws.cell(row=total_row, column=4).value = f"=SUM(D4:D{3 + num_rows})"
        ws.cell(row=total_row, column=4).font = BOLD_FONT
        ws.cell(row=total_row, column=4).number_format = '$#,##0.00'

        ws.cell(row=total_row + 1, column=3, value="Entries with Receipts:").font = BOLD_FONT
        ws.cell(row=total_row + 1, column=4).value = f'=COUNTIF(E4:E{3 + num_rows},"Y")'

        ws.cell(row=total_row + 2, column=3, value="Entries Missing Receipts:").font = BOLD_FONT
        ws.cell(row=total_row + 2, column=4).value = f'=COUNTIF(E4:E{3 + num_rows},"N")'
        ws.cell(row=total_row + 2, column=4).font = RED_FONT

        # Category breakdown
        ws.cell(row=total_row + 4, column=3, value="Category Breakdown:").font = SUBTITLE_FONT
        for ci, cat in enumerate(cat_list):
            r = total_row + 5 + ci
            ws.cell(row=r, column=3, value=cat).font = NORMAL_FONT
            ws.cell(row=r, column=4).value = f'=SUMIF(B4:B{3 + num_rows},"{cat}",D4:D{3 + num_rows})'
            ws.cell(row=r, column=4).number_format = '$#,##0.00'

    # Annual Summary
    ws = wb.create_sheet("Annual Summary", 1)
    ws.sheet_properties.tabColor = AMBER

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Annual Tax Deduction Summary - 2026"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ROYAL_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    set_col_widths(ws, [28, 18, 18, 18, 18])

    # Quarter totals
    ws.cell(row=3, column=1, value="Quarter").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=2, value="Total Deductions").font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL
    ws.cell(row=3, column=3, value="Entries").font = HEADER_FONT
    ws.cell(row=3, column=3).fill = HEADER_FILL

    q_names = ["Q1 (Jan-Mar)", "Q2 (Apr-Jun)", "Q3 (Jul-Sep)", "Q4 (Oct-Dec)"]
    total_ref = 4 + 60 + 1
    for qi, q_name in enumerate(q_names):
        row = 4 + qi
        ws.cell(row=row, column=1, value=q_name).font = BOLD_FONT
        ws.cell(row=row, column=2).value = f"='{q_name}'!D{total_ref}"
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).value = f"=COUNTA('{q_name}'!A4:A63)"

    style_data_rows(ws, 4, 7, 3)

    # Annual total
    ws.cell(row=9, column=1, value="ANNUAL TOTAL").font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    for c in range(1, 4):
        ws.cell(row=9, column=c).fill = PatternFill(start_color=ROYAL_BLUE, end_color=ROYAL_BLUE, fill_type="solid")
        ws.cell(row=9, column=c).font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws.cell(row=9, column=2).value = "=SUM(B4:B7)"
    ws.cell(row=9, column=2).number_format = '$#,##0.00'
    ws.cell(row=9, column=3).value = "=SUM(C4:C7)"

    # Tax savings estimate
    ws.cell(row=11, column=1, value="Estimated Tax Rate:").font = BOLD_FONT
    ws.cell(row=11, column=2, value=0.25).number_format = '0%'
    ws.cell(row=11, column=2).font = AMBER_FONT

    ws.cell(row=12, column=1, value="Estimated Tax Savings:").font = Font(name="Calibri", bold=True, size=14, color="228B22")
    ws.cell(row=12, column=2).value = "=B9*B11"
    ws.cell(row=12, column=2).number_format = '$#,##0.00'
    ws.cell(row=12, column=2).font = Font(name="Calibri", bold=True, size=14, color="228B22")

    # Category summary across all quarters
    ws.cell(row=14, column=1, value="Category Breakdown (Annual)").font = SUBTITLE_FONT
    ws.cell(row=15, column=1, value="Category").font = HEADER_FONT
    ws.cell(row=15, column=1).fill = HEADER_FILL
    ws.cell(row=15, column=2, value="Total").font = HEADER_FONT
    ws.cell(row=15, column=2).fill = HEADER_FILL

    cat_row_offset = 5  # offset from total_ref+4
    for ci, cat in enumerate(cat_list):
        row = 16 + ci
        ws.cell(row=row, column=1, value=cat).font = NORMAL_FONT
        # Sum from all 4 quarter tabs
        parts = []
        for q_name in q_names:
            parts.append(f"'{q_name}'!D{total_ref + 5 + ci}")
        ws.cell(row=row, column=2).value = f"={'+'.join(parts)}"
        ws.cell(row=row, column=2).number_format = '$#,##0.00'

    style_data_rows(ws, 16, 16 + len(cat_list) - 1, 2)

    fp = os.path.join(OUTPUT_DIR, "Tax_Deduction_Tracker_2026.xlsx")
    wb.save(fp)
    print(f"  Created: {fp}")


def main():
    print("=" * 60)
    print("Small Business Planner & Finance Tracker Bundle")
    print("Generating 8 Excel Spreadsheets...")
    print("=" * 60)

    print("\n1/8 Business Income Tracker...")
    create_income_tracker()

    print("2/8 Business Expense Tracker...")
    create_expense_tracker()

    print("3/8 Profit & Loss Statement...")
    create_profit_loss()

    print("4/8 Invoice Tracker...")
    create_invoice_tracker()

    print("5/8 Client CRM Tracker...")
    create_client_crm()

    print("6/8 Social Media Content Calendar...")
    create_social_media_calendar()

    print("7/8 Inventory Tracker...")
    create_inventory_tracker()

    print("8/8 Tax Deduction Tracker...")
    create_tax_deduction_tracker()

    print("\n" + "=" * 60)
    print("All 8 spreadsheets created successfully!")
    print("=" * 60)


if __name__ == "__main__":
    main()
