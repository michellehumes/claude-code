"""
Life OS v2 Spreadsheet Builder

Main class for creating and populating the 54-tab Google Sheets system.
"""

import os
import sys
from typing import Optional, Dict, List, Any
from tqdm import tqdm

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.settings import TABS, COLORS, NUMBER_FORMATS

try:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False
    print("Warning: Google API libraries not installed. Run: pip install google-auth google-api-python-client")


class SpreadsheetBuilder:
    """Build and populate the Life OS v2 spreadsheet"""

    def __init__(self, credentials_path: Optional[str] = None):
        self.credentials_path = credentials_path or "config/credentials.json"
        self.service = None
        self.spreadsheet_id = None
        self.sheet_ids = {}

    def authenticate(self) -> bool:
        """Authenticate with Google Sheets API"""
        if not GOOGLE_API_AVAILABLE:
            print("Google API libraries not available")
            return False

        try:
            scopes = [
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"
            ]

            if os.path.exists(self.credentials_path):
                credentials = Credentials.from_service_account_file(
                    self.credentials_path, scopes=scopes
                )
                self.service = build('sheets', 'v4', credentials=credentials)
                print("✓ Authenticated with Google Sheets API")
                return True
            else:
                print(f"✗ Credentials file not found: {self.credentials_path}")
                return False

        except Exception as e:
            print(f"✗ Authentication failed: {e}")
            return False

    def create_spreadsheet(self, title: str = "Life OS v2") -> Optional[str]:
        """Create a new spreadsheet with all 54 tabs"""
        if not self.service:
            print("Not authenticated. Call authenticate() first.")
            return None

        try:
            # Create spreadsheet with all tabs
            spreadsheet_body = {
                "properties": {
                    "title": title,
                    "locale": "en_US",
                    "timeZone": "America/New_York"
                },
                "sheets": []
            }

            # Add all 54 tabs
            for i, tab in enumerate(TABS):
                sheet = {
                    "properties": {
                        "sheetId": i,
                        "title": tab["name"],
                        "index": i,
                        "tabColor": tab["color"],
                        "gridProperties": {
                            "frozenRowCount": 1
                        }
                    }
                }
                spreadsheet_body["sheets"].append(sheet)

            # Create the spreadsheet
            spreadsheet = self.service.spreadsheets().create(
                body=spreadsheet_body
            ).execute()

            self.spreadsheet_id = spreadsheet["spreadsheetId"]

            # Store sheet IDs for later use
            for sheet in spreadsheet["sheets"]:
                self.sheet_ids[sheet["properties"]["title"]] = sheet["properties"]["sheetId"]

            print(f"✓ Created spreadsheet: {title}")
            print(f"  ID: {self.spreadsheet_id}")
            print(f"  URL: https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}")

            return self.spreadsheet_id

        except HttpError as e:
            print(f"✗ Failed to create spreadsheet: {e}")
            return None

    def open_spreadsheet(self, spreadsheet_id: str) -> bool:
        """Open an existing spreadsheet"""
        if not self.service:
            print("Not authenticated. Call authenticate() first.")
            return False

        try:
            spreadsheet = self.service.spreadsheets().get(
                spreadsheetId=spreadsheet_id
            ).execute()

            self.spreadsheet_id = spreadsheet_id

            # Store sheet IDs
            for sheet in spreadsheet["sheets"]:
                self.sheet_ids[sheet["properties"]["title"]] = sheet["properties"]["sheetId"]

            print(f"✓ Opened spreadsheet: {spreadsheet['properties']['title']}")
            return True

        except HttpError as e:
            print(f"✗ Failed to open spreadsheet: {e}")
            return False

    def populate_tab(self, tab_name: str, headers: List[str], data: List[List[Any]],
                     start_row: int = 1) -> bool:
        """Populate a tab with data"""
        if not self.service or not self.spreadsheet_id:
            print("Not ready. Create or open a spreadsheet first.")
            return False

        try:
            # Prepare data with headers
            all_data = [headers] + data

            # Update the sheet
            range_name = f"'{tab_name}'!A{start_row}"

            self.service.spreadsheets().values().update(
                spreadsheetId=self.spreadsheet_id,
                range=range_name,
                valueInputOption="USER_ENTERED",
                body={"values": all_data}
            ).execute()

            return True

        except HttpError as e:
            print(f"✗ Failed to populate {tab_name}: {e}")
            return False

    def apply_formatting(self, tab_name: str, formatting_config: Dict) -> bool:
        """Apply formatting to a tab"""
        if not self.service or not self.spreadsheet_id:
            return False

        if tab_name not in self.sheet_ids:
            return False

        sheet_id = self.sheet_ids[tab_name]
        requests = []

        try:
            # Header formatting
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": 1
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": COLORS["header_bg"],
                            "textFormat": {
                                "foregroundColor": COLORS["header_text"],
                                "bold": True,
                                "fontSize": 11
                            },
                            "horizontalAlignment": "CENTER"
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                }
            })

            # Freeze header row
            requests.append({
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": sheet_id,
                        "gridProperties": {
                            "frozenRowCount": 1
                        }
                    },
                    "fields": "gridProperties.frozenRowCount"
                }
            })

            # Apply currency formatting if specified
            if "currency_cols" in formatting_config:
                for col in formatting_config["currency_cols"]:
                    requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": 1,
                                "startColumnIndex": col,
                                "endColumnIndex": col + 1
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "numberFormat": {
                                        "type": "CURRENCY",
                                        "pattern": NUMBER_FORMATS["currency"]
                                    }
                                }
                            },
                            "fields": "userEnteredFormat.numberFormat"
                        }
                    })

            # Apply percentage formatting if specified
            if "percentage_col" in formatting_config:
                col = formatting_config["percentage_col"]
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "startColumnIndex": col,
                            "endColumnIndex": col + 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {
                                    "type": "PERCENT",
                                    "pattern": NUMBER_FORMATS["percentage"]
                                }
                            }
                        },
                        "fields": "userEnteredFormat.numberFormat"
                    }
                })

            # Apply date formatting if specified
            if "date_col" in formatting_config:
                col = formatting_config["date_col"]
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "startColumnIndex": col,
                            "endColumnIndex": col + 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {
                                    "type": "DATE",
                                    "pattern": NUMBER_FORMATS["date"]
                                }
                            }
                        },
                        "fields": "userEnteredFormat.numberFormat"
                    }
                })

            # Execute all formatting requests
            if requests:
                self.service.spreadsheets().batchUpdate(
                    spreadsheetId=self.spreadsheet_id,
                    body={"requests": requests}
                ).execute()

            return True

        except HttpError as e:
            print(f"✗ Failed to format {tab_name}: {e}")
            return False

    def set_column_widths(self, tab_name: str, widths: List[int]) -> bool:
        """Set column widths for a tab"""
        if not self.service or not self.spreadsheet_id:
            return False

        if tab_name not in self.sheet_ids:
            return False

        sheet_id = self.sheet_ids[tab_name]
        requests = []

        try:
            for i, width in enumerate(widths):
                requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": i,
                            "endIndex": i + 1
                        },
                        "properties": {
                            "pixelSize": width
                        },
                        "fields": "pixelSize"
                    }
                })

            if requests:
                self.service.spreadsheets().batchUpdate(
                    spreadsheetId=self.spreadsheet_id,
                    body={"requests": requests}
                ).execute()

            return True

        except HttpError as e:
            print(f"✗ Failed to set column widths for {tab_name}: {e}")
            return False

    def add_data_validation(self, tab_name: str, validation_config: Dict) -> bool:
        """Add data validation (dropdowns) to a tab"""
        if not self.service or not self.spreadsheet_id:
            return False

        if tab_name not in self.sheet_ids:
            return False

        sheet_id = self.sheet_ids[tab_name]
        requests = []

        try:
            for field_name, config in validation_config.items():
                col = config["column"]
                values = config.get("values", [])

                if values:
                    requests.append({
                        "setDataValidation": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": 1,
                                "startColumnIndex": col,
                                "endColumnIndex": col + 1
                            },
                            "rule": {
                                "condition": {
                                    "type": "ONE_OF_LIST",
                                    "values": [{"userEnteredValue": v} for v in values]
                                },
                                "showCustomUi": True,
                                "strict": True
                            }
                        }
                    })

            if requests:
                self.service.spreadsheets().batchUpdate(
                    spreadsheetId=self.spreadsheet_id,
                    body={"requests": requests}
                ).execute()

            return True

        except HttpError as e:
            print(f"✗ Failed to add validation for {tab_name}: {e}")
            return False

    def build_all_tabs(self, transaction_data=None, oura_data=None) -> Dict[str, bool]:
        """Build all 54 tabs with data and formatting"""
        from src.tabs.financial import FinancialTabBuilder
        from src.tabs.job_search import JobSearchTabBuilder
        from src.tabs.goals import GoalsTabBuilder
        from src.tabs.household import HouseholdTabBuilder
        from src.tabs.shelzys import ShelzysTabBuilder
        from src.tabs.fertility import FertilityTabBuilder
        from src.tabs.executive import ExecutiveTabBuilder
        from src.tabs.oura import OuraTabBuilder

        results = {}

        print("\nBuilding tabs...")

        # Build each tab group
        tab_builders = [
            ("Financial (1-11)", FinancialTabBuilder, {"transaction_data": transaction_data}),
            ("Job Search (12-18)", JobSearchTabBuilder, {}),
            ("Goals (19-22)", GoalsTabBuilder, {}),
            ("Household (23-25)", HouseholdTabBuilder, {}),
            ("Shelzy's (26-34)", ShelzysTabBuilder, {}),
            ("Fertility (35-40)", FertilityTabBuilder, {}),
            ("Executive (41-46)", ExecutiveTabBuilder, {}),
            ("Oura (47-54)", OuraTabBuilder, {"oura_data": oura_data}),
        ]

        for group_name, BuilderClass, kwargs in tqdm(tab_builders, desc="Tab Groups"):
            try:
                builder = BuilderClass(self.service, self.spreadsheet_id)
                group_results = builder.build_all_tabs(**kwargs) if kwargs else builder.build_all_tabs()

                # Populate each tab
                for tab_key, tab_config in group_results.items():
                    if isinstance(tab_config, dict) and 'tab_name' in tab_config:
                        tab_name = tab_config['tab_name']

                        # Populate data
                        success = self.populate_tab(
                            tab_name,
                            tab_config.get('headers', []),
                            tab_config.get('data', [])
                        )

                        if success:
                            # Apply formatting
                            if 'formatting' in tab_config:
                                self.apply_formatting(tab_name, tab_config['formatting'])

                            # Set column widths
                            if 'column_widths' in tab_config:
                                self.set_column_widths(tab_name, tab_config['column_widths'])

                            # Add validation
                            if 'validation' in tab_config:
                                self.add_data_validation(tab_name, tab_config['validation'])

                        results[tab_name] = success

                print(f"  ✓ {group_name}")

            except Exception as e:
                print(f"  ✗ {group_name}: {e}")
                results[group_name] = False

        return results

    def get_spreadsheet_url(self) -> Optional[str]:
        """Get the URL of the current spreadsheet"""
        if self.spreadsheet_id:
            return f"https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}"
        return None


class OfflineSpreadsheetBuilder:
    """Build spreadsheet data without Google API (for testing/export)"""

    def __init__(self):
        self.tabs_data = {}

    def build_all_tabs(self, transaction_data=None, oura_data=None) -> Dict[str, Any]:
        """Build all tab configurations"""
        from src.tabs.financial import FinancialTabBuilder
        from src.tabs.job_search import JobSearchTabBuilder
        from src.tabs.goals import GoalsTabBuilder
        from src.tabs.household import HouseholdTabBuilder
        from src.tabs.shelzys import ShelzysTabBuilder
        from src.tabs.fertility import FertilityTabBuilder
        from src.tabs.executive import ExecutiveTabBuilder
        from src.tabs.oura import OuraTabBuilder

        print("Building tab configurations (offline mode)...")

        # Build each group
        builders = [
            FinancialTabBuilder(None, None),
            JobSearchTabBuilder(None, None),
            GoalsTabBuilder(None, None),
            HouseholdTabBuilder(None, None),
            ShelzysTabBuilder(None, None),
            FertilityTabBuilder(None, None),
            ExecutiveTabBuilder(None, None),
            OuraTabBuilder(None, None),
        ]

        for builder in builders:
            try:
                if hasattr(builder, 'build_all_tabs'):
                    results = builder.build_all_tabs()
                    for tab_key, tab_config in results.items():
                        if isinstance(tab_config, dict) and 'tab_name' in tab_config:
                            self.tabs_data[tab_config['tab_name']] = tab_config
            except Exception as e:
                print(f"  Warning: {e}")

        print(f"✓ Built {len(self.tabs_data)} tab configurations")
        return self.tabs_data

    def export_to_csv(self, output_dir: str = "data/export") -> List[str]:
        """Export all tabs to CSV files"""
        import csv
        import os

        os.makedirs(output_dir, exist_ok=True)
        exported = []

        for tab_name, config in self.tabs_data.items():
            if 'headers' in config and 'data' in config:
                # Clean filename
                filename = tab_name.replace('/', '-').replace(' ', '_')
                for emoji in ['📋', '📊', '💳', '📈', '💰', '✅', '🎁', '🩺', '📅',
                              '💼', '🎯', '🏠', '🌸', '🔴']:
                    filename = filename.replace(emoji, '').strip()

                filepath = os.path.join(output_dir, f"{filename}.csv")

                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(config['headers'])
                    for row in config['data']:
                        writer.writerow(row)

                exported.append(filepath)

        print(f"✓ Exported {len(exported)} CSV files to {output_dir}")
        return exported

    def export_to_excel(self, output_path: str = "data/export/LifeOS_v2.xlsx") -> bool:
        """Export all tabs to a single Excel file"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            for tab_name, config in self.tabs_data.items():
                if 'headers' in config and 'data' in config:
                    # Create worksheet (limit name to 31 chars for Excel)
                    ws_name = tab_name[:31]
                    ws = wb.create_sheet(title=ws_name)

                    # Write headers
                    for col, header in enumerate(config['headers'], 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4285F4', end_color='4285F4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')

                    # Write data
                    for row_idx, row_data in enumerate(config['data'], 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    # Freeze header row
                    ws.freeze_panes = 'A2'

            wb.save(output_path)
            print(f"✓ Exported to {output_path}")
            return True

        except Exception as e:
            print(f"✗ Failed to export to Excel: {e}")
            return False
