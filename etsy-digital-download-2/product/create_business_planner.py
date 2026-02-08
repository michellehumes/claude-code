"""
Create a Small Business Planner & Starter Kit Spreadsheet Bundle
for Etsy Digital Download - Product #2
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.worksheet.datavalidation import DataValidation

# ── Color Palette (Warm, professional, modern) ──
NAVY = "1B2A4A"
STEEL_BLUE = "2E4057"
DUSTY_BLUE = "6B8CAE"
PALE_BLUE = "D6E4F0"
ICE_BLUE = "EBF2FA"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "E0E0E0"
DARK_TEXT = "212121"
CORAL = "E8806A"
WARM_GOLD = "D4A853"
SOFT_PEACH = "F5E6D3"
MINT = "7BC8A4"

def header_font(size=12, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=11, bold=False, color=DARK_TEXT):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    side = Side(style="thin", color=MED_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def currency_fmt():
    return '"$"#,##0.00'

def pct_fmt():
    return '0.0%'

def style_header_row(ws, row, cols, bg_color=NAVY):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = fill(bg_color)
        cell.alignment = center_align()
        cell.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    bg = LIGHT_GRAY if alt else WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.fill = fill(bg)
        cell.alignment = center_align()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title_block(ws, title, subtitle, merge_end_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = fill(PALE_BLUE)
    ws.row_dimensions[1].height = 45

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end_col)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color=STEEL_BLUE)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.fill = fill(PALE_BLUE)
    ws.row_dimensions[2].height = 25


# ═══════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────
#  SHEET 1: BUSINESS DASHBOARD
# ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_properties.tabColor = NAVY

add_title_block(ws1, "SMALL BUSINESS DASHBOARD", "Your business at a glance — update monthly for best results", 9)

# KPI Cards
row = 4
ws1.cell(row=row, column=1, value="KEY METRICS").font = header_font(14, True, NAVY)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
ws1.cell(row=row, column=1).fill = fill(ICE_BLUE)
ws1.cell(row=row, column=1).alignment = center_align()

kpi_labels = ["Monthly Revenue", "Monthly Expenses", "Net Profit", "Profit Margin",
              "Total Clients", "Active Projects", "Items in Inventory", "Avg Order Value", "YTD Revenue"]
row = 5
for i, label in enumerate(kpi_labels, 1):
    ws1.cell(row=row, column=i, value=label)
style_header_row(ws1, row, 9, STEEL_BLUE)

row = 6
for i in range(1, 10):
    cell = ws1.cell(row=row, column=i)
    if i <= 3 or i == 8 or i == 9:
        cell.number_format = currency_fmt()
    elif i == 4:
        cell.number_format = pct_fmt()
    cell.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    cell.fill = fill(WHITE)
    cell.border = thin_border()
    cell.alignment = center_align()
ws1.row_dimensions[6].height = 40

# Monthly Revenue Tracking
row = 8
ws1.cell(row=row, column=1, value="MONTHLY REVENUE & EXPENSES").font = header_font(14, True, NAVY)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
ws1.cell(row=row, column=1).fill = fill(ICE_BLUE)
ws1.cell(row=row, column=1).alignment = center_align()

row = 9
month_headers = ["Month", "Revenue", "Expenses", "Net Profit", "Profit Margin", "Orders", "New Clients", "Avg Order Value", "Notes"]
for i, h in enumerate(month_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 9)

months = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

for idx, month in enumerate(months):
    r = 10 + idx
    ws1.cell(row=r, column=1, value=month)
    ws1.cell(row=r, column=2).number_format = currency_fmt()
    ws1.cell(row=r, column=3).number_format = currency_fmt()
    ws1.cell(row=r, column=4, value=f"=IF(B{r}=\"\",\"\",B{r}-C{r})").number_format = currency_fmt()
    ws1.cell(row=r, column=5, value=f"=IF(B{r}=0,\"\",D{r}/B{r})").number_format = pct_fmt()
    ws1.cell(row=r, column=8, value=f"=IF(F{r}=0,\"\",B{r}/F{r})").number_format = currency_fmt()
    style_data_row(ws1, r, 9, idx % 2 == 1)

# Totals
total_r = 22
ws1.cell(row=total_r, column=1, value="TOTALS / AVG")
ws1.cell(row=total_r, column=2, value="=SUM(B10:B21)").number_format = currency_fmt()
ws1.cell(row=total_r, column=3, value="=SUM(C10:C21)").number_format = currency_fmt()
ws1.cell(row=total_r, column=4, value="=SUM(D10:D21)").number_format = currency_fmt()
ws1.cell(row=total_r, column=5, value="=IF(B22=0,\"\",D22/B22)").number_format = pct_fmt()
ws1.cell(row=total_r, column=6, value="=SUM(F10:F21)")
ws1.cell(row=total_r, column=7, value="=SUM(G10:G21)")
ws1.cell(row=total_r, column=8, value="=IF(F22=0,\"\",B22/F22)").number_format = currency_fmt()
style_header_row(ws1, total_r, 9, NAVY)

# Bar chart
bar = BarChart()
bar.type = "col"
bar.title = "Monthly Revenue vs Expenses"
bar.style = 10
bar.y_axis.title = "Amount ($)"
data = Reference(ws1, min_col=1, min_row=9, max_col=3, max_row=21)
bar.add_data(data, from_rows=False, titles_from_data=True)
cats = Reference(ws1, min_col=1, min_row=10, max_row=21)
bar.set_categories(cats)
bar.width = 28
bar.height = 14
ws1.add_chart(bar, "A24")

set_col_widths(ws1, [16, 16, 16, 14, 14, 12, 14, 16, 22])

# ──────────────────────────────────────────────────
#  SHEET 2: CLIENT / CUSTOMER TRACKER
# ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Client Tracker")
ws2.sheet_properties.tabColor = DUSTY_BLUE

add_title_block(ws2, "CLIENT & CUSTOMER TRACKER", "Keep all your client info organized in one place", 10)

row = 4
headers2 = ["Client Name", "Email", "Phone", "Business/Company", "Service/Product",
            "Date Acquired", "Total Revenue", "Last Contact", "Status", "Notes"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, 10)

for r in range(5, 55):
    ws2.cell(row=r, column=6).number_format = "MM/DD/YYYY"
    ws2.cell(row=r, column=7).number_format = currency_fmt()
    ws2.cell(row=r, column=8).number_format = "MM/DD/YYYY"
    style_data_row(ws2, r, 10, (r - 5) % 2 == 1)

# Status dropdown
status_dv = DataValidation(type="list", formula1='"Lead,Active,Completed,On Hold,Lost,VIP,Repeat"', allow_blank=True)
ws2.add_data_validation(status_dv)
status_dv.add("I5:I54")

set_col_widths(ws2, [22, 28, 16, 22, 22, 14, 14, 14, 12, 30])

# ──────────────────────────────────────────────────
#  SHEET 3: INVOICE TRACKER
# ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Invoice Tracker")
ws3.sheet_properties.tabColor = WARM_GOLD

add_title_block(ws3, "INVOICE TRACKER", "Track every invoice — never miss a payment again", 10)

row = 4
headers3 = ["Invoice #", "Client Name", "Description", "Date Sent", "Due Date",
            "Amount", "Amount Paid", "Balance Due", "Status", "Payment Method"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 10)

for r in range(5, 55):
    ws3.cell(row=r, column=4).number_format = "MM/DD/YYYY"
    ws3.cell(row=r, column=5).number_format = "MM/DD/YYYY"
    ws3.cell(row=r, column=6).number_format = currency_fmt()
    ws3.cell(row=r, column=7).number_format = currency_fmt()
    ws3.cell(row=r, column=8, value=f"=IF(F{r}=\"\",\"\",F{r}-G{r})").number_format = currency_fmt()
    style_data_row(ws3, r, 10, (r - 5) % 2 == 1)

inv_status_dv = DataValidation(type="list", formula1='"Draft,Sent,Viewed,Paid,Partial,Overdue,Cancelled"', allow_blank=True)
ws3.add_data_validation(inv_status_dv)
inv_status_dv.add("I5:I54")

pay_dv = DataValidation(type="list", formula1='"Cash,Check,Credit Card,Bank Transfer,PayPal,Venmo,Zelle,Stripe,Square,Other"', allow_blank=True)
ws3.add_data_validation(pay_dv)
pay_dv.add("J5:J54")

# Summary section
sum_r = 56
ws3.cell(row=sum_r, column=1, value="INVOICE SUMMARY").font = header_font(14, True, NAVY)
ws3.merge_cells(start_row=sum_r, start_column=1, end_row=sum_r, end_column=4)
ws3.cell(row=sum_r, column=1).fill = fill(ICE_BLUE)
ws3.cell(row=sum_r, column=1).alignment = center_align()

summary_items = [
    ("Total Invoiced", "=SUM(F5:F54)"),
    ("Total Collected", "=SUM(G5:G54)"),
    ("Total Outstanding", "=SUM(H5:H54)"),
    ("Collection Rate", "=IF(SUM(F5:F54)=0,0,SUM(G5:G54)/SUM(F5:F54))"),
]
for idx, (label, formula) in enumerate(summary_items):
    r = sum_r + 1 + idx
    ws3.cell(row=r, column=1, value=label).font = body_font(11, True)
    ws3.cell(row=r, column=1).fill = fill(PALE_BLUE)
    ws3.cell(row=r, column=1).border = thin_border()
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    val = ws3.cell(row=r, column=3, value=formula)
    if idx == 3:
        val.number_format = pct_fmt()
    else:
        val.number_format = currency_fmt()
    val.font = body_font(12, True, NAVY)
    val.border = thin_border()
    val.alignment = center_align()

set_col_widths(ws3, [14, 22, 28, 14, 14, 14, 14, 14, 12, 16])

# ──────────────────────────────────────────────────
#  SHEET 4: PROJECT / ORDER TRACKER
# ──────────────────────────────────────────────────
ws4 = wb.create_sheet("Project Tracker")
ws4.sheet_properties.tabColor = CORAL

add_title_block(ws4, "PROJECT & ORDER TRACKER", "Manage every project and order from start to finish", 10)

row = 4
headers4 = ["Project/Order", "Client", "Start Date", "Deadline", "Days Left",
            "Budget/Price", "Cost", "Profit", "Status", "Priority"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 10)

for r in range(5, 55):
    ws4.cell(row=r, column=3).number_format = "MM/DD/YYYY"
    ws4.cell(row=r, column=4).number_format = "MM/DD/YYYY"
    ws4.cell(row=r, column=5, value=f'=IF(D{r}="","",MAX(0,D{r}-TODAY()))')
    ws4.cell(row=r, column=6).number_format = currency_fmt()
    ws4.cell(row=r, column=7).number_format = currency_fmt()
    ws4.cell(row=r, column=8, value=f'=IF(F{r}="","",F{r}-G{r})').number_format = currency_fmt()
    style_data_row(ws4, r, 10, (r - 5) % 2 == 1)

proj_status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Review,Completed,On Hold,Cancelled"', allow_blank=True)
ws4.add_data_validation(proj_status_dv)
proj_status_dv.add("I5:I54")

priority_dv = DataValidation(type="list", formula1='"Low,Medium,High,Urgent"', allow_blank=True)
ws4.add_data_validation(priority_dv)
priority_dv.add("J5:J54")

set_col_widths(ws4, [26, 20, 14, 14, 12, 14, 14, 14, 14, 12])

# ──────────────────────────────────────────────────
#  SHEET 5: SOCIAL MEDIA CONTENT CALENDAR
# ──────────────────────────────────────────────────
ws5 = wb.create_sheet("Content Calendar")
ws5.sheet_properties.tabColor = MINT

add_title_block(ws5, "SOCIAL MEDIA CONTENT CALENDAR", "Plan, schedule, and track all your content in one place", 9)

row = 4
headers5 = ["Date", "Day", "Platform", "Content Type", "Topic / Caption Idea",
            "Hashtags", "Status", "Engagement", "Notes"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, 9)

for r in range(5, 105):
    ws5.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    ws5.cell(row=r, column=2, value=f'=IF(A{r}="","",TEXT(A{r},"dddd"))')
    style_data_row(ws5, r, 9, (r - 5) % 2 == 1)

platform_dv = DataValidation(type="list", formula1='"Instagram Post,Instagram Story,Instagram Reel,TikTok,Facebook,Pinterest,Twitter/X,LinkedIn,YouTube,Blog,Email,Other"', allow_blank=True)
ws5.add_data_validation(platform_dv)
platform_dv.add("C5:C104")

content_dv = DataValidation(type="list", formula1='"Photo,Video,Carousel,Story,Reel,Text Post,Blog Post,Newsletter,Infographic,Tutorial,Behind the Scenes,Testimonial,Promotion,Giveaway"', allow_blank=True)
ws5.add_data_validation(content_dv)
content_dv.add("D5:D104")

sm_status_dv = DataValidation(type="list", formula1='"Idea,Writing,Designing,Scheduled,Published,Boosted"', allow_blank=True)
ws5.add_data_validation(sm_status_dv)
sm_status_dv.add("G5:G104")

# Content ideas section
ideas_row = 107
ws5.cell(row=ideas_row, column=1, value="CONTENT IDEAS BANK").font = header_font(14, True, NAVY)
ws5.merge_cells(start_row=ideas_row, start_column=1, end_row=ideas_row, end_column=9)
ws5.cell(row=ideas_row, column=1).fill = fill(ICE_BLUE)
ws5.cell(row=ideas_row, column=1).alignment = center_align()

content_ideas = [
    "Behind-the-scenes of your process", "Customer testimonial/review spotlight",
    "Before & after transformation", "Day in the life", "FAQ answered",
    "Product/service feature highlight", "Industry tip or hack", "Team/personal introduction",
    "Seasonal/holiday themed post", "User-generated content repost",
    "Milestone celebration", "Sneak peek of upcoming launch",
    "Educational carousel", "Trending audio + your spin", "Collaboration shoutout",
    "Poll or question sticker", "Mini tutorial or how-to", "Packing/shipping an order",
    "Workspace or studio tour", "Motivational/inspirational quote"
]
for idx, idea in enumerate(content_ideas):
    r = ideas_row + 1 + idx
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws5.cell(row=r, column=1, value=f"  {idx+1}. {idea}").font = body_font(10)
    ws5.cell(row=r, column=1).alignment = left_align()
    ws5.cell(row=r, column=1).fill = fill(LIGHT_GRAY if idx % 2 == 0 else WHITE)

set_col_widths(ws5, [14, 14, 18, 18, 35, 30, 14, 14, 25])

# ──────────────────────────────────────────────────
#  SHEET 6: BUSINESS INCOME & EXPENSE LOG
# ──────────────────────────────────────────────────
ws6 = wb.create_sheet("Income & Expenses")
ws6.sheet_properties.tabColor = "4A7C59"

add_title_block(ws6, "BUSINESS INCOME & EXPENSE LOG", "Track every dollar in and out — essential for tax time", 9)

# Income section
row = 4
ws6.cell(row=row, column=1, value="INCOME LOG").font = header_font(14, True, NAVY)
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
ws6.cell(row=row, column=1).fill = fill(ICE_BLUE)
ws6.cell(row=row, column=1).alignment = center_align()

row = 5
inc_headers = ["Date", "Description", "Client", "Category", "Amount", "Payment Method", "Invoice #", "Tax Deductible?", "Notes"]
for i, h in enumerate(inc_headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 9, STEEL_BLUE)

for r in range(6, 56):
    ws6.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    ws6.cell(row=r, column=5).number_format = currency_fmt()
    style_data_row(ws6, r, 9, (r - 6) % 2 == 1)

inc_cat_dv = DataValidation(type="list", formula1='"Product Sales,Service Revenue,Consulting,Freelance,Affiliate Income,Sponsorship,Wholesale,Other Income"', allow_blank=True)
ws6.add_data_validation(inc_cat_dv)
inc_cat_dv.add("D6:D55")

inc_total_r = 56
ws6.cell(row=inc_total_r, column=1, value="TOTAL INCOME")
ws6.cell(row=inc_total_r, column=5, value="=SUM(E6:E55)").number_format = currency_fmt()
style_header_row(ws6, inc_total_r, 9, NAVY)

# Expense section
exp_start = 58
ws6.cell(row=exp_start, column=1, value="EXPENSE LOG").font = header_font(14, True, CORAL)
ws6.merge_cells(start_row=exp_start, start_column=1, end_row=exp_start, end_column=9)
ws6.cell(row=exp_start, column=1).fill = fill(SOFT_PEACH)
ws6.cell(row=exp_start, column=1).alignment = center_align()

row = exp_start + 1
exp_headers = ["Date", "Description", "Vendor/Payee", "Category", "Amount", "Payment Method", "Receipt?", "Tax Deductible?", "Notes"]
for i, h in enumerate(exp_headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 9, CORAL)

for r in range(exp_start + 2, exp_start + 52):
    ws6.cell(row=r, column=1).number_format = "MM/DD/YYYY"
    ws6.cell(row=r, column=5).number_format = currency_fmt()
    style_data_row(ws6, r, 9, (r - exp_start - 2) % 2 == 1)

exp_cat_dv = DataValidation(type="list", formula1='"Supplies/Materials,Shipping,Marketing/Ads,Software/Tools,Rent/Office,Utilities,Insurance,Professional Services,Travel,Meals/Entertainment,Equipment,Inventory,Bank Fees,Other"', allow_blank=True)
ws6.add_data_validation(exp_cat_dv)
exp_cat_dv.add(f"D{exp_start+2}:D{exp_start+51}")

yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws6.add_data_validation(yn_dv)
yn_dv.add(f"G{exp_start+2}:G{exp_start+51}")
yn_dv.add(f"H{exp_start+2}:H{exp_start+51}")
yn_dv.add("H6:H55")

exp_total_r = exp_start + 52
ws6.cell(row=exp_total_r, column=1, value="TOTAL EXPENSES")
ws6.cell(row=exp_total_r, column=5, value=f"=SUM(E{exp_start+2}:E{exp_total_r-1})").number_format = currency_fmt()
style_header_row(ws6, exp_total_r, 9, CORAL)

# Profit summary
profit_r = exp_total_r + 2
ws6.cell(row=profit_r, column=1, value="NET PROFIT").font = header_font(16, True, NAVY)
ws6.merge_cells(start_row=profit_r, start_column=1, end_row=profit_r, end_column=4)
ws6.cell(row=profit_r, column=1).fill = fill(PALE_BLUE)
ws6.cell(row=profit_r, column=1).alignment = center_align()
ws6.cell(row=profit_r, column=5, value=f"=E{inc_total_r}-E{exp_total_r}").number_format = currency_fmt()
ws6.cell(row=profit_r, column=5).font = Font(name="Calibri", size=16, bold=True, color=NAVY)
ws6.cell(row=profit_r, column=5).fill = fill(PALE_BLUE)
ws6.cell(row=profit_r, column=5).border = thin_border()
ws6.cell(row=profit_r, column=5).alignment = center_align()

set_col_widths(ws6, [14, 28, 22, 22, 14, 16, 14, 14, 25])

# ──────────────────────────────────────────────────
#  SHEET 7: INVENTORY MANAGER
# ──────────────────────────────────────────────────
ws7 = wb.create_sheet("Inventory")
ws7.sheet_properties.tabColor = "8B6DB0"

add_title_block(ws7, "INVENTORY MANAGER", "Track stock levels, costs, and reorder points", 10)

row = 4
headers7 = ["SKU/ID", "Product Name", "Category", "Qty on Hand", "Reorder Point",
            "Reorder Status", "Cost per Unit", "Sell Price", "Profit per Unit", "Total Stock Value"]
for i, h in enumerate(headers7, 1):
    ws7.cell(row=row, column=i, value=h)
style_header_row(ws7, row, 10)

for r in range(5, 55):
    ws7.cell(row=r, column=6, value=f'=IF(D{r}="","",IF(D{r}<=E{r},"REORDER NOW",IF(D{r}<=E{r}*1.5,"Low Stock","In Stock")))')
    ws7.cell(row=r, column=7).number_format = currency_fmt()
    ws7.cell(row=r, column=8).number_format = currency_fmt()
    ws7.cell(row=r, column=9, value=f'=IF(H{r}="","",H{r}-G{r})').number_format = currency_fmt()
    ws7.cell(row=r, column=10, value=f'=IF(D{r}="","",D{r}*G{r})').number_format = currency_fmt()
    style_data_row(ws7, r, 10, (r - 5) % 2 == 1)

inv_total = 55
ws7.cell(row=inv_total, column=1, value="TOTALS")
ws7.cell(row=inv_total, column=4, value="=SUM(D5:D54)")
ws7.cell(row=inv_total, column=10, value="=SUM(J5:J54)").number_format = currency_fmt()
style_header_row(ws7, inv_total, 10, NAVY)

set_col_widths(ws7, [14, 28, 18, 14, 14, 16, 14, 14, 14, 16])

# ──────────────────────────────────────────────────
#  SHEET 8: QUARTERLY GOALS
# ──────────────────────────────────────────────────
ws8 = wb.create_sheet("Goals & Planning")
ws8.sheet_properties.tabColor = WARM_GOLD

add_title_block(ws8, "QUARTERLY GOALS & ACTION PLANNER", "Set goals, break them into actions, and track your progress", 8)

quarters = ["Q1 (Jan-Mar)", "Q2 (Apr-Jun)", "Q3 (Jul-Sep)", "Q4 (Oct-Dec)"]
current_row = 4

for qi, quarter in enumerate(quarters):
    ws8.cell(row=current_row, column=1, value=quarter).font = header_font(14, True, NAVY)
    ws8.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    ws8.cell(row=current_row, column=1).fill = fill(ICE_BLUE)
    ws8.cell(row=current_row, column=1).alignment = center_align()
    current_row += 1

    goal_headers = ["Goal", "Category", "Key Actions / Milestones", "Deadline", "Metric/KPI", "Target", "Actual", "Status"]
    for i, h in enumerate(goal_headers, 1):
        ws8.cell(row=current_row, column=i, value=h)
    style_header_row(ws8, current_row, 8, STEEL_BLUE)
    current_row += 1

    for g in range(5):
        r = current_row
        ws8.cell(row=r, column=4).number_format = "MM/DD/YYYY"
        style_data_row(ws8, r, 8, g % 2 == 1)
        current_row += 1

    goal_cat_dv = DataValidation(type="list", formula1='"Revenue,Marketing,Product,Operations,Personal Growth,Customer Service,Financial,Other"', allow_blank=True)
    ws8.add_data_validation(goal_cat_dv)
    goal_cat_dv.add(f"B{current_row-5}:B{current_row-1}")

    goal_status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,On Track,Behind,Completed,Dropped"', allow_blank=True)
    ws8.add_data_validation(goal_status_dv)
    goal_status_dv.add(f"H{current_row-5}:H{current_row-1}")

    current_row += 1  # blank row

# Annual vision
ws8.cell(row=current_row, column=1, value="ANNUAL VISION & BIG PICTURE").font = header_font(14, True, NAVY)
ws8.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
ws8.cell(row=current_row, column=1).fill = fill(ICE_BLUE)
ws8.cell(row=current_row, column=1).alignment = center_align()
current_row += 1

vision_prompts = [
    "Annual Revenue Target:",
    "Number of Clients/Customers Goal:",
    "Top 3 Products/Services to Focus On:",
    "Biggest Challenge to Overcome:",
    "One New Skill to Learn:",
    "Marketing Strategy Focus:",
    "Personal Work-Life Balance Goal:",
]
for prompt in vision_prompts:
    ws8.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws8.cell(row=current_row, column=1, value=prompt).font = body_font(11, True)
    ws8.cell(row=current_row, column=1).fill = fill(PALE_BLUE)
    ws8.cell(row=current_row, column=1).border = thin_border()
    ws8.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=8)
    ws8.cell(row=current_row, column=3).border = thin_border()
    ws8.cell(row=current_row, column=3).fill = fill(WHITE)
    current_row += 1

set_col_widths(ws8, [22, 16, 35, 14, 18, 14, 14, 14])

# ──────────────────────────────────────────────────
#  SHEET 9: HOW TO USE
# ──────────────────────────────────────────────────
ws9 = wb.create_sheet("How to Use")
ws9.sheet_properties.tabColor = PALE_BLUE

add_title_block(ws9, "HOW TO USE YOUR SMALL BUSINESS PLANNER", "Quick-start guide to get the most from every tab", 6)

instructions = [
    ("GETTING STARTED", [
        "1. Save a copy of this file to your computer or Google Drive before editing.",
        "2. Start with the Dashboard tab to get familiar with the layout.",
        "3. Fill in your Client Tracker first — this is the foundation of your business records.",
        "4. Then set up your Quarterly Goals to give your business direction.",
    ]),
    ("DASHBOARD TAB", [
        "- This is your business command center — fill in the monthly revenue & expense totals.",
        "- The chart updates automatically to show your revenue vs expenses trend.",
        "- Update the top KPI row with current metrics for a quick health check.",
        "- TIP: Update this at the end of each month for the best overview.",
    ]),
    ("CLIENT TRACKER TAB", [
        "- Add every client/customer with their contact info and status.",
        "- Use the Status dropdown (Lead, Active, Completed, VIP, etc.) to segment clients.",
        "- Track total revenue per client to identify your most valuable relationships.",
        "- TIP: Review monthly and reach out to clients you haven't contacted recently.",
    ]),
    ("INVOICE TRACKER TAB", [
        "- Log every invoice with amounts, dates, and payment status.",
        "- The Balance Due column auto-calculates what's still owed.",
        "- Use the summary section at the bottom to see your collection rate.",
        "- TIP: Follow up on any invoice that's been 'Sent' for more than 14 days.",
    ]),
    ("PROJECT TRACKER TAB", [
        "- Track every project or order from start to finish.",
        "- The Days Left column automatically counts down to your deadline.",
        "- Profit per project is auto-calculated (Budget minus Cost).",
        "- Use Priority levels (Low/Medium/High/Urgent) to focus your daily work.",
    ]),
    ("CONTENT CALENDAR TAB", [
        "- Plan your social media posts across all platforms in one view.",
        "- Use the dropdown menus for Platform, Content Type, and Status.",
        "- The Day column auto-fills based on the date you enter.",
        "- Scroll down for 20 ready-to-use content ideas to beat writer's block!",
    ]),
    ("INCOME & EXPENSE LOG TAB", [
        "- Log every business transaction — crucial for tax time!",
        "- Mark items as Tax Deductible to make filing easier.",
        "- Net Profit auto-calculates at the bottom.",
        "- TIP: Enter transactions weekly so they don't pile up.",
    ]),
    ("INVENTORY TAB", [
        "- Track all products with SKUs, quantities, and pricing.",
        "- Set reorder points — the status column alerts you when stock is low.",
        "- Profit per unit and total stock value are auto-calculated.",
        "- Perfect for Etsy sellers, craft businesses, and product-based shops.",
    ]),
    ("GOALS & PLANNING TAB", [
        "- Set 5 goals per quarter with categories, deadlines, and KPIs.",
        "- Track actual vs target to measure your progress.",
        "- Use the Annual Vision section to set your big-picture direction.",
        "- TIP: Review goals at the start of each month and adjust as needed.",
    ]),
    ("PRO TIPS FOR BUSINESS SUCCESS", [
        "- Spend 30 minutes every Monday updating your tracker — consistency is key.",
        "- Use the Dashboard as your weekly 'business health check.'",
        "- Keep your Client Tracker updated — your network is your net worth.",
        "- Plan content 2 weeks ahead — it reduces stress and improves quality.",
        "- Save receipts digitally (photo/scan) and log expenses the same day.",
        "- Remember: Tracking your numbers is the first step to growing them!",
    ]),
]

current_row = 4
for section_title, items in instructions:
    ws9.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws9.cell(row=current_row, column=1, value=section_title).font = header_font(13, True, NAVY)
    ws9.cell(row=current_row, column=1).fill = fill(ICE_BLUE)
    ws9.cell(row=current_row, column=1).alignment = left_align()
    ws9.row_dimensions[current_row].height = 28
    current_row += 1
    for item in items:
        ws9.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws9.cell(row=current_row, column=1, value=f"  {item}").font = body_font(10)
        ws9.cell(row=current_row, column=1).alignment = left_align()
        current_row += 1
    current_row += 1

set_col_widths(ws9, [80, 15, 15, 15, 15, 15])

# ── Move How to Use to front ──
wb.move_sheet("How to Use", offset=-8)

# ── Save ──
output_path = "/home/user/claude-code/etsy-digital-download-2/product/Small_Business_Planner_Starter_Kit_2026.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
