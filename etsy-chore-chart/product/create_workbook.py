#!/usr/bin/env python3
"""
Kids Chore Chart System — Excel Workbook Generator
Creates a polished, structured parenting tool with:
  - Main Chore Chart (daily + weekly)
  - Weekly Summary with progress bars
  - Reward Tracker with auto-calculated goals
  - Instructions Guide
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# ── COLOR SYSTEM ──────────────────────────────────────────────────────────────
# Soft, structured palette — playful but controlled
COLORS = {
    "primary":        "4A90D9",   # Calm blue — headers, accents
    "primary_light":  "D6E4F0",   # Light blue — alternating rows
    "secondary":      "F5A623",   # Warm amber — progress, highlights
    "secondary_light":"FFF3D6",   # Light amber — reward zones
    "success":        "27AE60",   # Green — completed / reward earned
    "success_light":  "D5F5E3",   # Light green — success backgrounds
    "incomplete":     "E0E0E0",   # Neutral gray — not done
    "incomplete_text":"999999",   # Gray text
    "danger":         "E74C3C",   # Red — missed / alert
    "header_bg":      "2C3E50",   # Dark slate — main headers
    "header_text":    "FFFFFF",   # White text on headers
    "sub_header_bg":  "34495E",   # Slightly lighter slate
    "body_bg":        "FAFBFC",   # Off-white background
    "input_bg":       "FFFFF0",   # Ivory — editable input cells
    "border":         "BDC3C7",   # Soft border
    "white":          "FFFFFF",
    "black":          "000000",
    "dark_text":      "2C3E50",
}

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

no_border = Border(
    left=Side(style=None),
    right=Side(style=None),
    top=Side(style=None),
    bottom=Side(style=None),
)

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def header_font(size=12, bold=True, color=COLORS["header_text"]):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=11, bold=False, color=COLORS["dark_text"]):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False, indent=1):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap, indent=indent)

def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format

def style_range(ws, row, col_start, col_end, **kwargs):
    for c in range(col_start, col_end + 1):
        apply_style(ws.cell(row=row, column=c), **kwargs)

def merge_and_style(ws, start_row, start_col, end_row, end_col, value, **kwargs):
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=end_row, end_column=end_col
    )
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    apply_style(cell, **kwargs)
    # Apply border to all merged cells
    brd = kwargs.get("border", thin_border)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = brd


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MAIN CHORE CHART
# ══════════════════════════════════════════════════════════════════════════════
def build_chore_chart(wb):
    ws = wb.active
    ws.title = "Chore Chart"
    ws.sheet_properties.tabColor = COLORS["primary"]

    # Page setup
    ws.sheet_format.defaultRowHeight = 22
    ws.sheet_format.defaultColWidth = 14

    # Column widths
    col_widths = {1: 4, 2: 28, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 16, 11: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Remove gridlines
    ws.sheet_view.showGridLines = False

    # ── TITLE BAR (Row 1-2) ──
    merge_and_style(ws, 1, 1, 2, 11,
        "KIDS CHORE CHART SYSTEM",
        font=header_font(20, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    # ── CHILD INFO SECTION (Row 4) ──
    row = 4
    merge_and_style(ws, row, 1, row, 2,
        "CHILD'S NAME:",
        font=header_font(11, True, COLORS["dark_text"]),
        fill=make_fill(COLORS["primary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    merge_and_style(ws, row, 3, row, 5,
        "",
        font=body_font(12, True),
        fill=make_fill(COLORS["input_bg"]),
        alignment=center_align(),
        border=thin_border
    )
    merge_and_style(ws, row, 6, row, 7,
        "WEEK OF:",
        font=header_font(11, True, COLORS["dark_text"]),
        fill=make_fill(COLORS["primary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    merge_and_style(ws, row, 8, row, 9,
        "",
        font=body_font(12, True),
        fill=make_fill(COLORS["input_bg"]),
        alignment=center_align(),
        border=thin_border
    )
    merge_and_style(ws, row, 10, row, 11,
        "POINTS GOAL:",
        font=header_font(11, True, COLORS["dark_text"]),
        fill=make_fill(COLORS["primary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )

    # Points goal input — Row 5, col 10-11
    row = 5
    merge_and_style(ws, row, 10, row, 11,
        35,
        font=body_font(14, True, COLORS["secondary"]),
        fill=make_fill(COLORS["input_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    # ── SECTION HEADER: DAILY CHORES (Row 6) ──
    row = 6
    ws.row_dimensions[row].height = 8
    row = 7
    merge_and_style(ws, row, 1, row, 11,
        "DAILY CHORES",
        font=header_font(13, True),
        fill=make_fill(COLORS["primary"]),
        alignment=center_align(),
        border=thin_border
    )

    # ── COLUMN HEADERS (Row 8) ──
    row = 8
    ws.row_dimensions[row].height = 30
    headers = ["#", "CHORE", "MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN", "DONE", "POINTS"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_style(cell,
            font=header_font(10, True),
            fill=make_fill(COLORS["sub_header_bg"]),
            alignment=center_align(),
            border=thin_border
        )

    # ── DAILY CHORE ROWS (Rows 9-18 = 10 chores) ──
    daily_chores = [
        "Make Bed", "Brush Teeth (AM)", "Brush Teeth (PM)",
        "Pick Up Toys", "Set the Table", "Clear Dishes",
        "Put Clothes Away", "Pack School Bag", "Feed Pet",
        "Tidy Bedroom"
    ]

    # Dropdown validation: ✓ or blank
    dv_check = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
    dv_check.error = "Enter ✓ or leave blank"
    dv_check.errorTitle = "Invalid"
    ws.add_data_validation(dv_check)

    for idx, chore in enumerate(daily_chores):
        r = 9 + idx
        ws.row_dimensions[r].height = 26
        is_alt = idx % 2 == 1
        row_fill = make_fill(COLORS["primary_light"]) if is_alt else make_fill(COLORS["white"])

        # Row number
        cell = ws.cell(row=r, column=1, value=idx + 1)
        apply_style(cell, font=body_font(10, color=COLORS["incomplete_text"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Chore name (editable)
        cell = ws.cell(row=r, column=2, value=chore)
        apply_style(cell, font=body_font(11), fill=make_fill(COLORS["input_bg"]),
                    alignment=left_align(), border=thin_border)

        # Day checkboxes (Mon-Sun = cols 3-9)
        for dc in range(3, 10):
            cell = ws.cell(row=r, column=dc, value="")
            apply_style(cell, font=body_font(14, True, COLORS["success"]),
                        fill=row_fill, alignment=center_align(), border=thin_border)
            dv_check.add(cell)

        # DONE count formula
        cell = ws.cell(row=r, column=10)
        cell.value = f'=COUNTIF(C{r}:I{r},"✓")'
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # POINTS (1 point per check)
        cell = ws.cell(row=r, column=11)
        cell.value = f'=J{r}'
        apply_style(cell, font=body_font(11, True, COLORS["secondary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

    # ── SPACER ROW ──
    spacer_row = 19
    ws.row_dimensions[spacer_row].height = 8

    # ── SECTION HEADER: WEEKLY CHORES (Row 20) ──
    row = 20
    merge_and_style(ws, row, 1, row, 11,
        "WEEKLY CHORES",
        font=header_font(13, True),
        fill=make_fill(COLORS["primary"]),
        alignment=center_align(),
        border=thin_border
    )

    # Column headers again
    row = 21
    ws.row_dimensions[row].height = 30
    weekly_headers = ["#", "CHORE", "MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN", "DONE", "POINTS"]
    for i, h in enumerate(weekly_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_style(cell,
            font=header_font(10, True),
            fill=make_fill(COLORS["sub_header_bg"]),
            alignment=center_align(),
            border=thin_border
        )

    # Weekly chores (5 rows)
    weekly_chores = [
        "Clean Room Thoroughly", "Take Out Trash",
        "Help with Laundry", "Vacuum / Sweep",
        "Organize Bookshelf"
    ]

    for idx, chore in enumerate(weekly_chores):
        r = 22 + idx
        ws.row_dimensions[r].height = 26
        is_alt = idx % 2 == 1
        row_fill = make_fill(COLORS["primary_light"]) if is_alt else make_fill(COLORS["white"])

        cell = ws.cell(row=r, column=1, value=idx + 1)
        apply_style(cell, font=body_font(10, color=COLORS["incomplete_text"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        cell = ws.cell(row=r, column=2, value=chore)
        apply_style(cell, font=body_font(11), fill=make_fill(COLORS["input_bg"]),
                    alignment=left_align(), border=thin_border)

        for dc in range(3, 10):
            cell = ws.cell(row=r, column=dc, value="")
            apply_style(cell, font=body_font(14, True, COLORS["success"]),
                        fill=row_fill, alignment=center_align(), border=thin_border)
            dv_check.add(cell)

        cell = ws.cell(row=r, column=10)
        cell.value = f'=COUNTIF(C{r}:I{r},"✓")'
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        cell = ws.cell(row=r, column=11)
        cell.value = f'=J{r}*2'   # Weekly chores worth 2 points each
        apply_style(cell, font=body_font(11, True, COLORS["secondary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

    # ── SPACER ──
    ws.row_dimensions[27].height = 10

    # ── SUMMARY SECTION (Row 28+) ──
    row = 28
    merge_and_style(ws, row, 1, row, 11,
        "WEEKLY SUMMARY",
        font=header_font(13, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    row = 29
    ws.row_dimensions[row].height = 28

    # Total Chores Completed
    merge_and_style(ws, row, 1, row, 3,
        "Total Completed:",
        font=body_font(11, True),
        fill=make_fill(COLORS["primary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    merge_and_style(ws, row, 4, row, 5,
        None,
        font=body_font(14, True, COLORS["primary"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=row, column=4).value = '=SUM(J9:J18,J22:J26)'

    # Total Points
    merge_and_style(ws, row, 6, row, 8,
        "Total Points:",
        font=body_font(11, True),
        fill=make_fill(COLORS["secondary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    merge_and_style(ws, row, 9, row, 11,
        None,
        font=body_font(14, True, COLORS["secondary"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=row, column=9).value = '=SUM(K9:K18,K22:K26)'

    # Progress row
    row = 30
    ws.row_dimensions[row].height = 28

    merge_and_style(ws, row, 1, row, 3,
        "Progress:",
        font=body_font(11, True),
        fill=make_fill(COLORS["primary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    merge_and_style(ws, row, 4, row, 5,
        None,
        font=body_font(14, True, COLORS["primary"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=row, column=4).value = '=IF(J5=0,0,SUM(K9:K18,K22:K26)/J5)'
    ws.cell(row=30, column=4).number_format = '0%'

    # Reward Status
    merge_and_style(ws, row, 6, row, 8,
        "Reward Status:",
        font=body_font(11, True),
        fill=make_fill(COLORS["secondary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    merge_and_style(ws, row, 9, row, 11,
        None,
        font=body_font(14, True),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=row, column=9).value = '=IF(SUM(K9:K18,K22:K26)>=J5,"REWARD EARNED!","Keep Going!")'

    # Conditional formatting for reward earned
    ws.conditional_formatting.add(
        f'I{row}:K{row}',
        CellIsRule(
            operator='equal',
            formula=['"REWARD EARNED!"'],
            fill=make_fill(COLORS["success_light"]),
            font=Font(bold=True, color=COLORS["success"], size=14)
        )
    )

    # ── PROGRESS BAR ROW (Row 31-32) ──
    row = 31
    ws.row_dimensions[row].height = 6
    row = 32
    ws.row_dimensions[row].height = 28
    merge_and_style(ws, row, 1, row, 3,
        "Progress Bar:",
        font=body_font(11, True),
        fill=make_fill(COLORS["primary_light"]),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border
    )
    # We'll use conditional data bar formatting across cols 4-11
    for c in range(4, 12):
        cell = ws.cell(row=row, column=c)
        cell.value = f'=IF(J5=0,0,SUM(K9:K18,K22:K26)/J5)'
        cell.number_format = '0%'
        apply_style(cell, font=body_font(1, color=COLORS["white"]),
                    fill=make_fill(COLORS["white"]),
                    alignment=center_align(), border=thin_border)

    ws.conditional_formatting.add(
        f'D32:K32',
        DataBarRule(
            start_type='num', start_value=0,
            end_type='num', end_value=1,
            color=COLORS["primary"]
        )
    )

    # ── NOTES SECTION (Row 34) ──
    row = 34
    merge_and_style(ws, row, 1, row, 11,
        "NOTES & REWARD DESCRIPTION",
        font=header_font(11, True),
        fill=make_fill(COLORS["sub_header_bg"]),
        alignment=center_align(),
        border=thin_border
    )
    row = 35
    ws.row_dimensions[row].height = 60
    merge_and_style(ws, row, 1, row, 11,
        "",
        font=body_font(11),
        fill=make_fill(COLORS["input_bg"]),
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1),
        border=thin_border
    )

    # Freeze panes below header
    ws.freeze_panes = "A9"

    # Print area
    ws.print_area = "A1:K35"

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — WEEKLY SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
def build_weekly_summary(wb):
    ws = wb.create_sheet("Weekly Summary")
    ws.sheet_properties.tabColor = COLORS["secondary"]
    ws.sheet_view.showGridLines = False

    col_widths = {1: 4, 2: 22, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── TITLE ──
    merge_and_style(ws, 1, 1, 2, 10,
        "WEEKLY PROGRESS SUMMARY",
        font=header_font(18, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    # ── DAY-BY-DAY BREAKDOWN ──
    row = 4
    merge_and_style(ws, row, 1, row, 10,
        "DAY-BY-DAY COMPLETION RATE",
        font=header_font(12, True),
        fill=make_fill(COLORS["primary"]),
        alignment=center_align(),
        border=thin_border
    )

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    row = 5
    headers = ["", "DAY", "DAILY DONE", "DAILY TOTAL", "WEEKLY DONE", "WEEKLY TOTAL", "ALL DONE", "ALL TOTAL", "% COMPLETE", "STATUS"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_style(cell, font=header_font(9, True),
                    fill=make_fill(COLORS["sub_header_bg"]),
                    alignment=center_align(), border=thin_border)

    for idx, day in enumerate(days):
        r = 6 + idx
        ws.row_dimensions[r].height = 26
        is_alt = idx % 2 == 1
        row_fill = make_fill(COLORS["primary_light"]) if is_alt else make_fill(COLORS["white"])
        day_col_idx = idx + 3  # Mon=C(3), Tue=D(4), etc.
        day_col = get_column_letter(day_col_idx)

        cell = ws.cell(row=r, column=1, value=idx+1)
        apply_style(cell, font=body_font(10, color=COLORS["incomplete_text"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        cell = ws.cell(row=r, column=2, value=day)
        apply_style(cell, font=body_font(11, True), fill=row_fill,
                    alignment=left_align(), border=thin_border)

        # Daily done (count of ✓ in daily chores for this day)
        cell = ws.cell(row=r, column=3)
        cell.value = f'=COUNTIF(\'Chore Chart\'!{day_col}9:{day_col}18,"✓")'
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Daily total
        cell = ws.cell(row=r, column=4, value=10)
        apply_style(cell, font=body_font(11), fill=row_fill,
                    alignment=center_align(), border=thin_border)

        # Weekly done
        cell = ws.cell(row=r, column=5)
        cell.value = f'=COUNTIF(\'Chore Chart\'!{day_col}22:{day_col}26,"✓")'
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Weekly total
        cell = ws.cell(row=r, column=6, value=5)
        apply_style(cell, font=body_font(11), fill=row_fill,
                    alignment=center_align(), border=thin_border)

        # All done
        cell = ws.cell(row=r, column=7)
        cell.value = f'=C{r}+E{r}'
        apply_style(cell, font=body_font(11, True, COLORS["secondary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # All total
        cell = ws.cell(row=r, column=8)
        cell.value = f'=D{r}+F{r}'
        apply_style(cell, font=body_font(11), fill=row_fill,
                    alignment=center_align(), border=thin_border)

        # Percent complete
        cell = ws.cell(row=r, column=9)
        cell.value = f'=IF(H{r}=0,0,G{r}/H{r})'
        cell.number_format = '0%'
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Status
        cell = ws.cell(row=r, column=10)
        cell.value = f'=IF(I{r}>=0.8,"Great!",IF(I{r}>=0.5,"Good","Needs Work"))'
        apply_style(cell, font=body_font(10, True), fill=row_fill,
                    alignment=center_align(), border=thin_border)

    # Conditional formatting on status column
    ws.conditional_formatting.add(
        'J6:J12',
        CellIsRule(operator='equal', formula=['"Great!"'],
                   fill=make_fill(COLORS["success_light"]),
                   font=Font(bold=True, color=COLORS["success"]))
    )
    ws.conditional_formatting.add(
        'J6:J12',
        CellIsRule(operator='equal', formula=['"Needs Work"'],
                   fill=make_fill("FDEDEC"),
                   font=Font(bold=True, color=COLORS["danger"]))
    )

    # Data bars on percent column
    ws.conditional_formatting.add(
        'I6:I12',
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=1,
                    color=COLORS["primary"])
    )

    # ── WEEKLY TOTALS ──
    row = 14
    merge_and_style(ws, row, 1, row, 10,
        "OVERALL WEEKLY STATS",
        font=header_font(12, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    stats = [
        ("Total Check-Marks", '=SUM(G6:G12)', None),
        ("Total Possible", '=SUM(H6:H12)', None),
        ("Weekly Completion Rate", '=IF(SUM(H6:H12)=0,0,SUM(G6:G12)/SUM(H6:H12))', '0%'),
        ("Total Points Earned", "='Chore Chart'!I29", None),
        ("Points Goal", "='Chore Chart'!J5", None),
        ("Reward Progress", "=IF('Chore Chart'!J5=0,0,'Chore Chart'!I29/'Chore Chart'!J5)", '0%'),
    ]

    for i, (label, formula, fmt) in enumerate(stats):
        r = 15 + i
        ws.row_dimensions[r].height = 28
        is_alt = i % 2 == 1
        row_fill = make_fill(COLORS["secondary_light"]) if is_alt else make_fill(COLORS["white"])

        merge_and_style(ws, r, 1, r, 5,
            label,
            font=body_font(11, True),
            fill=row_fill,
            alignment=Alignment(horizontal="right", vertical="center", indent=1),
            border=thin_border
        )
        merge_and_style(ws, r, 6, r, 10,
            None,
            font=body_font(14, True, COLORS["primary"]),
            fill=row_fill,
            alignment=center_align(),
            border=thin_border
        )
        ws.cell(row=r, column=6).value = formula
        if fmt:
            ws.cell(row=r, column=6).number_format = fmt

    # ── BAR CHART ──
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Daily Completion"
    chart.y_axis.title = "Chores Completed"
    chart.x_axis.title = "Day"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 15

    data = Reference(ws, min_col=7, min_row=5, max_row=12)
    cats = Reference(ws, min_col=2, min_row=6, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 22
    chart.height = 12

    # Color the bars
    series = chart.series[0]
    series.graphicalProperties.solidFill = COLORS["primary"]

    ws.add_chart(chart, "A22")

    ws.freeze_panes = "A6"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — REWARD TRACKER
# ══════════════════════════════════════════════════════════════════════════════
def build_reward_tracker(wb):
    ws = wb.create_sheet("Reward Tracker")
    ws.sheet_properties.tabColor = COLORS["success"]
    ws.sheet_view.showGridLines = False

    col_widths = {1: 4, 2: 24, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── TITLE ──
    merge_and_style(ws, 1, 1, 2, 8,
        "REWARD TRACKER",
        font=header_font(18, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    # ── CURRENT WEEK STATUS ──
    row = 4
    merge_and_style(ws, row, 1, row, 8,
        "THIS WEEK'S REWARD STATUS",
        font=header_font(12, True),
        fill=make_fill(COLORS["secondary"]),
        alignment=center_align(),
        border=thin_border
    )

    # Points Earned
    row = 5
    ws.row_dimensions[row].height = 35
    merge_and_style(ws, row, 1, row, 3,
        "Points Earned This Week:",
        font=body_font(12, True),
        fill=make_fill(COLORS["secondary_light"]),
        alignment=Alignment(horizontal="right", vertical="center", indent=1),
        border=thin_border
    )
    merge_and_style(ws, row, 4, row, 5,
        None,
        font=Font(name="Calibri", size=22, bold=True, color=COLORS["primary"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=5, column=4).value = "='Chore Chart'!I29"

    merge_and_style(ws, row, 6, row, 6,
        "of",
        font=body_font(12),
        fill=make_fill(COLORS["secondary_light"]),
        alignment=center_align(),
        border=thin_border
    )
    merge_and_style(ws, row, 7, row, 8,
        None,
        font=Font(name="Calibri", size=22, bold=True, color=COLORS["secondary"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=5, column=7).value = "='Chore Chart'!J5"

    # Progress percentage
    row = 6
    ws.row_dimensions[row].height = 35
    merge_and_style(ws, row, 1, row, 3,
        "Progress Toward Reward:",
        font=body_font(12, True),
        fill=make_fill(COLORS["secondary_light"]),
        alignment=Alignment(horizontal="right", vertical="center", indent=1),
        border=thin_border
    )
    merge_and_style(ws, row, 4, row, 5,
        None,
        font=Font(name="Calibri", size=22, bold=True, color=COLORS["primary"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=6, column=4).value = "=IF('Chore Chart'!J5=0,0,'Chore Chart'!I29/'Chore Chart'!J5)"
    ws.cell(row=6, column=4).number_format = '0%'

    merge_and_style(ws, row, 6, row, 8,
        None,
        font=body_font(14, True, COLORS["success"]),
        fill=make_fill(COLORS["white"]),
        alignment=center_align(),
        border=thin_border
    )
    ws.cell(row=6, column=6).value = "=IF('Chore Chart'!I29>='Chore Chart'!J5,\"REWARD EARNED!\",\"Keep Working!\")"

    ws.conditional_formatting.add(
        'F6:H6',
        CellIsRule(operator='equal', formula=['"REWARD EARNED!"'],
                   fill=make_fill(COLORS["success_light"]),
                   font=Font(bold=True, color=COLORS["success"], size=16))
    )

    # Progress bar
    row = 7
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 3,
        "Progress Bar:",
        font=body_font(11, True),
        fill=make_fill(COLORS["secondary_light"]),
        alignment=Alignment(horizontal="right", vertical="center", indent=1),
        border=thin_border
    )
    for c in range(4, 9):
        cell = ws.cell(row=row, column=c)
        cell.value = "=IF('Chore Chart'!J5=0,0,'Chore Chart'!I29/'Chore Chart'!J5)"
        cell.number_format = '0%'
        apply_style(cell, font=body_font(1, color=COLORS["white"]),
                    fill=make_fill(COLORS["white"]),
                    alignment=center_align(), border=thin_border)

    ws.conditional_formatting.add(
        'D7:H7',
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=1,
                    color=COLORS["success"])
    )

    # ── REWARD BANK (Multi-week tracker) ──
    row = 9
    merge_and_style(ws, row, 1, row, 8,
        "REWARD BANK — MULTI-WEEK TRACKER",
        font=header_font(12, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    row = 10
    bank_headers = ["#", "WEEK", "POINTS EARNED", "POINTS GOAL", "% ACHIEVED", "REWARD EARNED?", "REWARD GIVEN", "NOTES"]
    for i, h in enumerate(bank_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_style(cell, font=header_font(9, True),
                    fill=make_fill(COLORS["sub_header_bg"]),
                    alignment=center_align(), border=thin_border)

    for idx in range(8):
        r = 11 + idx
        ws.row_dimensions[r].height = 26
        is_alt = idx % 2 == 1
        row_fill = make_fill(COLORS["secondary_light"]) if is_alt else make_fill(COLORS["white"])

        cell = ws.cell(row=r, column=1, value=idx + 1)
        apply_style(cell, font=body_font(10, color=COLORS["incomplete_text"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Week label (editable)
        cell = ws.cell(row=r, column=2, value=f"Week {idx + 1}")
        apply_style(cell, font=body_font(11), fill=make_fill(COLORS["input_bg"]),
                    alignment=center_align(), border=thin_border)

        # Points earned (input)
        cell = ws.cell(row=r, column=3, value="")
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=make_fill(COLORS["input_bg"]),
                    alignment=center_align(), border=thin_border)

        # Points goal (input)
        cell = ws.cell(row=r, column=4, value="")
        apply_style(cell, font=body_font(11), fill=make_fill(COLORS["input_bg"]),
                    alignment=center_align(), border=thin_border)

        # % achieved
        cell = ws.cell(row=r, column=5)
        cell.value = f'=IF(D{r}="","",(IF(D{r}=0,0,C{r}/D{r})))'
        cell.number_format = '0%'
        apply_style(cell, font=body_font(11, True, COLORS["primary"]),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Reward earned?
        cell = ws.cell(row=r, column=6)
        cell.value = f'=IF(D{r}="","",IF(C{r}>=D{r},"YES","NO"))'
        apply_style(cell, font=body_font(11, True),
                    fill=row_fill, alignment=center_align(), border=thin_border)

        # Reward given (checkbox)
        cell = ws.cell(row=r, column=7, value="")
        apply_style(cell, font=body_font(11, True, COLORS["success"]),
                    fill=make_fill(COLORS["input_bg"]),
                    alignment=center_align(), border=thin_border)

        # Notes
        cell = ws.cell(row=r, column=8, value="")
        apply_style(cell, font=body_font(10),
                    fill=make_fill(COLORS["input_bg"]),
                    alignment=left_align(), border=thin_border)

    # Conditional formatting on reward earned
    ws.conditional_formatting.add(
        'F11:F18',
        CellIsRule(operator='equal', formula=['"YES"'],
                   fill=make_fill(COLORS["success_light"]),
                   font=Font(bold=True, color=COLORS["success"]))
    )
    ws.conditional_formatting.add(
        'F11:F18',
        CellIsRule(operator='equal', formula=['"NO"'],
                   fill=make_fill("FDEDEC"),
                   font=Font(bold=True, color=COLORS["danger"]))
    )

    # Data bars on % column
    ws.conditional_formatting.add(
        'E11:E18',
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=1,
                    color=COLORS["secondary"])
    )

    # ── CUMULATIVE STATS ──
    row = 20
    merge_and_style(ws, row, 1, row, 8,
        "CUMULATIVE STATS",
        font=header_font(12, True),
        fill=make_fill(COLORS["primary"]),
        alignment=center_align(),
        border=thin_border
    )

    cum_stats = [
        ("Total Points Earned (All Weeks)", '=SUM(C11:C18)'),
        ("Total Rewards Earned", '=COUNTIF(F11:F18,"YES")'),
        ("Average Weekly Completion", '=IF(COUNTA(E11:E18)=0,0,AVERAGE(E11:E18))'),
    ]

    for i, (label, formula) in enumerate(cum_stats):
        r = 21 + i
        ws.row_dimensions[r].height = 28
        is_alt = i % 2 == 1
        row_fill = make_fill(COLORS["secondary_light"]) if is_alt else make_fill(COLORS["white"])

        merge_and_style(ws, r, 1, r, 5, label,
            font=body_font(11, True), fill=row_fill,
            alignment=Alignment(horizontal="right", vertical="center", indent=1),
            border=thin_border)
        merge_and_style(ws, r, 6, r, 8, None,
            font=body_font(14, True, COLORS["primary"]),
            fill=row_fill, alignment=center_align(), border=thin_border)
        ws.cell(row=r, column=6).value = formula
        if "Average" in label:
            ws.cell(row=r, column=6).number_format = '0%'

    ws.freeze_panes = "A11"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — INSTRUCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = COLORS["incomplete_text"]
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 80

    # Title
    merge_and_style(ws, 1, 1, 2, 2,
        "HOW TO USE YOUR CHORE CHART SYSTEM",
        font=header_font(16, True),
        fill=make_fill(COLORS["header_bg"]),
        alignment=center_align(),
        border=thin_border
    )

    instructions = [
        ("GETTING STARTED", [
            "Enter your child's name in the 'Child's Name' field on the Chore Chart tab.",
            "Enter the current week's start date in the 'Week Of' field.",
            "Set a Points Goal — this is the number of points needed to earn a reward.",
            "Customize chore names by clicking on any chore in the yellow-tinted cells.",
        ]),
        ("TRACKING DAILY CHORES", [
            "Each day, mark completed chores with a ✓ (checkmark) in the day's column.",
            "Use the dropdown or type ✓ directly into the cell.",
            "Daily chores earn 1 point per completion.",
            "The 'DONE' column auto-counts how many days each chore was completed.",
        ]),
        ("TRACKING WEEKLY CHORES", [
            "Weekly chores are bigger tasks done once or a few times per week.",
            "Mark them the same way — with a ✓ on the day completed.",
            "Weekly chores earn 2 points per completion (double daily chores).",
        ]),
        ("UNDERSTANDING THE SUMMARY", [
            "Total Completed: Total number of individual chore check-marks.",
            "Total Points: Sum of all points (daily = 1pt, weekly = 2pt each).",
            "Progress: Shows percentage toward your points goal.",
            "Reward Status: Automatically shows 'REWARD EARNED!' when goal is met.",
            "Progress Bar: Visual representation of progress toward the reward.",
        ]),
        ("USING THE WEEKLY SUMMARY TAB", [
            "This tab auto-calculates daily completion rates for each day of the week.",
            "See which days your child is most/least consistent.",
            "A bar chart shows at-a-glance daily performance.",
            "Overall stats track total completion rate and reward progress.",
        ]),
        ("USING THE REWARD TRACKER TAB", [
            "The top section links to this week's Chore Chart data automatically.",
            "The Reward Bank below lets you log results from multiple weeks.",
            "Enter each week's points earned and goal to track long-term progress.",
            "Cumulative stats show total rewards earned and average completion.",
        ]),
        ("TIPS FOR SUCCESS", [
            "Set realistic goals — start with a lower point threshold and increase.",
            "Let your child choose their reward to increase motivation.",
            "Review the chart together each evening as a routine.",
            "Celebrate progress, not just perfection — the progress bar helps!",
            "Print the Chore Chart tab for fridge display if desired.",
        ]),
        ("EDITABLE VS LOCKED CELLS", [
            "Yellow-tinted cells = Editable (chore names, check marks, notes, names).",
            "White/colored cells with formulas = Auto-calculated (do not edit).",
            "If you accidentally overwrite a formula, use Ctrl+Z to undo.",
        ]),
    ]

    current_row = 4
    for section_title, items in instructions:
        merge_and_style(ws, current_row, 1, current_row, 2,
            section_title,
            font=header_font(12, True),
            fill=make_fill(COLORS["primary"]),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=thin_border
        )
        current_row += 1

        for item in items:
            ws.row_dimensions[current_row].height = 24
            cell = ws.cell(row=current_row, column=1)
            apply_style(cell, fill=make_fill(COLORS["white"]), border=thin_border)

            cell = ws.cell(row=current_row, column=2, value=f"  {item}")
            apply_style(cell, font=body_font(10),
                        fill=make_fill(COLORS["white"]),
                        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                        border=thin_border)
            current_row += 1

        current_row += 1  # spacer

    ws.freeze_panes = "A4"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN — BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    build_chore_chart(wb)
    build_weekly_summary(wb)
    build_reward_tracker(wb)
    build_instructions(wb)

    output_path = "/home/user/claude-code/etsy-chore-chart/product/Kids_Chore_Chart_System.xlsx"
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")


if __name__ == "__main__":
    main()
