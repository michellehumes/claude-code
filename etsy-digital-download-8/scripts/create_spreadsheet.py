#!/usr/bin/env python3
"""Create Home Organization & Cleaning Planner 2026 Excel spreadsheet."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from copy import copy
from datetime import datetime, date

OUTPUT = "/home/user/claude-code/etsy-digital-download-8/product/Home_Organization_Cleaning_Planner_2026.xlsx"

# ── Colors ──────────────────────────────────────────────────────────────
FOREST   = "276749"
SAGE     = "68D391"
WARM_TERRA = "C05621"
CREAM    = "FFFFF0"
SOFT_SAGE = "C6F6D5"
LINEN    = "FAF5FF"
DUSTY_PINK = "FEB2B2"
DARK_TEXT = "1A202C"
WHITE    = "FFFFFF"

# ── Reusable Styles ────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=FOREST, end_color=FOREST, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11, color=DARK_TEXT)
subheader_fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
body_font = Font(name="Calibri", size=11, color=DARK_TEXT)
title_font = Font(name="Calibri", bold=True, size=16, color=FOREST)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=WARM_TERRA)
row_fill_1 = PatternFill(start_color=CREAM, end_color=CREAM, fill_type="solid")
row_fill_2 = PatternFill(start_color=SOFT_SAGE, end_color=SOFT_SAGE, fill_type="solid")
linen_fill = PatternFill(start_color=LINEN, end_color=LINEN, fill_type="solid")
pink_fill = PatternFill(start_color=DUSTY_PINK, end_color=DUSTY_PINK, fill_type="solid")
terra_font = Font(name="Calibri", bold=True, size=12, color=WARM_TERRA)

thin_border = Border(
    left=Side(style="thin", color=SAGE),
    right=Side(style="thin", color=SAGE),
    top=Side(style="thin", color=SAGE),
    bottom=Side(style="thin", color=SAGE),
)
header_border = Border(
    left=Side(style="thin", color=FOREST),
    right=Side(style="thin", color=FOREST),
    top=Side(style="medium", color=FOREST),
    bottom=Side(style="medium", color=FOREST),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
wrap_align = Alignment(wrap_text=True, vertical="top")

def style_header_row(ws, row, num_cols, extra_font=None, extra_fill=None):
    f = extra_font or header_font
    fi = extra_fill or header_fill
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = f
        c.fill = fi
        c.alignment = center_align
        c.border = header_border

def style_body_cell(ws, row, col, alt=False):
    c = ws.cell(row=row, column=col)
    c.font = body_font
    c.border = thin_border
    c.alignment = center_align
    if alt:
        c.fill = row_fill_2 if row % 2 == 0 else row_fill_1
    return c

def add_title_block(ws, title, subtitle, merge_end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end_col)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = subtitle_font
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

def apply_alt_rows(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            style_body_cell(ws, r, col, alt=True)

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════
# TAB 1: Weekly Cleaning Schedule
# ════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Weekly Cleaning Schedule"
ws1.sheet_properties.tabColor = FOREST

add_title_block(ws1, "🏠 WEEKLY CLEANING SCHEDULE 2026", "Track daily cleaning tasks by room — check off as you go!", 11)

headers1 = ["Room / Area", "Monday", "Tuesday", "Wednesday", "Thursday",
            "Friday", "Saturday", "Sunday", "Weekly Deep Clean Task",
            "Frequency Notes", "Priority"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, len(headers1))

rooms_data = [
    ("Kitchen — Counters & Stovetop", "Wipe counters, clean stovetop", "Daily wipe-down"),
    ("Kitchen — Floors", "Sweep & mop floors", "Daily sweep, mop 2x/week"),
    ("Kitchen — Dishes & Sink", "Wash dishes, scrub sink", "After every meal"),
    ("Kitchen — Appliances", "Clean microwave, wipe fridge", "Weekly deep clean"),
    ("Bathroom — Master", "Scrub toilet, clean mirror", "Daily quick wipe"),
    ("Bathroom — Guest", "Wipe surfaces, restock supplies", "2x per week"),
    ("Bathroom — Kids", "Clean tub, organize toiletries", "Daily pickup"),
    ("Master Bedroom", "Make bed, dust surfaces", "Daily bed-making"),
    ("Kids Bedroom 1", "Pick up toys, make bed", "Daily routine"),
    ("Kids Bedroom 2", "Organize desk, vacuum", "Every other day"),
    ("Guest Bedroom", "Dust, change sheets if needed", "Before guests / monthly"),
    ("Living Room", "Vacuum, fluff pillows, dust TV", "Daily tidy"),
    ("Dining Room", "Wipe table, vacuum under chairs", "After meals"),
    ("Home Office", "Organize desk, dust electronics", "Daily declutter"),
    ("Laundry Room", "Run loads, clean lint trap", "Daily loads"),
    ("Hallways & Stairs", "Vacuum, dust railings", "2x per week"),
    ("Garage", "Sweep, organize tools", "Weekly"),
    ("Front Porch / Entryway", "Sweep, wipe door handle", "Daily"),
    ("Basement", "Check for moisture, organize", "Biweekly"),
    ("Attic / Storage", "Dust, check inventory", "Monthly"),
]

check_dv = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
check_dv.prompt = "Select ✓ when done"
check_dv.promptTitle = "Task Status"
ws1.add_data_validation(check_dv)

priority_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws1.add_data_validation(priority_dv)

for idx, (room, deep_task, freq) in enumerate(rooms_data):
    r = 5 + idx
    ws1.cell(row=r, column=1, value=room)
    for col in range(2, 9):  # Mon-Sun
        cell = ws1.cell(row=r, column=col, value="")
        check_dv.add(cell)
    ws1.cell(row=r, column=9, value=deep_task)
    ws1.cell(row=r, column=10, value=freq)
    ws1.cell(row=r, column=11, value="")
    priority_dv.add(ws1.cell(row=r, column=11))

# Pre-fill some checkmarks for visual appeal
sample_checks = [
    (5, 2), (5, 3), (5, 5), (5, 7),
    (6, 2), (6, 4), (6, 6),
    (7, 2), (7, 3), (7, 4), (7, 5), (7, 6),
    (8, 3), (8, 5), (8, 7),
    (9, 2), (9, 3), (9, 4), (9, 5), (9, 6), (9, 7), (9, 8),
    (10, 4), (10, 7),
    (11, 2), (11, 5),
    (12, 2), (12, 4), (12, 6),
]
for (sr, sc) in sample_checks:
    ws1.cell(row=sr, column=sc, value="✓")

apply_alt_rows(ws1, 5, 24, len(headers1))
# Set some priorities
ws1.cell(row=5, column=11, value="High")
ws1.cell(row=9, column=11, value="High")
ws1.cell(row=12, column=11, value="High")
ws1.cell(row=15, column=11, value="Medium")

col_widths1 = [30, 12, 12, 14, 14, 12, 12, 12, 32, 24, 12]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 2: Deep Cleaning Checklist
# ════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Deep Cleaning Checklist")
ws2.sheet_properties.tabColor = SAGE

add_title_block(ws2, "🧹 DEEP CLEANING CHECKLIST", "Never miss a deep-clean task — auto-calculates next due date!", 8)

headers2 = ["Room", "Task", "Frequency", "Last Done", "Next Due", "Status", "Assigned To", "Notes"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, len(headers2))

freq_dv = DataValidation(type="list", formula1='"Weekly,Biweekly,Monthly,Quarterly,Annually"', allow_blank=True)
ws2.add_data_validation(freq_dv)

status_dv = DataValidation(type="list", formula1='"Done,Due,Overdue,Scheduled"', allow_blank=True)
ws2.add_data_validation(status_dv)

deep_tasks = [
    # Kitchen (15)
    ("Kitchen", "Deep clean oven interior", "Monthly", "2026-01-10"),
    ("Kitchen", "Clean behind refrigerator", "Quarterly", "2025-12-15"),
    ("Kitchen", "Degrease range hood & filter", "Monthly", "2026-01-20"),
    ("Kitchen", "Descale coffee maker & kettle", "Monthly", "2026-01-05"),
    ("Kitchen", "Clean & organize pantry", "Monthly", "2026-01-15"),
    ("Kitchen", "Sanitize trash cans", "Biweekly", "2026-01-25"),
    ("Kitchen", "Deep clean dishwasher", "Monthly", "2026-01-08"),
    ("Kitchen", "Wipe cabinet fronts & handles", "Monthly", "2026-01-12"),
    ("Kitchen", "Clean light fixtures", "Quarterly", "2025-11-01"),
    ("Kitchen", "Defrost freezer", "Quarterly", "2025-12-20"),
    ("Kitchen", "Clean grout & backsplash", "Monthly", "2026-01-18"),
    ("Kitchen", "Polish stainless steel appliances", "Weekly", "2026-02-01"),
    ("Kitchen", "Organize spice cabinet", "Quarterly", "2025-12-01"),
    ("Kitchen", "Clean under sink & organize", "Monthly", "2026-01-22"),
    ("Kitchen", "Sharpen knives & clean block", "Quarterly", "2025-11-15"),
    # Bathroom (12)
    ("Bathroom", "Scrub tile grout", "Monthly", "2026-01-15"),
    ("Bathroom", "Deep clean shower head (vinegar soak)", "Monthly", "2026-01-10"),
    ("Bathroom", "Wash shower curtain & liner", "Monthly", "2026-01-20"),
    ("Bathroom", "Clean exhaust fan", "Quarterly", "2025-12-01"),
    ("Bathroom", "Organize medicine cabinet", "Quarterly", "2025-12-15"),
    ("Bathroom", "Replace toothbrush heads", "Quarterly", "2025-12-20"),
    ("Bathroom", "Deep clean toilet (incl. base & behind)", "Biweekly", "2026-01-28"),
    ("Bathroom", "Wash bath mats & rugs", "Biweekly", "2026-01-25"),
    ("Bathroom", "Clean mirror edges & frame", "Monthly", "2026-01-18"),
    ("Bathroom", "Descale faucets", "Monthly", "2026-01-05"),
    ("Bathroom", "Organize under-sink storage", "Quarterly", "2025-11-01"),
    ("Bathroom", "Replace caulking if needed", "Annually", "2025-06-01"),
    # Bedroom (10)
    ("Bedroom", "Flip/rotate mattress", "Quarterly", "2025-12-15"),
    ("Bedroom", "Wash pillows & duvet", "Quarterly", "2025-12-01"),
    ("Bedroom", "Clean under bed", "Monthly", "2026-01-15"),
    ("Bedroom", "Dust ceiling fan blades", "Monthly", "2026-01-10"),
    ("Bedroom", "Vacuum mattress", "Monthly", "2026-01-20"),
    ("Bedroom", "Wash curtains/blinds", "Quarterly", "2025-11-01"),
    ("Bedroom", "Organize closet & donate", "Quarterly", "2025-12-20"),
    ("Bedroom", "Clean light switches & outlets", "Monthly", "2026-01-22"),
    ("Bedroom", "Dust baseboards", "Monthly", "2026-01-18"),
    ("Bedroom", "Clean window tracks", "Quarterly", "2025-10-01"),
    # Living Room (10)
    ("Living Room", "Deep clean upholstery/sofa", "Quarterly", "2025-12-15"),
    ("Living Room", "Shampoo carpets/rugs", "Quarterly", "2025-12-01"),
    ("Living Room", "Clean windows inside & out", "Quarterly", "2025-11-01"),
    ("Living Room", "Dust all bookshelves & decor", "Monthly", "2026-01-12"),
    ("Living Room", "Clean TV screen & electronics", "Biweekly", "2026-01-25"),
    ("Living Room", "Vacuum under furniture", "Monthly", "2026-01-15"),
    ("Living Room", "Clean air vents & registers", "Quarterly", "2025-12-01"),
    ("Living Room", "Polish wood furniture", "Monthly", "2026-01-20"),
    ("Living Room", "Wash throw blankets & pillows", "Monthly", "2026-01-10"),
    ("Living Room", "Clean fireplace (if applicable)", "Annually", "2025-09-01"),
    # Other Areas (13)
    ("Laundry Room", "Clean washing machine (vinegar cycle)", "Monthly", "2026-01-15"),
    ("Laundry Room", "Deep clean dryer vent", "Quarterly", "2025-12-01"),
    ("Laundry Room", "Organize supplies & shelves", "Monthly", "2026-01-20"),
    ("Home Office", "Clean keyboard & mouse", "Biweekly", "2026-01-28"),
    ("Home Office", "Organize files & shred papers", "Monthly", "2026-01-10"),
    ("Home Office", "Dust & clean monitor", "Biweekly", "2026-01-25"),
    ("Garage", "Sweep & organize", "Monthly", "2026-01-05"),
    ("Garage", "Check for pests", "Quarterly", "2025-12-01"),
    ("Entryway", "Clean door & hardware", "Monthly", "2026-01-22"),
    ("Entryway", "Wash welcome mat", "Monthly", "2026-01-18"),
    ("Whole House", "Replace HVAC filters", "Quarterly", "2025-12-15"),
    ("Whole House", "Test smoke & CO detectors", "Monthly", "2026-01-01"),
    ("Whole House", "Deep clean all baseboards", "Quarterly", "2025-11-15"),
]

freq_days = {"Weekly": 7, "Biweekly": 14, "Monthly": 30, "Quarterly": 90, "Annually": 365}

for idx, (room, task, freq, last_done) in enumerate(deep_tasks):
    r = 5 + idx
    ws2.cell(row=r, column=1, value=room)
    ws2.cell(row=r, column=2, value=task)
    fc = ws2.cell(row=r, column=3, value=freq)
    freq_dv.add(fc)
    ld = ws2.cell(row=r, column=4, value=datetime.strptime(last_done, "%Y-%m-%d").date())
    ld.number_format = "YYYY-MM-DD"
    # Next Due formula
    formula_map = {
        "Weekly": 7, "Biweekly": 14, "Monthly": 30, "Quarterly": 91, "Annually": 365
    }
    next_cell = ws2.cell(row=r, column=5)
    next_cell.value = f'=IF(D{r}="","",D{r}+CHOOSE(MATCH(C{r},{{"Weekly","Biweekly","Monthly","Quarterly","Annually"}},0),7,14,30,91,365))'
    next_cell.number_format = "YYYY-MM-DD"

    sc = ws2.cell(row=r, column=6, value="Due")
    status_dv.add(sc)
    ws2.cell(row=r, column=7, value="")
    ws2.cell(row=r, column=8, value="")

apply_alt_rows(ws2, 5, 64, len(headers2))

# Conditional formatting for status
ws2.conditional_formatting.add(f'F5:F64',
    CellIsRule(operator='equal', formula=['"Overdue"'],
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
              font=Font(color="C53030", bold=True)))
ws2.conditional_formatting.add(f'F5:F64',
    CellIsRule(operator='equal', formula=['"Done"'],
              fill=PatternFill(start_color=SOFT_SAGE, end_color=SOFT_SAGE, fill_type="solid"),
              font=Font(color=FOREST, bold=True)))

col_widths2 = [18, 42, 14, 14, 14, 12, 14, 24]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 3: Declutter Tracker
# ════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Declutter Tracker")
ws3.sheet_properties.tabColor = WARM_TERRA

add_title_block(ws3, "📦 DECLUTTER TRACKER", "Track your decluttering journey room by room — celebrate your progress!", 10)

headers3 = ["Room", "Area / Category", "Items to Sort", "Keep", "Donate", "Trash", "Sell", "Date Completed", "Photos Taken?", "Notes"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, len(headers3))

photo_dv = DataValidation(type="list", formula1='"Yes - Before & After,Yes - Before Only,Yes - After Only,No,Planned"', allow_blank=True)
ws3.add_data_validation(photo_dv)

declutter_data = [
    ("Kitchen", "Pantry & Food Storage", 45, 30, 5, 10, 0, "2026-01-10", "Yes - Before & After", "Expired items removed"),
    ("Kitchen", "Tupperware & Containers", 35, 15, 8, 12, 0, "2026-01-12", "Yes - Before & After", "Matched lids to containers"),
    ("Kitchen", "Under Sink", 20, 12, 3, 5, 0, "2026-01-15", "Yes - Before Only", "Organized cleaning supplies"),
    ("Kitchen", "Utensil Drawers", 55, 35, 10, 10, 0, "2026-01-18", "Yes - Before & After", "Kept only daily-use items"),
    ("Kitchen", "Small Appliances", 12, 8, 2, 0, 2, "", "No", ""),
    ("Bathroom", "Medicine Cabinet", 30, 15, 0, 15, 0, "2026-01-20", "Yes - Before & After", "Checked all expiry dates"),
    ("Bathroom", "Under Vanity", 25, 15, 5, 5, 0, "2026-01-22", "Yes - After Only", "Added organizer bins"),
    ("Bathroom", "Linen Closet", 40, 25, 10, 5, 0, "", "Planned", ""),
    ("Master Bedroom", "Closet — Hanging Clothes", 80, 45, 25, 5, 5, "2026-01-25", "Yes - Before & After", "Seasonal rotation done"),
    ("Master Bedroom", "Dresser Drawers", 60, 40, 12, 8, 0, "2026-01-26", "Yes - Before & After", "Folded KonMari style"),
    ("Master Bedroom", "Nightstands", 15, 10, 2, 3, 0, "2026-01-27", "No", "Quick sort"),
    ("Master Bedroom", "Shoes & Accessories", 35, 20, 10, 0, 5, "", "Planned", ""),
    ("Kids Room", "Toys & Games", 75, 40, 20, 10, 5, "", "Planned", "Involve kids in decisions"),
    ("Kids Room", "Books", 50, 35, 15, 0, 0, "", "No", ""),
    ("Kids Room", "Clothing", 65, 40, 20, 5, 0, "", "Planned", "Check sizes"),
    ("Living Room", "Bookshelves", 45, 30, 10, 5, 0, "2026-02-01", "Yes - Before & After", "Organized by category"),
    ("Living Room", "Entertainment Center", 25, 18, 4, 3, 0, "", "No", ""),
    ("Living Room", "Magazine & Paper Pile", 30, 5, 0, 25, 0, "2026-02-02", "No", "Recycled most"),
    ("Home Office", "Desk & Supplies", 40, 25, 5, 10, 0, "", "Planned", ""),
    ("Home Office", "Filing Cabinet", 100, 60, 0, 40, 0, "", "Planned", "Shred sensitive docs"),
    ("Home Office", "Tech & Cables", 30, 15, 5, 5, 5, "", "No", ""),
    ("Garage", "Tools & Hardware", 50, 35, 5, 5, 5, "", "Planned", ""),
    ("Garage", "Seasonal Items", 25, 20, 3, 2, 0, "", "No", ""),
    ("Garage", "Sports Equipment", 20, 12, 5, 0, 3, "", "Planned", ""),
    ("Entryway", "Coat Closet", 30, 20, 8, 2, 0, "", "Planned", ""),
    ("Basement", "Storage Boxes", 40, 25, 8, 7, 0, "", "Planned", "Label everything"),
    ("Basement", "Holiday Decorations", 35, 30, 3, 2, 0, "", "No", ""),
    ("Attic", "Memorabilia & Keepsakes", 30, 25, 0, 5, 0, "", "Planned", "Scan photos digitally"),
    ("Laundry Room", "Supplies & Products", 20, 12, 3, 5, 0, "", "No", ""),
    ("Whole House", "Duplicate Cleaning Supplies", 15, 8, 2, 5, 0, "", "No", "Consolidate"),
]

for idx, row_data in enumerate(declutter_data):
    r = 5 + idx
    for col_idx, val in enumerate(row_data):
        c = ws3.cell(row=r, column=col_idx + 1, value=val)
        if col_idx == 7 and val:  # Date
            try:
                c.value = datetime.strptime(val, "%Y-%m-%d").date()
                c.number_format = "YYYY-MM-DD"
            except:
                pass
        if col_idx == 8:
            photo_dv.add(c)

apply_alt_rows(ws3, 5, 34, len(headers3))

# Summary row
summary_row = 36
ws3.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
ws3.cell(row=summary_row, column=1, value="TOTALS").font = Font(name="Calibri", bold=True, size=12, color=FOREST)
ws3.cell(row=summary_row, column=3, value=f"=SUM(C5:C34)").font = Font(bold=True, size=11, color=FOREST)
ws3.cell(row=summary_row, column=4, value=f"=SUM(D5:D34)").font = Font(bold=True, size=11, color=FOREST)
ws3.cell(row=summary_row, column=5, value=f"=SUM(E5:E34)").font = Font(bold=True, size=11, color=FOREST)
ws3.cell(row=summary_row, column=6, value=f"=SUM(F5:F34)").font = Font(bold=True, size=11, color=FOREST)
ws3.cell(row=summary_row, column=7, value=f"=SUM(G5:G34)").font = Font(bold=True, size=11, color=FOREST)
for col in range(1, 11):
    ws3.cell(row=summary_row, column=col).fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
    ws3.cell(row=summary_row, column=col).border = header_border

col_widths3 = [18, 28, 14, 10, 10, 10, 10, 16, 22, 28]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 4: Home Inventory
# ════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Home Inventory")
ws4.sheet_properties.tabColor = "68D391"

add_title_block(ws4, "🏡 HOME INVENTORY", "Track all household items — essential for insurance claims & moving!", 11)

headers4 = ["Room", "Item", "Category", "Purchase Date", "Purchase Price",
            "Current Value", "Serial #", "Warranty Expiry", "Insurance Claimed?", "Photo Ref", "Notes"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, len(headers4))

cat_dv = DataValidation(type="list", formula1='"Furniture,Electronics,Appliances,Decor,Kitchen,Clothing,Other,Jewelry,Sporting Goods,Tools"', allow_blank=True)
ws4.add_data_validation(cat_dv)

ins_dv = DataValidation(type="list", formula1='"Yes,No,Pending,N/A"', allow_blank=True)
ws4.add_data_validation(ins_dv)

inventory_items = [
    ("Living Room", "Sectional Sofa", "Furniture", "2024-03-15", 2499.99, 1800.00, "", "2029-03-15", "No", "IMG_001", "Gray fabric"),
    ("Living Room", "65\" Smart TV", "Electronics", "2024-06-20", 899.99, 650.00, "SN-TV88421", "2027-06-20", "No", "IMG_002", "Samsung QLED"),
    ("Living Room", "Coffee Table", "Furniture", "2023-11-01", 349.99, 280.00, "", "", "No", "IMG_003", "Walnut wood"),
    ("Living Room", "Floor Lamp", "Decor", "2024-01-10", 129.99, 100.00, "", "", "No", "IMG_004", "Arc style"),
    ("Living Room", "Bookshelf", "Furniture", "2023-08-15", 299.99, 240.00, "", "", "No", "IMG_005", "5-tier oak"),
    ("Living Room", "Sound Bar", "Electronics", "2025-01-05", 449.99, 400.00, "SN-SB90122", "2028-01-05", "No", "IMG_006", "Sonos Beam"),
    ("Living Room", "Area Rug 8x10", "Decor", "2024-05-01", 599.99, 450.00, "", "", "No", "IMG_007", "Persian style"),
    ("Kitchen", "Refrigerator", "Appliances", "2023-06-01", 1899.99, 1500.00, "SN-RF33201", "2028-06-01", "No", "IMG_008", "LG French Door"),
    ("Kitchen", "Dishwasher", "Appliances", "2023-06-01", 749.99, 550.00, "SN-DW44102", "2028-06-01", "No", "IMG_009", "Bosch"),
    ("Kitchen", "Stand Mixer", "Kitchen", "2024-02-14", 349.99, 300.00, "SN-MX55003", "2029-02-14", "No", "IMG_010", "KitchenAid"),
    ("Kitchen", "Espresso Machine", "Kitchen", "2025-03-01", 599.99, 550.00, "SN-ES66104", "2028-03-01", "No", "IMG_011", "Breville"),
    ("Kitchen", "Microwave", "Appliances", "2024-08-10", 249.99, 180.00, "SN-MW77205", "2027-08-10", "No", "IMG_012", "Over-the-range"),
    ("Kitchen", "Knife Set", "Kitchen", "2024-01-20", 199.99, 170.00, "", "2034-01-20", "No", "IMG_013", "Wüsthof 12-pc"),
    ("Kitchen", "Dining Table + 6 Chairs", "Furniture", "2023-09-15", 1299.99, 1000.00, "", "", "No", "IMG_014", "Farmhouse style"),
    ("Master Bedroom", "King Mattress", "Furniture", "2024-01-02", 1499.99, 1200.00, "", "2034-01-02", "No", "IMG_015", "Purple Hybrid"),
    ("Master Bedroom", "Bed Frame", "Furniture", "2024-01-02", 599.99, 480.00, "", "", "No", "IMG_016", "Upholstered"),
    ("Master Bedroom", "Dresser", "Furniture", "2023-07-20", 799.99, 600.00, "", "", "No", "IMG_017", "6-drawer oak"),
    ("Master Bedroom", "Nightstands (pair)", "Furniture", "2023-07-20", 399.99, 300.00, "", "", "No", "IMG_018", "Matching set"),
    ("Master Bedroom", "Ceiling Fan", "Other", "2023-05-01", 249.99, 180.00, "", "", "No", "IMG_019", "52-inch"),
    ("Home Office", "Standing Desk", "Furniture", "2024-09-01", 699.99, 560.00, "", "2029-09-01", "No", "IMG_020", "Electric adj."),
    ("Home Office", "Office Chair", "Furniture", "2024-09-01", 449.99, 360.00, "", "2036-09-01", "No", "IMG_021", "Herman Miller"),
    ("Home Office", "MacBook Pro 16\"", "Electronics", "2025-06-01", 2499.99, 2200.00, "SN-MB88306", "2028-06-01", "No", "IMG_022", "M4 Pro"),
    ("Home Office", "Monitor 27\"", "Electronics", "2024-09-15", 549.99, 440.00, "SN-MN99407", "2027-09-15", "No", "IMG_023", "4K IPS"),
    ("Home Office", "Printer", "Electronics", "2024-03-01", 299.99, 200.00, "SN-PR10508", "2027-03-01", "No", "IMG_024", "HP LaserJet"),
    ("Kids Room", "Bunk Bed", "Furniture", "2024-04-15", 899.99, 700.00, "", "", "No", "IMG_025", "Twin/Full"),
    ("Kids Room", "Desk", "Furniture", "2024-08-01", 249.99, 200.00, "", "", "No", "IMG_026", "With hutch"),
    ("Kids Room", "Tablet", "Electronics", "2025-12-25", 329.99, 300.00, "SN-TB11609", "2027-12-25", "No", "IMG_027", "iPad 10th gen"),
    ("Garage", "Lawn Mower", "Tools", "2024-04-01", 449.99, 350.00, "SN-LM12710", "2027-04-01", "No", "IMG_028", "Electric"),
    ("Garage", "Power Drill Set", "Tools", "2023-06-15", 199.99, 140.00, "SN-PD13811", "2026-06-15", "No", "IMG_029", "DeWalt 20V"),
    ("Garage", "Snow Blower", "Tools", "2023-10-01", 799.99, 600.00, "SN-SB14912", "2028-10-01", "No", "IMG_030", "2-stage"),
    ("Laundry", "Washer", "Appliances", "2023-06-01", 899.99, 680.00, "SN-WS15013", "2028-06-01", "No", "IMG_031", "Samsung front"),
    ("Laundry", "Dryer", "Appliances", "2023-06-01", 849.99, 640.00, "SN-DR16114", "2028-06-01", "No", "IMG_032", "Samsung"),
    ("Bathroom", "Towel Warmer", "Other", "2025-01-15", 149.99, 130.00, "", "2028-01-15", "No", "IMG_033", "Wall-mounted"),
    ("Guest Room", "Queen Bed Set", "Furniture", "2024-02-01", 999.99, 780.00, "", "", "No", "IMG_034", "Frame+mattress"),
    ("Patio", "Outdoor Dining Set", "Furniture", "2024-05-15", 1199.99, 900.00, "", "", "No", "IMG_035", "6-piece teak"),
]

for idx, item_data in enumerate(inventory_items):
    r = 5 + idx
    ws4.cell(row=r, column=1, value=item_data[0])
    ws4.cell(row=r, column=2, value=item_data[1])
    cat_cell = ws4.cell(row=r, column=3, value=item_data[2])
    cat_dv.add(cat_cell)
    if item_data[3]:
        dc = ws4.cell(row=r, column=4, value=datetime.strptime(item_data[3], "%Y-%m-%d").date())
        dc.number_format = "YYYY-MM-DD"
    ws4.cell(row=r, column=5, value=item_data[4]).number_format = '$#,##0.00'
    ws4.cell(row=r, column=6, value=item_data[5]).number_format = '$#,##0.00'
    ws4.cell(row=r, column=7, value=item_data[6])
    if item_data[7]:
        we = ws4.cell(row=r, column=8, value=datetime.strptime(item_data[7], "%Y-%m-%d").date())
        we.number_format = "YYYY-MM-DD"
    else:
        ws4.cell(row=r, column=8, value="")
    ins_cell = ws4.cell(row=r, column=9, value=item_data[8])
    ins_dv.add(ins_cell)
    ws4.cell(row=r, column=10, value=item_data[9])
    ws4.cell(row=r, column=11, value=item_data[10])

# Fill remaining rows (up to 100)
for r in range(5 + len(inventory_items), 105):
    for col in range(1, 12):
        ws4.cell(row=r, column=col, value="")
    cat_dv.add(ws4.cell(row=r, column=3))
    ins_dv.add(ws4.cell(row=r, column=9))

apply_alt_rows(ws4, 5, 104, len(headers4))

# Totals
total_row = 106
ws4.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
ws4.cell(row=total_row, column=1, value="TOTAL INVENTORY VALUE").font = Font(name="Calibri", bold=True, size=13, color=FOREST)
ws4.cell(row=total_row, column=5, value="=SUM(E5:E104)").number_format = '$#,##0.00'
ws4.cell(row=total_row, column=5).font = Font(bold=True, size=12, color=FOREST)
ws4.cell(row=total_row, column=6, value="=SUM(F5:F104)").number_format = '$#,##0.00'
ws4.cell(row=total_row, column=6).font = Font(bold=True, size=12, color=FOREST)
for col in range(1, 12):
    ws4.cell(row=total_row, column=col).fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
    ws4.cell(row=total_row, column=col).border = header_border

col_widths4 = [16, 26, 16, 14, 15, 15, 18, 16, 16, 12, 22]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 5: Home Maintenance
# ════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Home Maintenance")
ws5.sheet_properties.tabColor = WARM_TERRA

add_title_block(ws5, "🔧 HOME MAINTENANCE SCHEDULE", "Stay ahead of repairs — auto-calculates next due dates!", 10)

headers5 = ["Task", "Category", "Frequency", "Last Completed", "Next Due",
            "Est. Cost", "Professional?", "Contractor", "Phone", "Notes"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=4, column=i, value=h)
style_header_row(ws5, 4, len(headers5))

maint_cat_dv = DataValidation(type="list",
    formula1='"HVAC,Plumbing,Electrical,Exterior,Appliances,Safety,Seasonal,Interior,Landscaping"',
    allow_blank=True)
ws5.add_data_validation(maint_cat_dv)

maint_freq_dv = DataValidation(type="list",
    formula1='"Monthly,Quarterly,Semi-Annually,Annually,As Needed"',
    allow_blank=True)
ws5.add_data_validation(maint_freq_dv)

pro_dv = DataValidation(type="list", formula1='"Yes,No,Optional"', allow_blank=True)
ws5.add_data_validation(pro_dv)

maint_tasks = [
    ("Replace HVAC filters", "HVAC", "Quarterly", "2025-12-15", 25, "No", "", "", "Use MERV 11+"),
    ("Service furnace", "HVAC", "Annually", "2025-10-01", 150, "Yes", "ABC Heating", "555-0101", "Before winter"),
    ("Service A/C unit", "HVAC", "Annually", "2025-04-01", 150, "Yes", "ABC Heating", "555-0101", "Before summer"),
    ("Clean air ducts", "HVAC", "Annually", "2025-06-15", 400, "Yes", "DuctPro", "555-0102", "Every 3-5 yrs ideal"),
    ("Check thermostat batteries", "HVAC", "Semi-Annually", "2025-11-01", 10, "No", "", "", ""),
    ("Flush water heater", "Plumbing", "Annually", "2025-09-01", 0, "Optional", "Joe's Plumbing", "555-0103", "Remove sediment"),
    ("Check for leaks under sinks", "Plumbing", "Quarterly", "2025-12-01", 0, "No", "", "", "Kitchen + baths"),
    ("Clean gutters & downspouts", "Exterior", "Semi-Annually", "2025-11-15", 200, "Optional", "GutterGuard Co", "555-0104", "Spring + Fall"),
    ("Inspect roof", "Exterior", "Annually", "2025-09-01", 300, "Yes", "RoofRight LLC", "555-0105", "Check for damage"),
    ("Power wash siding & deck", "Exterior", "Annually", "2025-05-15", 250, "Optional", "", "", "Rent or hire"),
    ("Seal driveway", "Exterior", "Annually", "2025-08-01", 100, "No", "", "", "Every 2-3 years"),
    ("Touch up exterior paint", "Exterior", "Annually", "2025-06-01", 150, "No", "", "", "Check for peeling"),
    ("Inspect windows & caulking", "Exterior", "Annually", "2025-10-15", 50, "No", "", "", "Re-caulk as needed"),
    ("Clean dryer vent", "Appliances", "Quarterly", "2025-12-01", 0, "No", "", "", "Fire prevention"),
    ("Clean garbage disposal", "Appliances", "Monthly", "2026-01-15", 0, "No", "", "", "Ice + citrus method"),
    ("Descale water appliances", "Appliances", "Quarterly", "2025-12-15", 10, "No", "", "", "Coffee maker, kettle"),
    ("Clean refrigerator coils", "Appliances", "Semi-Annually", "2025-10-01", 0, "No", "", "", "Improve efficiency"),
    ("Test smoke detectors", "Safety", "Monthly", "2026-01-01", 0, "No", "", "", "Replace batteries yearly"),
    ("Test CO detectors", "Safety", "Monthly", "2026-01-01", 0, "No", "", "", "Replace unit every 7yrs"),
    ("Check fire extinguishers", "Safety", "Annually", "2025-07-01", 0, "No", "", "", "Check pressure gauge"),
    ("Inspect electrical panel", "Electrical", "Annually", "2025-08-01", 0, "Optional", "SparkSafe Electric", "555-0106", "Look for rust/damage"),
    ("Test GFCI outlets", "Electrical", "Monthly", "2026-01-15", 0, "No", "", "", "Press test button"),
    ("Check outdoor lighting", "Electrical", "Quarterly", "2025-12-01", 20, "No", "", "", "Replace bulbs"),
    ("Winterize outdoor faucets", "Seasonal", "Annually", "2025-10-15", 0, "No", "", "", "Before first freeze"),
    ("Fertilize lawn", "Seasonal", "Quarterly", "2025-10-01", 50, "No", "", "", "Spring, Summer, Fall"),
    ("Trim trees & bushes", "Landscaping", "Semi-Annually", "2025-11-01", 200, "Optional", "GreenThumb", "555-0107", "Spring + late summer"),
    ("Aerate lawn", "Landscaping", "Annually", "2025-09-15", 100, "Optional", "GreenThumb", "555-0107", "Fall is best"),
    ("Clean & store patio furniture", "Seasonal", "Annually", "2025-10-20", 0, "No", "", "", "Before winter"),
    ("Check attic insulation", "Interior", "Annually", "2025-09-01", 0, "No", "", "", "R-38 minimum"),
    ("Deep clean carpets", "Interior", "Semi-Annually", "2025-11-01", 200, "Optional", "FreshClean Pro", "555-0108", "Rent or hire"),
    ("Re-grout tile as needed", "Interior", "As Needed", "2025-06-01", 50, "No", "", "", "Kitchen + bath"),
    ("Lubricate garage door", "Exterior", "Semi-Annually", "2025-10-01", 10, "No", "", "", "Hinges + tracks"),
    ("Inspect sump pump", "Plumbing", "Quarterly", "2025-12-15", 0, "No", "", "", "Pour water to test"),
    ("Snake main drain", "Plumbing", "Annually", "2025-07-01", 150, "Yes", "Joe's Plumbing", "555-0103", "Preventive"),
    ("Inspect foundation", "Exterior", "Annually", "2025-05-01", 0, "No", "", "", "Look for cracks"),
    ("Seal grout lines", "Interior", "Annually", "2025-08-01", 30, "No", "", "", "Prevents staining"),
    ("Check washing machine hoses", "Appliances", "Semi-Annually", "2025-10-15", 20, "No", "", "", "Replace every 5yrs"),
    ("Service garage door opener", "Exterior", "Annually", "2025-09-01", 100, "Optional", "DoorTech", "555-0109", "Safety sensors"),
    ("Clean chimney/flue", "Interior", "Annually", "2025-09-01", 250, "Yes", "SweepMaster", "555-0110", "Before heating season"),
    ("Pest inspection", "Safety", "Annually", "2025-04-01", 150, "Yes", "BugFree Inc", "555-0111", "Spring is ideal"),
]

for idx, (task, cat, freq, last_done, cost, pro, contractor, phone, notes) in enumerate(maint_tasks):
    r = 5 + idx
    ws5.cell(row=r, column=1, value=task)
    cc = ws5.cell(row=r, column=2, value=cat)
    maint_cat_dv.add(cc)
    fc = ws5.cell(row=r, column=3, value=freq)
    maint_freq_dv.add(fc)
    ld = ws5.cell(row=r, column=4, value=datetime.strptime(last_done, "%Y-%m-%d").date())
    ld.number_format = "YYYY-MM-DD"
    # Next Due formula
    next_cell = ws5.cell(row=r, column=5)
    next_cell.value = f'=IF(D{r}="","",D{r}+CHOOSE(MATCH(C{r},{{"Monthly","Quarterly","Semi-Annually","Annually","As Needed"}},0),30,91,182,365,365))'
    next_cell.number_format = "YYYY-MM-DD"
    ws5.cell(row=r, column=6, value=cost).number_format = '$#,##0.00'
    pc = ws5.cell(row=r, column=7, value=pro)
    pro_dv.add(pc)
    ws5.cell(row=r, column=8, value=contractor)
    ws5.cell(row=r, column=9, value=phone)
    ws5.cell(row=r, column=10, value=notes)

apply_alt_rows(ws5, 5, 44, len(headers5))

# Cost total
total_row5 = 46
ws5.merge_cells(start_row=total_row5, start_column=1, end_row=total_row5, end_column=5)
ws5.cell(row=total_row5, column=1, value="ESTIMATED ANNUAL MAINTENANCE COST").font = Font(bold=True, size=13, color=FOREST)
ws5.cell(row=total_row5, column=6, value="=SUM(F5:F44)").number_format = '$#,##0.00'
ws5.cell(row=total_row5, column=6).font = Font(bold=True, size=12, color=FOREST)
for col in range(1, 11):
    ws5.cell(row=total_row5, column=col).fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
    ws5.cell(row=total_row5, column=col).border = header_border

col_widths5 = [32, 16, 16, 16, 14, 14, 14, 20, 14, 28]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 6: Household Budget
# ════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Household Budget")
ws6.sheet_properties.tabColor = FOREST

add_title_block(ws6, "💰 HOUSEHOLD BUDGET", "Track monthly household expenses — see exactly where your money goes!", 5)

headers6 = ["Category", "Monthly Budget", "Actual Spent", "Difference", "% of Budget"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=4, column=i, value=h)
style_header_row(ws6, 4, len(headers6))

budget_items = [
    ("🏠 Mortgage / Rent", 2200.00, 2200.00),
    ("💡 Electric", 150.00, 142.50),
    ("🔥 Gas / Heating", 80.00, 95.00),
    ("💧 Water & Sewer", 65.00, 58.00),
    ("📡 Internet & Cable", 120.00, 119.99),
    ("📱 Phone (Family Plan)", 180.00, 180.00),
    ("🛡️ Homeowner's Insurance", 175.00, 175.00),
    ("🏥 Health Insurance", 450.00, 450.00),
    ("🛒 Groceries", 800.00, 875.00),
    ("🍽️ Dining Out / Takeout", 250.00, 310.00),
    ("🧹 Cleaning Supplies", 50.00, 42.00),
    ("🔧 Home Maintenance / Repairs", 200.00, 150.00),
    ("⛽ Transportation / Gas", 300.00, 285.00),
    ("🚗 Car Insurance", 200.00, 200.00),
    ("👶 Childcare / Activities", 400.00, 425.00),
    ("🐾 Pet Care", 100.00, 85.00),
    ("🎓 Education / Subscriptions", 75.00, 79.99),
    ("💊 Medical / Pharmacy", 100.00, 65.00),
    ("🎉 Entertainment / Hobbies", 150.00, 175.00),
    ("💰 Savings / Emergency Fund", 500.00, 500.00),
]

for idx, (cat, budget, actual) in enumerate(budget_items):
    r = 5 + idx
    ws6.cell(row=r, column=1, value=cat)
    ws6.cell(row=r, column=2, value=budget).number_format = '$#,##0.00'
    ws6.cell(row=r, column=3, value=actual).number_format = '$#,##0.00'
    ws6.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = '$#,##0.00'
    ws6.cell(row=r, column=5, value=f'=IF(B{r}=0,"",C{r}/B{r})').number_format = '0.0%'

apply_alt_rows(ws6, 5, 24, len(headers6))

# Conditional formatting for difference (negative = over budget)
ws6.conditional_formatting.add('D5:D24',
    CellIsRule(operator='lessThan', formula=['0'],
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
              font=Font(color="C53030", bold=True)))
ws6.conditional_formatting.add('D5:D24',
    CellIsRule(operator='greaterThan', formula=['0'],
              fill=PatternFill(start_color=SOFT_SAGE, end_color=SOFT_SAGE, fill_type="solid"),
              font=Font(color=FOREST, bold=True)))

# Totals
total_r = 26
ws6.cell(row=total_r, column=1, value="MONTHLY TOTALS").font = Font(bold=True, size=13, color=FOREST)
ws6.cell(row=total_r, column=2, value="=SUM(B5:B24)").number_format = '$#,##0.00'
ws6.cell(row=total_r, column=2).font = Font(bold=True, size=12, color=FOREST)
ws6.cell(row=total_r, column=3, value="=SUM(C5:C24)").number_format = '$#,##0.00'
ws6.cell(row=total_r, column=3).font = Font(bold=True, size=12, color=FOREST)
ws6.cell(row=total_r, column=4, value="=B26-C26").number_format = '$#,##0.00'
ws6.cell(row=total_r, column=4).font = Font(bold=True, size=12, color=FOREST)
ws6.cell(row=total_r, column=5, value='=IF(B26=0,"",C26/B26)').number_format = '0.0%'
ws6.cell(row=total_r, column=5).font = Font(bold=True, size=12, color=FOREST)
for col in range(1, 6):
    ws6.cell(row=total_r, column=col).fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
    ws6.cell(row=total_r, column=col).border = header_border

# Annual projection
ws6.merge_cells(start_row=28, start_column=1, end_row=28, end_column=2)
ws6.cell(row=28, column=1, value="ANNUAL PROJECTION").font = Font(bold=True, size=12, color=WARM_TERRA)
ws6.cell(row=28, column=3, value="=C26*12").number_format = '$#,##0.00'
ws6.cell(row=28, column=3).font = Font(bold=True, size=12, color=WARM_TERRA)

col_widths6 = [32, 18, 18, 18, 14]
for i, w in enumerate(col_widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 7: Meal Planning & Groceries
# ════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Meal Planning & Groceries")
ws7.sheet_properties.tabColor = "68D391"

add_title_block(ws7, "🍽️ WEEKLY MEAL PLANNER & GROCERY LIST", "Plan meals for the week and auto-calculate your grocery budget!", 8)

# Meal grid
meal_headers = ["", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
for i, h in enumerate(meal_headers, 1):
    ws7.cell(row=4, column=i, value=h)
style_header_row(ws7, 4, len(meal_headers))

meals = ["Breakfast", "Snack AM", "Lunch", "Snack PM", "Dinner", "Dessert"]
sample_meals = {
    "Breakfast": ["Oatmeal & Fruit", "Eggs & Toast", "Smoothie Bowl", "Pancakes", "Yogurt Parfait", "Avocado Toast", "French Toast"],
    "Snack AM": ["Apple & PB", "Trail Mix", "Cheese Stick", "Granola Bar", "Banana", "Veggies & Hummus", "Muffin"],
    "Lunch": ["Chicken Salad", "Turkey Wrap", "Soup & Bread", "Grain Bowl", "Leftover Stir Fry", "BLT Sandwich", "Pizza Bagels"],
    "Snack PM": ["Crackers & Cheese", "Fruit Cup", "Popcorn", "Smoothie", "Rice Cakes", "Energy Balls", "Veggies"],
    "Dinner": ["Grilled Salmon", "Pasta Bolognese", "Chicken Stir Fry", "Taco Night", "Sheet Pan Chicken", "Homemade Pizza", "Pot Roast"],
    "Dessert": ["Fruit Salad", "", "Brownies", "", "Ice Cream", "Cookies", "Apple Crisp"],
}

for idx, meal in enumerate(meals):
    r = 5 + idx
    ws7.cell(row=r, column=1, value=meal).font = Font(bold=True, size=11, color=FOREST)
    ws7.cell(row=r, column=1).fill = PatternFill(start_color=SOFT_SAGE, end_color=SOFT_SAGE, fill_type="solid")
    for day in range(7):
        ws7.cell(row=r, column=2 + day, value=sample_meals[meal][day])

apply_alt_rows(ws7, 5, 10, 8)

# Grocery List Section
ws7.merge_cells(start_row=13, start_column=1, end_row=13, end_column=8)
ws7.cell(row=13, column=1, value="📝 GROCERY LIST").font = Font(bold=True, size=14, color=FOREST)
ws7.cell(row=13, column=1).fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
ws7.cell(row=13, column=1).alignment = Alignment(horizontal="center")

grocery_headers = ["Category", "Item", "Quantity", "Unit", "Est. Price", "Got It?", "Store", "Notes"]
for i, h in enumerate(grocery_headers, 1):
    ws7.cell(row=14, column=i, value=h)
style_header_row(ws7, 14, 8)

groc_cat_dv = DataValidation(type="list",
    formula1='"Produce,Dairy & Eggs,Meat & Seafood,Bakery,Pantry,Frozen,Beverages,Snacks,Household,Personal Care"',
    allow_blank=True)
ws7.add_data_validation(groc_cat_dv)

unit_dv = DataValidation(type="list",
    formula1='"each,lb,oz,bunch,bag,box,pack,can,bottle,gallon,dozen,container"',
    allow_blank=True)
ws7.add_data_validation(unit_dv)

got_dv = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
ws7.add_data_validation(got_dv)

groceries = [
    ("Produce", "Bananas", 1, "bunch", 1.29, "✓", "Grocery", "Organic"),
    ("Produce", "Baby Spinach", 1, "bag", 3.49, "✓", "Grocery", "For smoothies & salads"),
    ("Produce", "Avocados", 4, "each", 4.00, "", "Grocery", "Ripe"),
    ("Produce", "Tomatoes", 1, "lb", 2.99, "", "Grocery", ""),
    ("Produce", "Onions", 1, "bag", 3.49, "✓", "Grocery", "Yellow"),
    ("Produce", "Lemons", 4, "each", 2.00, "", "Grocery", ""),
    ("Produce", "Apples", 1, "bag", 4.99, "", "Grocery", "Honeycrisp"),
    ("Produce", "Bell Peppers", 3, "each", 3.00, "", "Grocery", "Mixed colors"),
    ("Dairy & Eggs", "Eggs (large)", 1, "dozen", 3.99, "✓", "Grocery", "Free range"),
    ("Dairy & Eggs", "Whole Milk", 1, "gallon", 4.29, "✓", "Grocery", "Organic"),
    ("Dairy & Eggs", "Greek Yogurt", 1, "container", 5.49, "", "Grocery", "Plain"),
    ("Dairy & Eggs", "Shredded Cheese", 1, "bag", 3.99, "", "Grocery", "Mexican blend"),
    ("Dairy & Eggs", "Butter", 1, "box", 4.99, "✓", "Grocery", "Unsalted"),
    ("Meat & Seafood", "Chicken Breast", 2, "lb", 9.98, "", "Grocery", "Boneless skinless"),
    ("Meat & Seafood", "Salmon Fillets", 1, "lb", 12.99, "", "Grocery", "Wild caught"),
    ("Meat & Seafood", "Ground Beef", 1, "lb", 6.99, "", "Grocery", "90% lean"),
    ("Meat & Seafood", "Turkey Deli Meat", 1, "pack", 5.49, "", "Grocery", "Low sodium"),
    ("Bakery", "Whole Wheat Bread", 1, "each", 3.99, "✓", "Grocery", ""),
    ("Bakery", "Tortillas", 1, "pack", 3.49, "", "Grocery", "Flour, large"),
    ("Bakery", "Pizza Dough", 1, "each", 2.99, "", "Grocery", "From bakery"),
    ("Pantry", "Pasta (Penne)", 1, "box", 1.79, "✓", "Grocery", ""),
    ("Pantry", "Marinara Sauce", 1, "bottle", 3.49, "", "Grocery", ""),
    ("Pantry", "Rice (Jasmine)", 1, "bag", 4.99, "", "Grocery", ""),
    ("Pantry", "Olive Oil", 1, "bottle", 8.99, "", "Grocery", "Extra virgin"),
    ("Pantry", "Oatmeal", 1, "container", 4.49, "", "Grocery", "Rolled oats"),
    ("Frozen", "Frozen Fruit Mix", 1, "bag", 4.99, "", "Grocery", "For smoothies"),
    ("Frozen", "Ice Cream", 1, "container", 5.99, "", "Grocery", "Vanilla"),
    ("Beverages", "Coffee Beans", 1, "bag", 12.99, "", "Grocery", "Medium roast"),
    ("Household", "Paper Towels", 1, "pack", 8.99, "", "Grocery", ""),
    ("Household", "Dish Soap", 1, "bottle", 3.49, "", "Grocery", ""),
]

for idx, (cat, item, qty, unit, price, got, store, notes) in enumerate(groceries):
    r = 15 + idx
    cc = ws7.cell(row=r, column=1, value=cat)
    groc_cat_dv.add(cc)
    ws7.cell(row=r, column=2, value=item)
    ws7.cell(row=r, column=3, value=qty)
    uc = ws7.cell(row=r, column=4, value=unit)
    unit_dv.add(uc)
    ws7.cell(row=r, column=5, value=price).number_format = '$#,##0.00'
    gc = ws7.cell(row=r, column=6, value=got)
    got_dv.add(gc)
    ws7.cell(row=r, column=7, value=store)
    ws7.cell(row=r, column=8, value=notes)

apply_alt_rows(ws7, 15, 44, 8)

# Grocery total
gtotal_r = 46
ws7.merge_cells(start_row=gtotal_r, start_column=1, end_row=gtotal_r, end_column=4)
ws7.cell(row=gtotal_r, column=1, value="ESTIMATED GROCERY TOTAL").font = Font(bold=True, size=13, color=FOREST)
ws7.cell(row=gtotal_r, column=5, value="=SUM(E15:E44)").number_format = '$#,##0.00'
ws7.cell(row=gtotal_r, column=5).font = Font(bold=True, size=13, color=FOREST)
for col in range(1, 9):
    ws7.cell(row=gtotal_r, column=col).fill = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
    ws7.cell(row=gtotal_r, column=col).border = header_border

col_widths7 = [18, 24, 10, 12, 14, 10, 12, 24]
for i, w in enumerate(col_widths7, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════
# TAB 8: How to Use
# ════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("How to Use")
ws8.sheet_properties.tabColor = DUSTY_PINK

ws8.merge_cells("A1:H1")
ws8.cell(row=1, column=1, value="🏠 HOME ORGANIZATION & CLEANING PLANNER 2026 — USER GUIDE").font = Font(bold=True, size=18, color=FOREST)
ws8.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws8.cell(row=1, column=1).fill = PatternFill(start_color=SOFT_SAGE, end_color=SOFT_SAGE, fill_type="solid")
ws8.row_dimensions[1].height = 50

sections = [
    ("", ""),
    ("📋 GETTING STARTED", ""),
    ("", "Welcome to your Home Organization & Cleaning Planner! This spreadsheet is designed to help you manage every aspect of your home — from daily cleaning to long-term maintenance."),
    ("", ""),
    ("🧹 TAB 1: WEEKLY CLEANING SCHEDULE", ""),
    ("", "• Each row = a room or area in your home (20 pre-loaded)"),
    ("", "• Use the dropdown (✓) to check off daily tasks Mon-Sun"),
    ("", "• The 'Weekly Deep Clean Task' column reminds you of bigger weekly jobs"),
    ("", "• Set Priority (High/Medium/Low) to focus on what matters most"),
    ("", "• TIP: Do a quick 15-min tidy each evening to stay ahead!"),
    ("", ""),
    ("🧽 TAB 2: DEEP CLEANING CHECKLIST", ""),
    ("", "• 60 pre-loaded deep cleaning tasks organized by room"),
    ("", "• 'Next Due' auto-calculates based on frequency + last done date"),
    ("", "• Use 'Assigned To' for family chore distribution"),
    ("", "• Status: Done / Due / Overdue / Scheduled"),
    ("", "• TIP: Tackle 2-3 deep cleaning tasks per week to stay on track"),
    ("", ""),
    ("📦 TAB 3: DECLUTTER TRACKER", ""),
    ("", "• Track Keep / Donate / Trash / Sell counts for each area"),
    ("", "• Before & After photos help motivate you — track if taken!"),
    ("", "• Totals auto-calculate at the bottom"),
    ("", "• TIP: Use the 'one in, one out' rule to maintain progress"),
    ("", ""),
    ("🏡 TAB 4: HOME INVENTORY", ""),
    ("", "• Essential for insurance claims! Document 100+ items"),
    ("", "• Track purchase price, current value, serial numbers, warranties"),
    ("", "• Total inventory value auto-calculates"),
    ("", "• TIP: Take photos of receipts and serial numbers for each item"),
    ("", ""),
    ("🔧 TAB 5: HOME MAINTENANCE", ""),
    ("", "• 40 pre-loaded maintenance tasks with auto-calculated due dates"),
    ("", "• Track contractor info, costs, and whether professional help is needed"),
    ("", "• Annual cost estimate at the bottom"),
    ("", "• TIP: Set calendar reminders for seasonal tasks"),
    ("", ""),
    ("💰 TAB 6: HOUSEHOLD BUDGET", ""),
    ("", "• 20 common household expense categories"),
    ("", "• Green = under budget, Red = over budget (auto-highlighted)"),
    ("", "• Annual projection included"),
    ("", "• TIP: Review weekly and adjust categories as needed"),
    ("", ""),
    ("🍽️ TAB 7: MEAL PLANNING & GROCERIES", ""),
    ("", "• Weekly meal grid with all meals + snacks"),
    ("", "• Grocery list with categories, quantities & estimated costs"),
    ("", "• Auto-total for grocery budget planning"),
    ("", "• TIP: Plan meals around what's on sale + what you already have"),
    ("", ""),
    ("💡 PRO CLEANING TIPS", ""),
    ("", "1. Clean top to bottom, left to right in every room"),
    ("", "2. Keep a cleaning caddy with essentials for quick access"),
    ("", "3. The '2-minute rule': if it takes less than 2 minutes, do it now"),
    ("", "4. Assign age-appropriate chores to all family members"),
    ("", "5. Declutter before deep cleaning — less stuff = easier cleaning"),
    ("", "6. Use baking soda + vinegar for natural cleaning solutions"),
    ("", "7. Microfiber cloths are your best friend — wash & reuse!"),
    ("", "8. Clean your cleaning tools monthly (vacuum filters, mop heads)"),
    ("", "9. Set a timer for 15 minutes — you'll be amazed what you can do"),
    ("", "10. Celebrate your wins! A clean home = a peaceful mind 🎉"),
    ("", ""),
    ("⚠️ IMPORTANT NOTES", ""),
    ("", "• This is a DIGITAL DOWNLOAD — no physical product will be shipped"),
    ("", "• Works in Excel, Google Sheets, and Numbers"),
    ("", "• Customize everything! Add/remove rows, change categories"),
    ("", "• Formulas are pre-built — just enter your data and watch the magic"),
    ("", "• For best results: save a backup copy before making changes"),
    ("", "• Print any tab for a physical reference (Letter or A4 size)"),
    ("", ""),
    ("❤️ THANK YOU!", ""),
    ("", "Thank you for purchasing the Home Organization & Cleaning Planner 2026!"),
    ("", "If you love it, please leave a ⭐⭐⭐⭐⭐ review — it helps so much!"),
    ("", "Questions? Message us through Etsy — we're happy to help!"),
]

for idx, (col_a, col_b) in enumerate(sections):
    r = 3 + idx
    if col_a:  # Section header
        ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws8.cell(row=r, column=1, value=col_a)
        c.font = Font(bold=True, size=14, color=FOREST)
        c.fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type="solid")
        ws8.row_dimensions[r].height = 30
    elif col_b:
        ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws8.cell(row=r, column=1, value=col_b)
        c.font = Font(size=11, color=DARK_TEXT)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws8.row_dimensions[r].height = 22

ws8.column_dimensions['A'].width = 100

# ═══════════════════ FINAL: Freeze panes & save ════════════════════════
ws1.freeze_panes = "B5"
ws2.freeze_panes = "A5"
ws3.freeze_panes = "A5"
ws4.freeze_panes = "A5"
ws5.freeze_panes = "A5"
ws6.freeze_panes = "A5"
ws7.freeze_panes = "A15"

wb.save(OUTPUT)
print(f"Spreadsheet saved to {OUTPUT}")
print(f"Tabs: {wb.sheetnames}")
