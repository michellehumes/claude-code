import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from copy import copy

# ── Color Palette ──
INDIGO = "4C51BF"
VIOLET = "805AD5"
TEAL = "319795"
CORAL = "F56565"
SUNSHINE = "ECC94B"
PALE_INDIGO = "E8E3F3"
ICE_BLUE = "EBF4FF"
DARK_TEXT = "1A202C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F7F7FB"
MED_GRAY = "CBD5E0"

# ── Reusable styles ──
header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=INDIGO, end_color=INDIGO, fill_type="solid")
sub_header_font = Font(name="Calibri", bold=True, size=11, color=DARK_TEXT)
sub_header_fill = PatternFill(start_color=PALE_INDIGO, end_color=PALE_INDIGO, fill_type="solid")
body_font = Font(name="Calibri", size=11, color=DARK_TEXT)
title_font = Font(name="Calibri", bold=True, size=16, color=INDIGO)
section_font = Font(name="Calibri", bold=True, size=13, color=VIOLET)
alt_fill_1 = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
alt_fill_2 = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
violet_fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
coral_fill = PatternFill(start_color=CORAL, end_color=CORAL, fill_type="solid")
sunshine_fill = PatternFill(start_color=SUNSHINE, end_color=SUNSHINE, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, fill=None):
    f = fill or header_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = center_align
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        fill = alt_fill_1 if (r - start_row) % 2 == 0 else alt_fill_2
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# TAB 1: Semester Dashboard
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Semester Dashboard"
ws.sheet_properties.tabColor = INDIGO

set_col_widths(ws, [3, 22, 20, 18, 18, 18, 18, 18, 3])

# Title
ws.merge_cells("B2:H2")
ws["B2"] = "STUDENT ACADEMIC PLANNER 2026"
ws["B2"].font = Font(name="Calibri", bold=True, size=20, color=INDIGO)
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B3:H3")
ws["B3"] = "Semester Dashboard"
ws["B3"].font = Font(name="Calibri", bold=True, size=14, color=VIOLET)
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

# Summary Cards Row
card_labels = ["Semester", "GPA Goal", "Current GPA", "Credits This Sem", "Total Credits"]
card_cols = [2, 3, 4, 5, 6]
for i, (lbl, col) in enumerate(zip(card_labels, card_cols)):
    cell = ws.cell(row=5, column=col, value=lbl)
    cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    fills = [header_fill, PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid"),
             PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid"),
             header_fill, PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")]
    cell.fill = fills[i]
    cell.alignment = center_align
    cell.border = thin_border

# Sample data row
sample_vals = ["Fall 2026", 3.75, "", 16, 48]
for i, (val, col) in enumerate(zip(sample_vals, card_cols)):
    cell = ws.cell(row=6, column=col, value=val)
    cell.font = Font(name="Calibri", bold=True, size=14, color=INDIGO)
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = PatternFill(start_color=PALE_INDIGO, end_color=PALE_INDIGO, fill_type="solid")

# Current GPA formula (average from Grade Calculator)
ws["D6"] = '=IF(COUNTA(\'Grade Calculator\'!B5:B12)>0, AVERAGE(\'Grade Calculator\'!K5:K12), 0)'
ws["D6"].number_format = '0.00'

# Classes Quick View
ws.merge_cells("B8:H8")
ws["B8"] = "CLASSES QUICK VIEW"
ws["B8"].font = section_font
ws["B8"].alignment = Alignment(horizontal="center")
ws["B8"].fill = PatternFill(start_color=PALE_INDIGO, end_color=PALE_INDIGO, fill_type="solid")

class_headers = ["#", "Class Name", "Course Code", "Credits", "Current Grade", "Letter", "Status"]
for i, h in enumerate(class_headers, 2):
    ws.cell(row=9, column=i, value=h)
style_header_row(ws, 9, 8)

classes_data = [
    [1, "Intro to Psychology", "PSY 101", 3, 92, "A-", "On Track"],
    [2, "Calculus II", "MATH 202", 4, 87, "B+", "On Track"],
    [3, "English Composition", "ENG 110", 3, 95, "A", "On Track"],
    [4, "Biology I", "BIO 150", 4, 78, "C+", "Needs Work"],
    [5, "Art History", "ART 201", 3, 88, "B+", "On Track"],
]
for r, row_data in enumerate(classes_data, 10):
    for c, val in enumerate(row_data, 2):
        ws.cell(row=r, column=c, value=val)
style_data_rows(ws, 10, 17, 8)

# Upcoming Deadlines
ws.merge_cells("B19:H19")
ws["B19"] = "UPCOMING DEADLINES"
ws["B19"].font = section_font
ws["B19"].alignment = Alignment(horizontal="center")
ws["B19"].fill = PatternFill(start_color=PALE_INDIGO, end_color=PALE_INDIGO, fill_type="solid")

dl_headers = ["#", "Assignment", "Class", "Due Date", "Days Left", "Priority", "Status"]
for i, h in enumerate(dl_headers, 2):
    ws.cell(row=20, column=i, value=h)
style_header_row(ws, 20, 8)

deadlines = [
    [1, "Essay Draft", "ENG 110", "2026-02-15", 6, "High", "In Progress"],
    [2, "Problem Set 5", "MATH 202", "2026-02-16", 7, "Med", "Not Started"],
    [3, "Lab Report", "BIO 150", "2026-02-18", 9, "High", "In Progress"],
    [4, "Chapter Reading", "PSY 101", "2026-02-20", 11, "Low", "Not Started"],
    [5, "Midterm Study", "ART 201", "2026-02-25", 16, "Urgent", "Not Started"],
]
for r, row_data in enumerate(deadlines, 21):
    for c, val in enumerate(row_data, 2):
        ws.cell(row=r, column=c, value=val)
style_data_rows(ws, 21, 28, 8)

# Conditional formatting for priority
red_font = Font(color=CORAL, bold=True)
ws.conditional_formatting.add("G21:G28",
    CellIsRule(operator="equal", formula=['"Urgent"'], font=red_font,
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")))

ws.row_dimensions[2].height = 35
ws.row_dimensions[3].height = 25

# ═══════════════════════════════════════════════
# TAB 2: Class Schedule
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("Class Schedule")
ws2.sheet_properties.tabColor = VIOLET

headers2 = ["Class Name", "Course Code", "Professor", "Room", "Days", "Start Time",
            "End Time", "Credits", "Office Hours", "TA Contact", "Notes"]
set_col_widths(ws2, [22, 14, 18, 12, 12, 12, 12, 10, 18, 22, 20])

ws2.merge_cells("A1:K1")
ws2["A1"] = "CLASS SCHEDULE — FALL 2026"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

for i, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=i, value=h)
style_header_row(ws2, 2, 11)

# Days dropdown
days_dv = DataValidation(type="list", formula1='"MWF,TTh,MW,Daily,Online"', allow_blank=True)
days_dv.prompt = "Select days"
ws2.add_data_validation(days_dv)

schedule_data = [
    ["Intro to Psychology", "PSY 101", "Dr. Martinez", "Hall 204", "MWF", "9:00 AM", "9:50 AM", 3, "Tue 2-4 PM", "jsmith@univ.edu", "Bring textbook"],
    ["Calculus II", "MATH 202", "Prof. Chen", "Sci 301", "TTh", "10:30 AM", "11:45 AM", 4, "Mon 3-5 PM", "kwong@univ.edu", "Graphing calc req."],
    ["English Composition", "ENG 110", "Dr. Williams", "Lib 105", "MWF", "1:00 PM", "1:50 PM", 3, "Wed 1-3 PM", "alee@univ.edu", ""],
    ["Biology I", "BIO 150", "Prof. Patel", "Bio 210", "TTh", "2:00 PM", "3:15 PM", 4, "Fri 10-12 PM", "mross@univ.edu", "Lab Thursdays"],
    ["Art History", "ART 201", "Dr. Nakamura", "Art 102", "MW", "4:00 PM", "5:15 PM", 3, "Thu 1-3 PM", "", "Museum visit Feb"],
]
for r, row_data in enumerate(schedule_data, 3):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    days_dv.add(ws2.cell(row=r, column=5))

style_data_rows(ws2, 3, 10, 11)
for r in range(8, 11):
    days_dv.add(ws2.cell(row=r, column=5))

# ═══════════════════════════════════════════════
# TAB 3: Assignment Tracker
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("Assignment Tracker")
ws3.sheet_properties.tabColor = TEAL

headers3 = ["Class", "Assignment Name", "Type", "Date Assigned", "Due Date",
            "Days Left", "Weight %", "Grade", "Status", "Priority", "Notes"]
set_col_widths(ws3, [20, 24, 14, 14, 14, 12, 10, 10, 14, 12, 22])

ws3.merge_cells("A1:K1")
ws3["A1"] = "ASSIGNMENT TRACKER"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

for i, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=i, value=h)
style_header_row(ws3, 2, 11)

# Dropdowns
type_dv = DataValidation(type="list", formula1='"Homework,Essay,Project,Lab,Quiz,Exam,Presentation,Other"', allow_blank=True)
status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Submitted,Graded"', allow_blank=True)
priority_dv = DataValidation(type="list", formula1='"Low,Med,High,Urgent"', allow_blank=True)
ws3.add_data_validation(type_dv)
ws3.add_data_validation(status_dv)
ws3.add_data_validation(priority_dv)

assignments = [
    ["PSY 101", "Chapter 3 Reading", "Homework", "2026-02-01", "2026-02-08", None, 5, 95, "Graded", "Low", ""],
    ["MATH 202", "Problem Set 4", "Homework", "2026-02-03", "2026-02-10", None, 5, 88, "Graded", "Med", "Show work"],
    ["ENG 110", "Persuasive Essay Draft", "Essay", "2026-02-05", "2026-02-15", None, 15, "", "In Progress", "High", "MLA format"],
    ["BIO 150", "Lab Report: Osmosis", "Lab", "2026-02-06", "2026-02-18", None, 10, "", "In Progress", "High", "Include graphs"],
    ["ART 201", "Museum Response Paper", "Essay", "2026-02-07", "2026-02-20", None, 10, "", "Not Started", "Med", "Visit art museum"],
    ["MATH 202", "Problem Set 5", "Homework", "2026-02-10", "2026-02-16", None, 5, "", "Not Started", "Med", "Ch 7-8"],
    ["PSY 101", "Research Proposal", "Project", "2026-02-10", "2026-03-01", None, 20, "", "Not Started", "High", "APA format"],
    ["BIO 150", "Midterm Exam", "Exam", "2026-02-01", "2026-02-25", None, 25, "", "Not Started", "Urgent", "Ch 1-8"],
    ["ENG 110", "Peer Review", "Other", "2026-02-15", "2026-02-22", None, 5, "", "Not Started", "Low", "Review 2 essays"],
    ["ART 201", "Midterm Presentation", "Presentation", "2026-02-10", "2026-02-28", None, 20, "", "Not Started", "High", "10 min + slides"],
]

for r, row_data in enumerate(assignments, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if c == 4 or c == 5:
            cell.number_format = 'YYYY-MM-DD'

# Days Left formula
for r in range(3, 103):
    cell = ws3.cell(row=r, column=6)
    cell.value = f'=IF(E{r}="","",MAX(0,E{r}-TODAY()))'
    cell.number_format = '0'
    type_dv.add(ws3.cell(row=r, column=3))
    status_dv.add(ws3.cell(row=r, column=9))
    priority_dv.add(ws3.cell(row=r, column=10))

style_data_rows(ws3, 3, 102, 11)

# Conditional formatting for days left
ws3.conditional_formatting.add("F3:F102",
    CellIsRule(operator="lessThanOrEqual", formula=["3"],
              font=Font(color=WHITE, bold=True),
              fill=PatternFill(start_color=CORAL, end_color=CORAL, fill_type="solid")))
ws3.conditional_formatting.add("F3:F102",
    CellIsRule(operator="between", formula=["4", "7"],
              font=Font(color=DARK_TEXT, bold=True),
              fill=PatternFill(start_color=SUNSHINE, end_color=SUNSHINE, fill_type="solid")))

# ═══════════════════════════════════════════════
# TAB 4: Grade Calculator
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet("Grade Calculator")
ws4.sheet_properties.tabColor = CORAL

set_col_widths(ws4, [4, 22, 14, 10, 14, 10, 14, 10, 14, 10, 14, 10, 16, 12, 10])

ws4.merge_cells("A1:O1")
ws4["A1"] = "GRADE CALCULATOR"
ws4["A1"].font = title_font
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

# Sub-header row
ws4.merge_cells("A3:A4")
ws4["A3"] = "#"
ws4["B3"] = "Class Name"
ws4.merge_cells("B3:B4")

categories = ["Homework", "Quizzes", "Midterm", "Final", "Projects", "Participation"]
weights = [20, 10, 20, 25, 20, 5]

col = 3
for cat, w in zip(categories, weights):
    ws4.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws4.cell(row=3, column=col, value=f"{cat} ({w}%)")
    ws4.cell(row=4, column=col, value="Grade")
    ws4.cell(row=4, column=col+1, value="Wtd")
    col += 2

ws4.cell(row=3, column=15, value="Weighted Avg")
ws4.merge_cells("O3:O4")
ws4["O3"] = "Letter"
ws4.merge_cells("O3:O4")

# Actually let me redo the header more carefully
# Columns: A=#, B=Class, C=HW Grade, D=HW Wtd, E=Quiz Grade, F=Quiz Wtd, 
# G=Mid Grade, H=Mid Wtd, I=Final Grade, J=Final Wtd, K=Proj Grade, L=Proj Wtd,
# M=Part Grade, N=Part Wtd, O=Weighted Avg, P=Letter, Q=GPA

set_col_widths(ws4, [4, 22, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 14, 10, 10])

ws4.merge_cells("A1:Q1")
ws4["A1"] = "GRADE CALCULATOR"
ws4["A1"].font = title_font
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Row 3: category names merged across 2 cols
cats_info = [
    ("Homework", 20, 3), ("Quizzes", 10, 5), ("Midterm", 20, 7),
    ("Final", 25, 9), ("Projects", 20, 11), ("Participation", 5, 13)
]
for cat_name, wt, start_col in cats_info:
    ws4.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col+1)
    cell = ws4.cell(row=3, column=start_col, value=f"{cat_name} ({wt}%)")
    cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    cell.fill = violet_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws4.cell(row=3, column=start_col+1).fill = violet_fill
    ws4.cell(row=3, column=start_col+1).border = thin_border

ws4.cell(row=3, column=1, value="#")
ws4.cell(row=3, column=2, value="Class Name")
ws4.cell(row=3, column=15, value="Wtd Avg")
ws4.cell(row=3, column=16, value="Letter")
ws4.cell(row=3, column=17, value="GPA Pts")

for c in [1, 2, 15, 16, 17]:
    ws4.merge_cells(start_row=3, start_column=c, end_row=4, end_column=c)
    cell = ws4.cell(row=3, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Row 4: Grade / Wtd sub-headers
for _, _, start_col in cats_info:
    for dc, lbl in [(0, "Grade"), (1, "Wtd")]:
        cell = ws4.cell(row=4, column=start_col+dc, value=lbl)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border

# Classes data
class_names = ["Intro to Psychology", "Calculus II", "English Composition", "Biology I",
               "Art History", "", "", ""]
sample_grades = [
    [92, 88, 85, 0, 90, 95],
    [85, 82, 78, 0, 88, 90],
    [95, 90, 88, 0, 92, 98],
    [78, 72, 0, 0, 80, 85],
    [88, 85, 0, 0, 90, 92],
]

for r_idx in range(8):
    row = r_idx + 5
    ws4.cell(row=row, column=1, value=r_idx+1)
    ws4.cell(row=row, column=2, value=class_names[r_idx] if r_idx < len(class_names) else "")

    if r_idx < len(sample_grades):
        for g_idx, (_, wt, start_col) in enumerate(cats_info):
            grade_val = sample_grades[r_idx][g_idx]
            if grade_val > 0:
                ws4.cell(row=row, column=start_col, value=grade_val)
            # Weighted formula
            grade_col = get_column_letter(start_col)
            ws4.cell(row=row, column=start_col+1).value = \
                f'=IF({grade_col}{row}="",0,{grade_col}{row}*{wt/100})'
            ws4.cell(row=row, column=start_col+1).number_format = '0.00'

    # Formulas for empty rows too
    for g_idx, (_, wt, start_col) in enumerate(cats_info):
        grade_col = get_column_letter(start_col)
        if ws4.cell(row=row, column=start_col+1).value is None:
            ws4.cell(row=row, column=start_col+1).value = \
                f'=IF({grade_col}{row}="",0,{grade_col}{row}*{wt/100})'
            ws4.cell(row=row, column=start_col+1).number_format = '0.00'

    # Weighted Average = sum of all weighted columns
    wtd_cols = [get_column_letter(sc+1) for _, _, sc in cats_info]
    wtd_formula = "+".join([f'{c}{row}' for c in wtd_cols])
    ws4.cell(row=row, column=15).value = f'=IF(B{row}="","",{wtd_formula})'
    ws4.cell(row=row, column=15).number_format = '0.00'

    # Letter grade
    ws4.cell(row=row, column=16).value = (
        f'=IF(O{row}="","",IF(O{row}>=93,"A",IF(O{row}>=90,"A-",'
        f'IF(O{row}>=87,"B+",IF(O{row}>=83,"B",IF(O{row}>=80,"B-",'
        f'IF(O{row}>=77,"C+",IF(O{row}>=73,"C",IF(O{row}>=70,"C-",'
        f'IF(O{row}>=67,"D+",IF(O{row}>=60,"D","F")))))))))))'
    )

    # GPA points
    ws4.cell(row=row, column=17).value = (
        f'=IF(P{row}="","",IF(P{row}="A",4.0,IF(P{row}="A-",3.7,'
        f'IF(P{row}="B+",3.3,IF(P{row}="B",3.0,IF(P{row}="B-",2.7,'
        f'IF(P{row}="C+",2.3,IF(P{row}="C",2.0,IF(P{row}="C-",1.7,'
        f'IF(P{row}="D+",1.3,IF(P{row}="D",1.0,0)))))))))))'
    )
    ws4.cell(row=row, column=17).number_format = '0.00'

style_data_rows(ws4, 5, 12, 17)

# Semester GPA
ws4.merge_cells("A14:B14")
ws4["A14"] = "SEMESTER GPA"
ws4["A14"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
ws4["A14"].fill = PatternFill(start_color=INDIGO, end_color=INDIGO, fill_type="solid")
ws4["A14"].alignment = center_align
ws4["C14"] = '=IF(COUNT(Q5:Q12)>0,AVERAGE(Q5:Q12),0)'
ws4["C14"].font = Font(name="Calibri", bold=True, size=16, color=INDIGO)
ws4["C14"].number_format = '0.00'
ws4["C14"].alignment = center_align

# ═══════════════════════════════════════════════
# TAB 5: Study Planner
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet("Study Planner")
ws5.sheet_properties.tabColor = SUNSHINE

set_col_widths(ws5, [14, 16, 20, 24, 12, 16, 12])

ws5.merge_cells("A1:G1")
ws5["A1"] = "WEEKLY STUDY PLANNER"
ws5["A1"].font = title_font
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 40

headers5 = ["Day", "Time Block", "Class/Subject", "Study Task", "Duration (hrs)", "Location", "Completed"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=2, column=i, value=h)
style_header_row(ws5, 2, 7)

# Dropdowns
day_dv = DataValidation(type="list", formula1='"Monday,Tuesday,Wednesday,Thursday,Friday,Saturday,Sunday"', allow_blank=True)
block_dv = DataValidation(type="list", formula1='"Morning,Afternoon,Evening"', allow_blank=True)
yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws5.add_data_validation(day_dv)
ws5.add_data_validation(block_dv)
ws5.add_data_validation(yn_dv)

study_data = [
    ["Monday", "Morning", "MATH 202", "Review Ch 7 problems", 1.5, "Library 2nd Floor", "Yes"],
    ["Monday", "Afternoon", "PSY 101", "Read Chapter 4", 1, "Dorm Room", "Yes"],
    ["Monday", "Evening", "ENG 110", "Write essay outline", 2, "Coffee Shop", "No"],
    ["Tuesday", "Morning", "BIO 150", "Review lab procedures", 1, "Science Lab", "No"],
    ["Tuesday", "Afternoon", "ART 201", "Museum notes review", 1.5, "Art Building", "No"],
    ["Tuesday", "Evening", "MATH 202", "Practice integrals", 2, "Library 2nd Floor", "No"],
    ["Wednesday", "Morning", "PSY 101", "Flashcards Ch 3-4", 1, "Student Center", "No"],
    ["Wednesday", "Afternoon", "ENG 110", "Draft body paragraphs", 2.5, "Library 3rd Floor", "No"],
    ["Wednesday", "Evening", "BIO 150", "Study guide Ch 5", 1.5, "Dorm Room", "No"],
    ["Thursday", "Morning", "ART 201", "Slide analysis notes", 1, "Art Building", "No"],
    ["Thursday", "Afternoon", "MATH 202", "Problem Set 5", 2, "Math Tutoring Lab", "No"],
    ["Thursday", "Evening", "PSY 101", "Research proposal draft", 2, "Library 2nd Floor", "No"],
    ["Friday", "Morning", "BIO 150", "Lab report writing", 2, "Science Library", "No"],
    ["Friday", "Afternoon", "ENG 110", "Peer review prep", 1, "Dorm Room", "No"],
    ["Saturday", "Morning", "MATH 202", "Review weak topics", 2, "Coffee Shop", "No"],
    ["Saturday", "Afternoon", "BIO 150", "Midterm study group", 3, "Library Group Room", "No"],
    ["Sunday", "Afternoon", "All Classes", "Week ahead planning", 1, "Dorm Room", "No"],
    ["Sunday", "Evening", "PSY 101", "Quizlet review", 1, "Dorm Room", "No"],
]

for r, row_data in enumerate(study_data, 3):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)

for r in range(3, 45):
    day_dv.add(ws5.cell(row=r, column=1))
    block_dv.add(ws5.cell(row=r, column=2))
    yn_dv.add(ws5.cell(row=r, column=7))

style_data_rows(ws5, 3, 44, 7)

# Exam Schedule Section
ws5.merge_cells("A46:G46")
ws5["A46"] = "EXAM SCHEDULE"
ws5["A46"].font = section_font
ws5["A46"].alignment = Alignment(horizontal="center")
ws5["A46"].fill = PatternFill(start_color=PALE_INDIGO, end_color=PALE_INDIGO, fill_type="solid")

exam_headers = ["Class", "Exam Type", "Date", "Time", "Location", "Study Plan", "Materials Needed"]
for i, h in enumerate(exam_headers, 1):
    ws5.cell(row=47, column=i, value=h)
style_header_row(ws5, 47, 7)

exams = [
    ["BIO 150", "Midterm", "2026-02-25", "2:00 PM", "Bio 210", "Study guide + flashcards", "Calculator, pencils"],
    ["PSY 101", "Midterm", "2026-03-05", "9:00 AM", "Hall 204", "Review notes Ch 1-6", "Blue book, pen"],
    ["MATH 202", "Midterm", "2026-03-10", "10:30 AM", "Sci 301", "Practice problems", "Graphing calculator"],
    ["ENG 110", "Midterm Essay", "2026-03-12", "1:00 PM", "Lib 105", "Review essay structure", "Laptop"],
    ["ART 201", "Midterm", "2026-03-15", "4:00 PM", "Art 102", "Memorize artworks", "Printed study guide"],
]
for r, row_data in enumerate(exams, 48):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)
style_data_rows(ws5, 48, 55, 7)

# ═══════════════════════════════════════════════
# TAB 6: Expense Tracker
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet("Expense Tracker")
ws6.sheet_properties.tabColor = TEAL

set_col_widths(ws6, [14, 28, 18, 12, 16, 22])

ws6.merge_cells("A1:F1")
ws6["A1"] = "STUDENT EXPENSE TRACKER"
ws6["A1"].font = title_font
ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 40

headers6 = ["Date", "Description", "Category", "Amount", "Payment Method", "Notes"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=2, column=i, value=h)
style_header_row(ws6, 2, 6)

cat_dv = DataValidation(type="list",
    formula1='"Textbooks,Supplies,Food,Transportation,Housing,Entertainment,Subscriptions,Other"',
    allow_blank=True)
ws6.add_data_validation(cat_dv)

expenses = [
    ["2026-02-01", "Psychology Textbook (used)", "Textbooks", 45.99, "Debit Card", "Found on eBay"],
    ["2026-02-01", "Notebook + pens bundle", "Supplies", 12.50, "Cash", ""],
    ["2026-02-02", "Campus meal plan top-up", "Food", 150.00, "Debit Card", "February allocation"],
    ["2026-02-03", "Bus pass (monthly)", "Transportation", 35.00, "Debit Card", "Student discount"],
    ["2026-02-05", "Graphing calculator", "Supplies", 89.99, "Credit Card", "For MATH 202"],
    ["2026-02-07", "Coffee shop study snacks", "Food", 8.75, "Cash", ""],
    ["2026-02-08", "Movie night with friends", "Entertainment", 14.00, "Venmo", ""],
    ["2026-02-09", "Spotify Student", "Subscriptions", 5.99, "Credit Card", "Monthly"],
    ["2026-02-10", "Lab coat for Bio", "Supplies", 22.00, "Debit Card", ""],
    ["2026-02-12", "Uber to museum", "Transportation", 12.50, "Venmo", "ART 201 field trip"],
]

for r, row_data in enumerate(expenses, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws6.cell(row=r, column=c, value=val)
        if c == 4:
            cell.number_format = '$#,##0.00'

for r in range(3, 103):
    cat_dv.add(ws6.cell(row=r, column=3))

style_data_rows(ws6, 3, 102, 6)

# Monthly / Semester totals
ws6.merge_cells("A104:C104")
ws6["A104"] = "MONTHLY TOTAL"
ws6["A104"].font = Font(name="Calibri", bold=True, size=12, color=WHITE)
ws6["A104"].fill = violet_fill
ws6["A104"].alignment = center_align
ws6["D104"] = '=SUM(D3:D102)'
ws6["D104"].font = Font(name="Calibri", bold=True, size=13, color=VIOLET)
ws6["D104"].number_format = '$#,##0.00'
ws6["D104"].alignment = center_align

ws6.merge_cells("A105:C105")
ws6["A105"] = "SEMESTER TOTAL"
ws6["A105"].font = Font(name="Calibri", bold=True, size=12, color=WHITE)
ws6["A105"].fill = header_fill
ws6["A105"].alignment = center_align
ws6["D105"] = '=D104'
ws6["D105"].font = Font(name="Calibri", bold=True, size=13, color=INDIGO)
ws6["D105"].number_format = '$#,##0.00'
ws6["D105"].alignment = center_align

# ═══════════════════════════════════════════════
# TAB 7: Semester Goals
# ═══════════════════════════════════════════════
ws7 = wb.create_sheet("Semester Goals")
ws7.sheet_properties.tabColor = CORAL

set_col_widths(ws7, [28, 14, 14, 14, 28, 28])

ws7.merge_cells("A1:F1")
ws7["A1"] = "SEMESTER GOALS"
ws7["A1"].font = title_font
ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 40

headers7 = ["Goal", "Category", "Deadline", "Status", "Action Steps", "Progress Notes"]
for i, h in enumerate(headers7, 1):
    ws7.cell(row=2, column=i, value=h)
style_header_row(ws7, 2, 6)

goal_cat_dv = DataValidation(type="list",
    formula1='"Academic,Career,Health,Social,Financial,Personal"', allow_blank=True)
goal_status_dv = DataValidation(type="list",
    formula1='"Not Started,In Progress,On Track,At Risk,Completed"', allow_blank=True)
ws7.add_data_validation(goal_cat_dv)
ws7.add_data_validation(goal_status_dv)

goals = [
    ["Achieve 3.5+ semester GPA", "Academic", "2026-05-15", "In Progress", "Attend all classes, use tutoring", "Currently at 3.4"],
    ["Land summer internship", "Career", "2026-04-01", "In Progress", "Apply to 10+ companies, update resume", "3 applications sent"],
    ["Exercise 3x per week", "Health", "2026-05-15", "On Track", "Gym MWF mornings, join intramural team", "Week 3 consistent"],
    ["Join 2 student clubs", "Social", "2026-03-01", "In Progress", "Research clubs, attend info sessions", "Joined debate club"],
    ["Save $500 by semester end", "Financial", "2026-05-15", "In Progress", "Budget $50/wk spending, tutoring job", "Saved $120 so far"],
    ["Read 5 books for pleasure", "Personal", "2026-05-15", "Not Started", "30 min reading before bed", "Book list created"],
    ["Improve public speaking", "Career", "2026-05-15", "In Progress", "Toastmasters, class presentations", "First speech done"],
    ["Maintain healthy sleep schedule", "Health", "2026-05-15", "At Risk", "Bed by 11 PM, no screens after 10", "Struggling on weekends"],
]

for r, row_data in enumerate(goals, 3):
    for c, val in enumerate(row_data, 1):
        ws7.cell(row=r, column=c, value=val)

for r in range(3, 23):
    goal_cat_dv.add(ws7.cell(row=r, column=2))
    goal_status_dv.add(ws7.cell(row=r, column=4))

style_data_rows(ws7, 3, 22, 6)

# Conditional formatting for goal status
ws7.conditional_formatting.add("D3:D22",
    CellIsRule(operator="equal", formula=['"Completed"'],
              font=Font(color=WHITE, bold=True),
              fill=PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")))
ws7.conditional_formatting.add("D3:D22",
    CellIsRule(operator="equal", formula=['"At Risk"'],
              font=Font(color=WHITE, bold=True),
              fill=PatternFill(start_color=CORAL, end_color=CORAL, fill_type="solid")))

# ═══════════════════════════════════════════════
# TAB 8: How to Use
# ═══════════════════════════════════════════════
ws8 = wb.create_sheet("How to Use")
ws8.sheet_properties.tabColor = PALE_INDIGO

set_col_widths(ws8, [5, 60, 5])

ws8.merge_cells("A1:C1")
ws8["A1"] = "HOW TO USE YOUR STUDENT ACADEMIC PLANNER"
ws8["A1"].font = title_font
ws8["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws8.row_dimensions[1].height = 45

instructions = [
    ("GETTING STARTED", [
        "1. Start with the CLASS SCHEDULE tab — enter all your classes for the semester.",
        "2. Fill in the SEMESTER DASHBOARD with your semester name, GPA goal, and credits.",
        "3. The Current GPA will auto-calculate from your Grade Calculator entries.",
    ]),
    ("TRACKING ASSIGNMENTS", [
        "1. Use the ASSIGNMENT TRACKER for every assignment, quiz, and exam.",
        "2. The 'Days Left' column auto-calculates — watch for red (urgent!) items.",
        "3. Use dropdown menus for Type, Status, and Priority for easy filtering.",
        "4. Update the Status as you progress through each assignment.",
    ]),
    ("CALCULATING GRADES", [
        "1. Enter your grades in the GRADE CALCULATOR as you receive them.",
        "2. Weighted averages calculate automatically based on standard weights.",
        "3. Letter grades and GPA points update in real-time.",
        "4. Your Semester GPA is shown at the bottom of the calculator.",
    ]),
    ("STUDY PLANNING", [
        "1. Plan your week in the STUDY PLANNER — block out dedicated study time.",
        "2. Use the Exam Schedule section to prepare well in advance.",
        "3. Mark tasks as completed to track your consistency.",
    ]),
    ("MANAGING EXPENSES", [
        "1. Log every purchase in the EXPENSE TRACKER.",
        "2. Use categories to see where your money goes.",
        "3. Monthly and semester totals help you stay on budget.",
    ]),
    ("STUDY TIPS FOR SUCCESS", [
        "Use the Pomodoro Technique: 25 min study, 5 min break.",
        "Review notes within 24 hours of each class.",
        "Form study groups for difficult subjects.",
        "Visit office hours at least once per month.",
        "Start assignments the day they're assigned, even if just an outline.",
        "Sleep 7-8 hours — your brain consolidates learning during sleep!",
        "Use active recall (flashcards) over passive re-reading.",
        "Take breaks between subjects to avoid mental fatigue.",
    ]),
]

row = 3
for section_title, items in instructions:
    ws8.cell(row=row, column=2, value=section_title)
    ws8.cell(row=row, column=2).font = section_font
    ws8.cell(row=row, column=2).fill = PatternFill(start_color=PALE_INDIGO, end_color=PALE_INDIGO, fill_type="solid")
    ws8.row_dimensions[row].height = 28
    row += 1
    for item in items:
        ws8.cell(row=row, column=2, value=item)
        ws8.cell(row=row, column=2).font = body_font
        ws8.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        ws8.row_dimensions[row].height = 22
        row += 1
    row += 1

# Freeze panes on key sheets
ws.freeze_panes = "A5"
ws2.freeze_panes = "A3"
ws3.freeze_panes = "A3"
ws4.freeze_panes = "A5"
ws5.freeze_panes = "A3"
ws6.freeze_panes = "A3"
ws7.freeze_panes = "A3"

# Save
output_path = "/home/user/claude-code/etsy-digital-download-9/product/Student_Academic_Planner_2026.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to {output_path}")
print(f"Sheets: {wb.sheetnames}")
