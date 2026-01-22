#!/usr/bin/env python3
"""Create a Finance Tracker Excel file with sample data and formulas."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import FormulaRule

# Create workbook
wb = Workbook()

# ============ TRANSACTIONS SHEET ============
ws_trans = wb.active
ws_trans.title = "Transactions"

# Headers
headers = ["Date", "Merchant", "Category", "Account", "Amount", "Owner", "Notes"]
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, header in enumerate(headers, 1):
    cell = ws_trans.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Sample transaction data
transactions = [
    ("2026-01-09", "Spotify", "Music", "Venture Card", -11.99, "Shared", ""),
    ("2026-01-09", "Geico", "Car Insurance", "Venture Card", -152.51, "Shared", ""),
    ("2026-01-09", "BMW", "Car Payment", "Checking", -751.00, "Shared", ""),
    ("2026-01-09", "Target", "Shopping", "Target Card", -130.36, "Shared", ""),
    ("2026-01-09", "OpenAI", "Subscriptions", "Freedom Card", -217.75, "Shared", ""),
    ("2026-01-08", "Sagtown Coffee", "Delivery & Takeout", "Amex Gold", -26.14, "Shared", ""),
    ("2026-01-08", "E Hampton", "Restaurants & Bars", "Amex Platinum", -169.65, "Shared", ""),
    ("2026-01-07", "Stop & Shop", "Groceries", "Amex Gold", -20.04, "Shared", ""),
    ("2026-01-07", "Addepar", "Paychecks", "Checking", 5005.08, "Gray", "Salary"),
    ("2026-01-06", "Bridgehampton Wine", "Groceries", "USAA Visa", -45.68, "Shared", ""),
    ("2026-01-06", "Citarella", "Groceries", "Amex Gold", -14.15, "Shared", ""),
    ("2026-01-06", "U-Haul", "Storage Units", "Venture Card", -43.00, "Shared", ""),
    ("2026-01-05", "Amazon", "Shopping", "Amazon Prime Card", -57.11, "Shared", ""),
    ("2026-01-05", "Stop & Shop", "Groceries", "Amex Gold", -69.31, "Shared", ""),
    ("2026-01-05", "Costco", "Gas & Car Wash", "Venture Card", -25.63, "Shared", ""),
    ("2026-01-05", "Costco", "Shopping", "Venture Card", -146.81, "Shared", ""),
    ("2026-01-05", "Apple", "Subscriptions", "Venture Card", -2.99, "Shared", ""),
    ("2026-01-04", "Amazon", "Shopping", "Amazon Prime Card", -26.97, "Shared", ""),
    ("2026-01-04", "Villa Italian", "Delivery & Takeout", "Sapphire Reserve", -31.60, "Shared", ""),
    ("2026-01-04", "Citarella", "Groceries", "Amex Gold", -28.91, "Shared", ""),
    ("2026-01-03", "Google", "Subscriptions", "Freedom Card", -27.21, "Shared", ""),
    ("2026-01-03", "Hulu", "Subscriptions", "Marriott Card", -12.99, "Shared", ""),
    ("2026-01-03", "UBER EATS", "Delivery & Takeout", "Amex Platinum", -33.03, "Shared", ""),
    ("2026-01-03", "PSEG", "Electric & Gas", "Venture Card", -264.97, "Shared", ""),
    ("2026-01-02", "Amazon", "Shopping", "Amazon Prime Card", 20.34, "Shared", "Refund"),
    ("2026-01-02", "Nuuly", "Shopping", "Sapphire Reserve", -363.57, "Shared", ""),
    ("2026-01-02", "Youtube TV", "Subscriptions", "Amex Platinum", -82.99, "Shared", ""),
    ("2026-01-02", "Breeze Airways", "Airline", "Amex Platinum", -175.00, "Shared", ""),
    ("2026-01-02", "OpenAI", "Subscriptions", "Venture Card", -21.78, "Shared", ""),
    ("2026-01-01", "Amazon", "Shopping", "Amazon Prime Card", -14.07, "Shared", ""),
    ("2026-01-01", "Etsy", "Shopping", "Slate Card", -10.86, "Shared", ""),
    ("2026-01-01", "Breeze Airways", "Airline", "Amex Platinum", -1504.34, "Shared", ""),
    ("2026-01-01", "American Airlines", "Airline", "Amex Platinum", -315.48, "Shared", ""),
    ("2026-10-01", "Harbor Rent", "Rent", "Checking", -7500.00, "Shared", "Monthly rent"),
]

# Add data
for row, trans in enumerate(transactions, 2):
    for col, value in enumerate(trans, 1):
        cell = ws_trans.cell(row=row, column=col, value=value)
        if col == 5:  # Amount column
            cell.number_format = '"$"#,##0.00'
            if value < 0:
                cell.font = Font(color="DC2626")
            else:
                cell.font = Font(color="16A34A")

# Column widths
col_widths = [12, 25, 20, 20, 12, 10, 20]
for i, width in enumerate(col_widths, 1):
    ws_trans.column_dimensions[get_column_letter(i)].width = width

# Freeze header row
ws_trans.freeze_panes = "A2"

# ============ SUMMARY SHEET ============
ws_summary = wb.create_sheet("Summary")

# Title
ws_summary["A1"] = "FINANCIAL SUMMARY"
ws_summary["A1"].font = Font(size=18, bold=True, color="2563EB")
ws_summary.merge_cells("A1:D1")

# Summary metrics
metrics = [
    ("Total Income", f"=SUMIF(Transactions!E:E,\">0\",Transactions!E:E)"),
    ("Total Expenses", f"=ABS(SUMIF(Transactions!E:E,\"<0\",Transactions!E:E))"),
    ("Net Cash Flow", f"=SUM(Transactions!E:E)"),
    ("Total Transactions", f"=COUNTA(Transactions!A:A)-1"),
]

ws_summary["A3"] = "Metric"
ws_summary["B3"] = "Value"
ws_summary["A3"].font = Font(bold=True)
ws_summary["B3"].font = Font(bold=True)

for i, (label, formula) in enumerate(metrics, 4):
    ws_summary[f"A{i}"] = label
    ws_summary[f"B{i}"] = formula
    if "Income" in label:
        ws_summary[f"B{i}"].font = Font(color="16A34A", bold=True)
    elif "Expenses" in label:
        ws_summary[f"B{i}"].font = Font(color="DC2626", bold=True)
    ws_summary[f"B{i}"].number_format = '"$"#,##0.00'

ws_summary.column_dimensions["A"].width = 20
ws_summary.column_dimensions["B"].width = 15

# ============ CATEGORY BREAKDOWN SHEET ============
ws_cat = wb.create_sheet("By Category")

ws_cat["A1"] = "SPENDING BY CATEGORY"
ws_cat["A1"].font = Font(size=18, bold=True, color="2563EB")
ws_cat.merge_cells("A1:C1")

# Get unique categories and calculate totals
categories = {}
for trans in transactions:
    cat = trans[2]
    amount = trans[4]
    if amount < 0:  # Only expenses
        categories[cat] = categories.get(cat, 0) + abs(amount)

# Sort by amount
sorted_cats = sorted(categories.items(), key=lambda x: x[1], reverse=True)

ws_cat["A3"] = "Category"
ws_cat["B3"] = "Amount"
ws_cat["C3"] = "% of Total"
ws_cat["A3"].font = Font(bold=True)
ws_cat["B3"].font = Font(bold=True)
ws_cat["C3"].font = Font(bold=True)

total_expenses = sum(categories.values())

for i, (cat, amount) in enumerate(sorted_cats, 4):
    ws_cat[f"A{i}"] = cat
    ws_cat[f"B{i}"] = amount
    ws_cat[f"B{i}"].number_format = '"$"#,##0.00'
    ws_cat[f"C{i}"] = amount / total_expenses if total_expenses > 0 else 0
    ws_cat[f"C{i}"].number_format = "0.0%"

ws_cat.column_dimensions["A"].width = 25
ws_cat.column_dimensions["B"].width = 15
ws_cat.column_dimensions["C"].width = 12

# Add Pie Chart
pie = PieChart()
pie.title = "Spending by Category"
labels = Reference(ws_cat, min_col=1, min_row=4, max_row=3 + len(sorted_cats))
data = Reference(ws_cat, min_col=2, min_row=3, max_row=3 + len(sorted_cats))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 15
pie.height = 10
ws_cat.add_chart(pie, "E3")

# ============ ACCOUNT BREAKDOWN SHEET ============
ws_acc = wb.create_sheet("By Account")

ws_acc["A1"] = "SPENDING BY ACCOUNT"
ws_acc["A1"].font = Font(size=18, bold=True, color="2563EB")
ws_acc.merge_cells("A1:D1")

# Get unique accounts
accounts = {}
for trans in transactions:
    acc = trans[3]
    amount = trans[4]
    if acc not in accounts:
        accounts[acc] = {"income": 0, "expenses": 0, "count": 0}
    if amount > 0:
        accounts[acc]["income"] += amount
    else:
        accounts[acc]["expenses"] += abs(amount)
    accounts[acc]["count"] += 1

ws_acc["A3"] = "Account"
ws_acc["B3"] = "Income"
ws_acc["C3"] = "Expenses"
ws_acc["D3"] = "Net"
ws_acc["E3"] = "Transactions"
for col in ["A", "B", "C", "D", "E"]:
    ws_acc[f"{col}3"].font = Font(bold=True)

for i, (acc, data) in enumerate(accounts.items(), 4):
    ws_acc[f"A{i}"] = acc
    ws_acc[f"B{i}"] = data["income"]
    ws_acc[f"B{i}"].number_format = '"$"#,##0.00'
    ws_acc[f"B{i}"].font = Font(color="16A34A")
    ws_acc[f"C{i}"] = data["expenses"]
    ws_acc[f"C{i}"].number_format = '"$"#,##0.00'
    ws_acc[f"C{i}"].font = Font(color="DC2626")
    ws_acc[f"D{i}"] = data["income"] - data["expenses"]
    ws_acc[f"D{i}"].number_format = '"$"#,##0.00'
    ws_acc[f"E{i}"] = data["count"]

ws_acc.column_dimensions["A"].width = 22
ws_acc.column_dimensions["B"].width = 12
ws_acc.column_dimensions["C"].width = 12
ws_acc.column_dimensions["D"].width = 12
ws_acc.column_dimensions["E"].width = 14

# ============ INSTRUCTIONS SHEET ============
ws_help = wb.create_sheet("Instructions")

ws_help["A1"] = "HOW TO USE THIS FINANCE TRACKER"
ws_help["A1"].font = Font(size=18, bold=True, color="2563EB")

instructions = [
    "",
    "1. ADD YOUR TRANSACTIONS",
    "   - Go to the 'Transactions' sheet",
    "   - Add your transactions in the rows below the existing data",
    "   - Use negative amounts for expenses, positive for income",
    "",
    "2. IMPORT FROM MONARCH/MINT",
    "   - Export CSV from your finance app",
    "   - Copy and paste the data into the Transactions sheet",
    "   - Make sure columns match: Date, Merchant, Category, Account, Amount",
    "",
    "3. VIEW YOUR SUMMARY",
    "   - 'Summary' sheet shows total income, expenses, and net",
    "   - 'By Category' shows where your money goes (with chart)",
    "   - 'By Account' shows activity per account",
    "",
    "4. TIPS",
    "   - Use consistent category names for accurate reporting",
    "   - The Summary sheet updates automatically with formulas",
    "   - Sort the Transactions sheet by clicking column headers",
]

for i, line in enumerate(instructions, 3):
    ws_help[f"A{i}"] = line

ws_help.column_dimensions["A"].width = 60

# Save the workbook
output_path = "/home/user/claude-code/finance-tracker/Finance_Tracker.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
