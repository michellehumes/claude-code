#!/usr/bin/env python3
"""
2026 Teacher Planner & Grade Book - Excel Spreadsheet Template
For: Shelzy's Designs (Etsy)
Academic Year: September 2026 - June 2027
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import date, timedelta
import calendar

# ── Color Palette (Navy Blue Professional Theme) ────────────────────────
NAVY        = "1B2A4A"
DARK_NAVY   = "0F1D35"
MID_BLUE    = "2E4A7A"
LIGHT_BLUE  = "D6E4F0"
ACCENT_BLUE = "4472C4"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D9D9D9"
DARK_TEXT    = "1B2A4A"
RED_ACCENT  = "C0392B"
GREEN_ACCENT = "27AE60"

# ── Reusable Styles ─────────────────────────────────────────────────────
font_title = Font(name="Calibri", size=18, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=12, bold=True, color=WHITE)
font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_subheader = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_body = Font(name="Calibri", size=10, color=DARK_TEXT)
font_body_bold = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_small = Font(name="Calibri", size=9, color=DARK_TEXT)
font_instruction = Font(name="Calibri", size=10, italic=True, color=MID_BLUE)

fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_dark_navy = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
fill_mid_blue = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
fill_light_blue = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_accent = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_light_gray = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
fill_med_gray = PatternFill(start_color=MED_GRAY, end_color=MED_GRAY, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

thin_side = Side(style="thin", color=MED_GRAY)
thick_side = Side(style="medium", color=NAVY)
border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_bottom_navy = Border(bottom=Side(style="medium", color=NAVY))
border_all_navy = Border(
    left=Side(style="thin", color=NAVY),
    right=Side(style="thin", color=NAVY),
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="thin", color=NAVY),
)


def style_cell(ws, row, col, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Apply multiple styles to a cell."""
    cell = ws.cell(row=row, column=col)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def style_range(ws, row_start, row_end, col_start, col_end, **kwargs):
    """Apply styles to a rectangular range."""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            style_cell(ws, r, c, **kwargs)


def create_title_banner(ws, row, col_start, col_end, title, subtitle=None):
    """Create a styled title banner spanning columns."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font = font_title
    cell.fill = fill_dark_navy
    cell.alignment = align_center
    # Fill merged area
    for c in range(col_start, col_end + 1):
        style_cell(ws, row, c, fill=fill_dark_navy)

    if subtitle:
        ws.merge_cells(
            start_row=row + 1, start_column=col_start,
            end_row=row + 1, end_column=col_end
        )
        cell2 = ws.cell(row=row + 1, column=col_start, value=subtitle)
        cell2.font = font_subtitle
        cell2.fill = fill_navy
        cell2.alignment = align_center
        for c in range(col_start, col_end + 1):
            style_cell(ws, row + 1, c, fill=fill_navy)


# ═════════════════════════════════════════════════════════════════════════
# SHEET 1: CLASS ROSTER
# ═════════════════════════════════════════════════════════════════════════
def create_class_roster(wb):
    ws = wb.active
    ws.title = "Class Roster"
    ws.sheet_properties.tabColor = NAVY

    # Column widths
    col_widths = {
        1: 5,    # #
        2: 22,   # Student Name
        3: 12,   # Student ID
        4: 12,   # Grade Level
        5: 14,   # Date of Birth
        6: 22,   # Parent/Guardian 1
        7: 18,   # Phone 1
        8: 28,   # Email 1
        9: 22,   # Parent/Guardian 2
        10: 18,  # Phone 2
        11: 28,  # Email 2
        12: 18,  # Emergency Contact
        13: 18,  # Emergency Phone
        14: 15,  # Medical/Allergies
        15: 25,  # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Row heights
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 30

    # Title banner
    create_title_banner(ws, 1, 1, 15,
                        "CLASS ROSTER",
                        "2026-2027 Academic Year")

    # Spacer row
    for c in range(1, 16):
        style_cell(ws, 3, c, fill=fill_white)

    # Teacher info row
    ws.merge_cells("A4:B4")
    ws.cell(row=4, column=1, value="Teacher:").font = font_body_bold
    ws.cell(row=4, column=1).fill = fill_light_blue
    ws.cell(row=4, column=1).alignment = align_left
    style_cell(ws, 4, 2, fill=fill_light_blue)
    ws.merge_cells("C4:D4")
    ws.cell(row=4, column=3, value="School:").font = font_body_bold
    ws.cell(row=4, column=3).fill = fill_light_blue
    style_cell(ws, 4, 4, fill=fill_light_blue)
    ws.merge_cells("E4:F4")
    ws.cell(row=4, column=5, value="Subject:").font = font_body_bold
    ws.cell(row=4, column=5).fill = fill_light_blue
    style_cell(ws, 4, 6, fill=fill_light_blue)
    ws.merge_cells("G4:H4")
    ws.cell(row=4, column=7, value="Room #:").font = font_body_bold
    ws.cell(row=4, column=7).fill = fill_light_blue
    style_cell(ws, 4, 8, fill=fill_light_blue)
    ws.merge_cells("I4:J4")
    ws.cell(row=4, column=9, value="Period/Block:").font = font_body_bold
    ws.cell(row=4, column=9).fill = fill_light_blue
    style_cell(ws, 4, 10, fill=fill_light_blue)
    for c in range(11, 16):
        style_cell(ws, 4, c, fill=fill_light_blue)

    # Spacer
    ws.row_dimensions[5].height = 6

    # Column headers (row 6)
    headers = [
        "#", "Student Name", "Student ID", "Grade", "Date of Birth",
        "Parent/Guardian 1", "Phone", "Email",
        "Parent/Guardian 2", "Phone", "Email",
        "Emergency Contact", "Emergency Phone",
        "Medical/Allergies", "Notes"
    ]
    ws.row_dimensions[6].height = 32
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin

    # Student rows (35 rows: rows 7-41)
    for row_num in range(7, 42):
        student_num = row_num - 6
        ws.row_dimensions[row_num].height = 24

        # Number column
        cell = ws.cell(row=row_num, column=1, value=student_num)
        cell.font = font_body_bold
        cell.alignment = align_center

        # Alternate row colors
        row_fill = fill_light_gray if student_num % 2 == 0 else fill_white

        for col in range(1, 16):
            c = ws.cell(row=row_num, column=col)
            c.fill = row_fill
            c.border = border_thin
            c.font = font_body
            c.alignment = align_left if col > 1 else align_center

    # Summary row
    summary_row = 43
    ws.row_dimensions[42].height = 6
    ws.row_dimensions[summary_row].height = 26
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws.cell(row=summary_row, column=1, value="Total Students:").font = font_body_bold
    ws.cell(row=summary_row, column=1).fill = fill_light_blue
    ws.cell(row=summary_row, column=1).alignment = align_right
    style_cell(ws, summary_row, 2, fill=fill_light_blue)
    cell_total = ws.cell(row=summary_row, column=3,
                         value='=COUNTA(B7:B41)')
    cell_total.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    cell_total.fill = fill_light_blue
    cell_total.alignment = align_center

    # Footer
    footer_row = 45
    ws.merge_cells(f"A{footer_row}:O{footer_row}")
    ws.cell(row=footer_row, column=1,
            value="© Shelzy's Designs  |  2026-2027 Teacher Planner & Grade Book").font = \
        Font(name="Calibri", size=9, italic=True, color=MID_BLUE)
    ws.cell(row=footer_row, column=1).alignment = align_center

    # Print setup
    ws.print_area = "A1:O43"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5
    )
    ws.freeze_panes = "B7"

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET 2: GRADE BOOK
# ═════════════════════════════════════════════════════════════════════════
def create_grade_book(wb):
    ws = wb.create_sheet("Grade Book")
    ws.sheet_properties.tabColor = ACCENT_BLUE

    # Column layout:
    # A=# B=Name C-G=Assignments(5) H=Avg I-L=Quizzes(4) M=Avg
    # N-P=Tests(3) Q=Avg R=Participation S=Final Grade T=Letter
    col_widths = {
        1: 5, 2: 22,
        3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 12,  # Assignments + Avg
        9: 12, 10: 12, 11: 12, 12: 12, 13: 12,       # Quizzes + Avg
        14: 12, 15: 12, 16: 12, 17: 12,               # Tests + Avg
        18: 14,                                         # Participation
        19: 13, 20: 10                                  # Final + Letter
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Title
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22
    create_title_banner(ws, 1, 1, 20,
                        "GRADE BOOK",
                        "2026-2027 Academic Year")

    # Spacer
    ws.row_dimensions[3].height = 6

    # Weighting info (row 4)
    ws.row_dimensions[4].height = 26
    ws.merge_cells("A4:B4")
    ws.cell(row=4, column=1, value="GRADE WEIGHTS:").font = font_body_bold
    ws.cell(row=4, column=1).fill = fill_light_blue
    ws.cell(row=4, column=1).alignment = align_right
    style_cell(ws, 4, 2, fill=fill_light_blue)

    weights_info = [
        (3, "C4:D4", "Assignments: 30%"),
        (5, "E4:F4", "Quizzes: 20%"),
        (7, "G4:H4", "Tests: 35%"),
        (9, "I4:J4", "Participation: 15%"),
    ]
    # Simplified weight display
    ws.merge_cells("C4:E4")
    ws.cell(row=4, column=3, value="Assignments: 30%").font = font_body_bold
    ws.cell(row=4, column=3).fill = fill_light_blue
    ws.cell(row=4, column=3).alignment = align_center
    style_cell(ws, 4, 4, fill=fill_light_blue)
    style_cell(ws, 4, 5, fill=fill_light_blue)

    ws.merge_cells("F4:H4")
    ws.cell(row=4, column=6, value="Quizzes: 20%").font = font_body_bold
    ws.cell(row=4, column=6).fill = fill_light_blue
    ws.cell(row=4, column=6).alignment = align_center
    style_cell(ws, 4, 7, fill=fill_light_blue)
    style_cell(ws, 4, 8, fill=fill_light_blue)

    ws.merge_cells("I4:K4")
    ws.cell(row=4, column=9, value="Tests: 35%").font = font_body_bold
    ws.cell(row=4, column=9).fill = fill_light_blue
    ws.cell(row=4, column=9).alignment = align_center
    style_cell(ws, 4, 10, fill=fill_light_blue)
    style_cell(ws, 4, 11, fill=fill_light_blue)

    ws.merge_cells("L4:N4")
    ws.cell(row=4, column=12, value="Participation: 15%").font = font_body_bold
    ws.cell(row=4, column=12).fill = fill_light_blue
    ws.cell(row=4, column=12).alignment = align_center
    style_cell(ws, 4, 13, fill=fill_light_blue)
    style_cell(ws, 4, 14, fill=fill_light_blue)

    for c in range(15, 21):
        style_cell(ws, 4, c, fill=fill_light_blue)

    # Grading scale (row 5)
    ws.row_dimensions[5].height = 22
    ws.merge_cells("A5:T5")
    ws.cell(row=5, column=1,
            value="GRADING SCALE:   A = 90-100%   |   B = 80-89%   |   C = 70-79%   |   D = 60-69%   |   F = Below 60%").font = \
        Font(name="Calibri", size=10, italic=True, color=NAVY)
    ws.cell(row=5, column=1).alignment = align_center
    for c in range(1, 21):
        style_cell(ws, 5, c, fill=fill_white)

    ws.row_dimensions[6].height = 6

    # Category headers (row 7)
    ws.row_dimensions[7].height = 26
    ws.merge_cells("A7:B7")
    style_cell(ws, 7, 1, fill=fill_dark_navy, font=font_header, alignment=align_center)
    style_cell(ws, 7, 2, fill=fill_dark_navy)

    ws.merge_cells("C7:H7")
    ws.cell(row=7, column=3, value="ASSIGNMENTS (30%)").font = font_header
    ws.cell(row=7, column=3).fill = fill_navy
    ws.cell(row=7, column=3).alignment = align_center
    for c in range(4, 9):
        style_cell(ws, 7, c, fill=fill_navy)

    ws.merge_cells("I7:M7")
    ws.cell(row=7, column=9, value="QUIZZES (20%)").font = font_header
    ws.cell(row=7, column=9).fill = fill_mid_blue
    ws.cell(row=7, column=9).alignment = align_center
    for c in range(10, 14):
        style_cell(ws, 7, c, fill=fill_mid_blue)

    ws.merge_cells("N7:Q7")
    ws.cell(row=7, column=14, value="TESTS (35%)").font = font_header
    ws.cell(row=7, column=14).fill = fill_navy
    ws.cell(row=7, column=14).alignment = align_center
    for c in range(15, 18):
        style_cell(ws, 7, c, fill=fill_navy)

    ws.merge_cells("R7:R7")
    ws.cell(row=7, column=18, value="PART. (15%)").font = font_header
    ws.cell(row=7, column=18).fill = fill_mid_blue
    ws.cell(row=7, column=18).alignment = align_center

    ws.merge_cells("S7:T7")
    ws.cell(row=7, column=19, value="FINAL").font = font_header
    ws.cell(row=7, column=19).fill = fill_dark_navy
    ws.cell(row=7, column=19).alignment = align_center
    style_cell(ws, 7, 20, fill=fill_dark_navy)

    # Sub-headers (row 8)
    ws.row_dimensions[8].height = 28
    sub_headers = [
        "#", "Student Name",
        "Assign 1", "Assign 2", "Assign 3", "Assign 4", "Assign 5", "AVG",
        "Quiz 1", "Quiz 2", "Quiz 3", "Quiz 4", "AVG",
        "Test 1", "Test 2", "Test 3", "AVG",
        "Score",
        "Final %", "Grade"
    ]
    for i, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
        cell.alignment = align_center
        cell.border = border_thin
        # Color code by category
        if i <= 2:
            cell.fill = fill_dark_navy
        elif i <= 8:
            cell.fill = fill_navy
        elif i <= 13:
            cell.fill = fill_mid_blue
        elif i <= 17:
            cell.fill = fill_navy
        elif i == 18:
            cell.fill = fill_mid_blue
        else:
            cell.fill = fill_dark_navy

    # Student data rows (35 students: rows 9-43)
    for row_num in range(9, 44):
        student_num = row_num - 8
        ws.row_dimensions[row_num].height = 22

        # Number
        ws.cell(row=row_num, column=1, value=student_num).font = font_body_bold
        ws.cell(row=row_num, column=1).alignment = align_center

        # Linked student name from Roster
        ws.cell(row=row_num, column=2,
                value=f"=IF('Class Roster'!B{row_num - 2}<>\"\", 'Class Roster'!B{row_num - 2}, \"\")").font = font_body

        # Assignment Average (col H = 8): average of C-G
        c_col = get_column_letter(3)
        g_col = get_column_letter(7)
        ws.cell(row=row_num, column=8,
                value=f'=IF(COUNT({c_col}{row_num}:{g_col}{row_num})>0, AVERAGE({c_col}{row_num}:{g_col}{row_num}), "")').number_format = '0.0'

        # Quiz Average (col M = 13): average of I-L
        i_col = get_column_letter(9)
        l_col = get_column_letter(12)
        ws.cell(row=row_num, column=13,
                value=f'=IF(COUNT({i_col}{row_num}:{l_col}{row_num})>0, AVERAGE({i_col}{row_num}:{l_col}{row_num}), "")').number_format = '0.0'

        # Test Average (col Q = 17): average of N-P
        n_col = get_column_letter(14)
        p_col = get_column_letter(16)
        ws.cell(row=row_num, column=17,
                value=f'=IF(COUNT({n_col}{row_num}:{p_col}{row_num})>0, AVERAGE({n_col}{row_num}:{p_col}{row_num}), "")').number_format = '0.0'

        # Final Grade % (col S = 19): weighted average
        # H=Assign Avg (30%), M=Quiz Avg (20%), Q=Test Avg (35%), R=Part (15%)
        ws.cell(row=row_num, column=19,
                value=f'=IF(B{row_num}<>"", IF(COUNT(H{row_num},M{row_num},Q{row_num},R{row_num})>0, SUMPRODUCT(IF(ISNUMBER(H{row_num}),H{row_num}*0.3,0)+IF(ISNUMBER(M{row_num}),M{row_num}*0.2,0)+IF(ISNUMBER(Q{row_num}),Q{row_num}*0.35,0)+IF(ISNUMBER(R{row_num}),R{row_num}*0.15,0)), ""), "")').number_format = '0.0'

        # Letter Grade (col T = 20)
        ws.cell(row=row_num, column=20,
                value=f'=IF(ISNUMBER(S{row_num}), IF(S{row_num}>=90,"A", IF(S{row_num}>=80,"B", IF(S{row_num}>=70,"C", IF(S{row_num}>=60,"D","F")))), "")').font = font_body_bold

        # Alternate row colors and borders
        row_fill = fill_light_gray if student_num % 2 == 0 else fill_white
        for col in range(1, 21):
            c = ws.cell(row=row_num, column=col)
            if c.fill == PatternFill() or c.fill.start_color.rgb == "00000000":
                c.fill = row_fill
            else:
                c.fill = row_fill
            c.border = border_thin
            if col not in [2]:
                c.alignment = align_center
            else:
                c.alignment = align_left
            if not c.font or c.font == Font():
                c.font = font_body

        # Color the average columns slightly different
        for avg_col in [8, 13, 17]:
            style_cell(ws, row_num, avg_col,
                       fill=fill_light_blue, font=font_body_bold, alignment=align_center)
        # Final grade columns
        style_cell(ws, row_num, 19, fill=fill_light_blue, font=font_body_bold, alignment=align_center)
        style_cell(ws, row_num, 20, fill=fill_light_blue, font=font_body_bold, alignment=align_center)

    # Class statistics row
    stats_row = 45
    ws.row_dimensions[44].height = 6
    ws.row_dimensions[stats_row].height = 26
    ws.row_dimensions[stats_row + 1].height = 26
    ws.row_dimensions[stats_row + 2].height = 26

    # Class Average
    ws.merge_cells(f"A{stats_row}:B{stats_row}")
    ws.cell(row=stats_row, column=1, value="CLASS AVERAGE:").font = font_body_bold
    ws.cell(row=stats_row, column=1).fill = fill_navy
    ws.cell(row=stats_row, column=1).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(row=stats_row, column=1).alignment = align_right
    style_cell(ws, stats_row, 2, fill=fill_navy)

    for col in range(3, 21):
        cl = get_column_letter(col)
        if col in [8, 13, 17, 18, 19]:
            ws.cell(row=stats_row, column=col,
                    value=f'=IF(COUNT({cl}9:{cl}43)>0, AVERAGE({cl}9:{cl}43), "")').number_format = '0.0'
        elif col == 20:
            ws.cell(row=stats_row, column=col, value="").font = font_body
        elif col in [1, 2]:
            pass
        else:
            ws.cell(row=stats_row, column=col,
                    value=f'=IF(COUNT({cl}9:{cl}43)>0, AVERAGE({cl}9:{cl}43), "")').number_format = '0.0'
        style_cell(ws, stats_row, col, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)

    # Highest Grade
    ws.merge_cells(f"A{stats_row+1}:B{stats_row+1}")
    ws.cell(row=stats_row+1, column=1, value="HIGHEST:").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(row=stats_row+1, column=1).fill = fill_navy
    ws.cell(row=stats_row+1, column=1).alignment = align_right
    style_cell(ws, stats_row+1, 2, fill=fill_navy)
    ws.cell(row=stats_row+1, column=19,
            value=f'=IF(COUNT(S9:S43)>0, MAX(S9:S43), "")').number_format = '0.0'
    style_cell(ws, stats_row+1, 19, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)
    style_cell(ws, stats_row+1, 20, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)

    # Lowest Grade
    ws.merge_cells(f"A{stats_row+2}:B{stats_row+2}")
    ws.cell(row=stats_row+2, column=1, value="LOWEST:").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(row=stats_row+2, column=1).fill = fill_navy
    ws.cell(row=stats_row+2, column=1).alignment = align_right
    style_cell(ws, stats_row+2, 2, fill=fill_navy)
    ws.cell(row=stats_row+2, column=19,
            value=f'=IF(COUNT(S9:S43)>0, MIN(S9:S43), "")').number_format = '0.0'
    style_cell(ws, stats_row+2, 19, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)
    style_cell(ws, stats_row+2, 20, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)

    # Conditional formatting for letter grades
    red_font = Font(color=RED_ACCENT, bold=True)
    green_font = Font(color=GREEN_ACCENT, bold=True)
    ws.conditional_formatting.add(
        f"T9:T43",
        CellIsRule(operator="equal", formula=['"A"'],
                   font=Font(color="27AE60", bold=True),
                   fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"T9:T43",
        CellIsRule(operator="equal", formula=['"F"'],
                   font=Font(color="C0392B", bold=True),
                   fill=PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"))
    )

    # Footer
    footer_row = 49
    ws.merge_cells(f"A{footer_row}:T{footer_row}")
    ws.cell(row=footer_row, column=1,
            value="© Shelzy's Designs  |  2026-2027 Teacher Planner & Grade Book").font = \
        Font(name="Calibri", size=9, italic=True, color=MID_BLUE)
    ws.cell(row=footer_row, column=1).alignment = align_center

    # Print setup
    ws.print_area = "A1:T47"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        left=0.3, right=0.3, top=0.4, bottom=0.4
    )
    ws.freeze_panes = "C9"

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET 3: LESSON PLANNER
# ═════════════════════════════════════════════════════════════════════════
def create_lesson_planner(wb):
    ws = wb.create_sheet("Lesson Planner")
    ws.sheet_properties.tabColor = "2E4A7A"

    # Layout: A=Time, B=Monday, C=Tuesday, D=Wednesday, E=Thursday, F=Friday, G=Notes
    col_widths = {1: 14, 2: 24, 3: 24, 4: 24, 5: 24, 6: 24, 7: 22}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Title
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22
    create_title_banner(ws, 1, 1, 7,
                        "WEEKLY LESSON PLANNER",
                        "2026-2027 Academic Year  |  September 2026 - June 2027")

    ws.row_dimensions[3].height = 6

    # Instructions
    ws.row_dimensions[4].height = 22
    ws.merge_cells("A4:G4")
    ws.cell(row=4, column=1,
            value="Duplicate this weekly template for each week. Each week block includes Subject/Objective, Activities, Materials, Homework, and Notes rows.").font = font_instruction
    ws.cell(row=4, column=1).alignment = align_center
    for c in range(1, 8):
        style_cell(ws, 4, c, fill=fill_white)

    # Generate weeks from Sept 2026 to June 2027
    # Find first Monday of September 2026
    start_date = date(2026, 9, 1)
    # Adjust to Monday
    while start_date.weekday() != 0:
        start_date -= timedelta(days=1)
    # If that's in August, move to next Monday
    if start_date.month == 8:
        start_date += timedelta(days=7)

    end_date = date(2027, 6, 30)

    current_date = start_date
    current_row = 6
    week_count = 0

    while current_date <= end_date:
        week_count += 1
        friday = current_date + timedelta(days=4)

        # Week header
        ws.row_dimensions[current_row].height = 28
        week_label = f"WEEK {week_count}:  {current_date.strftime('%B %d')} - {friday.strftime('%B %d, %Y')}"
        ws.merge_cells(f"A{current_row}:G{current_row}")
        cell = ws.cell(row=current_row, column=1, value=week_label)
        cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        cell.fill = fill_navy
        cell.alignment = align_center
        for c in range(2, 8):
            style_cell(ws, current_row, c, fill=fill_navy)
        current_row += 1

        # Day headers
        ws.row_dimensions[current_row].height = 24
        day_headers = ["", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Notes"]
        dates = [
            "",
            current_date.strftime("%m/%d"),
            (current_date + timedelta(days=1)).strftime("%m/%d"),
            (current_date + timedelta(days=2)).strftime("%m/%d"),
            (current_date + timedelta(days=3)).strftime("%m/%d"),
            (current_date + timedelta(days=4)).strftime("%m/%d"),
            ""
        ]
        for i, (dh, dt) in enumerate(zip(day_headers, dates), 1):
            display = f"{dh}\n{dt}" if dt else dh
            cell = ws.cell(row=current_row, column=i, value=display)
            cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            cell.fill = fill_mid_blue
            cell.alignment = align_center
            cell.border = border_thin
        current_row += 1

        # Row categories for each week
        categories = [
            "Subject / Objective",
            "Activities",
            "Materials",
            "Homework",
            "Assessment / Notes"
        ]
        for cat in categories:
            ws.row_dimensions[current_row].height = 40
            cell = ws.cell(row=current_row, column=1, value=cat)
            cell.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
            cell.fill = fill_light_blue
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border_thin

            for col in range(2, 8):
                c = ws.cell(row=current_row, column=col)
                c.fill = fill_white
                c.border = border_thin
                c.font = font_small
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            current_row += 1

        # Spacer between weeks
        ws.row_dimensions[current_row].height = 8
        current_row += 1

        # Next week
        current_date += timedelta(days=7)

    # Footer
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws.cell(row=current_row, column=1,
            value="© Shelzy's Designs  |  2026-2027 Teacher Planner & Grade Book").font = \
        Font(name="Calibri", size=9, italic=True, color=MID_BLUE)
    ws.cell(row=current_row, column=1).alignment = align_center

    # Print setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.5
    )
    ws.freeze_panes = "A6"

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET 4: ATTENDANCE TRACKER
# ═════════════════════════════════════════════════════════════════════════
def create_attendance_tracker(wb):
    ws = wb.create_sheet("Attendance Tracker")
    ws.sheet_properties.tabColor = "4472C4"

    # We'll create a monthly attendance view
    # Columns: A=# B=Name, then up to 23 school days per month, then Totals
    # We'll do a single comprehensive sheet with all months

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22

    # Title
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22

    total_cols = 37  # generous: # + name + 31 days + Present + Absent + Rate
    last_col_letter = get_column_letter(total_cols)

    create_title_banner(ws, 1, 1, total_cols,
                        "ATTENDANCE TRACKER",
                        "2026-2027 Academic Year")

    ws.row_dimensions[3].height = 6

    # Legend
    ws.row_dimensions[4].height = 22
    ws.merge_cells(f"A4:{last_col_letter}4")
    ws.cell(row=4, column=1,
            value="KEY:   P = Present   |   A = Absent   |   T = Tardy   |   E = Excused Absence   |   H = Holiday/No School").font = \
        Font(name="Calibri", size=10, bold=True, color=NAVY)
    ws.cell(row=4, column=1).alignment = align_center
    for c in range(1, total_cols + 1):
        style_cell(ws, 4, c, fill=fill_light_blue)

    ws.row_dimensions[5].height = 6

    # Generate monthly sections
    months = [
        (2026, 9, "September"), (2026, 10, "October"), (2026, 11, "November"),
        (2026, 12, "December"), (2027, 1, "January"), (2027, 2, "February"),
        (2027, 3, "March"), (2027, 4, "April"), (2027, 5, "May"), (2027, 6, "June")
    ]

    current_row = 6

    for year, month, month_name in months:
        num_days = calendar.monthrange(year, month)[1]

        # Find school days (weekdays only)
        school_days = []
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            if d.weekday() < 5:  # Monday-Friday
                school_days.append(day)

        num_school_days = len(school_days)
        # Columns: A=# B=Name [school_days] Present Absent Tardy Rate
        total_month_cols = 2 + num_school_days + 4

        # Month header
        ws.row_dimensions[current_row].height = 28
        mc_end = get_column_letter(total_month_cols)
        ws.merge_cells(f"A{current_row}:{mc_end}{current_row}")
        ws.cell(row=current_row, column=1,
                value=f"{month_name.upper()} {year}").font = Font(name="Calibri", size=14, bold=True, color=WHITE)
        ws.cell(row=current_row, column=1).fill = fill_dark_navy
        ws.cell(row=current_row, column=1).alignment = align_center
        for c in range(2, total_month_cols + 1):
            style_cell(ws, current_row, c, fill=fill_dark_navy)
        current_row += 1

        # Day headers
        ws.row_dimensions[current_row].height = 30
        ws.cell(row=current_row, column=1, value="#").font = font_header
        ws.cell(row=current_row, column=1).fill = fill_navy
        ws.cell(row=current_row, column=1).alignment = align_center
        ws.cell(row=current_row, column=1).border = border_thin

        ws.cell(row=current_row, column=2, value="Student Name").font = font_header
        ws.cell(row=current_row, column=2).fill = fill_navy
        ws.cell(row=current_row, column=2).alignment = align_center
        ws.cell(row=current_row, column=2).border = border_thin

        for idx, day in enumerate(school_days):
            col = 3 + idx
            d = date(year, month, day)
            day_abbr = d.strftime("%a")[0]  # M, T, W, T, F
            cell = ws.cell(row=current_row, column=col,
                           value=f"{day_abbr}\n{month}/{day}")
            cell.font = Font(name="Calibri", size=8, bold=True, color=WHITE)
            cell.fill = fill_navy
            cell.alignment = align_center
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(col)].width = 6

        # Summary columns
        summary_labels = ["Present", "Absent", "Tardy", "Rate"]
        for i, label in enumerate(summary_labels):
            col = 3 + num_school_days + i
            cell = ws.cell(row=current_row, column=col, value=label)
            cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            cell.fill = fill_mid_blue
            cell.alignment = align_center
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(col)].width = 10
        current_row += 1

        # Student rows (35 students)
        first_data_row = current_row
        for s in range(1, 36):
            ws.row_dimensions[current_row].height = 20

            # Number
            ws.cell(row=current_row, column=1, value=s).font = font_body_bold
            ws.cell(row=current_row, column=1).alignment = align_center

            # Student name linked from roster
            roster_row = s + 6  # roster starts at row 7
            ws.cell(row=current_row, column=2,
                    value=f"=IF('Class Roster'!B{roster_row}<>\"\", 'Class Roster'!B{roster_row}, \"\")").font = font_body

            row_fill = fill_light_gray if s % 2 == 0 else fill_white

            for col in range(1, total_month_cols + 1):
                c = ws.cell(row=current_row, column=col)
                c.fill = row_fill
                c.border = border_thin
                if col > 2 and col <= 2 + num_school_days:
                    c.alignment = align_center
                    c.font = font_small

            # Present count formula: COUNTIF range = "P"
            day_start_col = get_column_letter(3)
            day_end_col = get_column_letter(2 + num_school_days)
            present_col = 3 + num_school_days
            absent_col = present_col + 1
            tardy_col = absent_col + 1
            rate_col = tardy_col + 1

            ws.cell(row=current_row, column=present_col,
                    value=f'=COUNTIF({day_start_col}{current_row}:{day_end_col}{current_row},"P")').number_format = '0'
            style_cell(ws, current_row, present_col, fill=fill_light_blue, font=font_body_bold, alignment=align_center)

            ws.cell(row=current_row, column=absent_col,
                    value=f'=COUNTIF({day_start_col}{current_row}:{day_end_col}{current_row},"A")+COUNTIF({day_start_col}{current_row}:{day_end_col}{current_row},"E")').number_format = '0'
            style_cell(ws, current_row, absent_col, fill=fill_light_blue, font=font_body_bold, alignment=align_center)

            ws.cell(row=current_row, column=tardy_col,
                    value=f'=COUNTIF({day_start_col}{current_row}:{day_end_col}{current_row},"T")').number_format = '0'
            style_cell(ws, current_row, tardy_col, fill=fill_light_blue, font=font_body_bold, alignment=align_center)

            # Attendance rate
            present_cl = get_column_letter(present_col)
            absent_cl = get_column_letter(absent_col)
            tardy_cl = get_column_letter(tardy_col)
            ws.cell(row=current_row, column=rate_col,
                    value=f'=IF(({present_cl}{current_row}+{absent_cl}{current_row}+{tardy_cl}{current_row})>0, {present_cl}{current_row}/({present_cl}{current_row}+{absent_cl}{current_row}+{tardy_cl}{current_row}), "")').number_format = '0.0%'
            style_cell(ws, current_row, rate_col, fill=fill_light_blue, font=font_body_bold, alignment=align_center)

            current_row += 1

        last_data_row = current_row - 1

        # Monthly totals row
        ws.row_dimensions[current_row].height = 24
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws.cell(row=current_row, column=1, value="MONTHLY TOTALS:").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        ws.cell(row=current_row, column=1).fill = fill_navy
        ws.cell(row=current_row, column=1).alignment = align_right
        style_cell(ws, current_row, 2, fill=fill_navy)

        # Total present/absent/tardy for the month
        for summary_offset, _ in enumerate(summary_labels[:3]):
            col = 3 + num_school_days + summary_offset
            cl = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                    value=f'=SUM({cl}{first_data_row}:{cl}{last_data_row})').number_format = '0'
            style_cell(ws, current_row, col, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)

        # Average rate
        rate_cl = get_column_letter(rate_col)
        ws.cell(row=current_row, column=rate_col,
                value=f'=IF(COUNT({rate_cl}{first_data_row}:{rate_cl}{last_data_row})>0, AVERAGE({rate_cl}{first_data_row}:{rate_cl}{last_data_row}), "")').number_format = '0.0%'
        style_cell(ws, current_row, rate_col, fill=fill_light_blue, font=font_body_bold, alignment=align_center, border=border_thin)

        current_row += 2  # space between months

    # Data validation for attendance cells - adding P/A/T/E/H dropdown
    dv = DataValidation(
        type="list",
        formula1='"P,A,T,E,H"',
        allow_blank=True
    )
    dv.error = "Please enter P, A, T, E, or H"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "P=Present, A=Absent, T=Tardy, E=Excused, H=Holiday"
    dv.promptTitle = "Attendance"
    ws.add_data_validation(dv)

    # Apply validation to all attendance data cells
    # We need to re-iterate to apply validation
    current_row_check = 6
    for year, month, month_name in months:
        num_days = calendar.monthrange(year, month)[1]
        school_days = []
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            if d.weekday() < 5:
                school_days.append(day)
        num_school_days = len(school_days)

        current_row_check += 1  # month header
        current_row_check += 1  # day headers

        for s in range(35):
            day_start = get_column_letter(3)
            day_end = get_column_letter(2 + num_school_days)
            dv.add(f"{day_start}{current_row_check}:{day_end}{current_row_check}")
            current_row_check += 1

        current_row_check += 1  # totals row
        current_row_check += 1  # spacer

    # Conditional formatting for attendance
    ws.conditional_formatting.add(
        f"C8:AK{current_row}",
        CellIsRule(operator="equal", formula=['"A"'],
                   font=Font(color="C0392B", bold=True),
                   fill=PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"C8:AK{current_row}",
        CellIsRule(operator="equal", formula=['"P"'],
                   font=Font(color="27AE60", bold=True),
                   fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"C8:AK{current_row}",
        CellIsRule(operator="equal", formula=['"T"'],
                   font=Font(color="E67E22", bold=True),
                   fill=PatternFill(start_color="FDF2E9", end_color="FDF2E9", fill_type="solid"))
    )

    # Footer
    footer_row = current_row + 1
    ws.merge_cells(f"A{footer_row}:{last_col_letter}{footer_row}")
    ws.cell(row=footer_row, column=1,
            value="© Shelzy's Designs  |  2026-2027 Teacher Planner & Grade Book").font = \
        Font(name="Calibri", size=9, italic=True, color=MID_BLUE)
    ws.cell(row=footer_row, column=1).alignment = align_center

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        left=0.3, right=0.3, top=0.4, bottom=0.4
    )
    ws.freeze_panes = "C8"

    return ws


# ═════════════════════════════════════════════════════════════════════════
# MAIN: Build the workbook
# ═════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    print("Creating Class Roster...")
    create_class_roster(wb)

    print("Creating Grade Book...")
    create_grade_book(wb)

    print("Creating Lesson Planner...")
    create_lesson_planner(wb)

    print("Creating Attendance Tracker...")
    create_attendance_tracker(wb)

    # Save
    output_path = "/home/user/claude-code/2026_Teacher_Planner_Grade_Book.xlsx"
    wb.save(output_path)
    print(f"\nDone! Saved to: {output_path}")


if __name__ == "__main__":
    main()
